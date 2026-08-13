from __future__ import annotations

import json
import tempfile
from pathlib import Path
import sys

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
BACKEND = ROOT / 'backend'
sys.path.insert(0, str(BACKEND))

from modules.plantillas_oficiales import generar_desde_plantilla_oficial, iter_plantillas_oficiales_para_generacion


def assert_true(value, message):
    if not value:
        raise AssertionError(message)


def run():
    app_text = (BACKEND / 'app.py').read_text(encoding='utf-8')
    js_text = (ROOT / 'frontend' / 'js' / 'app.js').read_text(encoding='utf-8')
    html_text = (ROOT / 'frontend' / 'index.html').read_text(encoding='utf-8')

    assert_true('def _alpha69_generar_ram_directo' in app_text, 'Falta generación directa RAM')
    assert_true("'formatos_seleccionados': 'ram'" in app_text, 'RAM directo no está aislado')
    assert_true('def _alpha69_buscar_ram_periodo' in app_text, 'Falta búsqueda por periodo')
    assert_true("if formato_norm == 'ram':" in app_text, 'El endpoint no separa RAM')
    assert_true('periodo-formatos-mes' in html_text and 'periodo-formatos-anio' in html_text, 'Falta selector de periodo')
    assert_true("formData.append('mes'" in js_text and "formData.append('anio'" in js_text, 'El proceso no envía periodo')
    assert_true('queryPeriodo' in js_text, 'La descarga no envía periodo')

    templates = BACKEND / 'seed_data' / 'templates_originales'
    july = [x for x in iter_plantillas_oficiales_para_generacion(templates, mes=7, anio=2026) if x.get('tipo') == 'ram']
    august = [x for x in iter_plantillas_oficiales_para_generacion(templates, mes=8, anio=2026) if x.get('tipo') == 'ram']
    assert_true(len(july) == 1 and str(july[0].get('version')) == '2', 'Julio debe seleccionar RAM V2 histórico')
    assert_true(len(august) == 1 and str(august[0].get('version')) == '3', 'Agosto debe seleccionar RAM V3')

    user = {
        'tipo_documento': 'RC', 'numero_documento': '000001',
        'primer_nombre': 'PRUEBA', 'primer_apellido': 'CONTROL',
        'fecha_nacimiento': '2025-01-01', 'fecha_ingreso': '2026-01-01',
    }
    with tempfile.TemporaryDirectory() as temp:
        out2 = Path(temp) / 'ram_v2.xlsx'
        out3 = Path(temp) / 'ram_v3.xlsx'
        generar_desde_plantilla_oficial('ram', {'metadata': {'anio': 2026, 'mes_numero': 7}, 'usuarios': [user]}, out2, templates)
        generar_desde_plantilla_oficial('ram', {'metadata': {'anio': 2026, 'mes_numero': 8}, 'usuarios': [user]}, out3, templates)
        assert_true(out2.exists() and out3.exists(), 'No se generaron ambas versiones RAM')
        wb2 = load_workbook(out2, read_only=True)
        wb3 = load_workbook(out3, read_only=True)
        try:
            assert_true('FORMATO RAM V2 HISTORICO' in wb2.sheetnames, 'RAM V2 inválido')
            assert_true('FORMATO RAM' in wb3.sheetnames, 'RAM V3 inválido')
        finally:
            wb2.close()
            wb3.close()

    print(json.dumps({'ok': True, 'checks': [
        'Periodo visible y enviado',
        'Descarga estricta por UDS y periodo',
        'RAM V2 histórico hasta julio de 2026',
        'RAM V3 desde agosto de 2026',
        'Generación sintética de ambas versiones',
    ]}, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    run()
