from __future__ import annotations

from io import BytesIO
from pathlib import Path
import tempfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from services.listado_asistencia_usuarios_service import generate_list, replace_template


class Upload:
    filename = "planilla_oficial.xlsx"

    def __init__(self, content: bytes):
        self.content = content

    def save(self, path):
        Path(path).write_bytes(self.content)


def require(condition, message):
    if not condition:
        raise AssertionError(message)


def main():
    with tempfile.TemporaryDirectory() as temporary:
        data_dir = Path(temporary) / "data"
        source = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "ASISTENCIA OFICIAL"
        sheet.append(["PLANILLA OFICIAL DE ASISTENCIA"])
        sheet.append(["N°", "Nombre completo", "Documento", "Cargo", "UDS", "Telefono", "Firma", "Observaciones"])
        sheet.append(["", "", "", "", "", "", "", ""])
        for cell in sheet[3]:
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
            cell.font = Font(name="Arial", size=9)
        sheet.print_area = "A1:H20"
        workbook.save(source)

        info = replace_template(data_dir, Upload(source.getvalue()), tenant_id=7)
        fields = info["mapeo"]["campos"]
        require(fields["nombre"] == 2 and fields["documento"] == 3, "Mapeo incorrecto")
        require(fields["firma"] == 7, "No se detecto firma")

        users = [
            {"Nombre": "Ana Perez", "Documento": "1001", "Cargo": "Docente", "UDS": "UCA 1", "Telefono": "3001"},
            {"PrimerNombre": "Luis", "PrimerApellido": "Diaz", "NUI": "1002", "cargo": "Auxiliar"},
        ]
        output = Path(temporary) / "salida.xlsx"
        generate_list(data_dir, output, users, metadata={"unidad": "UCA 1"}, tenant_id=7)
        generated = load_workbook(output)
        result = generated["ASISTENCIA OFICIAL"]
        require(result["B3"].value == "Ana Perez" and result["C3"].value == "1001", "No diligencio primer usuario")
        require(result["B4"].value == "Luis Diaz" and result["C4"].value == "1002", "No creo segunda fila")
        require(result["E4"].value == "UCA 1", "No completo unidad desde metadatos")
        require(result["G3"].value in (None, "") and result["G4"].value in (None, ""), "La firma debe quedar vacia")
        require(result["B4"].fill.fgColor.rgb == result["B3"].fill.fgColor.rgb, "No conservo estilo de fila")
        require(result.page_setup.fitToWidth == 1, "No preparo impresion a una pagina de ancho")
        generated.close()
        print("OK: carga, mapeo, diligenciamiento por UDS e impresion")


if __name__ == "__main__":
    main()
