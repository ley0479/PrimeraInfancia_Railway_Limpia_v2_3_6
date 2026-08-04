from __future__ import annotations

import json
import tempfile
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
BACKEND = ROOT / 'backend'
sys.path.insert(0, str(BACKEND))

from services.ram_v3_service import generate_ram_v3, sha256_file

TEMPLATE = BACKEND / 'seed_data' / 'templates_originales' / 'oficiales' / 'plantilla_ram_oficial_v3.xlsx'
EXPECTED_HASH = '52ae6d192acd678d282d851bbc43d0ebd042efe4e338c3761a4f87f207b14fdf'


def assert_true(value, message):
    if not value:
        raise AssertionError(message)


def user(index: int):
    return {
        'tipo_documento': 'RC',
        'documento': f'{index:08d}',
        'primer_nombre': f'NOMBRE{index}',
        'primer_apellido': f'APELLIDO{index}',
        'fecha_nacimiento': '2025-08-15',
        'fecha_ingreso': '2026-08-01',
    }


def run():
    assert_true(sha256_file(TEMPLATE) == EXPECTED_HASH, 'Hash RAM V3 incorrecto')
    with tempfile.TemporaryDirectory() as temp:
        out = Path(temp) / 'ram_v3.xlsx'
        report = generate_ram_v3(
            TEMPLATE,
            out,
            [user(i) for i in range(1, 22)],
            2026,
            8,
            metadata={'unidad': 'BAJO PACURITA', 'mes_nombre': 'AGOSTO', 'anio': 2026},
            expected_sha256=EXPECTED_HASH,
        )
        assert_true(report['paginas_ram'] == 2, 'RAM V3 no paginó 21 participantes')
        wb = load_workbook(out, data_only=False)
        assert_true(wb.sheetnames == ['FORMATO RAM', 'FORMATO RAM 2', 'INSTRUCCIONES DILIGENCIAMIENTO'], 'Orden de hojas RAM V3 inválido')
        assert_true(wb['FORMATO RAM']['AG4'].value == 'Hoja 1 de 2', 'Página 1 incorrecta')
        assert_true(wb['FORMATO RAM 2']['AG4'].value == 'Hoja 2 de 2', 'Página 2 incorrecta')
        assert_true(wb['FORMATO RAM 2']['A15'].value == 21, 'Numeración de segunda página incorrecta')
        wb.close()
    assert_true(sha256_file(TEMPLATE) == EXPECTED_HASH, 'La plantilla RAM V3 fue alterada')
    print(json.dumps({'ok': True, 'checks': [
        'Hash maestro', 'Paginación 21 participantes', 'Instrucciones preservadas', 'Plantilla maestra intacta'
    ]}, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    run()
