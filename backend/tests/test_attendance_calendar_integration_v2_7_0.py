#!/usr/bin/env python3
"""Integración RAM oficial + Base Maestra/beneficiarios + Calendario."""
from __future__ import annotations

import sqlite3
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
BACKEND = ROOT / "backend"
sys.path.insert(0, str(BACKEND))

from generador_formatos import GeneradorFormatos  # noqa: E402
from modules.seguridad.tenant_context import tenant_context  # noqa: E402


def require(value, message: str) -> None:
    if not value:
        raise AssertionError(message)


def prepare(path: Path) -> None:
    conn = sqlite3.connect(path)
    conn.executescript(
        """
        CREATE TABLE fundaciones(id INTEGER PRIMARY KEY,nombre TEXT,nit TEXT,municipio TEXT,departamento TEXT,centro_zonal TEXT);
        INSERT INTO fundaciones VALUES(1,'FUNDACION PRUEBA','900000000-1','QUIBDO','CHOCO','QUIBDO');
        CREATE TABLE unidades(id INTEGER PRIMARY KEY,fundacion_id INTEGER,nombre TEXT,codigo TEXT,codigo_uds TEXT,direccion TEXT,telefono TEXT,municipio TEXT,modalidad TEXT);
        INSERT INTO unidades VALUES(1,1,'UCA PRUEBA','UCA-001','CUENTA-001','CALLE 1','3000000000','QUIBDO','PROPIA E INTERCULTURAL');
        CREATE TABLE th_personas(id INTEGER PRIMARY KEY,fundacion_id INTEGER,unidad TEXT,nombre TEXT,documento TEXT,telefono TEXT,cargo TEXT,estado TEXT,activo INTEGER);
        INSERT INTO th_personas VALUES(1,1,'UCA PRUEBA','DOCENTE PRUEBA','12345678','3111111111','DOCENTE','ACTIVO',1);
        CREATE TABLE beneficiarios(
          id INTEGER PRIMARY KEY,fundacion_id INTEGER,unidad TEXT,estado TEXT,documento TEXT,numero_documento TEXT,tipo_documento TEXT,
          nombres TEXT,apellidos TEXT,primer_nombre TEXT,segundo_nombre TEXT,primer_apellido TEXT,segundo_apellido TEXT,
          fecha_nacimiento TEXT,fecha_ingreso TEXT,fecha_retiro TEXT,motivo_retiro TEXT,nui TEXT
        );
        """
    )
    for index in range(1, 22):
        conn.execute(
            """INSERT INTO beneficiarios(
                id,fundacion_id,unidad,estado,documento,numero_documento,tipo_documento,nombres,apellidos,
                primer_nombre,primer_apellido,fecha_nacimiento,fecha_ingreso,nui
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                index, 1, "UCA PRUEBA", "ACTIVO", f"{index:08d}", f"{index:08d}", "RC",
                f"NOMBRE{index}", f"APELLIDO{index}", f"NOMBRE{index}", f"APELLIDO{index}",
                "2024-01-15", "2026-08-01", f"NUI{index}",
            ),
        )
    conn.commit()
    conn.close()


def run() -> None:
    with tempfile.TemporaryDirectory(prefix="pi-attendance-") as tmp:
        root = Path(tmp)
        db_path = root / "database.sqlite3"
        output = root / "output"
        output.mkdir()
        prepare(db_path)
        generator = GeneradorFormatos(
            str(db_path),
            str(BACKEND / "seed_data" / "templates_originales"),
            str(output),
        )
        with tenant_context(1, role="COORDINADOR", username="coord"):
            generated = Path(generator.generar_asistencia(8, 2026, "UCA PRUEBA"))
        require(generated.is_file(), "No se generó RAM oficial")
        workbook = load_workbook(generated, data_only=False)
        require(
            workbook.sheetnames == ["FORMATO RAM", "FORMATO RAM 2", "INSTRUCCIONES DILIGENCIAMIENTO"],
            f"Hojas inesperadas: {workbook.sheetnames}",
        )
        require(workbook["FORMATO RAM"]["A4"].value == "EAS o PDS: FUNDACION PRUEBA", "Encabezado EAS incorrecto")
        require(workbook["FORMATO RAM"]["F7"].value == "Código CUENTAME UDS: CUENTA-001", "Código UDS incorrecto")
        require(workbook["FORMATO RAM 2"]["A15"].value == 21, "Paginación de participante 21 incorrecta")
        daily = [workbook["FORMATO RAM"].cell(row, col).value for row in range(15, 35) for col in range(10, 35)]
        require(not any(value in {"A", "I"} for value in daily), "Se inventó asistencia diaria")
        workbook.close()

        conn = sqlite3.connect(db_path)
        synced = conn.execute(
            "SELECT COUNT(*) FROM calendario_entregables WHERE fundacion_id=1 AND modulo='RAM/RAN/Asistencia'"
        ).fetchone()[0]
        conn.close()
        require(synced == 1, "La generación RAM no sincronizó el calendario exactamente una vez")

    print("PASS test_attendance_calendar_integration_v2_7_0")


if __name__ == "__main__":
    run()
