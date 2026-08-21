from __future__ import annotations

import sys
import tempfile
import unittest
import csv
import json
import pandas as pd
from pathlib import Path

from openpyxl import Workbook

BACKEND = Path(__file__).resolve().parents[1]
if str(BACKEND) not in sys.path:
    sys.path.insert(0, str(BACKEND))

from services.data_import import UniversalMappingService
from services.data_import.normalizers import normalize_header, normalize_unit_code
from services.data_import.file_security import validate_tabular_source


HEADERS = [
    "Tipo de documento del beneficiario", "Documento del beneficiario",
    "Primer nombre del beneficiario", "Primer apellido del beneficiario", "Estado del beneficiario",
    "Nombre de la Regional de la Unidad de servicio", "Código del Municipio de la Unidad de servicio",
    "Nombre Municipio de la Unidad de servicio", "Nombre del Centro Zonal",
    "Código de la unidad de servicio", "Nombre de la unidad de servicio",
] + [f"Campo adicional {i}" for i in range(30)]


def build_fixture(path: Path, header_row: int = 1, shuffled: bool = False) -> None:
    workbook = Workbook()
    info = workbook.active
    info.title = "Instrucciones"
    info.append(["Archivo de intercambio; no contiene registros"])
    sheet = workbook.create_sheet("ICBFCUEBeneficiariosPIActivosRe")
    for _ in range(header_row - 1): sheet.append(["INSTITUTO COLOMBIANO DE BIENESTAR FAMILIAR"])
    headers = list(HEADERS)
    if shuffled: headers = headers[9:11] + headers[:9] + headers[11:]
    sheet.append(headers)
    for index in range(417):
        unit = index % 39
        values = {
            "Tipo de documento del beneficiario": "RC", "Documento del beneficiario": f"10{index:08d}",
            "Primer nombre del beneficiario": f"NOMBRE{index}", "Primer apellido del beneficiario": f"APELLIDO{index}",
            "Estado del beneficiario": "ACTIVO", "Nombre de la Regional de la Unidad de servicio": "Chocó",
            "Código del Municipio de la Unidad de servicio": "27001", "Nombre Municipio de la Unidad de servicio": "QUIBDÓ",
            "Nombre del Centro Zonal": "CENTRO ZONAL 1", "Código de la unidad de servicio": (f"0{unit:011d}" if unit % 2 else f"1{unit:012d}"),
            "Nombre de la unidad de servicio": f"UDS {unit + 1:02d}",
        }
        sheet.append([values.get(header, f"V{index}") for header in headers])
    workbook.save(path)


class UniversalMapperRegressionTests(unittest.TestCase):
    def analyze(self, header_row=1, shuffled=False):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "base_desconocida.xlsx"
            build_fixture(path, header_row, shuffled)
            return UniversalMappingService().analyze(str(path))

    def test_regression_regional_is_not_unit(self):
        result = self.analyze()
        self.assertEqual(result["selected_table"], "ICBFCUEBeneficiariosPIActivosRe")
        self.assertEqual(result["preview"]["header_row"], 1)
        self.assertEqual(result["mapping"]["unidad.codigo"]["selected"]["original_header"], "Código de la unidad de servicio")
        self.assertEqual(result["mapping"]["unidad.nombre"]["selected"]["original_header"], "Nombre de la unidad de servicio")
        self.assertEqual(result["mapping"]["regional.nombre"]["selected"]["original_header"], "Nombre de la Regional de la Unidad de servicio")
        self.assertEqual(result["mapping"]["municipio.codigo"]["selected"]["original_header"], "Código del Municipio de la Unidad de servicio")
        self.assertEqual(result["mapping"]["municipio.nombre"]["selected"]["original_header"], "Nombre Municipio de la Unidad de servicio")
        self.assertEqual(result["mapping"]["centro_zonal.nombre"]["selected"]["original_header"], "Nombre del Centro Zonal")
        self.assertEqual(result["units"]["count"], 39)
        self.assertEqual(result["units"]["missing_code"], 0)
        self.assertEqual(result["units"]["missing_name"], 0)
        self.assertNotIn("Chocó", {item["name"] for item in result["units"]["items"]})

    def test_reordered_columns_and_header_at_row_20(self):
        result = self.analyze(header_row=20, shuffled=True)
        self.assertEqual(result["preview"]["header_row"], 20)
        self.assertEqual(result["units"]["count"], 39)

    def test_normalization_and_codes(self):
        self.assertEqual(normalize_header(" Nombre\u00a0 de\n U.D.S. "), "nombre de u d s")
        self.assertEqual(normalize_unit_code("001234567890"), "001234567890")
        self.assertEqual(normalize_unit_code("1234567890123.0"), "1234567890123")
        self.assertEqual(normalize_unit_code("1.234E+12"), "1234000000000")

    def test_every_staged_field_keeps_provenance(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "source.xlsx"; build_fixture(path)
            service = UniversalMappingService(); result = service.analyze(str(path))
            rows = list(service.staging_rows(str(path), result, chunk_size=37))
            self.assertEqual(len(rows), 417)
            unit = rows[0]["provenance"]["unidad.nombre"]
            self.assertEqual(unit["original_header"], "Nombre de la unidad de servicio")
            self.assertEqual(unit["source_table"], "ICBFCUEBeneficiariosPIActivosRe")
            self.assertIn("score", unit)

    def test_manual_confirmation_overrides_ambiguous_alias(self):
        result = self.analyze()
        unit_name = next(c for c in result["preview"]["columns"] if c["original_header"] == "Nombre de la unidad de servicio")
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "source.xlsx"; build_fixture(path)
            confirmed = UniversalMappingService().analyze(str(path), confirmed={"unidad.nombre": unit_name["id"]})
            decision = confirmed["mapping"]["unidad.nombre"]
            self.assertEqual(decision["selected"]["column_id"], unit_name["id"])
            self.assertIn("administrator_confirmed", decision["selected"]["reasons"])

    def test_real_signature_is_checked(self):
        with tempfile.TemporaryDirectory() as folder:
            fake = Path(folder) / "fake.xlsx"; fake.write_text("not an excel", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "firma real"):
                validate_tabular_source(str(fake), ".xlsx")
            valid = Path(folder) / "valid.xlsx"; build_fixture(valid)
            self.assertTrue(validate_tabular_source(str(valid), ".xlsx")["signature_valid"])

    def test_single_unit_is_valid_even_when_territory_is_constant(self):
        with tempfile.TemporaryDirectory() as folder:
            path=Path(folder)/"single.xlsx"; build_fixture(path)
            workbook=__import__("openpyxl").load_workbook(path); sheet=workbook["ICBFCUEBeneficiariosPIActivosRe"]
            for row in range(2,419): sheet.cell(row,10).value="000000001234"; sheet.cell(row,11).value="UDS ÚNICA"
            workbook.save(path); workbook.close()
            result=UniversalMappingService().analyze(str(path))
            self.assertEqual(result["units"]["count"],1)
            self.assertEqual(result["units"]["items"][0]["code"],"000000001234")
            self.assertEqual(result["units"]["items"][0]["name"],"UDS ÚNICA")

    def test_missing_unit_code_is_not_invented(self):
        with tempfile.TemporaryDirectory() as folder:
            path=Path(folder)/"without-code.xlsx"; build_fixture(path)
            workbook=__import__("openpyxl").load_workbook(path); sheet=workbook["ICBFCUEBeneficiariosPIActivosRe"]
            sheet.cell(1,10).value="Referencia auxiliar"
            for row in range(2,419): sheet.cell(row,10).value=None
            workbook.save(path); workbook.close()
            result=UniversalMappingService().analyze(str(path))
            self.assertIsNone(result["mapping"]["unidad.codigo"]["selected"])
            self.assertEqual(result["units"]["count"],39)
            self.assertEqual(result["units"]["missing_code"],417)

    def test_duplicate_headers_keep_distinct_internal_ids(self):
        with tempfile.TemporaryDirectory() as folder:
            path=Path(folder)/"duplicate.xlsx"; build_fixture(path)
            workbook=__import__("openpyxl").load_workbook(path); sheet=workbook["ICBFCUEBeneficiariosPIActivosRe"]
            sheet.cell(1,12).value="Nombre de la unidad de servicio"
            workbook.save(path); workbook.close()
            result=UniversalMappingService().analyze(str(path))
            duplicates=[c for c in result["preview"]["columns"] if c["normalized_header"]=="nombre de la unidad de servicio"]
            self.assertEqual(len(duplicates),2); self.assertNotEqual(duplicates[0]["id"],duplicates[1]["id"])

    def test_csv_and_ndjson_use_same_canonical_mapper(self):
        headers=HEADERS[:11]
        record=["RC","1001","ANA","PEREZ","ACTIVO","Chocó","27001","QUIBDÓ","CZ 1","001234567890","UDS CSV"]
        with tempfile.TemporaryDirectory() as folder:
            csv_path=Path(folder)/"source.csv"
            with csv_path.open("w",newline="",encoding="utf-8") as stream: csv.writer(stream).writerows([headers,record])
            csv_result=UniversalMappingService().analyze(str(csv_path)); self.assertEqual(csv_result["units"]["items"][0]["name"],"UDS CSV")
            ndjson=Path(folder)/"source.ndjson"
            ndjson.write_text(json.dumps(dict(zip(headers,record)),ensure_ascii=False)+"\n",encoding="utf-8")
            json_result=UniversalMappingService().analyze(str(ndjson)); self.assertEqual(json_result["mapping"]["unidad.nombre"]["selected"]["original_header"],"Nombre de la unidad de servicio")

    def test_xlsm_and_ods_are_read_structurally(self):
        headers=HEADERS[:11]; record=["RC","1001","ANA","PEREZ","ACTIVO","Chocó","27001","QUIBDÓ","CZ 1","000012345678","UDS NATIVA"]
        with tempfile.TemporaryDirectory() as folder:
            xlsm=Path(folder)/"source.xlsm"; workbook=Workbook(); sheet=workbook.active; sheet.title="DATOS"; sheet.append(headers); sheet.append(record); workbook.save(xlsm)
            self.assertTrue(validate_tabular_source(str(xlsm),".xlsm")["signature_valid"])
            self.assertEqual(UniversalMappingService().analyze(str(xlsm))["units"]["items"][0]["code"],"000012345678")
            ods=Path(folder)/"source.ods"; pd.DataFrame([record],columns=headers).to_excel(ods,index=False,engine="odf")
            self.assertTrue(validate_tabular_source(str(ods),".ods")["signature_valid"])
            self.assertEqual(UniversalMappingService().analyze(str(ods))["units"]["items"][0]["name"],"UDS NATIVA")


if __name__ == "__main__":
    unittest.main()
