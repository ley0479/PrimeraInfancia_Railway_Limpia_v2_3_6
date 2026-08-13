"""Comparación estructural y visual de un RAM contra su plantilla oficial."""
from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from services.ram_v3_service import ATTENDANCE_COLUMNS, DATA_ROWS, RAM_SHEET


HEADER_VALUES = {"A4", "AG4", "A5", "D5", "F5", "H5", "U5", "A6", "D6", "F6", "I6", "A7", "F7", "K7", "A8", "F8", "T8"}
TOTAL_VALUES = {"AI35", "AJ35", "E36", "E37", "E38"}
AUTHORIZED_STYLE_CHANGES = {"H14", "I14", "E36", "E37", "E38"}


def _visual(cell) -> tuple[Any, ...]:
    return (
        copy(cell.font), copy(cell.fill), copy(cell.border), copy(cell.alignment),
        cell.number_format, copy(cell.protection), cell.quotePrefix,
    )


def _dimensions(ws) -> tuple[dict, dict]:
    columns = {key: (item.width, item.hidden, item.bestFit, item.outlineLevel, item.collapsed) for key, item in ws.column_dimensions.items()}
    rows = {key: (item.height, item.hidden, item.outlineLevel, item.collapsed) for key, item in ws.row_dimensions.items()}
    return columns, rows


def _header_footer(item) -> tuple:
    return tuple((part.text, part.font, part.size, part.color) for part in (item.left, item.center, item.right))


def _page(ws) -> tuple[Any, ...]:
    return (
        str(ws.print_area), str(ws.print_title_rows), str(ws.print_title_cols),
        copy(ws.page_margins), copy(ws.page_setup), copy(ws.print_options),
        _header_footer(ws.oddHeader), _header_footer(ws.evenHeader), _header_footer(ws.firstHeader),
        _header_footer(ws.oddFooter), _header_footer(ws.evenFooter), _header_footer(ws.firstFooter),
        ws.freeze_panes, copy(ws.sheet_properties), copy(ws.sheet_format),
    )


def _dynamic_value_coordinates() -> set[str]:
    coords = set(HEADER_VALUES) | set(TOTAL_VALUES)
    for row in DATA_ROWS:
        coords.update({f"{column}{row}" for column in ("B", "C", "D", "E", "F", "G", "H", "I", "AI", "AJ", "AK")})
        coords.update({f"{chr(64 + col)}{row}" if col <= 26 else f"A{chr(64 + col - 26)}{row}" for col in ATTENDANCE_COLUMNS})
    return coords


def compare_ram_visual_integrity(template_path: str | Path, generated_path: str | Path) -> dict[str, Any]:
    """Retorna PASS/FAIL; no altera ninguno de los dos libros."""
    source_wb = load_workbook(template_path, data_only=False, keep_links=True)
    output_wb = load_workbook(generated_path, data_only=False, keep_links=True)
    failures: list[str] = []
    try:
        if RAM_SHEET not in output_wb.sheetnames:
            return {"status": "FAIL", "failures": [f"Falta la hoja {RAM_SHEET}"]}
        source = source_wb[RAM_SHEET]
        generated = output_wb[RAM_SHEET]
        if {str(x) for x in source.merged_cells.ranges} != {str(x) for x in generated.merged_cells.ranges}:
            failures.append("combinaciones de celdas")
        if _dimensions(source) != _dimensions(generated):
            failures.append("anchos de columnas o alturas de filas")
        if _page(source) != _page(generated):
            failures.append("configuración de página, impresión, encabezado o pie")

        dynamic_values = _dynamic_value_coordinates()
        max_row = max(source.max_row, generated.max_row)
        max_col = max(source.max_column, generated.max_column)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                original = source.cell(row, col)
                current = generated.cell(row, col)
                coord = original.coordinate
                if coord not in dynamic_values and original.value != current.value:
                    failures.append(f"valor/fórmula estática {coord}")
                if coord not in AUTHORIZED_STYLE_CHANGES and _visual(original) != _visual(current):
                    failures.append(f"estilo {coord}")
                if len(failures) >= 100:
                    break
            if len(failures) >= 100:
                break
        return {
            "status": "PASS" if not failures else "FAIL",
            "failures": failures,
            "checked_static_cells": max_row * max_col - len(dynamic_values),
            "merges_preserved": "combinaciones de celdas" not in failures,
            "dimensions_preserved": "anchos de columnas o alturas de filas" not in failures,
            "print_preserved": "configuración de página, impresión, encabezado o pie" not in failures,
        }
    finally:
        source_wb.close()
        output_wb.close()
