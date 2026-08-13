"""Generador RAM histórico V2 sobre plantilla sanitizada.

Este servicio existe únicamente para periodos anteriores a la vigencia de RAM V3.
Escribe valores sin recrear estilos, bordes, combinaciones ni configuración de
impresión. No contiene catálogos ni datos personales embebidos.
"""
from __future__ import annotations

import calendar
import copy
import hashlib
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook


DATA_START_ROW = 15
DATA_END_ROW = 34
CAPACITY = DATA_END_ROW - DATA_START_ROW + 1
DATA_COLUMNS = list(range(1, 38))  # A:AK
ATTENDANCE_COLUMNS = tuple(range(10, 35))  # J:AH
MONTHS_ES = {1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"}


def sha256_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _value(data: dict[str, Any], *keys: str, default: Any = "") -> Any:
    for key in keys:
        value = data.get(key)
        if value not in (None, ""):
            return value
    return default


def _split_name(user: dict[str, Any]) -> tuple[str, str, str, str]:
    first = str(_value(user, "primer_nombre", "PrimerNombre", default="")).strip()
    second = str(_value(user, "segundo_nombre", "SegundoNombre", default="")).strip()
    surname1 = str(_value(user, "primer_apellido", "PrimerApellido", default="")).strip()
    surname2 = str(_value(user, "segundo_apellido", "SegundoApellido", default="")).strip()
    if any((first, second, surname1, surname2)):
        return first, second, surname1, surname2

    names = str(_value(user, "nombres", "Nombres", "nombre", "Nombre", default="")).split()
    surnames = str(_value(user, "apellidos", "Apellidos", default="")).split()
    if not surnames:
        full = str(_value(user, "nombre_completo", "NombreCompleto", default="")).split()
        if len(full) >= 4:
            names, surnames = full[:-2], full[-2:]
        elif len(full) >= 2:
            names, surnames = full[:-1], full[-1:]
    return (
        names[0] if names else "",
        " ".join(names[1:]) if len(names) > 1 else "",
        surnames[0] if surnames else "",
        " ".join(surnames[1:]) if len(surnames) > 1 else "",
    )


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw[:10], fmt).date()
        except ValueError:
            continue
    return None


def _age_on_first_day(user: dict[str, Any], year: int, month: int) -> tuple[Any, Any]:
    years = _value(user, "edad_anos", "EdadAnos", "edad_años", default="")
    months = _value(user, "edad_meses_resto", "EdadMesesResto", "meses", default="")
    if years not in (None, "") or months not in (None, ""):
        return years, months
    birth = _parse_date(_value(user, "fecha_nacimiento", "FechaNacimiento", default=""))
    if not birth:
        return "", ""
    ref = date(year, month, 1)
    total = (ref.year - birth.year) * 12 + ref.month - birth.month
    if ref.day < birth.day:
        total -= 1
    total = max(0, total)
    return total // 12, total % 12


def _clear_rows(ws) -> None:
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        for col in DATA_COLUMNS:
            ws.cell(row=row, column=col).value = None
        ws.cell(row=row, column=1).value = row - DATA_START_ROW + 1
    for coord in ("AI35", "AJ35", "E36", "E37", "E38"):
        ws[coord].value = None


def _set_headers(ws, metadata: dict[str, Any], year: int, month: int, page: int, pages: int) -> None:
    month_name = MONTHS_ES[month]
    values = {
        "A4": f"ENTIDAD ADMINISTRADORA DEL SERVICIO: {_value(metadata, 'entidad', 'fundacion', 'fundacion_nombre', default='')}",
        "A5": f"NIT: {_value(metadata, 'nit', 'nit_eas', default='')}",
        "D5": f"NÚMERO DE CONTRATO: {_value(metadata, 'contrato', 'numero_contrato', default='')}",
        "F5": f"Regional: {_value(metadata, 'regional', default='')}",
        "H5": f"Centro Zonal: {_value(metadata, 'centro_zonal', default='')}",
        "U5": f"Municipio: {_value(metadata, 'municipio', default='')}",
        "A6": f"MES: {month_name}",
        "D6": f"AÑO: {year}",
        "F6": f"Nombre Agente Educativo(a): {_value(metadata, 'agente_educativo', 'responsable_grupo', default='')}",
        "I6": f"CC: {_value(metadata, 'documento_agente', 'cc_agente', default='')}",
        "A7": f"MODALIDAD DE ATENCIÓN: {_value(metadata, 'modalidad', default='')}",
        "F7": f"Código CUÉNTAME UDS: {_value(metadata, 'codigo_cuentame', 'nui_uds', 'codigo_uds', 'codigo_unidad', default='')}",
        "K7": f"Nombre Unidad de Servicio / Unidad de Atención: {_value(metadata, 'unidad', 'nombre_uds', default='')}",
        "A8": f"SERVICIO DE ATENCIÓN: {_value(metadata, 'servicio_atencion', 'servicio', default='')}",
        "F8": f"Dirección UDS: {_value(metadata, 'direccion_uds', 'direccion', default='')}",
        "T8": f"Teléfono UDS: {_value(metadata, 'telefono_uds', 'telefono', default='')}",
        "AG4": f"Hoja {page} de {pages}",
    }
    for cell, value in values.items():
        ws[cell] = value


def _write_user(
    ws, row: int, index: int, user: dict[str, Any], year: int, month: int,
    attendance_provider: Callable[[dict[str, Any]], set[str]] | None = None,
    non_service_dates: set[str] | None = None,
) -> int:
    from services.ram_v3_service import normalize_document_type, participant_document_number, normalize_withdrawal_code
    first, second, surname1, surname2 = _split_name(user)
    age_years, age_months = _age_on_first_day(user, year, month)
    ws.cell(row=row, column=1).value = index
    ws.cell(row=row, column=2).value = normalize_document_type(_value(user, "tipo_documento", "TipoDocumento", "tipo_doc"))
    document, _ = participant_document_number(user)
    document_cell = ws.cell(row=row, column=3)
    document_cell.value = str(document or "")
    if document:
        document_cell.data_type = "s"
    ws.cell(row=row, column=4).value = first
    ws.cell(row=row, column=5).value = second
    ws.cell(row=row, column=6).value = surname1
    ws.cell(row=row, column=7).value = surname2
    ws.cell(row=row, column=8).value = age_years
    ws.cell(row=row, column=9).value = age_months

    allowed = {str(v).strip().lower() for v in (attendance_provider(user) or set())} if attendance_provider else set()
    excluded = non_service_dates or set()
    total = 0
    model = ws.cell(row=row, column=ATTENDANCE_COLUMNS[0])
    for day in range(1, calendar.monthrange(year, month)[1] + 1):
        current = date(year, month, day)
        if current.weekday() > 4:
            continue
        occurrence = ((day - 1) // 7)
        col = 10 + occurrence * 5 + current.weekday()
        if col not in ATTENDANCE_COLUMNS:
            continue
        cell = ws.cell(row=row, column=col)
        weekday = ("lunes", "martes", "miercoles", "jueves", "viernes")[current.weekday()]
        # A = asistencia. La marca histórica H no corresponde a la convención
        # solicitada para RAN/RAM y hacía ambiguo el control semanal.
        mark = "A" if weekday in allowed and current.isoformat() not in excluded else ""
        cell.value = mark
        if mark:
            cell.font = copy.copy(model.font)
            cell.alignment = copy.copy(model.alignment)
            total += 1
    ws.cell(row=row, column=35).value = total
    ws.cell(row=row, column=37).value = normalize_withdrawal_code(user)
    return total


def _restore_authorized_layout(ws) -> None:
    """Restaura únicamente alineaciones señaladas, sin cambiar dimensiones."""
    for coord in ("H14", "I14"):
        alignment = copy.copy(ws[coord].alignment)
        alignment.horizontal = "center"
        alignment.vertical = "center"
        ws[coord].alignment = alignment
    for coord in [*(ws.cell(14, col).coordinate for col in ATTENDANCE_COLUMNS), "AI12", "AJ12", "AK9"]:
        alignment = copy.copy(ws[coord].alignment)
        alignment.textRotation = 90
        alignment.horizontal = "center"
        alignment.vertical = "center"
        ws[coord].alignment = alignment


def _write_totals(ws, users: list[dict[str, Any]], attended: dict[int, int], year: int, month: int) -> None:
    from services.ram_v3_service import ram_age_group
    under_six = over_six = gestants = 0
    for index, user in enumerate(users):
        if attended.get(index, 0) <= 0:
            continue
        group = ram_age_group(user, year, month)
        if group == 0:
            under_six += 1
        elif group == 4:
            gestants += 1
        elif group in {1, 2, 3}:
            over_six += 1
    for coord, value in (("E36", under_six), ("E37", over_six), ("E38", gestants)):
        cell = ws[coord]
        cell.value = value
        font = copy.copy(cell.font)
        font.color = "000000"
        cell.font = font
        alignment = copy.copy(cell.alignment)
        alignment.horizontal = "center"
        alignment.vertical = "center"
        cell.alignment = alignment


def _copy_sheet_with_images(wb, source, title: str):
    target = wb.copy_worksheet(source)
    target.title = title
    for image in getattr(source, "_images", []):
        try:
            cloned = copy.copy(image)
            cloned.anchor = copy.copy(image.anchor)
            target.add_image(cloned)
        except Exception:
            continue
    return target


def generate_ram_historical(
    template_path: str | Path,
    output_path: str | Path,
    users: list[dict[str, Any]],
    year: int,
    month: int,
    *,
    metadata: dict[str, Any] | None = None,
    expected_sha256: str | None = None,
    attendance_provider: Callable[[dict[str, Any]], set[str]] | None = None,
    non_service_dates: Iterable[str] | None = None,
) -> dict[str, Any]:
    template_path = Path(template_path)
    output_path = Path(output_path)
    if expected_sha256 and sha256_file(template_path).lower() != str(expected_sha256).lower():
        raise ValueError("La plantilla RAM histórica no coincide con el hash administrado.")
    if not 1 <= int(month) <= 12:
        raise ValueError("Mes RAM histórico inválido.")

    wb = load_workbook(template_path, data_only=False)
    source = wb[wb.sheetnames[0]]
    from services.ram_v3_service import deduplicate_users, sort_users_for_ram
    users, warnings = deduplicate_users(users or [])
    users = sort_users_for_ram(users, int(year), int(month))
    pages = max(1, (len(users) + CAPACITY - 1) // CAPACITY)
    worksheets = [source]
    source.title = "FORMATO RAM V2 HISTORICO"
    for page in range(2, pages + 1):
        worksheets.append(_copy_sheet_with_images(wb, source, f"FORMATO RAM V2 HISTORICO {page}"))

    for page, ws in enumerate(worksheets, start=1):
        _clear_rows(ws)
        _set_headers(ws, dict(metadata or {}), int(year), int(month), page, pages)
        _restore_authorized_layout(ws)
        block = users[(page - 1) * CAPACITY: page * CAPACITY]
        attended: dict[int, int] = {}
        for offset, user in enumerate(block):
            attended[offset] = _write_user(
                ws, DATA_START_ROW + offset, (page - 1) * CAPACITY + offset + 1,
                user, int(year), int(month), attendance_provider,
                {str(value).strip() for value in (non_service_dates or []) if str(value).strip()},
            )
        _write_totals(ws, block, attended, int(year), int(month))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {
        "archivo": str(output_path),
        "version": "2",
        "total_participantes": len(users),
        "paginas_ram": pages,
        "warnings": warnings,
    }
