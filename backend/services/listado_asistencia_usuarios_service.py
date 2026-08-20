"""Plantilla Excel de asistencia de usuarios, aislada por fundacion.

La plantilla es aportada por cada corporacion. El servicio detecta la fila de
encabezados, conserva el libro original y escribe solamente los campos que pudo
mapear. Las columnas de firma y asistencia permanecen vacias para diligenciar
e imprimir el formato en campo.
"""
from __future__ import annotations

from copy import copy
from datetime import datetime
import json
from pathlib import Path
import re
import shutil
import unicodedata
import zipfile

from openpyxl import load_workbook

from modules.seguridad.tenant_context import current_tenant_id, tenant_storage_root


TEMPLATE_STEM = "plantilla_listado_asistencia_usuarios_oficial"
TEMPLATE_NAME = f"{TEMPLATE_STEM}.xlsx"
MAPPING_NAME = "mapeo_listado_asistencia_usuarios.json"
ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}

FIELD_ALIASES = {
    "numero": ("numero", "nro", "no", "n", "item"),
    "nombre": ("nombre", "nombres y apellidos", "nombre completo", "beneficiario", "usuario", "participante"),
    "documento": ("documento", "numero de documento", "identificacion", "cedula", "nui"),
    "tipo_documento": ("tipo documento", "tipo de documento", "td"),
    "cargo": ("cargo", "rol", "perfil"),
    "unidad": ("unidad", "uds", "uca", "unidad de servicio"),
    "telefono": ("telefono", "celular", "contacto"),
    "firma": ("firma", "firma del asistente"),
    "asistencia": ("asistencia", "asistio", "presente"),
    "observaciones": ("observacion", "observaciones", "novedad"),
}


def _tenant_id(value=None) -> int:
    try:
        parsed = int(value or current_tenant_id() or 1)
    except (TypeError, ValueError):
        parsed = 1
    return max(1, parsed)


def template_dir(data_dir, tenant_id=None) -> Path:
    path = tenant_storage_root(data_dir, _tenant_id(tenant_id)) / "official_templates" / "listado_asistencia_usuarios"
    path.mkdir(parents=True, exist_ok=True)
    (path / "backups").mkdir(exist_ok=True)
    return path


def _normal(value) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").strip().lower())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(re.sub(r"[^a-z0-9]+", " ", text).split())


def _field_for_header(value) -> str | None:
    text = _normal(value)
    if not text:
        return None
    for field, aliases in FIELD_ALIASES.items():
        if text in aliases or any(len(alias) >= 4 and alias in text for alias in aliases):
            return field
    return None


def detect_mapping(path: Path) -> dict:
    keep_vba = path.suffix.lower() == ".xlsm"
    workbook = load_workbook(path, read_only=False, data_only=False, keep_vba=keep_vba)
    best = None
    for sheet in workbook.worksheets:
        for row in range(1, min(sheet.max_row, 60) + 1):
            fields = {}
            for col in range(1, min(sheet.max_column, 80) + 1):
                field = _field_for_header(sheet.cell(row, col).value)
                if field and field not in fields:
                    fields[field] = col
            score = len(fields) + (3 if "nombre" in fields else 0) + (2 if "documento" in fields else 0)
            candidate = {"hoja": sheet.title, "fila_encabezado": row, "fila_datos": row + 1, "campos": fields, "puntaje": score}
            if best is None or score > best[0]:
                best = (score, candidate)
    workbook.close()
    if not best or best[0] < 4 or not ({"nombre", "documento"} & set(best[1]["campos"])):
        raise ValueError("No se pudo mapear la planilla. Debe tener encabezados como Nombre o Documento y al menos otra columna reconocible.")
    return best[1]


def _read_mapping(base: Path) -> dict | None:
    path = base / MAPPING_NAME
    try:
        return json.loads(path.read_text(encoding="utf-8")) if path.exists() else None
    except Exception:
        return None


def template_info(data_dir, tenant_id=None) -> dict:
    tid = _tenant_id(tenant_id)
    base = template_dir(data_dir, tid)
    mapping = _read_mapping(base)
    filename = str((mapping or {}).get("archivo") or TEMPLATE_NAME)
    path = base / filename
    return {
        "nombre": "Listado oficial de asistencia de usuarios",
        "archivo": TEMPLATE_NAME,
        "hoja": (mapping or {}).get("hoja", "Mapeo pendiente"),
        "version": "corporacion",
        "tipo_formato": "listado_asistencia_usuarios",
        "codigo_tipo": "listado_asistencia_usuarios",
        "preservar_estilos": True,
        "preservar_impresion": True,
        "existe": path.exists(),
        "tamano_bytes": path.stat().st_size if path.exists() else 0,
        "fecha_actualizacion": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds") if path.exists() else None,
        "ruta": str(path),
        "fundacion_id": tid,
        "mapeo": mapping,
    }


def replace_template(data_dir, uploaded_file, tenant_id=None) -> dict:
    filename = str(getattr(uploaded_file, "filename", "") or "")
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError("La planilla oficial de asistencia debe ser Excel .xlsx o .xlsm.")
    base = template_dir(data_dir, tenant_id)
    temporary = base / f".upload_{datetime.now().strftime('%Y%m%d%H%M%S%f')}{ext}"
    uploaded_file.save(temporary)
    try:
        if not zipfile.is_zipfile(temporary):
            raise ValueError("El archivo no es una planilla Excel valida.")
        mapping = detect_mapping(temporary)
        destination = base / f"{TEMPLATE_STEM}{ext}"
        if destination.exists():
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            shutil.copy2(destination, base / "backups" / f"{stamp}_{destination.name}")
            old_mapping = base / MAPPING_NAME
            if old_mapping.exists():
                shutil.copy2(old_mapping, base / "backups" / f"{stamp}_{MAPPING_NAME}")
        for obsolete in base.glob(f"{TEMPLATE_STEM}.*"):
            if obsolete != destination and obsolete.suffix.lower() in ALLOWED_EXTENSIONS:
                obsolete.unlink(missing_ok=True)
        shutil.copy2(temporary, destination)
        mapping["archivo"] = destination.name
        (base / MAPPING_NAME).write_text(json.dumps(mapping, ensure_ascii=False, indent=2), encoding="utf-8")
    finally:
        temporary.unlink(missing_ok=True)
    return template_info(data_dir, tenant_id)


def restore_template(data_dir, tenant_id=None) -> dict:
    base = template_dir(data_dir, tenant_id)
    candidates = sorted(
        (path for path in (base / "backups").glob(f"*_{TEMPLATE_STEM}.*") if path.suffix.lower() in ALLOWED_EXTENSIONS),
        key=lambda p: p.stat().st_mtime, reverse=True,
    )
    if not candidates:
        raise FileNotFoundError("No hay una planilla de asistencia anterior para restaurar.")
    selected = candidates[0]
    restored_name = f"{TEMPLATE_STEM}{selected.suffix.lower()}"
    shutil.copy2(selected, base / restored_name)
    prefix = selected.name[: -len(f'{TEMPLATE_STEM}{selected.suffix}')]
    mapping_backup = base / "backups" / f"{prefix}{MAPPING_NAME}"
    mapping = detect_mapping(base / restored_name)
    if mapping_backup.exists():
        shutil.copy2(mapping_backup, base / MAPPING_NAME)
    else:
        mapping["archivo"] = restored_name
        (base / MAPPING_NAME).write_text(json.dumps(mapping, ensure_ascii=False, indent=2), encoding="utf-8")
    return template_info(data_dir, tenant_id)


def _value(data, *keys):
    for key in keys:
        value = (data or {}).get(key)
        if value not in (None, ""):
            return str(value).strip()
    return ""


def _name(user) -> str:
    direct = _value(user, "Nombre", "nombre", "nombres", "NombreCompleto", "nombre_completo")
    if direct:
        return direct
    return " ".join(filter(None, (_value(user, "PrimerNombre", "primer_nombre"), _value(user, "SegundoNombre", "segundo_nombre"), _value(user, "PrimerApellido", "primer_apellido"), _value(user, "SegundoApellido", "segundo_apellido"))))


def _copy_row_style(sheet, source_row: int, target_row: int) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for col in range(1, sheet.max_column + 1):
        source = sheet.cell(source_row, col)
        target = sheet.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def generate_list(data_dir, output_path, users, metadata=None, tenant_id=None) -> str:
    info = template_info(data_dir, tenant_id)
    if not info["existe"]:
        raise FileNotFoundError("La corporacion aun no ha cargado su planilla oficial de asistencia de usuarios.")
    mapping = info.get("mapeo") or detect_mapping(Path(info["ruta"]))
    keep_vba = Path(info["ruta"]).suffix.lower() == ".xlsm"
    workbook = load_workbook(info["ruta"], keep_vba=keep_vba)
    sheet = workbook[mapping["hoja"]]
    start = int(mapping["fila_datos"])
    records = list(users or [])
    required = max(1, len(records))
    available = max(1, sheet.max_row - start + 1)
    if required > available:
        sheet.insert_rows(start + available, amount=required - available)
        for row in range(start + available, start + required):
            _copy_row_style(sheet, start, row)
    fields = mapping["campos"]
    meta = dict(metadata or {})
    for index in range(required):
        row = start + index
        user = records[index] if index < len(records) else {}
        values = {
            "numero": index + 1 if user else "",
            "nombre": _name(user),
            "documento": _value(user, "Documento", "documento", "NUI", "nui", "identificacion"),
            "tipo_documento": _value(user, "TipoDocumento", "tipo_documento"),
            "cargo": _value(user, "Cargo", "cargo", "Rol", "rol", "perfil"),
            "unidad": _value(user, "UDS", "uds", "UCA", "uca", "Unidad", "unidad") or _value(meta, "unidad", "uds", "uca"),
            "telefono": _value(user, "Telefono", "telefono", "Celular", "celular"),
            "firma": "",
            "asistencia": "",
            "observaciones": "",
        }
        for field, col in fields.items():
            sheet.cell(row, int(col)).value = values.get(field, "")
    if not sheet.print_area:
        last_col = sheet.max_column
        sheet.print_area = f"A1:{sheet.cell(start + required - 1, last_col).coordinate}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    workbook.close()
    return str(output)
