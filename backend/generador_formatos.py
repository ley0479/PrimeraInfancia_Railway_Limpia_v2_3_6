"""
Módulo para generación de formatos ICBF
"""
import os
import json
import sqlite3
import re
import unicodedata
from datetime import datetime, timedelta
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

from modules.print_master import aplicar_configuracion_impresion_libro
from modules.plantillas_oficiales import get_plantilla_oficial, generar_desde_plantilla_oficial
from modules.seguridad.tenant_context import current_tenant_id
from models import ConfiguracionSistema, EstadoUsuario


class GeneradorFormatos:
    """Genera formatos ICBF desde datos de la base de datos"""
    
    def __init__(self, db_path, templates_path, output_path):
        self.db_path = db_path
        self.templates_path = templates_path
        self.output_path = output_path
    
    def get_db_connection(self):
        """Obtiene conexión a BD"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _aplicar_impresion_y_guardar(self, ruta, tipo_formato):
        """Aplica la tabla maestra de impresión a un Excel ya generado."""
        try:
            wb = load_workbook(ruta)
            aplicar_configuracion_impresion_libro(wb, tipo_formato, source_name=os.path.basename(ruta))
            wb.save(ruta)
        except Exception as exc:
            print(f"No se pudo aplicar configuración de impresión {tipo_formato} a {ruta}: {exc}")


    def _sincronizar_calendario_entrega(self, tipo_formato, titulo, mes, año, unidad, ruta=None):
        """Marca en el Calendario Inteligente que un formato fue generado/entregado.

        No modifica el archivo oficial. Solo actualiza o crea el entregable operativo
        correspondiente al periodo del formato.
        """
        try:
            from modules.calendario_inteligente.repository import CalendarioInteligenteRepository
            repo = CalendarioInteligenteRepository(self.db_path, self.output_path)
            repo.init_schema()
            repo.sincronizar_entrega({
                'titulo': titulo,
                'fecha_entrega': f"{int(año):04d}-{int(mes):02d}-01",
                'modulo': tipo_formato,
                'tipo_formato': tipo_formato,
                'unidad': unidad or '',
                'archivo_evidencia': os.path.basename(ruta) if ruta else None,
                'observaciones': 'Sincronizado automáticamente al generar formato desde la plataforma.'
            })
        except Exception as exc:
            print(f"No se pudo sincronizar Calendario Inteligente ({tipo_formato}): {exc}")

    def _plantilla_oficial_disponible(self, tipo_formato):
        try:
            info = get_plantilla_oficial(self.templates_path, tipo_formato)
            return bool(info and info.get('existe'))
        except Exception:
            return False

    def _usuario_oficial(self, b):
        return {
            'NUI': b.get('nui') or b.get('documento') or '',
            'Documento': b.get('documento') or b.get('nui') or '',
            'TipoDocumento': b.get('tipo_documento') or b.get('tipo_doc') or 'RC',
            'Nombre': b.get('nombres') or '',
            'PrimerNombre': b.get('primer_nombre') or '',
            'SegundoNombre': b.get('segundo_nombre') or '',
            'PrimerApellido': b.get('primer_apellido') or '',
            'SegundoApellido': b.get('segundo_apellido') or '',
            'Acudiente': b.get('nombre_acudiente') or '',
            'DocumentoAcudiente': b.get('documento_acudiente') or '',
            'Parentesco': b.get('parentesco') or '',
            'Telefono': b.get('telefono') or b.get('celular') or '',
            'EdadMeses': b.get('edad_meses') or 0,
            'GrupoEdad': b.get('grupo_edad') or '',
            'TipoBeneficiario': b.get('tipo_beneficiario') or '',
        }

    def _metadata_oficial(self, mes, año, unidad, coordinador=None):
        """Construye encabezados oficiales desde la base de datos.

        Alpha16 corrige el problema de encabezados desactualizados: el formato
        no puede heredar UNIDAD DEMO 04, LUIS u otro dato de la plantilla. Los valores
        se toman por UDS exacta desde beneficiarios, unidades y talento humano.
        """
        meses = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']

        def norm(valor):
            texto = str(valor or '').strip().lower()
            texto = unicodedata.normalize('NFKD', texto)
            texto = ''.join(ch for ch in texto if not unicodedata.combining(ch))
            texto = texto.replace('ñ', 'n')
            texto = re.sub(r'[^a-z0-9]+', ' ', texto)
            return ' '.join(texto.split())

        def limpiar(valor):
            return str(valor or '').strip()

        def row_value(row, *keys):
            if not row:
                return ''
            for key in keys:
                try:
                    if key in row.keys() and row[key] not in (None, ''):
                        return row[key]
                except Exception:
                    try:
                        if key in row and row.get(key) not in (None, ''):
                            return row.get(key)
                    except Exception:
                        pass
            return ''

        def full_name(row):
            if not row:
                return ''
            direct = row_value(row, 'nombre', 'nombres_y_apellidos', 'Nombre', 'NOMBRE')
            if direct:
                return limpiar(direct).upper()
            return ' '.join(limpiar(row_value(row, k)) for k in ('nombres', 'apellidos') if limpiar(row_value(row, k))).upper()

        def unidades_de_row(row):
            valores = []
            for key in ('unidad', 'nombre', 'Nombre UDS', 'nombre_uds', 'unidad_servicio', 'comunidad'):
                val = row_value(row, key)
                if val:
                    valores.append(val)
            raw_unidades = row_value(row, 'unidades')
            if raw_unidades:
                try:
                    parsed = json.loads(raw_unidades)
                    if isinstance(parsed, list):
                        valores.extend(parsed)
                    elif isinstance(parsed, dict):
                        valores.extend(parsed.values())
                    else:
                        valores.append(parsed)
                except Exception:
                    valores.extend(re.split(r'[;,|/]+', str(raw_unidades)))
            return [limpiar(v).upper() for v in valores if limpiar(v)]

        unidad_norm = norm(unidad)

        def coincide_unidad(row):
            return any(norm(u) == unidad_norm for u in unidades_de_row(row))

        beneficiario_ref = {}
        unidad_db = {}
        talentos = []
        try:
            conn = self.get_db_connection()
            cursor = conn.cursor()

            try:
                filas_b = cursor.execute("SELECT * FROM beneficiarios WHERE unidad = ? LIMIT 1", (unidad,)).fetchall()
                if not filas_b:
                    filas_b = cursor.execute("SELECT * FROM beneficiarios LIMIT 2000").fetchall()
                for fila in filas_b:
                    data = dict(fila)
                    if not beneficiario_ref and (not unidad_norm or norm(data.get('unidad')) == unidad_norm):
                        beneficiario_ref = data
                        break
            except Exception:
                beneficiario_ref = {}

            try:
                for fila in cursor.execute("SELECT * FROM unidades").fetchall():
                    data = dict(fila)
                    if coincide_unidad(data):
                        unidad_db = data
                        break
            except Exception:
                unidad_db = {}

            for tabla in ('coordinadores', 'th_personas'):
                try:
                    filas = cursor.execute(f"SELECT * FROM {tabla}").fetchall()
                except Exception:
                    continue
                for fila in filas:
                    data = dict(fila)
                    estado = norm(row_value(data, 'estado'))
                    if estado in {'inactivo', 'retirado'}:
                        continue
                    activo = row_value(data, 'activo')
                    if str(activo).strip() in {'0', 'False', 'false'}:
                        continue
                    if coincide_unidad(data):
                        talentos.append(data)
            conn.close()
        except Exception:
            pass

        def score_talento(row):
            cargo = norm(row_value(row, 'cargo', 'tipo_equipo', 'perfil'))
            if 'agente' in cargo or 'docente' in cargo or 'educativo' in cargo:
                return 0
            if 'coordin' in cargo:
                return 1
            if 'suplente' in cargo or 'apoyo' in cargo or 'auxiliar' in cargo:
                return 2
            return 3

        talentos.sort(key=lambda item: (score_talento(item), norm(full_name(item))))
        responsable_row = talentos[0] if talentos else {}
        suplente_row = next((t for t in talentos if any(k in norm(row_value(t, 'cargo', 'tipo_equipo', 'perfil')) for k in ['suplente', 'apoyo', 'auxiliar'])), {})
        coordinador_row = next((t for t in talentos if 'coordin' in norm(row_value(t, 'cargo', 'tipo_equipo', 'perfil'))), {})

        responsable = full_name(responsable_row)
        if not responsable and coordinador:
            try:
                responsable = f"{coordinador['nombres']} {coordinador['apellidos']}".strip().upper()
            except Exception:
                responsable = ''

        regional = row_value(beneficiario_ref, 'regional', 'Regional') or 'CHOCÓ'
        centro_zonal = row_value(beneficiario_ref, 'centro_zonal', 'CentroZonal', 'Centro Zonal') or 'CZ Ciudad de prueba'
        municipio = row_value(beneficiario_ref, 'municipio', 'Municipio') or 'Ciudad de prueba'
        modalidad = row_value(beneficiario_ref, 'modalidad', 'Modalidad') or row_value(unidad_db, 'modalidad') or 'EDUCACIÓN INICIAL PROPIA DIARIA - PROPIA E INTERCULTURAL'
        servicio = row_value(beneficiario_ref, 'servicio_atencion', 'ServicioAtencion', 'servicio') or modalidad
        eas = row_value(beneficiario_ref, 'nombre_eas', 'NombreEAS', 'entidad_administradora') or 'FUNDACIÓN PACÍFICO VIVE'
        codigo_uds = row_value(beneficiario_ref, 'codigo_unidad_servicio', 'codigo_uds', 'CodigoUnidadServicio', 'codigo_unidad') or row_value(unidad_db, 'codigo_unidad_servicio', 'codigo_uds', 'codigo')
        direccion = row_value(unidad_db, 'direccion', 'Direccion') or row_value(beneficiario_ref, 'direccion_unidad', 'DireccionUnidad') or row_value(responsable_row, 'direccion')
        telefono_unidad = row_value(unidad_db, 'telefono', 'Telefono') or row_value(responsable_row, 'telefono', 'celular')
        contrato = row_value(beneficiario_ref, 'numero_contrato', 'NumeroContrato', 'contrato') or row_value(unidad_db, 'contrato') or row_value(responsable_row, 'contrato')
        barrio = row_value(beneficiario_ref, 'barrio', 'Barrio') or row_value(unidad_db, 'barrio')
        unidad_origen = row_value(beneficiario_ref, 'nombre_unidad_origen', 'NombreUnidadOrigen', 'nombre_punto_entrega_origen') or unidad
        codigo_origen = row_value(beneficiario_ref, 'codigo_unidad_origen', 'CodigoUnidadOrigen') or codigo_uds

        mes_nombre = meses[int(mes) - 1] if 1 <= int(mes) <= 12 else str(mes).upper()
        metadata = {
            'regional': limpiar(regional).upper(),
            'Regional': limpiar(regional).upper(),
            'centro_zonal': limpiar(centro_zonal).upper(),
            'CentroZonal': limpiar(centro_zonal).upper(),
            'municipio': limpiar(municipio).upper(),
            'Municipio': limpiar(municipio).upper(),
            'modalidad': limpiar(modalidad).upper(),
            'Modalidad': limpiar(modalidad).upper(),
            'servicio_atencion': limpiar(servicio).upper(),
            'ServicioAtencion': limpiar(servicio).upper(),
            'eas': limpiar(eas).upper(),
            'NombreEAS': limpiar(eas).upper(),
            'unidad': limpiar(unidad).upper(),
            'Unidad': limpiar(unidad).upper(),
            'unidad_origen': limpiar(unidad_origen).upper(),
            'NombreUnidadOrigen': limpiar(unidad_origen).upper(),
            'codigo_unidad': limpiar(codigo_uds),
            'codigo_uds': limpiar(codigo_uds),
            'CodigoUnidadServicio': limpiar(codigo_uds),
            'codigo_origen': limpiar(codigo_origen),
            'CodigoUnidadOrigen': limpiar(codigo_origen),
            'direccion': limpiar(direccion).upper(),
            'direccion_unidad': limpiar(direccion).upper(),
            'DireccionUnidad': limpiar(direccion).upper(),
            'barrio': limpiar(barrio).upper(),
            'Barrio': limpiar(barrio).upper(),
            'telefono': limpiar(telefono_unidad),
            'Telefono': limpiar(telefono_unidad),
            'telefono_docente': limpiar(row_value(responsable_row, 'telefono', 'celular')) or limpiar(telefono_unidad),
            'responsable': responsable,
            'docente': responsable,
            'agente_educativo': responsable,
            'cedula_docente': limpiar(row_value(responsable_row, 'documento', 'cedula', 'identificacion')),
            'suplente': full_name(suplente_row),
            'telefono_suplente': limpiar(row_value(suplente_row, 'telefono', 'celular')),
            'coordinador': full_name(coordinador_row) or limpiar(row_value(responsable_row, 'coordinador')),
            'contrato': limpiar(contrato),
            'NumeroContrato': limpiar(contrato),
            'mes': mes_nombre,
            'Mes': mes_nombre,
            'anio': año,
            'año': año,
            'year': año,
            'fecha_entrega': datetime(año, mes, 1).strftime('%d/%m/%Y'),
            'FechaEntrega': datetime(año, mes, 1).strftime('%d/%m/%Y'),
        }
        return metadata
    
    # ==================== ASISTENCIA ====================
    def generar_asistencia(self, mes, año, unidad):
        """Genera formato de asistencia para un mes"""
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        # Obtener beneficiarios de la unidad
        cursor.execute("""
            SELECT b.id, b.documento, b.nui, b.nombres, b.apellidos,
                   b.primer_nombre, b.segundo_nombre,
                   b.primer_apellido, b.segundo_apellido, b.fecha_carga
            FROM beneficiarios b
            WHERE b.unidad = ? AND b.estado = ?
            ORDER BY COALESCE(NULLIF(b.primer_nombre, ''), b.nombres),
                     COALESCE(NULLIF(b.primer_apellido, ''), b.apellidos)
        """, (unidad, EstadoUsuario.ACTIVO))
        
        beneficiarios = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        # Crear DataFrame
        df = pd.DataFrame()
        df['NUI'] = [b.get('nui', '') for b in beneficiarios]
        df['DOCUMENTO'] = [b['documento'] for b in beneficiarios]
        df['NOMBRES'] = [b['nombres'] for b in beneficiarios]
        df['APELLIDOS'] = [b['apellidos'] for b in beneficiarios]
        df['PRIMER NOMBRE'] = [b.get('primer_nombre', '') for b in beneficiarios]
        df['SEGUNDO NOMBRE'] = [b.get('segundo_nombre', '') for b in beneficiarios]
        df['PRIMER APELLIDO'] = [b.get('primer_apellido', '') for b in beneficiarios]
        df['SEGUNDO APELLIDO'] = [b.get('segundo_apellido', '') for b in beneficiarios]
        
        # Agregar columnas para días del mes
        import calendar
        _, dias_mes = calendar.monthrange(año, mes)
        for dia in range(1, dias_mes + 1):
            df[f'DIA_{dia}'] = ''  # Vacío para llenar manualmente
        
        df['TOTAL_ASISTENCIAS'] = ''
        df['OBSERVACIONES'] = ''
        
        # Guardar
        fecha_str = f"{año}{mes:02d}"
        nombre_archivo = f"ASISTENCIA_{unidad}_{fecha_str}.xlsx"
        ruta = os.path.join(self.output_path, nombre_archivo)
        
        df.to_excel(ruta, sheet_name='ASISTENCIA', index=False)
        self._aplicar_impresion_y_guardar(ruta, 'ram_ran')
        
        return ruta
    
    # ==================== BIENESTARINA ====================
    def generar_bienestarina(self, mes, año, unidad, bolsas_por_beneficiario=1):
        """Genera formato de entrega de Bienestarina"""
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        # Obtener beneficiarios activos de la unidad
        cursor.execute("""
            SELECT b.id, b.documento, b.nui, b.nombres, b.apellidos,
                   b.primer_nombre, b.segundo_nombre,
                   b.primer_apellido, b.segundo_apellido,
                   b.nombre_acudiente, b.documento_acudiente, b.parentesco,
                   b.fecha_nacimiento, b.fecha_carga
            FROM beneficiarios b
            WHERE b.unidad = ? AND b.estado = ?
            ORDER BY COALESCE(NULLIF(b.primer_nombre, ''), b.nombres),
                     COALESCE(NULLIF(b.primer_apellido, ''), b.apellidos)
        """, (unidad, EstadoUsuario.ACTIVO))
        
        beneficiarios = [dict(row) for row in cursor.fetchall()]
        
        # Obtener coordinador de la unidad
        cursor.execute("""
            SELECT c.nombres, c.apellidos
            FROM coordinadores c
            WHERE c.unidad = ? OR c.unidades LIKE ?
            LIMIT 1
        """, (unidad, f'%"{unidad}"%'))
        
        coordinador = cursor.fetchone()
        conn.close()

        if self._plantilla_oficial_disponible('bienestarina'):
            nombre_archivo = f"BIENESTARINA_{unidad}_{año}{mes:02d}.xlsx"
            ruta = os.path.join(self.output_path, nombre_archivo)
            datos = {
                'metadata': self._metadata_oficial(mes, año, unidad, coordinador),
                'usuarios': [self._usuario_oficial(b) for b in beneficiarios],
            }
            return generar_desde_plantilla_oficial('bienestarina', datos, ruta, self.templates_path)
        
        # Crear DataFrame
        df = pd.DataFrame()
        df['BENEFICIARIO'] = [b['nombres'] for b in beneficiarios]
        df['NUI'] = [b.get('nui', '') for b in beneficiarios]
        df['DOCUMENTO'] = [b['documento'] for b in beneficiarios]
        df['RESPONSABLE'] = ''
        df['PARENTESCO'] = ''
        df['PRIMER NOMBRE'] = [b.get('primer_nombre', '') for b in beneficiarios]
        df['SEGUNDO NOMBRE'] = [b.get('segundo_nombre', '') for b in beneficiarios]
        df['PRIMER APELLIDO'] = [b.get('primer_apellido', '') for b in beneficiarios]
        df['SEGUNDO APELLIDO'] = [b.get('segundo_apellido', '') for b in beneficiarios]
        df['ACUDIENTE'] = [b.get('nombre_acudiente', '') for b in beneficiarios]
        df['DOC ACUDIENTE'] = [b.get('documento_acudiente', '') for b in beneficiarios]
        df['PARENTESCO'] = [b.get('parentesco', '') for b in beneficiarios]
        df['MES ENTREGA'] = datetime(año, mes, 1).strftime('%B').upper()
        df['BOLSAS'] = [bolsas_por_beneficiario] * len(beneficiarios)
        df['FIRMA'] = ''
        df['HUELLA'] = ''
        df['OBSERVACIONES'] = ''
        
        # Metadatos
        fecha_str = datetime(año, mes, 1).strftime('%B %Y')
        
        # Guardar
        nombre_archivo = f"BIENESTARINA_{unidad}_{año}{mes:02d}.xlsx"
        ruta = os.path.join(self.output_path, nombre_archivo)
        
        # Crear workbook con estilos
        with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='ENTREGA', index=False)
            self._escribir_y_resaltar(df, writer, 'ENTREGA', beneficiarios)
            
            workbook = writer.book
            worksheet = writer.sheets['ENTREGA']
            
            # Agregar encabezado
            worksheet.insert_rows(1, 5)
            worksheet['A1'] = 'FORMATO DE ENTREGA - BIENESTARINA'
            worksheet['A1'].font = Font(bold=True, size=12)
            worksheet['A2'] = f'Período: {fecha_str}'
            worksheet['A3'] = f'Unidad: {unidad}'
            if coordinador:
                worksheet['A4'] = f"Coordinador: {coordinador['nombres']} {coordinador['apellidos']}"

            aplicar_configuracion_impresion_libro(workbook, 'bienestarina', source_name=nombre_archivo)
            
        return ruta

    def _escribir_y_resaltar(self, df, writer, sheet_name, raw_data):
        """Escribe el DataFrame sin alterar colores oficiales.

        Las versiones anteriores pintaban cambios recientes en verde. Eso era útil
        para auditoría, pero en formatos oficiales ICBF cambia la presentación y
        puede borrar/alterar colores exigidos. La trazabilidad se conserva en los
        datos y reportes, no en rellenos del Excel oficial.
        """
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    def _guardar_con_resaltado(self, df, ruta, sheet_name, raw_data):
        """Versión para to_excel directo (Asistencia)"""
        with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
            self._escribir_y_resaltar(df, writer, sheet_name, raw_data)
    
    # ==================== RAN (Registro Asistencia Nutrición) ====================
    def generar_ran(self, mes, año, unidad):
        """Genera RAN (Registro Asistencia Nutrición)"""
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        # Obtener beneficiarios por grupo etario
        grupos_etarios = {
            'GESTANTES': ("b.tipo_beneficiario = 'GESTANTE'", 0, 40),
            '0-5 MESES': ("b.tipo_beneficiario = 'NINO'", 0, 5),
            '6-11 MESES': ("b.tipo_beneficiario = 'NINO'", 6, 11),
            '1-2 AÑOS': ("b.tipo_beneficiario = 'NINO'", 12, 23),
            '3-5 AÑOS': ("b.tipo_beneficiario = 'NINO'", 24, 59),
        }
        
        resultados = {}
        
        for grupo, (tipo_sql, edad_min, edad_max) in grupos_etarios.items():
            cursor.execute(f"""
                SELECT COUNT(*) as total
                FROM beneficiarios b
                WHERE b.unidad = ? AND b.estado = ? AND {tipo_sql}
            """, (unidad, EstadoUsuario.ACTIVO))
            
            total = cursor.fetchone()['total']
            resultados[grupo] = total
        
        # Crear DataFrame
        df = pd.DataFrame(list(resultados.items()), columns=['GRUPO_ETARIO', 'CANTIDAD'])
        df['FECHA'] = datetime(año, mes, 1).strftime('%d/%m/%Y')
        df['RESPONSABLE'] = ''
        
        conn.close()
        
        # Guardar
        nombre_archivo = f"RAN_{unidad}_{año}{mes:02d}.xlsx"
        ruta = os.path.join(self.output_path, nombre_archivo)
        df.to_excel(ruta, sheet_name='RAN', index=False)
        self._aplicar_impresion_y_guardar(ruta, 'ram_ran')
        
        return ruta
    
    # ==================== RPP (Registro Procedencia Procedimiento) ====================
    def generar_rpp(self, mes, año, unidad):
        """Genera RPP desde plantilla oficial si está disponible."""
        conn = self.get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT b.*
            FROM beneficiarios b
            WHERE b.unidad = ? AND b.estado = ?
            ORDER BY COALESCE(NULLIF(b.nombres, ''), b.documento), b.documento
        """, (unidad, EstadoUsuario.ACTIVO))
        beneficiarios = [dict(row) for row in cursor.fetchall()]

        cursor.execute("""
            SELECT c.nombres, c.apellidos
            FROM coordinadores c
            WHERE c.unidad = ? OR c.unidades LIKE ?
            LIMIT 1
        """, (unidad, f'%"{unidad}"%'))
        coordinador = cursor.fetchone()
        conn.close()

        if self._plantilla_oficial_disponible('rpp'):
            nombre_archivo = f"RPP_{unidad}_{año}{mes:02d}.xlsx"
            ruta = os.path.join(self.output_path, nombre_archivo)
            datos = {
                'metadata': self._metadata_oficial(mes, año, unidad, coordinador),
                'usuarios': [self._usuario_oficial(b) for b in beneficiarios],
            }
            return generar_desde_plantilla_oficial('rpp', datos, ruta, self.templates_path)

        total = len(beneficiarios)
        df = pd.DataFrame({
            'CONCEPTO': ['BENEFICIARIOS ASISTIDOS', 'RACIONES ENTREGADAS', 'TOTAL KILOGRAMOS'],
            'CANTIDAD': [total, total, total * 0.5],
            'FECHA': datetime(año, mes, 1).strftime('%d/%m/%Y')
        })

        nombre_archivo = f"RPP_{unidad}_{año}{mes:02d}.xlsx"
        ruta = os.path.join(self.output_path, nombre_archivo)
        df.to_excel(ruta, sheet_name='RPP', index=False)
        self._aplicar_impresion_y_guardar(ruta, 'rpp')
        return ruta

    # ==================== NUTRICIÓN ====================
    def generar_nutricion(self, mes, año, unidad):
        """Genera reporte de nutrición"""
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        # Obtener mediciones del período
        fid = int(current_tenant_id(1) or 1)
        cursor.execute("""
            SELECT b.nombres, b.documento, pt.peso, pt.talla, pt.estado_nutricional
            FROM peso_talla pt
            JOIN beneficiarios b
              ON pt.beneficiario_id = b.id
             AND COALESCE(b.fundacion_id, 1) = ?
            WHERE b.unidad = ?
              AND COALESCE(pt.fundacion_id, 1) = ?
              AND strftime('%Y', pt.fecha_medicion) = ?
              AND strftime('%m', pt.fecha_medicion) = ?
            ORDER BY b.nombres
        """, (fid, unidad, fid, str(año), f"{mes:02d}"))
        
        registros = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        df = pd.DataFrame(registros)
        
        if not df.empty:
            df.rename(columns={
                'nombres': 'BENEFICIARIO',
                'documento': 'DOCUMENTO',
                'peso': 'PESO (KG)',
                'talla': 'TALLA (CM)',
                'estado_nutricional': 'ESTADO'
            }, inplace=True)
        
        # Guardar
        nombre_archivo = f"NUTRICION_{unidad}_{año}{mes:02d}.xlsx"
        ruta = os.path.join(self.output_path, nombre_archivo)
        df.to_excel(ruta, sheet_name='NUTRICION', index=False)
        
        return ruta
    
    # ==================== INFORME PEDAGÓGICO ====================
    def generar_informe_pedagogico(self, docente_id, mes, año):
        """Genera informe pedagógico del docente"""
        conn = self.get_db_connection()
        cursor = conn.cursor()
        
        # Obtener informe
        cursor.execute("""
            SELECT * FROM informes_pedagogicos
            WHERE docente_id = ? AND mes = ? AND año = ?
        """, (docente_id, mes, año))
        
        informe = cursor.fetchone()
        
        if not informe:
            conn.close()
            return None
        
        # Obtener datos del docente
        cursor.execute("SELECT * FROM docentes WHERE id = ?", (docente_id,))
        docente = cursor.fetchone()
        
        # Obtener evidencias
        cursor.execute("""
            SELECT * FROM evidencias
            WHERE informe_id = ?
            ORDER BY fecha_carga
        """, (informe['id'],))
        
        evidencias = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        # Crear informe en formato
        df = pd.DataFrame({
            'CAMPO': [
                'Docente', 'Unidad', 'Período',
                'Tema del Mes', 'Objetivos', 'Actividades Realizadas',
                'Resultados', 'Participación Familiar', 'Logros',
                'Dificultades', 'Recomendaciones', 'Total Evidencias'
            ],
            'CONTENIDO': [
                f"{docente['nombres']} {docente['apellidos']}",
                docente['unidad'],
                f"{mes}/{año}",
                informe['tema_mes'] or '',
                informe['objetivos'] or '',
                informe['actividades'] or '',
                informe['resultados'] or '',
                informe['participacion_familiar'] or '',
                informe['logros'] or '',
                informe['dificultades'] or '',
                informe['recomendaciones'] or '',
                len(evidencias)
            ]
        })
        
        # Guardar
        nombre_archivo = f"INFORME_PEDAGOGICO_DOCENTE{docente_id}_{año}{mes:02d}.xlsx"
        ruta = os.path.join(self.output_path, nombre_archivo)
        df.to_excel(ruta, sheet_name='INFORME', index=False)
        
        return ruta
    
    # ==================== GENERACIÓN MASIVA ====================
    def generar_mes_completo(self, mes, año, unidad):
        """Genera todos los formatos para un mes y unidad"""
        archivos_generados = {
            'asistencia': None,
            'bienestarina': None,
            'ran': None,
            'rpp': None,
            'nutricion': None
        }
        
        try:
            archivos_generados['asistencia'] = self.generar_asistencia(mes, año, unidad)
            self._sincronizar_calendario_entrega('RAM/RAN/Asistencia', 'Generación RAM/RAN/Asistencia', mes, año, unidad, archivos_generados['asistencia'])

            archivos_generados['bienestarina'] = self.generar_bienestarina(mes, año, unidad)
            self._sincronizar_calendario_entrega('Bienestarina', 'Generación Bienestarina', mes, año, unidad, archivos_generados['bienestarina'])

            archivos_generados['ran'] = self.generar_ran(mes, año, unidad)
            self._sincronizar_calendario_entrega('RAM/RAN/Asistencia', 'Generación RAN', mes, año, unidad, archivos_generados['ran'])

            archivos_generados['rpp'] = self.generar_rpp(mes, año, unidad)
            self._sincronizar_calendario_entrega('RPP', 'Generación RPP', mes, año, unidad, archivos_generados['rpp'])

            archivos_generados['nutricion'] = self.generar_nutricion(mes, año, unidad)
            self._sincronizar_calendario_entrega('Nutrición', 'Generación reporte nutricional', mes, año, unidad, archivos_generados['nutricion'])
        except Exception as e:
            print(f"Error generando formatos: {e}")
        
        return archivos_generados
