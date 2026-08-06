"""Sistema Integral de Salud y Nutrición 2.6.0.

Amplía el módulo histórico sin duplicar Base Maestra ni modificar valoraciones
anteriores. Los expedientes guardan únicamente referencias y atributos propios
del componente; las mediciones originales permanecen inmutables y su
validación profesional se versiona en una tabla separada.
"""
from __future__ import annotations

import hashlib
import json
import mimetypes
import os
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from flask import Blueprint, g, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from werkzeug.utils import secure_filename

from modules.seguridad.services import require_roles
from modules.seguridad.tenant_context import tenant_storage_root

from .schema import INTEGRAL_SCHEMA_SQL
from .services import now_iso

READ_ROLES = {"SUPERADMIN", "GERENTE", "COORDINADOR", "AUXILIAR_ADMINISTRATIVO", "NUTRICIONISTA"}
EDIT_ROLES = {"SUPERADMIN", "GERENTE", "COORDINADOR", "NUTRICIONISTA"}
COORDINATION_ROLES = {"SUPERADMIN", "GERENTE", "COORDINADOR"}
DOCUMENT_TYPES = (
    "AFILIACION_SGSSS",
    "VACUNACION_PAI",
    "VALORACION_INTEGRAL_SALUD",
    "SALUD_BUCAL",
    "TAMIZAJE_NEONATAL",
    "CONTROL_PRENATAL",
    "DISCAPACIDAD",
    "LACTANCIA_ALIMENTACION",
)
HEALTH_LINES = {
    "L1_ARTICULACION_SALUD",
    "L2_EDUCACION_SALUD_ALIMENTARIA",
    "L3_PREVENCION_ENFERMEDADES",
    "L4_ALIMENTOS_SANOS_SEGUROS",
    "L5_SEGUIMIENTO_NUTRICIONAL",
}
CLOSED_STATES = {"CERRADA", "CERRADO", "APROBADA", "APROBADO", "COMPLETADA", "COMPLETADO"}


def _norm(value: Any) -> str:
    return str(value or "").strip().upper().replace(" ", "_")


def _truthy(value: Any) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "si", "sí", "s", "x", "ok", "completo", "vigente", "activo"}


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def _user() -> dict[str, Any]:
    current = getattr(g, "current_user", None) or {}
    return {
        **current,
        "id": current.get("id"),
        "username": current.get("username") or current.get("email") or "sistema",
        "rol": _norm(current.get("rol")),
        "fundacion_id": int(current.get("fundacion_id") or 1),
    }


def _allowed_units(user: dict[str, Any]) -> list[str] | None:
    if user.get("rol") in COORDINATION_ROLES:
        return None
    raw = user.get("unidades")
    if isinstance(raw, str):
        try:
            parsed = json.loads(raw)
            raw = parsed
        except Exception:
            raw = [part.strip() for part in raw.split(",") if part.strip()]
    if not isinstance(raw, list):
        return []
    return [str(value).strip() for value in raw if str(value).strip()]


def _unit_key(value: Any) -> str:
    return " ".join(str(value or "").strip().upper().split())


def _can_access_unit(unit: Any, user: dict[str, Any]) -> bool:
    allowed = _allowed_units(user)
    if allowed is None:
        return True
    key = _unit_key(unit)
    return bool(key) and key in {_unit_key(value) for value in allowed}


def _require_unit(unit: Any, user: dict[str, Any]) -> str:
    text = str(unit or "").strip()
    allowed = _allowed_units(user)
    if not text and allowed is not None and len(allowed) == 1:
        text = allowed[0]
    if not _can_access_unit(text, user):
        raise PermissionError("No tienes permiso sobre esta UCA.")
    return text


class SaludNutricionIntegralService:
    def __init__(self, repo: Any, data_dir: str | os.PathLike[str]):
        self.repo = repo
        self.data_dir = Path(data_dir)

    def init_schema(self) -> None:
        self.repo.execute_script(INTEGRAL_SCHEMA_SQL)

    def _columns(self, table: str) -> set[str]:
        return set(self.repo.columns(table)) if self.repo.table_exists(table) else set()

    @staticmethod
    def _pick(row: dict[str, Any], *names: str, default: Any = None) -> Any:
        for name in names:
            if row.get(name) not in (None, ""):
                return row.get(name)
        return default

    def _participants(self, fundacion_id: int, unit: str | None = None) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        if self.repo.table_exists("master_ninos"):
            cols = self._columns("master_ninos")
            where = ["COALESCE(activo,1)=1"]
            params: list[Any] = []
            if "fundacion_id" in cols:
                where.append("fundacion_id=?")
                params.append(fundacion_id)
            if unit:
                where.append("unidad_servicio=?")
                params.append(unit)
            rows.extend(self.repo.fetch_all(f"SELECT * FROM master_ninos WHERE {' AND '.join(where)} ORDER BY documento", params))
        elif self.repo.table_exists("beneficiarios"):
            cols = self._columns("beneficiarios")
            where = ["COALESCE(estado,'ACTIVO')<>'RETIRADO'"] if "estado" in cols else ["1=1"]
            params = []
            if "fundacion_id" in cols:
                where.append("fundacion_id=?")
                params.append(fundacion_id)
            if unit and "unidad" in cols:
                where.append("unidad=?")
                params.append(unit)
            rows.extend(self.repo.fetch_all(f"SELECT * FROM beneficiarios WHERE {' AND '.join(where)} ORDER BY documento", params))

        if self.repo.table_exists("gestantes"):
            cols = self._columns("gestantes")
            where = ["COALESCE(estado,'ACTIVO')='ACTIVO'"] if "estado" in cols else ["1=1"]
            params = []
            if "fundacion_id" in cols:
                where.append("fundacion_id=?")
                params.append(fundacion_id)
            if unit and "unidad" in cols:
                where.append("unidad=?")
                params.append(unit)
            for row in self.repo.fetch_all(f"SELECT * FROM gestantes WHERE {' AND '.join(where)} ORDER BY documento", params):
                item = dict(row)
                item["_tipo_participante"] = "GESTANTE"
                rows.append(item)
        return rows

    def _uca_expedient_id(self, fundacion_id: int, unit: str | None) -> int | None:
        if not unit or not self.repo.table_exists("giu_expedientes_uca"):
            return None
        row = self.repo.fetch_one(
            "SELECT id FROM giu_expedientes_uca WHERE fundacion_id=? AND (unidad_nombre=? OR codigo_unidad=?) AND COALESCE(estado,'ACTIVO')<>'ELIMINADO' ORDER BY id DESC LIMIT 1",
            (fundacion_id, unit, unit),
        )
        return int(row["id"]) if row else None

    def sync_expedientes(self, fundacion_id: int, user: dict[str, Any], unit: str | None = None) -> dict[str, Any]:
        now = now_iso()
        created = updated = docs_created = 0
        rows = self._participants(fundacion_id, unit)
        with self.repo.connect() as conn:
            for source in rows:
                documento = str(self._pick(source, "documento", "nui", default="") or "").strip()
                if not documento:
                    continue
                unit_name = str(self._pick(source, "unidad_servicio", "unidad", default="") or "").strip()
                if unit and _unit_key(unit_name) != _unit_key(unit):
                    continue
                participant_type = source.get("_tipo_participante") or self._pick(source, "tipo_beneficiario", default="NINO_NINA")
                beneficiary_id = self._pick(source, "id")
                existing = conn.execute(
                    "SELECT id FROM sn_expedientes_integrales WHERE fundacion_id=? AND documento=?",
                    (fundacion_id, documento),
                ).fetchone()
                uca_exp = self._uca_expedient_id(fundacion_id, unit_name)
                if existing:
                    exp_id = int(existing[0])
                    conn.execute(
                        "UPDATE sn_expedientes_integrales SET beneficiario_id=?,tipo_participante=?,unidad_nombre=?,expediente_uca_id=?,actualizado_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?",
                        (beneficiary_id, participant_type, unit_name, uca_exp, user.get("id"), now, fundacion_id, exp_id),
                    )
                    updated += 1
                else:
                    cur = conn.execute(
                        "INSERT INTO sn_expedientes_integrales(fundacion_id,beneficiario_id,documento,tipo_participante,unidad_nombre,expediente_uca_id,responsable_id,responsable_nombre,estado,creado_por,fecha_creacion,actualizado_por,fecha_actualizacion) VALUES(?,?,?,?,?,?,?,?, 'ACTIVO',?,?,?,?)",
                        (fundacion_id, beneficiary_id, documento, participant_type, unit_name, uca_exp, user.get("id"), user.get("username"), user.get("id"), now, user.get("id"), now),
                    )
                    exp_id = int(cur.lastrowid)
                    created += 1

                source_flags = {
                    "AFILIACION_SGSSS": self._pick(source, "carne_salud", "afiliacion_salud"),
                    "VACUNACION_PAI": self._pick(source, "vacunas", "carne_vacunas"),
                    "VALORACION_INTEGRAL_SALUD": self._pick(source, "control_crecimiento", "carne_crecimiento"),
                    "SALUD_BUCAL": self._pick(source, "salud_bucal"),
                    "TAMIZAJE_NEONATAL": self._pick(source, "tamizaje_neonatal"),
                    "CONTROL_PRENATAL": self._pick(source, "control_prenatal"),
                    "DISCAPACIDAD": self._pick(source, "certificado_discapacidad", "diagnostico_discapacidad"),
                    "LACTANCIA_ALIMENTACION": self._pick(source, "lactancia", "alimentacion_complementaria"),
                }
                for doc_type in DOCUMENT_TYPES:
                    is_gestante = _norm(participant_type) == "GESTANTE"
                    no_aplica = (doc_type == "CONTROL_PRENATAL" and not is_gestante) or (doc_type == "TAMIZAJE_NEONATAL" and is_gestante)
                    state = "NO_APLICA" if no_aplica else ("VIGENTE" if _truthy(source_flags.get(doc_type)) else "PENDIENTE")
                    before = conn.execute(
                        "SELECT id FROM sn_documentos_salud WHERE fundacion_id=? AND expediente_id=? AND tipo_documento=? AND activo=1",
                        (fundacion_id, exp_id, doc_type),
                    ).fetchone()
                    if before:
                        if state == "VIGENTE":
                            conn.execute(
                                "UPDATE sn_documentos_salud SET estado=?,actualizado_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=? AND estado IN ('PENDIENTE','NO_REGISTRADO')",
                                (state, user.get("id"), now, fundacion_id, int(before[0])),
                            )
                    else:
                        conn.execute(
                            "INSERT INTO sn_documentos_salud(fundacion_id,expediente_id,tipo_documento,estado,activo,creado_por,fecha_creacion,actualizado_por,fecha_actualizacion) VALUES(?,?,?,?,1,?,?,?,?)",
                            (fundacion_id, exp_id, doc_type, state, user.get("id"), now, user.get("id"), now),
                        )
                        docs_created += 1
            conn.commit()
        self.repo.log("SINCRONIZAR_EXPEDIENTES_SALUD", "sn_expedientes_integrales", None, usuario=user.get("username", "sistema"), nuevos={"creados": created, "actualizados": updated, "documentos": docs_created})
        return {"fuentes": len(rows), "creados": created, "actualizados": updated, "documentos_creados": docs_created}

    def list_expedientes(self, fundacion_id: int, user: dict[str, Any], filters: dict[str, Any] | None = None) -> list[dict[str, Any]]:
        filters = filters or {}
        where = ["e.fundacion_id=?", "COALESCE(e.estado,'ACTIVO')<>'ELIMINADO'"]
        params: list[Any] = [fundacion_id]
        unit = str(filters.get("unidad") or "").strip()
        if unit:
            where.append("e.unidad_nombre=?")
            params.append(unit)
        doc_state = str(filters.get("estado_documental") or "").strip()
        sql = f"""
            SELECT e.*,
              (SELECT COUNT(*) FROM sn_documentos_salud d WHERE d.fundacion_id=e.fundacion_id AND d.expediente_id=e.id AND d.activo=1) documentos_total,
              (SELECT COUNT(*) FROM sn_documentos_salud d WHERE d.fundacion_id=e.fundacion_id AND d.expediente_id=e.id AND d.activo=1 AND d.estado IN ('VIGENTE','VALIDADO','NO_APLICA')) documentos_al_dia,
              (SELECT COUNT(*) FROM sn_canalizaciones c WHERE c.fundacion_id=e.fundacion_id AND c.expediente_id=e.id AND c.estado NOT IN ('CERRADA','CERRADO')) canalizaciones_abiertas,
              (SELECT COUNT(*) FROM sn_alertas a WHERE a.fundacion_id=e.fundacion_id AND a.documento=e.documento AND COALESCE(a.atendida,0)=0) alertas_abiertas
            FROM sn_expedientes_integrales e
            WHERE {' AND '.join(where)}
            ORDER BY e.unidad_nombre,e.documento
            LIMIT 5000
        """
        rows = self.repo.fetch_all(sql, params)
        allowed = _allowed_units(user)
        if allowed is not None:
            keys = {_unit_key(value) for value in allowed}
            rows = [row for row in rows if _unit_key(row.get("unidad_nombre")) in keys]
        if doc_state == "PENDIENTE":
            rows = [row for row in rows if int(row.get("documentos_al_dia") or 0) < int(row.get("documentos_total") or 0)]
        return rows

    def expediente_detail(self, fundacion_id: int, expediente_id: int, user: dict[str, Any]) -> dict[str, Any]:
        row = self.repo.fetch_one("SELECT * FROM sn_expedientes_integrales WHERE fundacion_id=? AND id=?", (fundacion_id, expediente_id))
        if not row:
            raise LookupError("Expediente de salud no encontrado.")
        if not _can_access_unit(row.get("unidad_nombre"), user):
            raise PermissionError("No tienes permiso sobre esta UCA.")
        docs = self.repo.fetch_all("SELECT * FROM sn_documentos_salud WHERE fundacion_id=? AND expediente_id=? AND activo=1 ORDER BY tipo_documento", (fundacion_id, expediente_id))
        vals = self.repo.fetch_all(
            """
            SELECT v.*,vv.estado_validacion,vv.clasificacion_profesional,vv.observacion_profesional,vv.profesional_nombre,vv.fecha_validacion
            FROM sn_valoraciones v
            LEFT JOIN sn_valoracion_validaciones vv ON vv.valoracion_id=v.id AND vv.fundacion_id=v.fundacion_id AND vv.activo=1
            WHERE v.fundacion_id=? AND v.documento=? AND v.activo=1
            ORDER BY v.fecha_valoracion DESC,v.id DESC
            """,
            (fundacion_id, row["documento"]),
        )
        routes = self.repo.fetch_all("SELECT * FROM sn_canalizaciones WHERE fundacion_id=? AND expediente_id=? ORDER BY fecha_creacion DESC", (fundacion_id, expediente_id))
        return {"expediente": row, "documentos": docs, "valoraciones": vals, "canalizaciones": routes}

    def update_document(self, fundacion_id: int, document_id: int, data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        current = self.repo.fetch_one(
            "SELECT d.*,e.unidad_nombre FROM sn_documentos_salud d JOIN sn_expedientes_integrales e ON e.id=d.expediente_id AND e.fundacion_id=d.fundacion_id WHERE d.fundacion_id=? AND d.id=? AND d.activo=1",
            (fundacion_id, document_id),
        )
        if not current:
            raise LookupError("Documento de salud no encontrado.")
        if not _can_access_unit(current.get("unidad_nombre"), user):
            raise PermissionError("No tienes permiso sobre esta UCA.")
        allowed_states = {"PENDIENTE", "EN_TRAMITE", "VIGENTE", "VENCIDO", "VALIDADO", "NO_APLICA", "OBSERVADO"}
        state = _norm(data.get("estado") or current.get("estado"))
        if state not in allowed_states:
            raise ValueError("Estado documental no permitido.")
        now = now_iso()
        self.repo.execute_update(
            """
            UPDATE sn_documentos_salud SET estado=?,entidad_emisora=?,numero_referencia=?,fecha_documento=?,fecha_verificacion=?,fecha_vencimiento=?,soporte_modulo=?,soporte_id=?,soporte_ruta=?,observaciones=?,validado_por=?,fecha_validacion=?,actualizado_por=?,fecha_actualizacion=?
            WHERE fundacion_id=? AND id=? AND activo=1
            """,
            (
                state, data.get("entidad_emisora"), data.get("numero_referencia"), data.get("fecha_documento"),
                data.get("fecha_verificacion"), data.get("fecha_vencimiento"), data.get("soporte_modulo"), data.get("soporte_id"),
                data.get("soporte_ruta"), data.get("observaciones"), user.get("id") if state in {"VIGENTE", "VALIDADO", "NO_APLICA"} else None,
                now if state in {"VIGENTE", "VALIDADO", "NO_APLICA"} else None, user.get("id"), now, fundacion_id, document_id,
            ),
        )
        self.repo.log("ACTUALIZAR_DOCUMENTO_SALUD", "sn_documentos_salud", document_id, usuario=user.get("username", "sistema"), anteriores=current, nuevos=data)
        return self.repo.fetch_one("SELECT * FROM sn_documentos_salud WHERE fundacion_id=? AND id=?", (fundacion_id, document_id)) or {}

    def validate_valoracion(self, fundacion_id: int, valoracion_id: int, data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        valuation = self.repo.fetch_one("SELECT * FROM sn_valoraciones WHERE fundacion_id=? AND id=? AND activo=1", (fundacion_id, valoracion_id))
        if not valuation:
            raise LookupError("Valoración no encontrada.")
        if not _can_access_unit(valuation.get("unidad"), user):
            raise PermissionError("No tienes permiso sobre esta UCA.")
        state = _norm(data.get("estado_validacion") or "VALIDADA")
        if state not in {"PENDIENTE", "VALIDADA", "DEVUELTA", "ANULADA"}:
            raise ValueError("Estado de validación no permitido.")
        automatic = {
            "diagnostico_global": valuation.get("diagnostico_global"),
            "nivel_alerta": valuation.get("nivel_alerta"),
            "z_peso_edad": valuation.get("z_peso_edad"),
            "z_talla_edad": valuation.get("z_talla_edad"),
            "z_peso_talla": valuation.get("z_peso_talla"),
            "z_imc_edad": valuation.get("z_imc_edad"),
            "diag_braquial_edad": valuation.get("diag_braquial_edad"),
        }
        previous = self.repo.fetch_one("SELECT COALESCE(MAX(version),0) version FROM sn_valoracion_validaciones WHERE fundacion_id=? AND valoracion_id=?", (fundacion_id, valoracion_id)) or {"version": 0}
        version = int(previous.get("version") or 0) + 1
        now = now_iso()
        with self.repo.connect() as conn:
            conn.execute("UPDATE sn_valoracion_validaciones SET activo=0,actualizado_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND valoracion_id=? AND activo=1", (user.get("id"), now, fundacion_id, valoracion_id))
            cur = conn.execute(
                """
                INSERT INTO sn_valoracion_validaciones(fundacion_id,valoracion_id,clasificacion_automatica,reglas_version,estado_validacion,clasificacion_profesional,observacion_profesional,profesional_id,profesional_nombre,fecha_validacion,version,activo,creado_por,fecha_creacion,actualizado_por,fecha_actualizacion)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,1,?,?,?,?)
                """,
                (fundacion_id, valoracion_id, _json(automatic), str(data.get("reglas_version") or "RES-2465-2016/CONFIGURABLE"), state, data.get("clasificacion_profesional"), data.get("observacion_profesional"), user.get("id"), user.get("username"), now if state == "VALIDADA" else None, version, user.get("id"), now, user.get("id"), now),
            )
            validation_id = int(cur.lastrowid)
            conn.commit()
        self.repo.log("VALIDAR_VALORACION_SALUD", "sn_valoracion_validaciones", validation_id, documento=valuation.get("documento") or "", usuario=user.get("username", "sistema"), nuevos={"estado": state, "version": version})
        return self.repo.fetch_one("SELECT * FROM sn_valoracion_validaciones WHERE fundacion_id=? AND id=?", (fundacion_id, validation_id)) or {}

    def create_activity(self, fundacion_id: int, data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        line = _norm(data.get("linea_componente"))
        if line not in HEALTH_LINES:
            raise ValueError("Selecciona una de las cinco líneas del componente Salud y Nutrición.")
        unit = _require_unit(data.get("unidad_nombre") or data.get("unidad"), user)
        title = str(data.get("titulo") or "").strip()
        if not title:
            raise ValueError("El título de la actividad es obligatorio.")
        now = now_iso()
        with self.repo.connect() as conn:
            cur = conn.execute(
                """
                INSERT INTO sn_actividades_integrales(fundacion_id,expediente_uca_id,unidad_nombre,linea_componente,tipo_actividad,titulo,objetivo,metodologia,fecha_programada,fecha_ejecucion,hora_inicio,hora_fin,lugar,responsable_id,responsable_nombre,estado,resultados,conclusiones_profesionales,compromisos_generales,requiere_acta,requiere_listado,requiere_evidencias,creado_por,fecha_creacion,actualizado_por,fecha_actualizacion)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    fundacion_id, data.get("expediente_uca_id") or self._uca_expedient_id(fundacion_id, unit), unit, line,
                    _norm(data.get("tipo_actividad") or "JORNADA"), title, data.get("objetivo"), data.get("metodologia"),
                    data.get("fecha_programada"), data.get("fecha_ejecucion"), data.get("hora_inicio"), data.get("hora_fin"), data.get("lugar"),
                    data.get("responsable_id") or user.get("id"), data.get("responsable_nombre") or user.get("username"), _norm(data.get("estado") or "PROGRAMADA"),
                    data.get("resultados"), data.get("conclusiones_profesionales"), data.get("compromisos_generales"), int(bool(data.get("requiere_acta", True))),
                    int(bool(data.get("requiere_listado", True))), int(bool(data.get("requiere_evidencias", True))), user.get("id"), now, user.get("id"), now,
                ),
            )
            activity_id = int(cur.lastrowid)
            participants = data.get("participantes") or []
            if data.get("incluir_uca_completa"):
                participants = [{"documento": row.get("documento"), "nombre_completo": row.get("nombre_completo") or row.get("nombre")} for row in self.list_expedientes(fundacion_id, user, {"unidad": unit})]
            for part in participants:
                doc = str(part.get("documento") or "").strip()
                if not doc:
                    continue
                exp = conn.execute("SELECT id FROM sn_expedientes_integrales WHERE fundacion_id=? AND documento=?", (fundacion_id, doc)).fetchone()
                conn.execute(
                    "INSERT INTO sn_actividad_participantes(fundacion_id,actividad_id,expediente_id,documento,nombre_completo,convocado,asistio,firma_estado,creado_por,fecha_creacion,actualizado_por,fecha_actualizacion) VALUES(?,?,?,?,?,1,0,'PENDIENTE',?,?,?,?) ON CONFLICT(fundacion_id,actividad_id,documento) DO NOTHING",
                    (fundacion_id, activity_id, int(exp[0]) if exp else None, doc, part.get("nombre_completo"), user.get("id"), now, user.get("id"), now),
                )
            conn.commit()
        self.repo.log("CREAR_ACTIVIDAD_SALUD", "sn_actividades_integrales", activity_id, usuario=user.get("username", "sistema"), nuevos={"titulo": title, "linea": line, "unidad": unit})
        return self.activity_detail(fundacion_id, activity_id, user)

    def list_activities(self, fundacion_id: int, user: dict[str, Any], filters: dict[str, Any] | None = None) -> list[dict[str, Any]]:
        filters = filters or {}
        where = ["a.fundacion_id=?"]
        params: list[Any] = [fundacion_id]
        for key, column in (("unidad", "a.unidad_nombre"), ("estado", "a.estado"), ("linea", "a.linea_componente")):
            value = str(filters.get(key) or "").strip()
            if value:
                where.append(f"{column}=?")
                params.append(value if key == "unidad" else _norm(value))
        rows = self.repo.fetch_all(
            f"""
            SELECT a.*,
             (SELECT COUNT(*) FROM sn_actividad_participantes p WHERE p.fundacion_id=a.fundacion_id AND p.actividad_id=a.id) participantes_total,
             (SELECT COUNT(*) FROM sn_actividad_participantes p WHERE p.fundacion_id=a.fundacion_id AND p.actividad_id=a.id AND p.asistio=1) asistentes_total,
             (SELECT COUNT(*) FROM sn_evidencias_integrales e WHERE e.fundacion_id=a.fundacion_id AND e.actividad_id=a.id AND e.activo=1) evidencias_total,
             (SELECT COUNT(*) FROM sn_productos_actividad d WHERE d.fundacion_id=a.fundacion_id AND d.actividad_id=a.id AND d.activo=1) documentos_total
            FROM sn_actividades_integrales a WHERE {' AND '.join(where)} ORDER BY COALESCE(a.fecha_programada,''),a.id DESC LIMIT 2000
            """,
            params,
        )
        allowed = _allowed_units(user)
        if allowed is not None:
            keys = {_unit_key(value) for value in allowed}
            rows = [row for row in rows if _unit_key(row.get("unidad_nombre")) in keys]
        return rows

    def activity_detail(self, fundacion_id: int, activity_id: int, user: dict[str, Any]) -> dict[str, Any]:
        row = self.repo.fetch_one("SELECT * FROM sn_actividades_integrales WHERE fundacion_id=? AND id=?", (fundacion_id, activity_id))
        if not row:
            raise LookupError("Actividad de Salud y Nutrición no encontrada.")
        if not _can_access_unit(row.get("unidad_nombre"), user):
            raise PermissionError("No tienes permiso sobre esta UCA.")
        participants = self.repo.fetch_all("SELECT * FROM sn_actividad_participantes WHERE fundacion_id=? AND actividad_id=? ORDER BY nombre_completo", (fundacion_id, activity_id))
        docs = self.repo.fetch_all("SELECT * FROM sn_productos_actividad WHERE fundacion_id=? AND actividad_id=? AND activo=1 ORDER BY fecha_generacion DESC", (fundacion_id, activity_id))
        evidences = self.repo.fetch_all("SELECT id,tipo,titulo,nombre_original,mime_type,tamano_bytes,sha256,fecha_carga FROM sn_evidencias_integrales WHERE fundacion_id=? AND actividad_id=? AND activo=1 ORDER BY fecha_carga DESC", (fundacion_id, activity_id))
        return {"actividad": row, "participantes": participants, "documentos": docs, "evidencias": evidences}

    def update_activity(self, fundacion_id: int, activity_id: int, data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        current = self.activity_detail(fundacion_id, activity_id, user)["actividad"]
        state = _norm(data.get("estado") or current.get("estado"))
        if state in CLOSED_STATES and user.get("rol") not in COORDINATION_ROLES:
            raise PermissionError("El cierre requiere validación de coordinación.")
        if state in CLOSED_STATES and int(current.get("requiere_evidencias") or 0):
            evidence = self.repo.fetch_one("SELECT id FROM sn_evidencias_integrales WHERE fundacion_id=? AND actividad_id=? AND activo=1 LIMIT 1", (fundacion_id, activity_id))
            if not evidence:
                raise ValueError("No se puede cerrar la actividad sin evidencia.")
        now = now_iso()
        fields = {
            "objetivo": data.get("objetivo", current.get("objetivo")),
            "metodologia": data.get("metodologia", current.get("metodologia")),
            "fecha_programada": data.get("fecha_programada", current.get("fecha_programada")),
            "fecha_ejecucion": data.get("fecha_ejecucion", current.get("fecha_ejecucion")),
            "hora_inicio": data.get("hora_inicio", current.get("hora_inicio")),
            "hora_fin": data.get("hora_fin", current.get("hora_fin")),
            "lugar": data.get("lugar", current.get("lugar")),
            "estado": state,
            "resultados": data.get("resultados", current.get("resultados")),
            "conclusiones_profesionales": data.get("conclusiones_profesionales", current.get("conclusiones_profesionales")),
            "compromisos_generales": data.get("compromisos_generales", current.get("compromisos_generales")),
        }
        self.repo.execute_update(
            "UPDATE sn_actividades_integrales SET objetivo=?,metodologia=?,fecha_programada=?,fecha_ejecucion=?,hora_inicio=?,hora_fin=?,lugar=?,estado=?,resultados=?,conclusiones_profesionales=?,compromisos_generales=?,actualizado_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?",
            (*fields.values(), user.get("id"), now, fundacion_id, activity_id),
        )
        self.repo.log("ACTUALIZAR_ACTIVIDAD_SALUD", "sn_actividades_integrales", activity_id, usuario=user.get("username", "sistema"), anteriores=current, nuevos=fields)
        return self.activity_detail(fundacion_id, activity_id, user)

    def update_attendance(self, fundacion_id: int, activity_id: int, rows: list[dict[str, Any]], user: dict[str, Any]) -> dict[str, Any]:
        self.activity_detail(fundacion_id, activity_id, user)
        now = now_iso()
        with self.repo.connect() as conn:
            for row in rows:
                doc = str(row.get("documento") or "").strip()
                if not doc:
                    continue
                conn.execute(
                    "UPDATE sn_actividad_participantes SET asistio=?,firma_estado=?,observaciones=?,actualizado_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND actividad_id=? AND documento=?",
                    (int(bool(row.get("asistio"))), _norm(row.get("firma_estado") or "PENDIENTE"), row.get("observaciones"), user.get("id"), now, fundacion_id, activity_id, doc),
                )
            conn.commit()
        return self.activity_detail(fundacion_id, activity_id, user)

    def _tenant_dir(self, fundacion_id: int, *parts: str) -> Path:
        path = tenant_storage_root(self.data_dir, fundacion_id) / "salud_nutricion_integral"
        path = path.joinpath(*parts)
        path.mkdir(parents=True, exist_ok=True)
        return path

    def add_evidence(self, fundacion_id: int, file_obj: Any, data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        if not file_obj or not getattr(file_obj, "filename", None):
            raise ValueError("Selecciona un archivo.")
        activity_id = int(data.get("actividad_id") or 0) or None
        channel_id = int(data.get("canalizacion_id") or 0) or None
        expediente_id = int(data.get("expediente_id") or 0) or None
        if activity_id:
            self.activity_detail(fundacion_id, activity_id, user)
        if expediente_id:
            self.expediente_detail(fundacion_id, expediente_id, user)
        original = str(file_obj.filename)
        safe = secure_filename(original) or "evidencia"
        folder = self._tenant_dir(fundacion_id, "evidencias")
        token = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        target = folder / f"{token}_{safe}"
        file_obj.save(target)
        size = target.stat().st_size
        if size <= 0:
            target.unlink(missing_ok=True)
            raise ValueError("El archivo está vacío.")
        digest = _file_sha256(target)
        mime = getattr(file_obj, "mimetype", None) or mimetypes.guess_type(original)[0] or "application/octet-stream"
        now = now_iso()
        with self.repo.connect() as conn:
            cur = conn.execute(
                "INSERT INTO sn_evidencias_integrales(fundacion_id,actividad_id,canalizacion_id,expediente_id,tipo,titulo,nombre_original,nombre_guardado,ruta_archivo,mime_type,tamano_bytes,sha256,version,activo,cargado_por,fecha_carga) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,1,1,?,?)",
                (fundacion_id, activity_id, channel_id, expediente_id, _norm(data.get("tipo") or "SOPORTE"), data.get("titulo") or original, original, target.name, str(target), mime, size, digest, user.get("id"), now),
            )
            eid = int(cur.lastrowid)
            conn.commit()
        self.repo.log("CARGAR_EVIDENCIA_SALUD", "sn_evidencias_integrales", eid, usuario=user.get("username", "sistema"), nuevos={"sha256": digest, "actividad_id": activity_id})
        return {"id": eid, "nombre_original": original, "mime_type": mime, "tamano_bytes": size, "sha256": digest, "fecha_carga": now}

    def evidence_path(self, fundacion_id: int, evidence_id: int) -> tuple[Path, str, str] | None:
        row = self.repo.fetch_one("SELECT * FROM sn_evidencias_integrales WHERE fundacion_id=? AND id=? AND activo=1", (fundacion_id, evidence_id))
        if not row:
            return None
        path = Path(row["ruta_archivo"]).resolve()
        try:
            path.relative_to(tenant_storage_root(self.data_dir, fundacion_id).resolve())
        except ValueError:
            return None
        if not path.is_file() or _file_sha256(path) != row.get("sha256"):
            return None
        return path, row.get("nombre_original") or path.name, row.get("mime_type") or "application/octet-stream"

    def prepare_activity_documents(self, fundacion_id: int, activity_id: int, user: dict[str, Any], types: Iterable[str]) -> dict[str, Any]:
        detail = self.activity_detail(fundacion_id, activity_id, user)
        activity = detail["actividad"]
        folder = self._tenant_dir(fundacion_id, "documentos")
        token = datetime.now().strftime("%Y%m%d_%H%M%S")
        output: list[dict[str, Any]] = []
        for raw in types:
            kind = _norm(raw)
            if kind == "LISTADO_ASISTENCIA":
                path = folder / f"listado_asistencia_salud_{activity_id}_{token}.xlsx"
                self._write_attendance(path, detail)
                mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            elif kind in {"ACTA", "INFORME"}:
                path = folder / f"{kind.lower()}_salud_{activity_id}_{token}.pdf"
                self._write_activity_pdf(path, detail, kind)
                mime = "application/pdf"
            else:
                continue
            output.append(self._store_product(fundacion_id, path, kind, user, activity_id=activity_id, template_code=f"SN-{kind}", template_version="INTERNA-1"))
        return {"documentos": output}

    def generate_capture(self, fundacion_id: int, user: dict[str, Any], unit: str | None = None, period: str | None = None, formats: Iterable[str] = ("XLSX", "PDF")) -> dict[str, Any]:
        if unit:
            unit = _require_unit(unit, user)
        where = ["v.fundacion_id=?", "v.activo=1", "vv.activo=1", "vv.estado_validacion='VALIDADA'"]
        params: list[Any] = [fundacion_id]
        if unit:
            where.append("v.unidad=?")
            params.append(unit)
        if period:
            where.append("v.periodo=?")
            params.append(period)
        rows = self.repo.fetch_all(
            f"""
            SELECT v.*,vv.clasificacion_profesional,vv.observacion_profesional,vv.profesional_nombre,vv.fecha_validacion
            FROM sn_valoraciones v
            JOIN sn_valoracion_validaciones vv ON vv.valoracion_id=v.id AND vv.fundacion_id=v.fundacion_id
            WHERE {' AND '.join(where)}
              AND NOT EXISTS(SELECT 1 FROM sn_valoraciones x WHERE x.fundacion_id=v.fundacion_id AND x.documento=v.documento AND x.activo=1 AND (COALESCE(x.fecha_valoracion,'')>COALESCE(v.fecha_valoracion,'') OR (COALESCE(x.fecha_valoracion,'')=COALESCE(v.fecha_valoracion,'') AND x.id>v.id)))
            ORDER BY v.unidad,v.nombre_completo
            """,
            params,
        )
        if not rows:
            raise ValueError("No hay valoraciones profesionales validadas para generar CAPTURE.")
        folder = self._tenant_dir(fundacion_id, "capture")
        token = datetime.now().strftime("%Y%m%d_%H%M%S")
        products: list[dict[str, Any]] = []
        for raw in formats:
            fmt = _norm(raw)
            if fmt == "XLSX":
                path = folder / f"CAPTURE_{period or 'TODOS'}_{token}.xlsx"
                self._write_capture_xlsx(path, rows, unit, period)
                products.append(self._store_product(fundacion_id, path, "CAPTURE_XLSX", user, template_code="CAPTURE", template_version="BORRADOR-CONTROLADO-1"))
            elif fmt == "PDF":
                path = folder / f"CAPTURE_{period or 'TODOS'}_{token}.pdf"
                self._write_capture_pdf(path, rows, unit, period)
                products.append(self._store_product(fundacion_id, path, "CAPTURE_PDF", user, template_code="CAPTURE", template_version="BORRADOR-CONTROLADO-1"))
        return {"total_registros": len(rows), "productos": products, "advertencia": "Borrador controlado: debe compararse con la plantilla oficial vigente antes de uso institucional."}

    def _store_product(self, fundacion_id: int, path: Path, kind: str, user: dict[str, Any], *, activity_id: int | None = None, expediente_id: int | None = None, template_code: str | None = None, template_version: str | None = None) -> dict[str, Any]:
        mime = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        digest = _file_sha256(path)
        now = now_iso()
        with self.repo.connect() as conn:
            cur = conn.execute(
                "INSERT INTO sn_productos_actividad(fundacion_id,actividad_id,expediente_id,tipo_producto,nombre_archivo,ruta_archivo,mime_type,tamano_bytes,sha256,plantilla_codigo,plantilla_version,estado,generado_por,fecha_generacion,activo) VALUES(?,?,?,?,?,?,?,?,?,?,?,'BORRADOR',?,?,1)",
                (fundacion_id, activity_id, expediente_id, kind, path.name, str(path), mime, path.stat().st_size, digest, template_code, template_version, user.get("id"), now),
            )
            product_id = int(cur.lastrowid)
            conn.commit()
        return {"id": product_id, "tipo_producto": kind, "nombre_archivo": path.name, "mime_type": mime, "tamano_bytes": path.stat().st_size, "sha256": digest, "estado": "BORRADOR", "fecha_generacion": now}

    def product_path(self, fundacion_id: int, product_id: int) -> tuple[Path, str, str] | None:
        row = self.repo.fetch_one("SELECT * FROM sn_productos_actividad WHERE fundacion_id=? AND id=? AND activo=1", (fundacion_id, product_id))
        if not row:
            return None
        path = Path(row["ruta_archivo"]).resolve()
        try:
            path.relative_to(tenant_storage_root(self.data_dir, fundacion_id).resolve())
        except ValueError:
            return None
        if not path.is_file() or _file_sha256(path) != row.get("sha256"):
            return None
        return path, row.get("nombre_archivo") or path.name, row.get("mime_type") or "application/octet-stream"

    def create_canalization(self, fundacion_id: int, data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        expediente_id = int(data.get("expediente_id") or 0)
        exp = self.expediente_detail(fundacion_id, expediente_id, user)["expediente"]
        reason = str(data.get("motivo") or "").strip()
        route = _norm(data.get("tipo_ruta"))
        if not reason or not route:
            raise ValueError("Tipo de ruta y motivo son obligatorios.")
        now = now_iso()
        with self.repo.connect() as conn:
            cur = conn.execute(
                """
                INSERT INTO sn_canalizaciones(fundacion_id,expediente_id,valoracion_id,alerta_id,unidad_nombre,tipo_ruta,motivo,prioridad,entidad_destino,contacto_entidad,fecha_activacion,fecha_limite,estado,responsable_id,responsable_nombre,creado_por,fecha_creacion,actualizado_por,fecha_actualizacion)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?, 'ABIERTA',?,?,?,?,?,?)
                """,
                (fundacion_id, expediente_id, data.get("valoracion_id"), data.get("alerta_id"), exp.get("unidad_nombre"), route, reason, _norm(data.get("prioridad") or "MEDIA"), data.get("entidad_destino"), data.get("contacto_entidad"), data.get("fecha_activacion") or date.today().isoformat(), data.get("fecha_limite"), data.get("responsable_id") or user.get("id"), data.get("responsable_nombre") or user.get("username"), user.get("id"), now, user.get("id"), now),
            )
            cid = int(cur.lastrowid)
            conn.commit()
        self.repo.log("CREAR_CANALIZACION_SALUD", "sn_canalizaciones", cid, documento=exp.get("documento") or "", usuario=user.get("username", "sistema"), nuevos={"tipo": route, "prioridad": data.get("prioridad")})
        return self.repo.fetch_one("SELECT * FROM sn_canalizaciones WHERE fundacion_id=? AND id=?", (fundacion_id, cid)) or {}

    def add_followup(self, fundacion_id: int, entity_type: str, entity_id: int, data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        entity_type = _norm(entity_type)
        if entity_type not in {"CANALIZACION", "ACTIVIDAD", "EXPEDIENTE"}:
            raise ValueError("Tipo de seguimiento no permitido.")
        now = now_iso()
        with self.repo.connect() as conn:
            cur = conn.execute(
                "INSERT INTO sn_seguimientos_integrales(fundacion_id,entidad_tipo,entidad_id,fecha_seguimiento,actuacion,resultado,proximo_seguimiento,responsable_id,responsable_nombre,evidencia_referencia,creado_por,fecha_creacion) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                (fundacion_id, entity_type, entity_id, data.get("fecha_seguimiento") or date.today().isoformat(), str(data.get("actuacion") or "").strip(), data.get("resultado"), data.get("proximo_seguimiento"), user.get("id"), user.get("username"), data.get("evidencia_referencia"), user.get("id"), now),
            )
            sid = int(cur.lastrowid)
            if entity_type == "CANALIZACION" and data.get("proximo_seguimiento"):
                conn.execute("UPDATE sn_canalizaciones SET fecha_limite=?,actualizado_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?", (data.get("proximo_seguimiento"), user.get("id"), now, fundacion_id, entity_id))
            conn.commit()
        return self.repo.fetch_one("SELECT * FROM sn_seguimientos_integrales WHERE fundacion_id=? AND id=?", (fundacion_id, sid)) or {}

    def close_canalization(self, fundacion_id: int, channel_id: int, data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        row = self.repo.fetch_one("SELECT * FROM sn_canalizaciones WHERE fundacion_id=? AND id=?", (fundacion_id, channel_id))
        if not row:
            raise LookupError("Canalización no encontrada.")
        if not _can_access_unit(row.get("unidad_nombre"), user):
            raise PermissionError("No tienes permiso sobre esta UCA.")
        result = str(data.get("resultado_cierre") or "").strip()
        evidence = str(data.get("evidencia_cierre") or "").strip()
        if not result or not evidence:
            raise ValueError("El cierre exige resultado profesional y evidencia.")
        now = now_iso()
        self.repo.execute_update(
            "UPDATE sn_canalizaciones SET estado='CERRADA',resultado_cierre=?,evidencia_cierre=?,fecha_cierre=?,cerrado_por=?,actualizado_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?",
            (result, evidence, date.today().isoformat(), user.get("id"), user.get("id"), now, fundacion_id, channel_id),
        )
        return self.repo.fetch_one("SELECT * FROM sn_canalizaciones WHERE fundacion_id=? AND id=?", (fundacion_id, channel_id)) or {}

    def dashboard(self, fundacion_id: int, user: dict[str, Any], unit: str | None = None) -> dict[str, Any]:
        exps = self.list_expedientes(fundacion_id, user, {"unidad": unit} if unit else {})
        acts = self.list_activities(fundacion_id, user, {"unidad": unit} if unit else {})
        exp_ids = {int(row["id"]) for row in exps}
        docs = self.repo.fetch_all("SELECT * FROM sn_documentos_salud WHERE fundacion_id=? AND activo=1", (fundacion_id,))
        docs = [row for row in docs if int(row.get("expediente_id") or 0) in exp_ids]
        channels = self.repo.fetch_all("SELECT * FROM sn_canalizaciones WHERE fundacion_id=? ORDER BY fecha_creacion DESC", (fundacion_id,))
        channels = [row for row in channels if int(row.get("expediente_id") or 0) in exp_ids]
        validations = self.repo.fetch_all("SELECT * FROM sn_valoracion_validaciones WHERE fundacion_id=? AND activo=1", (fundacion_id,))
        return {
            "resumen": {
                "expedientes": len(exps),
                "documentos_pendientes": sum(1 for row in docs if _norm(row.get("estado")) not in {"VIGENTE", "VALIDADO", "NO_APLICA"}),
                "sin_vacunas": sum(1 for row in docs if row.get("tipo_documento") == "VACUNACION_PAI" and _norm(row.get("estado")) not in {"VIGENTE", "VALIDADO", "NO_APLICA"}),
                "sin_afiliacion": sum(1 for row in docs if row.get("tipo_documento") == "AFILIACION_SGSSS" and _norm(row.get("estado")) not in {"VIGENTE", "VALIDADO", "NO_APLICA"}),
                "sin_valoracion_integral": sum(1 for row in docs if row.get("tipo_documento") == "VALORACION_INTEGRAL_SALUD" and _norm(row.get("estado")) not in {"VIGENTE", "VALIDADO", "NO_APLICA"}),
                "actividades_pendientes": sum(1 for row in acts if _norm(row.get("estado")) not in CLOSED_STATES),
                "canalizaciones_abiertas": sum(1 for row in channels if _norm(row.get("estado")) not in {"CERRADA", "CERRADO"}),
                "valoraciones_pendientes_validacion": sum(1 for row in validations if _norm(row.get("estado_validacion")) != "VALIDADA"),
            },
            "expedientes": exps[:250],
            "actividades": acts[:250],
            "canalizaciones": channels[:250],
        }

    def _write_activity_pdf(self, path: Path, detail: dict[str, Any], kind: str) -> None:
        styles = getSampleStyleSheet()
        activity = detail["actividad"]
        title = "ACTA DE ACTIVIDAD DE SALUD Y NUTRICIÓN" if kind == "ACTA" else "BORRADOR DE INFORME DE SALUD Y NUTRICIÓN"
        doc = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=1.5 * cm, leftMargin=1.5 * cm, topMargin=1.4 * cm, bottomMargin=1.4 * cm)
        story = [Paragraph(title, styles["Title"]), Paragraph("BORRADOR PARA REVISIÓN Y APROBACIÓN PROFESIONAL", styles["Heading2"]), Spacer(1, 10)]
        rows = [
            ["Campo", "Información"], ["Línea", activity.get("linea_componente") or ""], ["Tipo", activity.get("tipo_actividad") or ""],
            ["Título", activity.get("titulo") or ""], ["UCA", activity.get("unidad_nombre") or ""], ["Fecha programada", activity.get("fecha_programada") or ""],
            ["Fecha ejecutada", activity.get("fecha_ejecucion") or "Pendiente"], ["Responsable", activity.get("responsable_nombre") or ""],
            ["Objetivo", activity.get("objetivo") or "Pendiente de diligenciar"], ["Metodología", activity.get("metodologia") or "Pendiente de diligenciar"],
            ["Resultados", activity.get("resultados") or "Pendiente de diligenciar por el profesional"],
            ["Conclusiones", activity.get("conclusiones_profesionales") or "Pendiente de diligenciar por el profesional"],
            ["Compromisos", activity.get("compromisos_generales") or "Pendiente de diligenciar"],
        ]
        table = Table([[Paragraph(str(cell), styles["BodyText"]) for cell in row] for row in rows], colWidths=[4.2 * cm, 12.8 * cm])
        table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#047857")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), .35, colors.grey), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
        story.extend([table, Spacer(1, 10), Paragraph("La plataforma no inventa resultados ni reemplaza el criterio profesional. Revise y apruebe antes de uso oficial.", styles["Italic"])])
        doc.build(story)

    def _write_attendance(self, path: Path, detail: dict[str, Any]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencia"
        ws.append(["LISTADO DE ASISTENCIA - BORRADOR"])
        ws.merge_cells("A1:G1")
        ws["A1"].font = Font(bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="047857")
        activity = detail["actividad"]
        ws.append(["Actividad", activity.get("titulo"), "UCA", activity.get("unidad_nombre"), "Fecha", activity.get("fecha_ejecucion") or activity.get("fecha_programada"), "Responsable"])
        ws.append(["Documento", "Nombre", "Convocado", "Asistió", "Firma", "Observaciones", "Validación"])
        for row in detail["participantes"]:
            ws.append([row.get("documento"), row.get("nombre_completo"), "Sí" if row.get("convocado") else "No", "Sí" if row.get("asistio") else "No", row.get("firma_estado"), row.get("observaciones"), "Pendiente de revisión"])
        for column in "ABCDEFG":
            ws.column_dimensions[column].width = 24
        wb.save(path)

    def _capture_headers(self) -> list[str]:
        return ["UCA", "Documento", "Nombre", "Fecha nacimiento", "Edad meses", "Sexo", "Fecha valoración", "Peso kg", "Talla cm", "IMC", "Perímetro braquial cm", "Perímetro cefálico cm", "Z P/E", "Z T/E", "Z P/T", "Z IMC/E", "Clasificación automática", "Clasificación profesional", "Profesional", "Fecha validación", "Observaciones profesionales"]

    def _capture_row(self, row: dict[str, Any]) -> list[Any]:
        return [row.get("unidad"), row.get("documento"), row.get("nombre_completo"), row.get("fecha_nacimiento"), row.get("edad_meses"), row.get("sexo"), row.get("fecha_valoracion"), row.get("peso_kg"), row.get("talla_cm"), row.get("imc"), row.get("perimetro_braquial_cm"), row.get("perimetro_cefalico_cm"), row.get("z_peso_edad"), row.get("z_talla_edad"), row.get("z_peso_talla"), row.get("z_imc_edad"), row.get("diagnostico_global"), row.get("clasificacion_profesional") or row.get("diagnostico_global"), row.get("profesional_nombre"), row.get("fecha_validacion"), row.get("observacion_profesional")]

    def _write_capture_xlsx(self, path: Path, rows: list[dict[str, Any]], unit: str | None, period: str | None) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "CAPTURE"
        ws.append(["FORMATO CAPTURE - BORRADOR CONTROLADO"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self._capture_headers()))
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
        ws["A1"].fill = PatternFill("solid", fgColor="0F766E")
        ws.append([f"UCA: {unit or 'Todas'} · Periodo: {period or 'Todos'} · Solo valoraciones con validación profesional"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(self._capture_headers()))
        ws.append(self._capture_headers())
        for cell in ws[3]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="115E59")
            cell.alignment = Alignment(wrap_text=True)
        for row in rows:
            ws.append(self._capture_row(row))
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = ws.dimensions
        for idx in range(1, len(self._capture_headers()) + 1):
            ws.column_dimensions[ws.cell(3, idx).column_letter].width = 18
        wb.save(path)

    def _write_capture_pdf(self, path: Path, rows: list[dict[str, Any]], unit: str | None, period: str | None) -> None:
        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=.8 * cm, leftMargin=.8 * cm, topMargin=.8 * cm, bottomMargin=.8 * cm)
        story = [Paragraph("CAPTURE — BORRADOR CONTROLADO", styles["Title"]), Paragraph(f"UCA: {unit or 'Todas'} · Periodo: {period or 'Todos'} · Registros: {len(rows)}", styles["BodyText"]), Spacer(1, 8)]
        headers = ["UCA", "Documento", "Nombre", "Fecha", "Peso", "Talla", "PB", "Automática", "Profesional", "Validó"]
        data = [headers]
        for row in rows:
            data.append([row.get("unidad"), row.get("documento"), row.get("nombre_completo"), row.get("fecha_valoracion"), row.get("peso_kg"), row.get("talla_cm"), row.get("perimetro_braquial_cm"), row.get("diagnostico_global"), row.get("clasificacion_profesional") or row.get("diagnostico_global"), row.get("profesional_nombre")])
        table = Table([[Paragraph(str(cell or ""), styles["BodyText"]) for cell in row] for row in data], repeatRows=1, colWidths=[2.3*cm,2.5*cm,4.2*cm,2.2*cm,1.4*cm,1.4*cm,1.4*cm,3.1*cm,3.1*cm,2.8*cm])
        table.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#115E59")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), .25, colors.grey), ("VALIGN", (0,0), (-1,-1), "TOP"), ("FONTSIZE", (0,0), (-1,-1), 7)]))
        story.extend([table, Spacer(1, 8), Paragraph("Documento generado con información registrada y validada. Debe compararse con la plantilla oficial vigente antes de su presentación institucional.", styles["Italic"])])
        doc.build(story)


def register_integral_routes(bp: Blueprint, repo: Any, data_dir: str) -> SaludNutricionIntegralService:
    service = SaludNutricionIntegralService(repo, data_dir)
    service.init_schema()

    @bp.route("/integral/salud", methods=["GET"])
    def integral_health():
        return jsonify({"status": "ok", "module": "salud_nutricion_integral", "schema_version": 1, "version": "2.6.0"}), 200

    @bp.route("/integral/dashboard", methods=["GET"])
    @require_roles(*READ_ROLES)
    def integral_dashboard():
        user = _user()
        unit = request.args.get("unidad") or None
        if unit and not _can_access_unit(unit, user):
            return jsonify({"error": "No tienes permiso sobre esta UCA."}), 403
        return jsonify(service.dashboard(user["fundacion_id"], user, unit)), 200

    @bp.route("/integral/expedientes/sincronizar", methods=["POST"])
    @require_roles(*EDIT_ROLES)
    def sync_integral_records():
        user = _user(); data = request.get_json(silent=True) or {}
        try:
            unit = data.get("unidad") or data.get("unidad_nombre")
            if unit:
                unit = _require_unit(unit, user)
            result = service.sync_expedientes(user["fundacion_id"], user, unit)
            return jsonify({"message": "Expedientes de salud referenciados sin duplicar Base Maestra.", "resultado": result}), 200
        except PermissionError as exc:
            return jsonify({"error": str(exc)}), 403
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/integral/expedientes", methods=["GET"])
    @require_roles(*READ_ROLES)
    def integral_records():
        user = _user()
        return jsonify({"expedientes": service.list_expedientes(user["fundacion_id"], user, dict(request.args))}), 200

    @bp.route("/integral/expedientes/<int:expediente_id>", methods=["GET"])
    @require_roles(*READ_ROLES)
    def integral_record_detail(expediente_id: int):
        user = _user()
        try:
            return jsonify(service.expediente_detail(user["fundacion_id"], expediente_id, user)), 200
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except PermissionError as exc:
            return jsonify({"error": str(exc)}), 403

    @bp.route("/integral/documentos/<int:document_id>", methods=["PATCH"])
    @require_roles(*EDIT_ROLES)
    def update_integral_document(document_id: int):
        user = _user()
        try:
            item = service.update_document(user["fundacion_id"], document_id, request.get_json(silent=True) or {}, user)
            return jsonify({"message": "Estado documental actualizado.", "documento": item}), 200
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except PermissionError as exc:
            return jsonify({"error": str(exc)}), 403
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/integral/valoraciones/<int:valoracion_id>/validar", methods=["POST"])
    @require_roles(*EDIT_ROLES)
    def validate_integral_valuation(valoracion_id: int):
        user = _user()
        try:
            item = service.validate_valoracion(user["fundacion_id"], valoracion_id, request.get_json(silent=True) or {}, user)
            return jsonify({"message": "Validación profesional registrada sin modificar la medición histórica.", "validacion": item}), 201
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except PermissionError as exc:
            return jsonify({"error": str(exc)}), 403
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/integral/actividades", methods=["GET", "POST"])
    @require_roles(*READ_ROLES)
    def integral_activities():
        user = _user()
        if request.method == "POST":
            if user.get("rol") not in EDIT_ROLES:
                return jsonify({"error": "Tu rol no puede crear actividades de Salud y Nutrición."}), 403
            try:
                item = service.create_activity(user["fundacion_id"], request.get_json(silent=True) or {}, user)
                return jsonify({"message": "Actividad creada. Los documentos quedan en borrador hasta revisión.", "actividad": item}), 201
            except PermissionError as exc:
                return jsonify({"error": str(exc)}), 403
            except Exception as exc:
                return jsonify({"error": str(exc)}), 400
        return jsonify({"actividades": service.list_activities(user["fundacion_id"], user, dict(request.args))}), 200

    @bp.route("/integral/actividades/<int:activity_id>", methods=["GET", "PATCH"])
    @require_roles(*READ_ROLES)
    def integral_activity_detail(activity_id: int):
        user = _user()
        try:
            if request.method == "PATCH":
                if user.get("rol") not in EDIT_ROLES:
                    return jsonify({"error": "Tu rol solo puede consultar."}), 403
                item = service.update_activity(user["fundacion_id"], activity_id, request.get_json(silent=True) or {}, user)
                return jsonify({"message": "Actividad actualizada.", "actividad": item}), 200
            return jsonify(service.activity_detail(user["fundacion_id"], activity_id, user)), 200
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except PermissionError as exc:
            return jsonify({"error": str(exc)}), 403
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/integral/actividades/<int:activity_id>/asistencia", methods=["PATCH"])
    @require_roles(*EDIT_ROLES)
    def integral_attendance(activity_id: int):
        user = _user(); data = request.get_json(silent=True) or {}
        try:
            return jsonify({"message": "Asistencia actualizada.", "actividad": service.update_attendance(user["fundacion_id"], activity_id, data.get("participantes") or [], user)}), 200
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except PermissionError as exc:
            return jsonify({"error": str(exc)}), 403
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/integral/actividades/<int:activity_id>/documentos", methods=["POST"])
    @require_roles(*EDIT_ROLES)
    def integral_activity_documents(activity_id: int):
        user = _user(); data = request.get_json(silent=True) or {}
        try:
            result = service.prepare_activity_documents(user["fundacion_id"], activity_id, user, data.get("tipos") or ["ACTA", "LISTADO_ASISTENCIA", "INFORME"])
            return jsonify({"message": "Documentos preparados como borradores.", "resultado": result}), 201
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except PermissionError as exc:
            return jsonify({"error": str(exc)}), 403
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/integral/capture", methods=["POST"])
    @require_roles(*EDIT_ROLES)
    def integral_capture():
        user = _user(); data = request.get_json(silent=True) or {}
        try:
            result = service.generate_capture(user["fundacion_id"], user, data.get("unidad"), data.get("periodo"), data.get("formatos") or ["XLSX", "PDF"])
            return jsonify({"message": "CAPTURE preparado como borrador controlado.", "resultado": result}), 201
        except PermissionError as exc:
            return jsonify({"error": str(exc)}), 403
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/integral/productos/<int:product_id>/descargar", methods=["GET"])
    @require_roles(*READ_ROLES)
    def download_integral_product(product_id: int):
        user = _user(); item = service.product_path(user["fundacion_id"], product_id)
        if not item:
            return jsonify({"error": "Producto no disponible o integridad inválida."}), 404
        path, name, mime = item
        return send_file(path, as_attachment=True, download_name=name, mimetype=mime)

    @bp.route("/integral/evidencias", methods=["POST"])
    @require_roles(*EDIT_ROLES)
    def upload_integral_evidence():
        user = _user()
        try:
            item = service.add_evidence(user["fundacion_id"], request.files.get("file"), dict(request.form), user)
            return jsonify({"message": "Evidencia cargada.", "evidencia": item}), 201
        except PermissionError as exc:
            return jsonify({"error": str(exc)}), 403
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/integral/evidencias/<int:evidence_id>/descargar", methods=["GET"])
    @require_roles(*READ_ROLES)
    def download_integral_evidence(evidence_id: int):
        user = _user(); item = service.evidence_path(user["fundacion_id"], evidence_id)
        if not item:
            return jsonify({"error": "Evidencia no disponible o integridad inválida."}), 404
        path, name, mime = item
        return send_file(path, as_attachment=True, download_name=name, mimetype=mime)

    @bp.route("/integral/canalizaciones", methods=["GET", "POST"])
    @require_roles(*READ_ROLES)
    def integral_channels():
        user = _user()
        if request.method == "POST":
            if user.get("rol") not in EDIT_ROLES:
                return jsonify({"error": "Tu rol no puede activar rutas."}), 403
            try:
                item = service.create_canalization(user["fundacion_id"], request.get_json(silent=True) or {}, user)
                return jsonify({"message": "Canalización registrada para seguimiento humano.", "canalizacion": item}), 201
            except LookupError as exc:
                return jsonify({"error": str(exc)}), 404
            except PermissionError as exc:
                return jsonify({"error": str(exc)}), 403
            except Exception as exc:
                return jsonify({"error": str(exc)}), 400
        rows = service.repo.fetch_all("SELECT * FROM sn_canalizaciones WHERE fundacion_id=? ORDER BY fecha_creacion DESC LIMIT 2000", (user["fundacion_id"],))
        rows = [row for row in rows if _can_access_unit(row.get("unidad_nombre"), user)]
        return jsonify({"canalizaciones": rows}), 200

    @bp.route("/integral/seguimientos/<entity_type>/<int:entity_id>", methods=["POST"])
    @require_roles(*EDIT_ROLES)
    def integral_followup(entity_type: str, entity_id: int):
        user = _user()
        try:
            item = service.add_followup(user["fundacion_id"], entity_type, entity_id, request.get_json(silent=True) or {}, user)
            return jsonify({"message": "Seguimiento registrado.", "seguimiento": item}), 201
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/integral/canalizaciones/<int:channel_id>/cerrar", methods=["POST"])
    @require_roles(*COORDINATION_ROLES)
    def close_integral_channel(channel_id: int):
        user = _user()
        try:
            item = service.close_canalization(user["fundacion_id"], channel_id, request.get_json(silent=True) or {}, user)
            return jsonify({"message": "Canalización cerrada con resultado y evidencia.", "canalizacion": item}), 200
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except PermissionError as exc:
            return jsonify({"error": str(exc)}), 403
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400

    return service
