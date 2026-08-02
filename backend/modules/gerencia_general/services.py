from __future__ import annotations

import io
import json
import os
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from flask import g, request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from .schema import GERENCIA_GENERAL_SCHEMA_SQL
from modules.sqlalchemy_compat import CoreConnection, CoreCompatRepository


ESTADOS_SUSCRIPCION_RIESGO = {'VENCIDA', 'SUSPENDIDA', 'CANCELADA'}
ESTADOS_ENTREGABLE_PENDIENTE = {'PENDIENTE', 'EN PROCESO', 'EN_PROCESO', 'EN_REVISION', 'EN REVISIÓN', 'DEVUELTO', 'VENCIDO'}
ESTADOS_ENTREGABLE_CERRADO = {'APROBADO', 'CARGADO', 'CUMPLIDO', 'RESUELTO', 'CERRADO'}
NIVELES_CRITICOS = {'CRITICO', 'CRÍTICO', 'ROJO', 'ALTA', 'CRITICA', 'CRÍTICA'}
NIVELES_RIESGO_NUTRICION = {'CRITICO', 'CRÍTICO', 'ROJO', 'RIESGO', 'ALTO', 'ALTA', 'DESNUTRICION', 'DESNUTRICIÓN'}


def now_iso() -> str:
    return datetime.now().isoformat(timespec='seconds')


def first_day_month(anio: int | None = None, mes: int | None = None) -> date:
    today = date.today()
    y = int(anio or today.year)
    m = int(mes or today.month)
    return date(y, m, 1)


def month_bounds(anio: int | None = None, mes: int | None = None) -> tuple[str, str, str]:
    start = first_day_month(anio, mes)
    end = date(start.year + 1, 1, 1) if start.month == 12 else date(start.year, start.month + 1, 1)
    return start.isoformat(), end.isoformat(), f'{start.year}-{start.month:02d}'


def safe_json(value: Any) -> str:
    try:
        return json.dumps(value, ensure_ascii=False, default=str)
    except Exception:
        return '{}'


def parse_json_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v) for v in value]
    text = str(value).strip()
    if not text:
        return []
    try:
        parsed = json.loads(text)
        if isinstance(parsed, list):
            return [str(v) for v in parsed]
        if isinstance(parsed, dict):
            return [str(k) for k, enabled in parsed.items() if enabled]
    except Exception:
        pass
    return [part.strip() for part in text.split(',') if part.strip()]


def row_to_dict(row: dict[str, Any] | None) -> dict[str, Any] | None:
    return dict(row) if row is not None else None


class GerenciaGeneralService:
    """Servicio de lectura ejecutiva y comercial sin alterar módulos operativos.

    El tablero solo consulta tablas existentes y crea tablas gg_* para auditoría propia.
    No modifica formatos ICBF ni cambia datos operativos.
    """

    def __init__(self, database_path: str, output_folder: str | None = None):
        self.database_path = database_path
        self.output_folder = output_folder or ''

    def connect(self) -> CoreConnection:
        return CoreConnection()

    def init_schema(self) -> None:
        conn = self.connect()
        conn.executescript(GERENCIA_GENERAL_SCHEMA_SQL)
        conn.commit()
        conn.close()

    def ctx(self) -> dict[str, Any]:
        try:
            from modules.seguridad.services import get_request_user_context
            ctx = get_request_user_context() or {}
        except Exception:
            ctx = {}
        user = getattr(g, 'current_user', None) or {}
        ctx.setdefault('usuario_id', user.get('id'))
        ctx.setdefault('username', user.get('username') or 'sistema')
        ctx.setdefault('rol', user.get('rol') or 'SUPERADMIN')
        ctx.setdefault('fundacion_id', user.get('fundacion_id') or 1)
        return ctx

    def is_superadmin(self) -> bool:
        return str(self.ctx().get('rol') or '').upper() == 'SUPERADMIN'

    def current_fundacion_id(self) -> int:
        try:
            return int(self.ctx().get('fundacion_id') or 1)
        except Exception:
            return 1

    def table_exists(self, cur: object, table: str) -> bool:
        try:
            return CoreCompatRepository().table_exists(table)
        except Exception:
            return False

    def columns(self, cur: object, table: str) -> set[str]:
        try:
            return CoreCompatRepository().columns(table)
        except Exception:
            return set()

    def scope_sql(self, alias: str = '', column: str = 'fundacion_id') -> tuple[str, tuple[Any, ...]]:
        if self.is_superadmin():
            return '', ()
        prefix = f'{alias}.' if alias else ''
        return f' AND {prefix}{column} = ?', (self.current_fundacion_id(),)

    def count(self, cur: object, sql: str, params: tuple = ()) -> int:
        try:
            row = cur.execute(sql, params).fetchone()
            if not row:
                return 0
            if isinstance(row, dict):
                return int((next(iter(row.values())) if row else 0) or 0)
            return int((row[0] if row else 0) or 0)
        except Exception:
            return 0

    def sum_value(self, cur: object, sql: str, params: tuple = ()) -> float:
        try:
            row = cur.execute(sql, params).fetchone()
            if not row:
                return 0.0
            if isinstance(row, dict):
                return float((next(iter(row.values())) if row else 0) or 0)
            return float((row[0] if row else 0) or 0)
        except Exception:
            return 0.0

    def log(self, accion: str, detalle: str = '', datos: Any = None) -> None:
        ctx = self.ctx()
        try:
            conn = self.connect()
            conn.execute(
                """
                INSERT INTO gg_auditoria
                (fundacion_id, usuario_id, accion, detalle, datos, ip, user_agent, fecha)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    ctx.get('fundacion_id'),
                    ctx.get('usuario_id'),
                    accion,
                    detalle,
                    safe_json(datos) if datos is not None else None,
                    getattr(request, 'remote_addr', None),
                    request.headers.get('User-Agent') if request else None,
                    now_iso(),
                ),
            )
            conn.commit()
            conn.close()
        except Exception:
            pass

    def dashboard(self, anio: int | None = None, mes: int | None = None) -> dict[str, Any]:
        inicio, fin, periodo = month_bounds(anio, mes)
        conn = self.connect()
        cur = conn.cursor()
        data = {
            'periodo': {'anio': int(periodo[:4]), 'mes': int(periodo[5:7]), 'inicio': inicio, 'fin': fin, 'periodo': periodo},
            'indicadores': {},
            'licencias': [],
            'alertas_pago': [],
            'alertas_criticas': [],
            'nutricion_riesgo': [],
            'coordinadores_bajo_cumplimiento': [],
            'unidades_cobertura_incompleta': [],
            'ingresos_mes_detalle': [],
            'creditos_detalle': [],
            'usuarios_ultimo_ingreso': [],
            'tickets_soporte': [],
            'tendencias': {},
            'resumen_ejecutivo': '',
        }

        data['licencias'] = self.licencias_fundaciones(cur)
        data['alertas_pago'] = self.alertas_pago(cur)
        data['alertas_criticas'] = self.alertas_criticas(cur, limit=20)
        data['nutricion_riesgo'] = self.casos_nutricion_riesgo(cur, limit=25)
        data['coordinadores_bajo_cumplimiento'] = self.coordinadores_bajo_cumplimiento(cur, periodo, limit=25)
        data['unidades_cobertura_incompleta'] = self.unidades_cobertura_incompleta(cur, limit=50)
        data['ingresos_mes_detalle'] = self.ingresos_mes_detalle(cur, inicio, fin)
        data['creditos_detalle'] = self.creditos_consumidos_detalle(cur, inicio, fin)
        data['usuarios_ultimo_ingreso'] = self.ultimos_ingresos(cur, limit=10)
        data['tickets_soporte'] = self.tickets_soporte(cur, limit=15)
        data['tendencias'] = {
            'ingresos': self.ingresos_ultimos_meses(cur, months=6),
            'creditos': self.creditos_ultimos_meses(cur, months=6),
        }

        data['indicadores'] = {
            'fundaciones_activas': self.fundaciones_activas(cur),
            'fundaciones_vencidas': self.fundaciones_vencidas(cur),
            'ingresos_mes': sum(float(r.get('valor_pagado') or 0) for r in data['ingresos_mes_detalle']),
            'creditos_consumidos': int(sum(abs(int(r.get('creditos') or 0)) for r in data['creditos_detalle'])),
            'usuarios_activos': self.usuarios_activos(cur),
            'formatos_generados': self.formatos_generados(inicio, fin),
            'entregables_pendientes': self.entregables_pendientes(cur, periodo),
            'alertas_criticas': len(data['alertas_criticas']),
            'casos_nutricionales_riesgo': len(data['nutricion_riesgo']),
            'coordinadores_bajo_cumplimiento': len(data['coordinadores_bajo_cumplimiento']),
            'unidades_cobertura_incompleta': len(data['unidades_cobertura_incompleta']),
            'tickets_abiertos': len([t for t in data['tickets_soporte'] if str(t.get('estado') or '').upper() not in {'CERRADO', 'RESUELTO', 'ANULADO'}]),
            'alertas_pago': len(data['alertas_pago']),
        }
        data['resumen_ejecutivo'] = self.resumen_ejecutivo(data)
        conn.close()
        self.log('CONSULTAR_DASHBOARD_GERENCIA', 'Consulta de tablero de gerencia general', {'periodo': periodo})
        return data

    def fundaciones_activas(self, cur: Any) -> int:
        if not self.table_exists(cur, 'fundaciones'):
            return 0
        scope, params = self.scope_sql('f')
        return self.count(
            cur,
            f"""
            SELECT COUNT(*) FROM fundaciones f
            WHERE UPPER(COALESCE(f.estado,'ACTIVA'))='ACTIVA'
              AND UPPER(COALESCE(f.suscripcion_estado,'ACTIVA')) IN ('ACTIVA','POR_VENCER') {scope}
            """,
            params,
        )

    def fundaciones_vencidas(self, cur: Any) -> int:
        if not self.table_exists(cur, 'fundaciones'):
            return 0
        scope, params = self.scope_sql('f')
        return self.count(
            cur,
            f"""
            SELECT COUNT(*) FROM fundaciones f
            WHERE (UPPER(COALESCE(f.estado,'')) IN ('SUSPENDIDA','CANCELADA')
               OR UPPER(COALESCE(f.suscripcion_estado,'')) IN ('VENCIDA','SUSPENDIDA','CANCELADA')
               OR (f.fecha_vencimiento IS NOT NULL AND date(f.fecha_vencimiento) < date('now')))
               {scope}
            """,
            params,
        )

    def usuarios_activos(self, cur: Any) -> int:
        if not self.table_exists(cur, 'usuarios_app'):
            return 0
        scope, params = self.scope_sql('u')
        return self.count(cur, f"SELECT COUNT(*) FROM usuarios_app u WHERE COALESCE(u.activo,1)=1 {scope}", params)

    def formatos_generados(self, inicio: str, fin: str) -> int:
        if not self.output_folder or not os.path.isdir(self.output_folder):
            return 0
        count = 0
        inicio_dt = datetime.fromisoformat(inicio)
        fin_dt = datetime.fromisoformat(fin)
        for root, _, files in os.walk(self.output_folder):
            for name in files:
                if not name.lower().endswith(('.xlsx', '.xls', '.xlsm', '.pdf', '.docx', '.zip')):
                    continue
                try:
                    mtime = datetime.fromtimestamp(os.path.getmtime(os.path.join(root, name)))
                    if inicio_dt <= mtime < fin_dt:
                        count += 1
                except Exception:
                    pass
        return count

    def entregables_pendientes(self, cur: Any, periodo: str) -> int:
        total = 0
        if self.table_exists(cur, 'gp_entregables'):
            scope, params = self.scope_sql('e') if 'fundacion_id' in self.columns(cur, 'gp_entregables') else ('', ())
            total += self.count(
                cur,
                f"""SELECT COUNT(*) FROM gp_entregables e
                    WHERE UPPER(COALESCE(e.estado,'PENDIENTE')) NOT IN ({','.join(['?'] * len(ESTADOS_ENTREGABLE_CERRADO))})
                    AND (COALESCE(e.periodo,'')=? OR substr(COALESCE(e.fecha_limite,''),1,7)=?) {scope}""",
                tuple(ESTADOS_ENTREGABLE_CERRADO) + (periodo, periodo) + params,
            )
        if self.table_exists(cur, 'entregables_operacion'):
            scope, params = self.scope_sql('e') if 'fundacion_id' in self.columns(cur, 'entregables_operacion') else ('', ())
            total += self.count(
                cur,
                f"""SELECT COUNT(*) FROM entregables_operacion e
                    WHERE UPPER(COALESCE(e.estado,'PENDIENTE')) NOT IN ('APROBADO','CARGADO','CUMPLIDO','CERRADO')
                    AND COALESCE(e.periodo,'')=? {scope}""",
                (periodo,) + params,
            )
        return total

    def licencias_fundaciones(self, cur: Any) -> list[dict[str, Any]]:
        if not self.table_exists(cur, 'fundaciones'):
            return []
        scope, params = self.scope_sql('f')
        rows = cur.execute(
            f"""
            SELECT
                f.id AS fundacion_id,
                f.nombre AS fundacion,
                f.nit,
                f.estado AS estado_fundacion,
                COALESCE(sf.estado, f.suscripcion_estado, 'ACTIVA') AS estado_suscripcion,
                COALESCE(sf.fecha_inicio, f.fecha_inicio) AS fecha_inicio,
                COALESCE(sf.fecha_vencimiento, f.fecha_vencimiento) AS fecha_vencimiento,
                COALESCE(sf.dias_gracia, 0) AS dias_gracia,
                COALESCE(sf.creditos_disponibles, f.creditos_disponibles, 0) AS creditos_disponibles,
                COALESCE(p.nombre, f.plan, 'SIN PLAN') AS plan,
                COALESCE(p.limite_usuarios, 0) AS usuarios_permitidos,
                COALESCE(p.limite_coordinadores, 0) AS coordinadores_permitidos,
                COALESCE(p.limite_unidades, 0) AS unidades_permitidas,
                COALESCE(sf.modulos_habilitados, p.modulos_habilitados, '[]') AS modulos_habilitados,
                COALESCE((SELECT COUNT(*) FROM usuarios_app u WHERE u.fundacion_id=f.id AND COALESCE(u.activo,1)=1),0) AS usuarios_activos,
                COALESCE((SELECT MAX(s.fecha_creacion) FROM sesiones_usuario s WHERE s.fundacion_id=f.id), '') AS ultimo_ingreso
            FROM fundaciones f
            LEFT JOIN suscripciones_fundacion sf ON sf.fundacion_id=f.id
            LEFT JOIN planes_suscripcion p ON p.id=COALESCE(sf.plan_id, f.plan_id)
            WHERE 1=1 {scope}
            ORDER BY
              CASE UPPER(COALESCE(sf.estado, f.suscripcion_estado,''))
                WHEN 'VENCIDA' THEN 0 WHEN 'SUSPENDIDA' THEN 1 WHEN 'CANCELADA' THEN 2 WHEN 'POR_VENCER' THEN 3 ELSE 4 END,
              f.nombre
            """,
            params,
        ).fetchall()
        result = []
        today = date.today()
        for r in rows:
            item = row_to_dict(r) or {}
            item['modulos_habilitados_lista'] = parse_json_list(item.get('modulos_habilitados'))
            item['dias_restantes'] = None
            item['estado_pago'] = item.get('estado_suscripcion') or 'ACTIVA'
            fv = item.get('fecha_vencimiento')
            if fv:
                try:
                    vence = date.fromisoformat(str(fv)[:10])
                    item['dias_restantes'] = (vence - today).days
                    estado = str(item.get('estado_suscripcion') or '').upper()
                    if estado in ESTADOS_SUSCRIPCION_RIESGO or item['dias_restantes'] < 0:
                        item['estado_pago'] = 'VENCIDA'
                    elif item['dias_restantes'] <= 7:
                        item['estado_pago'] = 'POR_VENCER'
                    else:
                        item['estado_pago'] = 'AL_DIA'
                except Exception:
                    pass
            result.append(item)
        return result

    def ingresos_mes_detalle(self, cur: Any, inicio: str, fin: str) -> list[dict[str, Any]]:
        if not self.table_exists(cur, 'pagos_suscripcion'):
            return []
        scope, params = self.scope_sql('p')
        rows = cur.execute(
            f"""
            SELECT p.id, p.fundacion_id, f.nombre AS fundacion, ps.nombre AS plan,
                   p.valor_pagado, p.metodo_pago, p.fecha_pago, p.fecha_vencimiento,
                   p.referencia_pago, p.observaciones
            FROM pagos_suscripcion p
            LEFT JOIN fundaciones f ON f.id=p.fundacion_id
            LEFT JOIN planes_suscripcion ps ON ps.id=p.plan_id
            WHERE substr(COALESCE(p.fecha_pago,p.fecha_creacion),1,10) >= ?
              AND substr(COALESCE(p.fecha_pago,p.fecha_creacion),1,10) < ? {scope}
            ORDER BY COALESCE(p.fecha_pago,p.fecha_creacion) DESC
            LIMIT 200
            """,
            (inicio, fin) + params,
        ).fetchall()
        return [row_to_dict(r) for r in rows]

    def creditos_consumidos_detalle(self, cur: Any, inicio: str, fin: str) -> list[dict[str, Any]]:
        if not self.table_exists(cur, 'movimientos_credito'):
            return []
        scope, params = self.scope_sql('m')
        rows = cur.execute(
            f"""
            SELECT m.id, m.fundacion_id, f.nombre AS fundacion, m.tipo, m.accion,
                   m.creditos, m.saldo_anterior, m.saldo_nuevo, m.descripcion, m.fecha_movimiento
            FROM movimientos_credito m
            LEFT JOIN fundaciones f ON f.id=m.fundacion_id
            WHERE (UPPER(COALESCE(m.tipo,''))='CONSUMO' OR COALESCE(m.creditos,0) < 0)
              AND substr(COALESCE(m.fecha_movimiento,''),1,10) >= ?
              AND substr(COALESCE(m.fecha_movimiento,''),1,10) < ? {scope}
            ORDER BY m.fecha_movimiento DESC
            LIMIT 200
            """,
            (inicio, fin) + params,
        ).fetchall()
        return [row_to_dict(r) for r in rows]

    def ultimos_ingresos(self, cur: Any, limit: int = 10) -> list[dict[str, Any]]:
        if not self.table_exists(cur, 'sesiones_usuario') or not self.table_exists(cur, 'usuarios_app'):
            return []
        scope, params = self.scope_sql('s')
        rows = cur.execute(
            f"""
            SELECT s.fecha_creacion, s.ip, u.username, u.nombre_completo, u.rol, f.nombre AS fundacion
            FROM sesiones_usuario s
            LEFT JOIN usuarios_app u ON u.id=s.usuario_id
            LEFT JOIN fundaciones f ON f.id=s.fundacion_id
            WHERE 1=1 {scope}
            ORDER BY s.fecha_creacion DESC
            LIMIT ?
            """,
            params + (limit,),
        ).fetchall()
        return [row_to_dict(r) for r in rows]

    def tickets_soporte(self, cur: Any, limit: int = 20) -> list[dict[str, Any]]:
        table = 'pc_tickets_soporte' if self.table_exists(cur, 'pc_tickets_soporte') else ('pc_tickets' if self.table_exists(cur, 'pc_tickets') else None)
        if not table:
            return []
        scope, params = self.scope_sql('t')
        cols = self.columns(cur, table)
        title_col = 'titulo' if 'titulo' in cols else 'asunto'
        modulo_col = 'modulo_origen' if 'modulo_origen' in cols else 'modulo'
        user_col = 'usuario_creador_id' if 'usuario_creador_id' in cols else 'creado_por_id'
        rows = cur.execute(
            f"""
            SELECT t.id, t.fundacion_id, f.nombre AS fundacion, t.{title_col} AS titulo,
                   t.categoria, t.prioridad, t.estado, t.{modulo_col} AS modulo,
                   t.fecha_creacion, t.fecha_actualizacion, u.username AS creado_por
            FROM {table} t
            LEFT JOIN fundaciones f ON f.id=t.fundacion_id
            LEFT JOIN usuarios_app u ON u.id=t.{user_col}
            WHERE 1=1 {scope}
            ORDER BY
              CASE UPPER(COALESCE(t.prioridad,'')) WHEN 'CRITICA' THEN 0 WHEN 'ALTA' THEN 1 WHEN 'MEDIA' THEN 2 ELSE 3 END,
              t.fecha_creacion DESC
            LIMIT ?
            """,
            params + (limit,),
        ).fetchall()
        return [row_to_dict(r) for r in rows]

    def alertas_pago(self, cur: Any) -> list[dict[str, Any]]:
        alerts = []
        # Usa tabla de alertas comercial si existe.
        if self.table_exists(cur, 'pc_alertas_pago'):
            scope, params = self.scope_sql('a')
            rows = cur.execute(
                f"""
                SELECT a.*, f.nombre AS fundacion
                FROM pc_alertas_pago a
                LEFT JOIN fundaciones f ON f.id=a.fundacion_id
                WHERE UPPER(COALESCE(a.estado,'ABIERTA')) NOT IN ('CERRADA','RESUELTA') {scope}
                ORDER BY a.fecha_creacion DESC
                LIMIT 50
                """,
                params,
            ).fetchall()
            alerts.extend([row_to_dict(r) for r in rows])
        # Genera alertas calculadas desde suscripción aunque no haya tabla.
        for item in self.licencias_fundaciones(cur):
            estado = str(item.get('estado_pago') or '').upper()
            if estado in {'VENCIDA', 'POR_VENCER'}:
                alerts.append({
                    'fundacion_id': item.get('fundacion_id'),
                    'fundacion': item.get('fundacion'),
                    'tipo': 'SUSCRIPCION_' + estado,
                    'nivel': 'ROJO' if estado == 'VENCIDA' else 'AMARILLO',
                    'mensaje': f"{item.get('fundacion')} tiene suscripción {estado.lower().replace('_', ' ')}. Vence: {item.get('fecha_vencimiento') or 'sin fecha'}.",
                    'fecha_creacion': now_iso(),
                })
            try:
                creditos = int(item.get('creditos_disponibles') or 0)
                if creditos <= 0:
                    alerts.append({
                        'fundacion_id': item.get('fundacion_id'),
                        'fundacion': item.get('fundacion'),
                        'tipo': 'CREDITOS_AGOTADOS',
                        'nivel': 'ROJO',
                        'mensaje': f"{item.get('fundacion')} no tiene créditos disponibles.",
                        'fecha_creacion': now_iso(),
                    })
                elif creditos <= 10:
                    alerts.append({
                        'fundacion_id': item.get('fundacion_id'),
                        'fundacion': item.get('fundacion'),
                        'tipo': 'CREDITOS_BAJOS',
                        'nivel': 'AMARILLO',
                        'mensaje': f"{item.get('fundacion')} tiene créditos bajos: {creditos}.",
                        'fecha_creacion': now_iso(),
                    })
            except Exception:
                pass
        return alerts[:50]

    def alertas_criticas(self, cur: Any, limit: int = 20) -> list[dict[str, Any]]:
        items: list[dict[str, Any]] = []
        if self.table_exists(cur, 'gp_alertas'):
            scope, params = self.scope_sql('a') if 'fundacion_id' in self.columns(cur, 'gp_alertas') else ('', ())
            rows = cur.execute(
                f"""
                SELECT a.id, a.fundacion_id, f.nombre AS fundacion, a.nivel, a.tipo, a.mensaje, a.fecha_alerta, a.fecha_creacion, 'Gestión Pedagógica' AS origen
                FROM gp_alertas a
                LEFT JOIN fundaciones f ON f.id=a.fundacion_id
                WHERE UPPER(COALESCE(a.nivel,'')) IN ({','.join(['?'] * len(NIVELES_CRITICOS))})
                  AND COALESCE(a.leida,0)=0 {scope}
                ORDER BY COALESCE(a.fecha_alerta,a.fecha_creacion) DESC
                LIMIT ?
                """,
                tuple(NIVELES_CRITICOS) + params + (limit,),
            ).fetchall()
            items.extend([row_to_dict(r) for r in rows])
        if self.table_exists(cur, 'sn_alertas'):
            scope, params = self.scope_sql('a') if 'fundacion_id' in self.columns(cur, 'sn_alertas') else ('', ())
            rows = cur.execute(
                f"""
                SELECT a.id, a.fundacion_id, f.nombre AS fundacion, a.nivel, a.tipo, a.mensaje, a.unidad, a.fecha_alerta, a.fecha_creacion, 'Salud y Nutrición' AS origen
                FROM sn_alertas a
                LEFT JOIN fundaciones f ON f.id=a.fundacion_id
                WHERE UPPER(COALESCE(a.nivel,'')) IN ({','.join(['?'] * len(NIVELES_CRITICOS))})
                  AND COALESCE(a.atendida,0)=0 {scope}
                ORDER BY COALESCE(a.fecha_alerta,a.fecha_creacion) DESC
                LIMIT ?
                """,
                tuple(NIVELES_CRITICOS) + params + (limit,),
            ).fetchall()
            items.extend([row_to_dict(r) for r in rows])
        if self.table_exists(cur, 'alertas'):
            cols = self.columns(cur, 'alertas')
            tipo_col = 'tipo' if 'tipo' in cols else ('tipo_alerta' if 'tipo_alerta' in cols else None)
            mensaje_col = 'mensaje' if 'mensaje' in cols else ('descripcion' if 'descripcion' in cols else ('detalles' if 'detalles' in cols else None))
            fecha_col = 'fecha_creacion' if 'fecha_creacion' in cols else ('fecha_generacion' if 'fecha_generacion' in cols else 'id')
            if tipo_col and mensaje_col and 'nivel' in cols:
                scope = ''
                params: tuple[Any, ...] = ()
                if not self.is_superadmin() and 'fundacion_id' in cols:
                    scope = ' AND a.fundacion_id=?'
                    params = (self.current_fundacion_id(),)
                rows = cur.execute(
                    f"""
                    SELECT a.id, f.nombre AS fundacion, a.nivel, a.{tipo_col} AS tipo,
                           a.{mensaje_col} AS mensaje, a.{fecha_col} AS fecha_creacion, 'Sistema' AS origen
                    FROM alertas a
                    LEFT JOIN fundaciones f ON f.id=a.fundacion_id
                    WHERE UPPER(COALESCE(a.nivel,'')) IN ({','.join(['?'] * len(NIVELES_CRITICOS))}) {scope}
                    ORDER BY a.{fecha_col} DESC
                    LIMIT ?
                    """,
                    tuple(NIVELES_CRITICOS) + params + (limit,),
                ).fetchall()
                items.extend([row_to_dict(r) for r in rows])
        items.sort(key=lambda x: str(x.get('fecha_creacion') or x.get('fecha_alerta') or ''), reverse=True)
        return items[:limit]

    def casos_nutricion_riesgo(self, cur: Any, limit: int = 25) -> list[dict[str, Any]]:
        if not self.table_exists(cur, 'sn_valoraciones'):
            return []
        scope, params = self.scope_sql('v') if 'fundacion_id' in self.columns(cur, 'sn_valoraciones') else ('', ())
        levels = tuple(NIVELES_RIESGO_NUTRICION)
        rows = cur.execute(
            f"""
            SELECT v.id, v.fundacion_id, f.nombre AS fundacion, v.documento, v.nombre_completo,
                   v.unidad, v.docente, v.edad_texto, v.sexo, v.peso_kg, v.talla_cm,
                   v.diagnostico_global, v.nivel_alerta, v.fecha_valoracion, v.proximo_control
            FROM sn_valoraciones v
            LEFT JOIN fundaciones f ON f.id=v.fundacion_id
            WHERE (UPPER(COALESCE(v.nivel_alerta,'')) IN ({','.join(['?'] * len(levels))})
               OR UPPER(COALESCE(v.diagnostico_global,'')) LIKE '%RIESGO%'
               OR UPPER(COALESCE(v.diagnostico_global,'')) LIKE '%DESNUTRIC%') {scope}
            ORDER BY COALESCE(v.fecha_valoracion, v.fecha_carga) DESC
            LIMIT ?
            """,
            levels + params + (limit,),
        ).fetchall()
        return [row_to_dict(r) for r in rows]

    def coordinadores_bajo_cumplimiento(self, cur: Any, periodo: str, limit: int = 25) -> list[dict[str, Any]]:
        result: dict[str, dict[str, Any]] = {}
        if self.table_exists(cur, 'gp_estado_cumplimiento'):
            scope, params = self.scope_sql('c') if 'fundacion_id' in self.columns(cur, 'gp_estado_cumplimiento') else ('', ())
            cols = self.columns(cur, 'gp_estado_cumplimiento')
            porcentaje_col = 'porcentaje_cumplimiento' if 'porcentaje_cumplimiento' in cols else ('cumplimiento' if 'cumplimiento' in cols else None)
            if porcentaje_col:
                rows = cur.execute(
                    f"""
                    SELECT c.*, co.nombre AS coordinador, f.nombre AS fundacion
                    FROM gp_estado_cumplimiento c
                    LEFT JOIN gp_coordinadores co ON co.id=c.coordinador_id
                    LEFT JOIN fundaciones f ON f.id=c.fundacion_id
                    WHERE COALESCE(c.{porcentaje_col},0) < 80
                      AND (COALESCE(c.periodo,'')=? OR substr(COALESCE(c.fecha_creacion,''),1,7)=?) {scope}
                    ORDER BY COALESCE(c.{porcentaje_col},0) ASC
                    LIMIT ?
                    """,
                    (periodo, periodo) + params + (limit,),
                ).fetchall()
                for r in rows:
                    d = row_to_dict(r) or {}
                    key = str(d.get('coordinador_id') or d.get('coordinador') or d.get('id'))
                    result[key] = {
                        'coordinador_id': d.get('coordinador_id'),
                        'coordinador': d.get('coordinador') or d.get('responsable') or 'Sin coordinador',
                        'fundacion': d.get('fundacion'),
                        'cumplimiento': d.get(porcentaje_col),
                        'pendientes': d.get('pendientes') or 0,
                        'vencidos': d.get('vencidos') or 0,
                        'origen': 'Estado de cumplimiento',
                    }
        if self.table_exists(cur, 'gp_entregables'):
            scope, params = self.scope_sql('e') if 'fundacion_id' in self.columns(cur, 'gp_entregables') else ('', ())
            rows = cur.execute(
                f"""
                SELECT e.coordinador_id, COALESCE(co.nombre, e.responsable, 'Sin coordinador') AS coordinador,
                       f.nombre AS fundacion,
                       COUNT(*) AS pendientes,
                       SUM(CASE WHEN date(COALESCE(e.fecha_limite,'2999-12-31')) < date('now') THEN 1 ELSE 0 END) AS vencidos
                FROM gp_entregables e
                LEFT JOIN gp_coordinadores co ON co.id=e.coordinador_id
                LEFT JOIN fundaciones f ON f.id=e.fundacion_id
                WHERE UPPER(COALESCE(e.estado,'PENDIENTE')) NOT IN ('APROBADO','CARGADO','CUMPLIDO','CERRADO')
                  AND (COALESCE(e.periodo,'')=? OR substr(COALESCE(e.fecha_limite,''),1,7)=?) {scope}
                GROUP BY e.coordinador_id, coordinador, f.nombre
                HAVING pendientes > 0
                ORDER BY vencidos DESC, pendientes DESC
                LIMIT ?
                """,
                (periodo, periodo) + params + (limit,),
            ).fetchall()
            for r in rows:
                d = row_to_dict(r) or {}
                key = str(d.get('coordinador_id') or d.get('coordinador'))
                current = result.get(key, {})
                current.update({
                    'coordinador_id': d.get('coordinador_id'),
                    'coordinador': d.get('coordinador') or current.get('coordinador') or 'Sin coordinador',
                    'fundacion': d.get('fundacion') or current.get('fundacion'),
                    'pendientes': int(d.get('pendientes') or 0),
                    'vencidos': int(d.get('vencidos') or 0),
                    'cumplimiento': current.get('cumplimiento', max(0, 100 - int(d.get('pendientes') or 0) * 10)),
                    'origen': 'Entregables pendientes',
                })
                result[key] = current
        return list(result.values())[:limit]

    def unidades_cobertura_incompleta(self, cur: Any, limit: int = 50) -> list[dict[str, Any]]:
        result = []
        if self.table_exists(cur, 'unidades'):
            scope, params = self.scope_sql('u') if 'fundacion_id' in self.columns(cur, 'unidades') else ('', ())
            rows = cur.execute(
                f"""
                SELECT u.id, u.fundacion_id, f.nombre AS fundacion, u.nombre AS unidad,
                       COALESCE(u.total_usuarios, 0) AS total_usuarios, COALESCE(u.total_gestantes, 0) AS total_gestantes,
                       u.fecha_actualizacion
                FROM unidades u
                LEFT JOIN fundaciones f ON f.id=u.fundacion_id
                WHERE COALESCE(u.total_usuarios,0) > 0 AND COALESCE(u.total_usuarios,0) < 20 {scope}
                ORDER BY COALESCE(u.total_usuarios,0) ASC, u.nombre
                LIMIT ?
                """,
                params + (limit,),
            ).fetchall()
            result.extend([row_to_dict(r) for r in rows])
        if not result and self.table_exists(cur, 'beneficiarios'):
            scope, params = self.scope_sql('b') if 'fundacion_id' in self.columns(cur, 'beneficiarios') else ('', ())
            rows = cur.execute(
                f"""
                SELECT b.unidad, b.fundacion_id, f.nombre AS fundacion, COUNT(*) AS total_usuarios
                FROM beneficiarios b
                LEFT JOIN fundaciones f ON f.id=b.fundacion_id
                WHERE UPPER(COALESCE(b.estado,'ACTIVO'))='ACTIVO' {scope}
                GROUP BY b.unidad, b.fundacion_id, f.nombre
                HAVING total_usuarios > 0 AND total_usuarios < 20
                ORDER BY total_usuarios ASC, b.unidad
                LIMIT ?
                """,
                params + (limit,),
            ).fetchall()
            result.extend([row_to_dict(r) for r in rows])
        return result[:limit]

    def ingresos_ultimos_meses(self, cur: Any, months: int = 6) -> list[dict[str, Any]]:
        if not self.table_exists(cur, 'pagos_suscripcion'):
            return []
        base = date.today().replace(day=1)
        data = []
        for i in range(months - 1, -1, -1):
            y = base.year
            m = base.month - i
            while m <= 0:
                y -= 1
                m += 12
            start = date(y, m, 1)
            end = date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)
            scope, params = self.scope_sql('p')
            total = self.sum_value(
                cur,
                f"""SELECT COALESCE(SUM(valor_pagado),0) FROM pagos_suscripcion p
                    WHERE substr(COALESCE(p.fecha_pago,p.fecha_creacion),1,10)>=?
                    AND substr(COALESCE(p.fecha_pago,p.fecha_creacion),1,10)<? {scope}""",
                (start.isoformat(), end.isoformat()) + params,
            )
            data.append({'periodo': f'{y}-{m:02d}', 'total': total})
        return data

    def creditos_ultimos_meses(self, cur: Any, months: int = 6) -> list[dict[str, Any]]:
        if not self.table_exists(cur, 'movimientos_credito'):
            return []
        base = date.today().replace(day=1)
        data = []
        for i in range(months - 1, -1, -1):
            y = base.year
            m = base.month - i
            while m <= 0:
                y -= 1
                m += 12
            start = date(y, m, 1)
            end = date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)
            scope, params = self.scope_sql('m')
            total = abs(self.sum_value(
                cur,
                f"""SELECT COALESCE(SUM(creditos),0) FROM movimientos_credito m
                    WHERE (UPPER(COALESCE(m.tipo,''))='CONSUMO' OR COALESCE(m.creditos,0)<0)
                    AND substr(COALESCE(m.fecha_movimiento,''),1,10)>=?
                    AND substr(COALESCE(m.fecha_movimiento,''),1,10)<? {scope}""",
                (start.isoformat(), end.isoformat()) + params,
            ))
            data.append({'periodo': f'{y}-{m:02d}', 'creditos': int(total)})
        return data

    def resumen_ejecutivo(self, data: dict[str, Any]) -> str:
        ind = data.get('indicadores', {})
        partes = []
        partes.append(f"El tablero consolida {ind.get('fundaciones_activas', 0)} fundación(es) activa(s) y {ind.get('fundaciones_vencidas', 0)} fundación(es) vencida(s) o suspendida(s).")
        partes.append(f"En el periodo se registran ingresos por {ind.get('ingresos_mes', 0):,.0f} y consumo de {ind.get('creditos_consumidos', 0)} crédito(s).")
        partes.append(f"La operación reporta {ind.get('entregables_pendientes', 0)} entregable(s) pendiente(s), {ind.get('alertas_criticas', 0)} alerta(s) crítica(s) y {ind.get('casos_nutricionales_riesgo', 0)} caso(s) nutricional(es) en riesgo.")
        partes.append(f"Se identifican {ind.get('unidades_cobertura_incompleta', 0)} unidad(es) con cobertura incompleta y {ind.get('coordinadores_bajo_cumplimiento', 0)} coordinador(es) con bajo cumplimiento o pendientes abiertos.")
        return ' '.join(partes)

    def export_excel(self, anio: int | None = None, mes: int | None = None) -> bytes:
        data = self.dashboard(anio, mes)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Indicadores'
        header_fill = PatternFill('solid', fgColor='1F2937')
        header_font = Font(color='FFFFFF', bold=True)
        ws.append(['Indicador', 'Valor'])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for key, value in data['indicadores'].items():
            ws.append([key.replace('_', ' ').title(), value])
        ws.append([])
        ws.append(['Resumen Ejecutivo', data['resumen_ejecutivo']])
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 70

        def add_sheet(name: str, rows: list[dict[str, Any]], columns: list[tuple[str, str]]):
            sheet = wb.create_sheet(name[:31])
            sheet.append([label for _, label in columns])
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            for row in rows:
                sheet.append([row.get(key, '') for key, _ in columns])
            for idx, (_, label) in enumerate(columns, start=1):
                sheet.column_dimensions[chr(64 + idx) if idx <= 26 else 'Z'].width = max(14, min(40, len(label) + 6))

        add_sheet('Licencias', data['licencias'], [
            ('fundacion', 'Fundación'), ('nit', 'NIT'), ('estado_fundacion', 'Estado fundación'),
            ('estado_suscripcion', 'Estado suscripción'), ('estado_pago', 'Estado pago'), ('plan', 'Plan'),
            ('fecha_inicio', 'Inicio'), ('fecha_vencimiento', 'Vencimiento'), ('dias_gracia', 'Días gracia'),
            ('usuarios_permitidos', 'Usuarios permitidos'), ('usuarios_activos', 'Usuarios activos'),
            ('creditos_disponibles', 'Créditos disponibles'), ('ultimo_ingreso', 'Último ingreso')
        ])
        add_sheet('Alertas Pago', data['alertas_pago'], [('fundacion', 'Fundación'), ('tipo', 'Tipo'), ('nivel', 'Nivel'), ('mensaje', 'Mensaje'), ('fecha_creacion', 'Fecha')])
        add_sheet('Alertas Criticas', data['alertas_criticas'], [('fundacion', 'Fundación'), ('origen', 'Origen'), ('nivel', 'Nivel'), ('tipo', 'Tipo'), ('mensaje', 'Mensaje'), ('fecha_creacion', 'Fecha')])
        add_sheet('Nutricion Riesgo', data['nutricion_riesgo'], [('fundacion', 'Fundación'), ('unidad', 'Unidad'), ('nombre_completo', 'Niño'), ('documento', 'Documento'), ('diagnostico_global', 'Diagnóstico'), ('nivel_alerta', 'Alerta'), ('fecha_valoracion', 'Fecha')])
        add_sheet('Unidades Incompletas', data['unidades_cobertura_incompleta'], [('fundacion', 'Fundación'), ('unidad', 'Unidad'), ('total_usuarios', 'Usuarios'), ('total_gestantes', 'Gestantes'), ('fecha_actualizacion', 'Actualización')])
        add_sheet('Coordinadores Bajo Cumplimiento', data['coordinadores_bajo_cumplimiento'], [('fundacion', 'Fundación'), ('coordinador', 'Coordinador'), ('cumplimiento', 'Cumplimiento'), ('pendientes', 'Pendientes'), ('vencidos', 'Vencidos'), ('origen', 'Origen')])
        add_sheet('Tickets', data['tickets_soporte'], [('fundacion', 'Fundación'), ('titulo', 'Ticket'), ('categoria', 'Categoría'), ('prioridad', 'Prioridad'), ('estado', 'Estado'), ('fecha_creacion', 'Fecha')])

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        self.log('EXPORTAR_GERENCIA_EXCEL', 'Exportación Excel Gerencia General', {'periodo': data['periodo']})
        return out.getvalue()

    def export_pdf(self, anio: int | None = None, mes: int | None = None) -> bytes:
        data = self.dashboard(anio, mes)
        out = io.BytesIO()
        doc = SimpleDocTemplate(out, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph('Tablero de Gerencia General - PrimeraInfancia', styles['Title']))
        story.append(Paragraph(f"Periodo: {data['periodo']['periodo']}", styles['Normal']))
        story.append(Spacer(1, 10))
        story.append(Paragraph('Resumen ejecutivo', styles['Heading2']))
        story.append(Paragraph(data['resumen_ejecutivo'], styles['BodyText']))
        story.append(Spacer(1, 10))
        story.append(Paragraph('Indicadores principales', styles['Heading2']))
        indicadores = [['Indicador', 'Valor']] + [[k.replace('_', ' ').title(), str(v)] for k, v in data['indicadores'].items()]
        t = Table(indicadores, colWidths=[260, 140])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(t)
        story.append(Spacer(1, 12))

        def add_table(title: str, rows: list[dict[str, Any]], cols: list[tuple[str, str]], max_rows: int = 12):
            story.append(Paragraph(title, styles['Heading2']))
            data_rows = [[label for _, label in cols]]
            for row in rows[:max_rows]:
                data_rows.append([str(row.get(key, '') or '')[:70] for key, _ in cols])
            if len(data_rows) == 1:
                data_rows.append(['Sin registros'] + [''] * (len(cols) - 1))
            table = Table(data_rows, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#334155')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(table)
            story.append(Spacer(1, 10))

        add_table('Licencias por fundación', data['licencias'], [('fundacion', 'Fundación'), ('estado_pago', 'Pago'), ('plan', 'Plan'), ('fecha_vencimiento', 'Vence'), ('creditos_disponibles', 'Créditos'), ('usuarios_activos', 'Usuarios')])
        add_table('Alertas de pago', data['alertas_pago'], [('fundacion', 'Fundación'), ('nivel', 'Nivel'), ('tipo', 'Tipo'), ('mensaje', 'Mensaje')])
        add_table('Unidades con cobertura incompleta', data['unidades_cobertura_incompleta'], [('fundacion', 'Fundación'), ('unidad', 'Unidad'), ('total_usuarios', 'Usuarios'), ('total_gestantes', 'Gestantes')])
        add_table('Casos nutricionales en riesgo', data['nutricion_riesgo'], [('unidad', 'Unidad'), ('nombre_completo', 'Niño'), ('documento', 'Documento'), ('diagnostico_global', 'Diagnóstico'), ('nivel_alerta', 'Alerta')])
        doc.build(story)
        self.log('EXPORTAR_GERENCIA_PDF', 'Exportación PDF Gerencia General', {'periodo': data['periodo']})
        return out.getvalue()
