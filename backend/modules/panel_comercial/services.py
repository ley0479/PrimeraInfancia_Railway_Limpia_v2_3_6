from __future__ import annotations

import io
import json
import os
from datetime import date, datetime
from typing import Any, Iterable

from flask import g, request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .schema import PANEL_COMERCIAL_SCHEMA_SQL, TICKET_ESTADOS, TICKET_PRIORIDADES, TICKET_CATEGORIAS
from modules.sqlalchemy_compat import CoreConnection, CoreCompatRepository


def now_iso() -> str:
    return datetime.now().isoformat(timespec='seconds')


def today_month() -> str:
    return date.today().strftime('%Y-%m')


_CORE_COMPAT = CoreCompatRepository()


def connect(database_path: str) -> CoreConnection:
    return CoreConnection()


def table_exists(cur: object, table: str) -> bool:
    return _CORE_COMPAT.table_exists(table)


def parse_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except Exception:
        return None


def row_dict(row: dict[str, Any] | None) -> dict[str, Any]:
    return dict(row) if row else {}


def rows_list(rows: Iterable[Any]) -> list[dict[str, Any]]:
    return [dict(r) for r in rows]


class PanelComercialService:
    """Panel comercial y soporte para controlar fundaciones/clientes.

    Es un módulo de lectura y soporte que consume las tablas ya existentes de
    fundaciones, usuarios, facturación, suscripciones, pagos y movimientos de
    crédito. No reemplaza el módulo de Suscripción y Pagos.
    """

    def __init__(self, database_path: str):
        self.database_path = database_path

    def init_schema(self) -> None:
        conn = connect(self.database_path)
        try:
            conn.executescript(PANEL_COMERCIAL_SCHEMA_SQL)
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def context(self) -> dict[str, Any]:
        user = getattr(g, 'current_user', None) or {}
        return {
            'usuario_id': user.get('id'),
            'username': user.get('username') or 'sistema',
            'rol': user.get('rol') or 'SUPERADMIN',
            'fundacion_id': int(user.get('fundacion_id') or 1),
        }

    def is_superadmin(self) -> bool:
        return self.context().get('rol') == 'SUPERADMIN'

    def is_manager_scope(self) -> bool:
        return self.context().get('rol') in {'SUPERADMIN', 'GERENTE'}

    def scope_clause(self, alias: str = '') -> tuple[str, list[Any]]:
        if self.is_superadmin():
            return '1=1', []
        field = f'{alias}.fundacion_id' if alias else 'fundacion_id'
        return f'{field} = ?', [self.context()['fundacion_id']]

    def fetch_all(self, sql: str, params: Iterable[Any] = ()) -> list[dict[str, Any]]:
        conn = connect(self.database_path)
        rows = rows_list(conn.execute(sql, tuple(params)).fetchall())
        conn.close()
        return rows

    def fetch_one(self, sql: str, params: Iterable[Any] = ()) -> dict[str, Any]:
        conn = connect(self.database_path)
        row = conn.execute(sql, tuple(params)).fetchone()
        conn.close()
        return row_dict(row)

    def execute(self, sql: str, params: Iterable[Any] = ()) -> int:
        conn = connect(self.database_path)
        cur = conn.cursor()
        cur.execute(sql, tuple(params))
        conn.commit()
        last_id = int(cur.lastrowid or 0)
        conn.close()
        return last_id

    def execute_update(self, sql: str, params: Iterable[Any] = ()) -> None:
        conn = connect(self.database_path)
        conn.execute(sql, tuple(params))
        conn.commit()
        conn.close()

    def log(self, accion: str, tabla: str | None = None, registro_id: int | None = None, antes: Any = None, despues: Any = None) -> None:
        ctx = self.context()
        try:
            self.execute(
                """
                INSERT INTO pc_auditoria
                (usuario_id, username, fundacion_id, accion, tabla_afectada, registro_id,
                 datos_anteriores, datos_nuevos, ip, user_agent, fecha)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    ctx.get('usuario_id'), ctx.get('username'), ctx.get('fundacion_id'), accion,
                    tabla, registro_id,
                    json.dumps(antes, ensure_ascii=False) if antes is not None else None,
                    json.dumps(despues, ensure_ascii=False) if despues is not None else None,
                    request.remote_addr if request else None,
                    request.headers.get('User-Agent') if request else None,
                    now_iso(),
                ),
            )
        except Exception:
            pass

    def _count_scalar(self, sql: str, params: Iterable[Any] = ()) -> int:
        try:
            row = self.fetch_one(sql, params)
            return int(row.get('total') or 0)
        except Exception:
            return 0

    def _sum_scalar(self, sql: str, params: Iterable[Any] = ()) -> float:
        try:
            row = self.fetch_one(sql, params)
            return float(row.get('total') or 0)
        except Exception:
            return 0.0

    def dashboard(self) -> dict[str, Any]:
        scope_f, params_f = self.scope_clause('f')
        scope_s, params_s = self.scope_clause('s')
        scope_p, params_p = self.scope_clause('p')
        scope_m, params_m = self.scope_clause('m')
        scope_u, params_u = self.scope_clause('u')
        scope_t, params_t = self.scope_clause('t')
        mes = today_month()

        stats = {
            'fundaciones_activas': self._count_scalar(f"SELECT COUNT(*) AS total FROM fundaciones f WHERE {scope_f} AND UPPER(COALESCE(f.estado,'ACTIVA'))='ACTIVA'", params_f),
            'fundaciones_vencidas': self._count_scalar(f"SELECT COUNT(*) AS total FROM suscripciones_fundacion s WHERE {scope_s} AND UPPER(COALESCE(s.estado,'')) IN ('VENCIDA','SUSPENDIDA','CANCELADA')", params_s),
            'fundaciones_por_vencer': self._count_scalar(f"SELECT COUNT(*) AS total FROM suscripciones_fundacion s WHERE {scope_s} AND UPPER(COALESCE(s.estado,''))='POR_VENCER'", params_s),
            'ingresos_mes': self._sum_scalar(f"SELECT COALESCE(SUM(p.valor_pagado),0) AS total FROM pagos_suscripcion p WHERE {scope_p} AND substr(COALESCE(p.fecha_pago,''),1,7)=?", params_p + [mes]),
            'creditos_consumidos_mes': self._sum_scalar(f"SELECT COALESCE(SUM(CASE WHEN UPPER(m.tipo)='CONSUMO' THEN ABS(m.creditos) ELSE 0 END),0) AS total FROM movimientos_credito m WHERE {scope_m} AND substr(COALESCE(m.fecha_movimiento,''),1,7)=?", params_m + [mes]),
            'usuarios_activos': self._count_scalar(f"SELECT COUNT(*) AS total FROM usuarios_app u WHERE {scope_u} AND COALESCE(u.activo,1)=1 AND UPPER(COALESCE(u.estado,'ACTIVO'))='ACTIVO'", params_u),
            'tickets_abiertos': self._count_scalar(f"SELECT COUNT(*) AS total FROM pc_tickets_soporte t WHERE {scope_t} AND UPPER(COALESCE(t.estado,'ABIERTO')) NOT IN ('RESUELTO','CERRADO','ANULADO')", params_t),
        }
        ultimo_ingreso = self.ultimo_ingreso()
        tickets_resumen = self.tickets_resumen()
        alertas_pago = self.alertas_pago()
        estado_suscripciones = self.estado_suscripciones(limit=50)
        consumo_creditos = self.consumo_creditos(limit=50)
        ingresos_recientes = self.ingresos_recientes(limit=30)
        fundaciones = self.fundaciones_resumen(limit=100)

        return {
            'periodo': mes,
            'stats': stats,
            'ultimo_ingreso': ultimo_ingreso,
            'tickets_resumen': tickets_resumen,
            'alertas_pago': alertas_pago,
            'estado_suscripciones': estado_suscripciones,
            'consumo_creditos': consumo_creditos,
            'ingresos_recientes': ingresos_recientes,
            'fundaciones': fundaciones,
            'catalogos': self.catalogos(),
        }

    def ultimo_ingreso(self) -> dict[str, Any]:
        scope_u, params = self.scope_clause('u')
        rows = self.fetch_all(
            f"""
            SELECT u.id, u.username, u.nombre_completo, u.rol, u.fecha_ultima_conexion,
                   f.nombre AS fundacion_nombre, u.fundacion_id
            FROM usuarios_app u
            LEFT JOIN fundaciones f ON f.id=u.fundacion_id
            WHERE {scope_u} AND u.fecha_ultima_conexion IS NOT NULL
            ORDER BY u.fecha_ultima_conexion DESC LIMIT 1
            """,
            params,
        )
        return rows[0] if rows else {}

    def tickets_resumen(self) -> dict[str, Any]:
        scope, params = self.scope_clause('t')
        rows = self.fetch_all(
            f"""
            SELECT UPPER(COALESCE(t.estado,'ABIERTO')) AS estado, COUNT(*) AS total
            FROM pc_tickets_soporte t WHERE {scope}
            GROUP BY UPPER(COALESCE(t.estado,'ABIERTO'))
            """,
            params,
        )
        prioridad = self.fetch_all(
            f"""
            SELECT UPPER(COALESCE(t.prioridad,'MEDIA')) AS prioridad, COUNT(*) AS total
            FROM pc_tickets_soporte t WHERE {scope}
            GROUP BY UPPER(COALESCE(t.prioridad,'MEDIA'))
            """,
            params,
        )
        return {'por_estado': rows, 'por_prioridad': prioridad}

    def alertas_pago(self) -> list[dict[str, Any]]:
        scope, params = self.scope_clause('s')
        rows = self.fetch_all(
            f"""
            SELECT s.*, f.nombre AS fundacion_nombre, p.nombre AS plan_nombre
            FROM suscripciones_fundacion s
            LEFT JOIN fundaciones f ON f.id=s.fundacion_id
            LEFT JOIN planes_suscripcion p ON p.id=s.plan_id
            WHERE {scope}
            ORDER BY s.fecha_vencimiento ASC
            """,
            params,
        )
        alertas: list[dict[str, Any]] = []
        hoy = date.today()
        for row in rows:
            venc = parse_date(row.get('fecha_vencimiento'))
            dias = (venc - hoy).days if venc else 0
            estado = str(row.get('estado') or '').upper()
            nombre = row.get('fundacion_nombre') or f"Fundación {row.get('fundacion_id')}"
            creditos = int(row.get('creditos_disponibles') or 0)
            if estado in {'VENCIDA', 'SUSPENDIDA', 'CANCELADA'} or (venc and dias < 0):
                alertas.append({
                    'nivel': 'ROJO', 'tipo': 'SUSCRIPCION_VENCIDA', 'fundacion_id': row.get('fundacion_id'),
                    'fundacion_nombre': nombre, 'mensaje': f'{nombre} tiene suscripción {estado or "vencida"}.',
                    'fecha_vencimiento': row.get('fecha_vencimiento'), 'dias_restantes': dias,
                })
            elif estado == 'POR_VENCER' or (venc and 0 <= dias <= 5):
                alertas.append({
                    'nivel': 'AMARILLO', 'tipo': 'SUSCRIPCION_POR_VENCER', 'fundacion_id': row.get('fundacion_id'),
                    'fundacion_nombre': nombre, 'mensaje': f'{nombre} vence en {dias} día(s).',
                    'fecha_vencimiento': row.get('fecha_vencimiento'), 'dias_restantes': dias,
                })
            if creditos <= 0:
                alertas.append({
                    'nivel': 'ROJO', 'tipo': 'SIN_CREDITOS', 'fundacion_id': row.get('fundacion_id'),
                    'fundacion_nombre': nombre, 'mensaje': f'{nombre} no tiene créditos disponibles.',
                    'creditos_disponibles': creditos,
                })
            elif creditos <= 10:
                alertas.append({
                    'nivel': 'AMARILLO', 'tipo': 'CREDITOS_BAJOS', 'fundacion_id': row.get('fundacion_id'),
                    'fundacion_nombre': nombre, 'mensaje': f'{nombre} tiene créditos bajos ({creditos}).',
                    'creditos_disponibles': creditos,
                })
        return alertas[:100]

    def estado_suscripciones(self, limit: int = 100) -> list[dict[str, Any]]:
        scope, params = self.scope_clause('s')
        return self.fetch_all(
            f"""
            SELECT s.id, s.fundacion_id, f.nombre AS fundacion_nombre, f.estado AS fundacion_estado,
                   s.estado AS suscripcion_estado, s.fecha_inicio, s.fecha_vencimiento, s.creditos_disponibles,
                   s.dias_gracia, p.nombre AS plan_nombre, p.precio_mensual,
                   (SELECT COUNT(*) FROM usuarios_app u WHERE u.fundacion_id=s.fundacion_id AND COALESCE(u.activo,1)=1) AS usuarios_activos,
                   (SELECT MAX(u.fecha_ultima_conexion) FROM usuarios_app u WHERE u.fundacion_id=s.fundacion_id) AS ultimo_ingreso
            FROM suscripciones_fundacion s
            LEFT JOIN fundaciones f ON f.id=s.fundacion_id
            LEFT JOIN planes_suscripcion p ON p.id=s.plan_id
            WHERE {scope}
            ORDER BY CASE s.estado WHEN 'VENCIDA' THEN 1 WHEN 'SUSPENDIDA' THEN 2 WHEN 'POR_VENCER' THEN 3 ELSE 4 END,
                     s.fecha_vencimiento ASC
            LIMIT ?
            """,
            params + [limit],
        )

    def fundaciones_resumen(self, limit: int = 100) -> list[dict[str, Any]]:
        scope, params = self.scope_clause('f')
        return self.fetch_all(
            f"""
            SELECT f.id, f.nombre, f.nit, f.representante, f.email, f.telefono, f.estado,
                   s.estado AS suscripcion_estado, s.fecha_vencimiento, s.creditos_disponibles,
                   p.nombre AS plan_nombre,
                   (SELECT COUNT(*) FROM usuarios_app u WHERE u.fundacion_id=f.id AND COALESCE(u.activo,1)=1) AS usuarios_activos,
                   (SELECT MAX(u.fecha_ultima_conexion) FROM usuarios_app u WHERE u.fundacion_id=f.id) AS ultimo_ingreso,
                   (SELECT COUNT(*) FROM pc_tickets_soporte t WHERE t.fundacion_id=f.id AND UPPER(COALESCE(t.estado,'ABIERTO')) NOT IN ('RESUELTO','CERRADO','ANULADO')) AS tickets_abiertos
            FROM fundaciones f
            LEFT JOIN suscripciones_fundacion s ON s.fundacion_id=f.id
            LEFT JOIN planes_suscripcion p ON p.id=s.plan_id
            WHERE {scope}
            ORDER BY f.nombre
            LIMIT ?
            """,
            params + [limit],
        )

    def consumo_creditos(self, limit: int = 100) -> list[dict[str, Any]]:
        scope, params = self.scope_clause('m')
        return self.fetch_all(
            f"""
            SELECT m.*, f.nombre AS fundacion_nombre
            FROM movimientos_credito m
            LEFT JOIN fundaciones f ON f.id=m.fundacion_id
            WHERE {scope}
            ORDER BY m.fecha_movimiento DESC, m.id DESC LIMIT ?
            """,
            params + [limit],
        )

    def ingresos_recientes(self, limit: int = 100) -> list[dict[str, Any]]:
        scope, params = self.scope_clause('p')
        return self.fetch_all(
            f"""
            SELECT p.*, f.nombre AS fundacion_nombre, ps.nombre AS plan_nombre
            FROM pagos_suscripcion p
            LEFT JOIN fundaciones f ON f.id=p.fundacion_id
            LEFT JOIN planes_suscripcion ps ON ps.id=p.plan_id
            WHERE {scope}
            ORDER BY p.fecha_pago DESC, p.id DESC LIMIT ?
            """,
            params + [limit],
        )

    def list_tickets(self, estado: str | None = None, limit: int = 300) -> list[dict[str, Any]]:
        scope, params = self.scope_clause('t')
        extra = ''
        if estado:
            extra = ' AND UPPER(t.estado)=?'
            params.append(estado.upper())
        return self.fetch_all(
            f"""
            SELECT t.*, f.nombre AS fundacion_nombre, u.username AS usuario_creador,
                   ua.username AS usuario_asignado
            FROM pc_tickets_soporte t
            LEFT JOIN fundaciones f ON f.id=t.fundacion_id
            LEFT JOIN usuarios_app u ON u.id=t.usuario_creador_id
            LEFT JOIN usuarios_app ua ON ua.id=t.usuario_asignado_id
            WHERE {scope} {extra}
            ORDER BY CASE UPPER(t.prioridad) WHEN 'CRITICA' THEN 1 WHEN 'ALTA' THEN 2 WHEN 'MEDIA' THEN 3 ELSE 4 END,
                     CASE UPPER(t.estado) WHEN 'ABIERTO' THEN 1 WHEN 'EN_PROCESO' THEN 2 ELSE 3 END,
                     t.fecha_creacion DESC
            LIMIT ?
            """,
            params + [limit],
        )

    def get_ticket(self, ticket_id: int) -> dict[str, Any]:
        ticket = self.fetch_one("SELECT * FROM pc_tickets_soporte WHERE id=?", (ticket_id,))
        if not ticket:
            raise ValueError('Ticket no encontrado.')
        if not self.is_superadmin() and int(ticket.get('fundacion_id') or 0) != self.context()['fundacion_id']:
            raise PermissionError('No tienes permiso para este ticket.')
        comentarios = self.fetch_all(
            """
            SELECT c.*, u.username, u.nombre_completo
            FROM pc_ticket_comentarios c
            LEFT JOIN usuarios_app u ON u.id=c.usuario_id
            WHERE c.ticket_id=? AND COALESCE(c.fundacion_id, ?) = ?
            ORDER BY c.fecha_creacion ASC
            """,
            (ticket_id, int(ticket.get('fundacion_id') or 1), int(ticket.get('fundacion_id') or 1)),
        )
        ticket['comentarios'] = comentarios
        return ticket

    def save_ticket(self, data: dict[str, Any], ticket_id: int | None = None) -> dict[str, Any]:
        ctx = self.context()
        now = now_iso()
        titulo = (data.get('titulo') or '').strip()
        if not titulo:
            raise ValueError('Título del ticket requerido.')
        fundacion_id = int(data.get('fundacion_id') or ctx['fundacion_id']) if self.is_superadmin() else ctx['fundacion_id']
        estado = (data.get('estado') or 'ABIERTO').upper()
        if estado not in TICKET_ESTADOS:
            estado = 'ABIERTO'
        prioridad = (data.get('prioridad') or 'MEDIA').upper()
        if prioridad not in TICKET_PRIORIDADES:
            prioridad = 'MEDIA'
        payload = {
            'fundacion_id': fundacion_id,
            'titulo': titulo,
            'descripcion': data.get('descripcion') or '',
            'categoria': data.get('categoria') or 'Soporte general',
            'prioridad': prioridad,
            'estado': estado,
            'modulo_origen': data.get('modulo_origen') or '',
            'usuario_asignado_id': int(data.get('usuario_asignado_id') or 0) or None,
            'observaciones': data.get('observaciones') or '',
        }
        if ticket_id:
            before = self.get_ticket(ticket_id)
            if not self.is_superadmin() and before.get('estado') in {'RESUELTO', 'CERRADO'}:
                raise PermissionError('No se puede editar un ticket cerrado.')
            fecha_cierre = now if estado in {'RESUELTO', 'CERRADO', 'ANULADO'} else None
            self.execute_update(
                """
                UPDATE pc_tickets_soporte
                SET titulo=?, descripcion=?, categoria=?, prioridad=?, estado=?, modulo_origen=?,
                    usuario_asignado_id=?, observaciones=?, fecha_cierre=COALESCE(?, fecha_cierre), fecha_actualizacion=?
                WHERE id=?
                """,
                (payload['titulo'], payload['descripcion'], payload['categoria'], payload['prioridad'], payload['estado'],
                 payload['modulo_origen'], payload['usuario_asignado_id'], payload['observaciones'], fecha_cierre, now, ticket_id),
            )
            after = self.get_ticket(ticket_id)
            self.log('EDITAR_TICKET_SOPORTE', 'pc_tickets_soporte', ticket_id, before, after)
            return after
        new_id = self.execute(
            """
            INSERT INTO pc_tickets_soporte
            (fundacion_id, titulo, descripcion, categoria, prioridad, estado, modulo_origen,
             usuario_creador_id, usuario_asignado_id, observaciones, fecha_creacion, fecha_actualizacion)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (payload['fundacion_id'], payload['titulo'], payload['descripcion'], payload['categoria'], payload['prioridad'], payload['estado'],
             payload['modulo_origen'], ctx.get('usuario_id'), payload['usuario_asignado_id'], payload['observaciones'], now, now),
        )
        ticket = self.get_ticket(new_id)
        self.log('CREAR_TICKET_SOPORTE', 'pc_tickets_soporte', new_id, despues=ticket)
        return ticket

    def add_comment(self, ticket_id: int, comentario: str) -> dict[str, Any]:
        ticket = self.get_ticket(ticket_id)
        comentario = (comentario or '').strip()
        if not comentario:
            raise ValueError('Comentario requerido.')
        new_id = self.execute(
            """
            INSERT INTO pc_ticket_comentarios
            (ticket_id, fundacion_id, usuario_id, comentario, fecha_creacion)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                ticket_id,
                int(ticket.get('fundacion_id') or self.context().get('fundacion_id') or 1),
                self.context().get('usuario_id'),
                comentario,
                now_iso(),
            ),
        )
        self.execute_update("UPDATE pc_tickets_soporte SET fecha_actualizacion=? WHERE id=?", (now_iso(), ticket_id))
        row = self.fetch_one("SELECT * FROM pc_ticket_comentarios WHERE id=?", (new_id,))
        self.log('COMENTAR_TICKET_SOPORTE', 'pc_ticket_comentarios', new_id, despues=row)
        return row

    def catalogos(self) -> dict[str, Any]:
        return {
            'estados_ticket': TICKET_ESTADOS,
            'prioridades_ticket': TICKET_PRIORIDADES,
            'categorias_ticket': TICKET_CATEGORIAS,
        }

    def export_excel(self) -> bytes:
        data = self.dashboard()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Panel comercial'
        title_fill = PatternFill('solid', fgColor='1F2937')
        head_fill = PatternFill('solid', fgColor='E5E7EB')
        white = Font(color='FFFFFF', bold=True)
        bold = Font(bold=True)
        ws['A1'] = 'PANEL COMERCIAL - PRIMERA INFANCIA'
        ws['A1'].font = white
        ws['A1'].fill = title_fill
        ws.merge_cells('A1:H1')
        row = 3
        ws.cell(row, 1, 'Indicador').font = bold
        ws.cell(row, 2, 'Valor').font = bold
        ws.cell(row, 1).fill = head_fill
        ws.cell(row, 2).fill = head_fill
        for key, value in data.get('stats', {}).items():
            row += 1
            ws.cell(row, 1, key.replace('_', ' ').title())
            ws.cell(row, 2, value)
        row += 3
        ws.cell(row, 1, 'Fundaciones / Suscripciones').font = bold
        row += 1
        headers = ['Fundación', 'Estado fundación', 'Estado suscripción', 'Plan', 'Vencimiento', 'Créditos', 'Usuarios activos', 'Último ingreso', 'Tickets abiertos']
        for col, h in enumerate(headers, 1):
            ws.cell(row, col, h).font = bold
            ws.cell(row, col).fill = head_fill
        for f in data.get('fundaciones', []):
            row += 1
            values = [f.get('nombre'), f.get('estado'), f.get('suscripcion_estado'), f.get('plan_nombre'), f.get('fecha_vencimiento'), f.get('creditos_disponibles'), f.get('usuarios_activos'), f.get('ultimo_ingreso'), f.get('tickets_abiertos')]
            for col, value in enumerate(values, 1):
                ws.cell(row, col, value)
        row += 3
        ws.cell(row, 1, 'Alertas de pago').font = bold
        row += 1
        headers = ['Nivel', 'Tipo', 'Fundación', 'Mensaje', 'Vencimiento', 'Días', 'Créditos']
        for col, h in enumerate(headers, 1):
            ws.cell(row, col, h).font = bold
            ws.cell(row, col).fill = head_fill
        for a in data.get('alertas_pago', []):
            row += 1
            values = [a.get('nivel'), a.get('tipo'), a.get('fundacion_nombre'), a.get('mensaje'), a.get('fecha_vencimiento'), a.get('dias_restantes'), a.get('creditos_disponibles')]
            for col, value in enumerate(values, 1):
                ws.cell(row, col, value)
        for column in ws.columns:
            max_len = 12
            letter = column[0].column_letter
            for cell in column:
                max_len = max(max_len, len(str(cell.value or '')) + 2)
            ws.column_dimensions[letter].width = min(max_len, 45)
        ws.freeze_panes = 'A4'
        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()
