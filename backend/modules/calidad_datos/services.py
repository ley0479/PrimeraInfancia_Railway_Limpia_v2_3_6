from __future__ import annotations

import csv
import io
import json
import math
import os
import re
import unicodedata
from datetime import datetime, timedelta
from typing import Any

from services.uds_catalog import normalize_unit as catalog_normalize_unit

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


SEVERITY_ORDER = {
    'CRITICA': 0,
    'ALTA': 1,
    'MEDIA': 2,
    'BAJA': 3,
}

TIPOS_HALLAZGO = {
    'DOCUMENTO_VACIO': 'Documentos vacíos',
    'DOCUMENTO_DUPLICADO': 'Documentos duplicados',
    'NINO_SIN_UNIDAD': 'Niños sin unidad',
    'NINO_SIN_ACUDIENTE': 'Niños sin acudiente',
    'NINO_SIN_TELEFONO': 'Niños sin teléfono',
    'NINO_SIN_FECHA_NACIMIENTO': 'Niños sin fecha de nacimiento',
    'UNIDAD_SIN_DOCENTE': 'Unidades sin docente',
    'DOCENTE_SIN_UNIDAD': 'Docentes sin unidad',
    'TALENTO_DUPLICADO': 'Talento humano duplicado',
    'EDAD_INCONSISTENTE': 'Edades inconsistentes',
    'BENEFICIARIO_FUERA_RANGO': 'Beneficiarios fuera de rango',
    'COLUMNA_OBLIGATORIA_FALTANTE': 'Columnas obligatorias faltantes',
}


def normalizar_texto(valor: Any) -> str:
    texto = str(valor or '').strip().lower()
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r'[^a-z0-9]+', ' ', texto)
    return ' '.join(texto.split())


def limpiar_valor(valor: Any) -> str:
    if valor is None:
        return ''
    try:
        if pd.isna(valor):
            return ''
    except Exception:
        pass
    texto = str(valor).strip()
    if texto.lower() in {'nan', 'nat', 'none', 'null', 'sin dato', 's/d'}:
        return ''
    return texto


def valor_upper(valor: Any) -> str:
    return limpiar_valor(valor).upper()


def buscar_columna(df: pd.DataFrame, alias: list[str]) -> str | None:
    columnas = [(normalizar_texto(col), col) for col in df.columns]
    alias_norm = [normalizar_texto(a) for a in alias]
    for objetivo in alias_norm:
        for normalizada, original in columnas:
            if normalizada == objetivo:
                return original
    for objetivo in alias_norm:
        for normalizada, original in columnas:
            if objetivo and objetivo in normalizada:
                return original
    return None


def serie(df: pd.DataFrame, columna: str | None, default: str = '') -> pd.Series:
    if columna and columna in df.columns:
        return df[columna].apply(limpiar_valor)
    return pd.Series([default] * len(df), index=df.index)


def parse_fecha(valor: Any) -> datetime | None:
    if valor is None:
        return None
    try:
        if pd.isna(valor):
            return None
    except Exception:
        pass
    if isinstance(valor, datetime):
        return valor
    texto = limpiar_valor(valor)
    if not texto:
        return None
    if re.fullmatch(r'\d+(\.\d+)?', texto):
        try:
            numero = float(texto)
            if numero > 20000:
                return pd.to_datetime(numero, unit='D', origin='1899-12-30').to_pydatetime()
        except Exception:
            pass
    for dayfirst in (True, False):
        try:
            fecha = pd.to_datetime(texto, errors='coerce', dayfirst=dayfirst)
            if pd.notna(fecha):
                return fecha.to_pydatetime()
        except Exception:
            pass
    return None


def edad_meses(fecha_nacimiento: Any, ref: datetime | None = None) -> int | None:
    fecha = parse_fecha(fecha_nacimiento)
    if not fecha:
        return None
    hoy = ref or datetime.now()
    meses = (hoy.year - fecha.year) * 12 + (hoy.month - fecha.month)
    if hoy.day < fecha.day:
        meses -= 1
    return meses


def inferir_edad_meses(valor: Any) -> int | None:
    texto = limpiar_valor(valor)
    if not texto:
        return None
    clave = normalizar_texto(texto)
    nums = re.findall(r'\d+', clave)
    if not nums:
        return None
    n = int(nums[0])
    if 'mes' in clave:
        return n
    if 'ano' in clave or 'anio' in clave or 'anos' in clave:
        return n * 12
    # En Cuéntame suele venir edad en años cuando es un número bajo.
    if 0 <= n <= 6:
        return n * 12
    return n


def normalize_unidad(valor: Any) -> str:
    return catalog_normalize_unit(valor, preserve_unknown=True)


def clasificar_cargo(cargo: Any) -> str:
    clave = normalizar_texto(cargo)
    if any(x in clave for x in ['coordinador', 'coordinadora']):
        return 'COORDINADOR'
    if any(x in clave for x in ['agente educativo', 'docente', 'educador', 'educadora']):
        return 'DOCENTE'
    if 'psicosocial' in clave or 'psicologo' in clave or 'psicologa' in clave:
        return 'PSICOSOCIAL'
    if 'enfermer' in clave or 'salud' in clave:
        return 'ENFERMERIA'
    if 'nutric' in clave:
        return 'NUTRICIONISTA'
    if 'pedagog' in clave:
        return 'PEDAGOGIA'
    if 'administr' in clave or 'auxiliar' in clave:
        return 'AUXILIAR_ADMINISTRATIVO'
    return valor_upper(cargo) or 'APOYO'


def read_tabular_file(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path.lower())[1]
    if ext in {'.xlsx', '.xls', '.xlsm', '.ods'}:
        hojas = pd.read_excel(path, sheet_name=None, dtype=str)
        frames = []
        for sheet, df in hojas.items():
            if df is None or df.empty:
                continue
            df = df.copy()
            df['_hoja_origen'] = sheet
            frames.append(df)
        if not frames:
            raise ValueError('El libro no contiene hojas con datos.')
        return pd.concat(frames, ignore_index=True, sort=False)

    if ext in {'.csv', '.txt', '.tsv', '.tab', '.dat'}:
        last_error = None
        for encoding in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
            try:
                if ext in {'.tsv', '.tab'}:
                    return pd.read_csv(path, sep='\t', dtype=str, encoding=encoding)
                return pd.read_csv(path, sep=None, dtype=str, encoding=encoding, engine='python')
            except Exception as exc:
                last_error = exc
        raise ValueError(f'No se pudo leer el archivo tabular: {last_error}')

    if ext in {'.html', '.htm'}:
        tablas = pd.read_html(path)
        if tablas:
            return tablas[0].astype(str)
        raise ValueError('El archivo HTML no contiene tablas.')

    if ext == '.json':
        df = pd.read_json(path, dtype=str)
        if df is not None and not df.empty:
            return df
        raise ValueError('JSON vacío o no tabular.')

    raise ValueError(f'Formato no soportado para análisis de calidad: {ext}')


def normalizar_base(df: pd.DataFrame) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, str]]:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    aliases = {
        'documento': ['Documento del beneficiario', 'Número de documento del beneficiario', 'Numero de documento del beneficiario', 'Documento', 'NUI', 'NUIP', 'Identificación', 'Identificacion', 'No documento', 'Nº DOC. IDENT.'],
        'tipo_documento': ['Tipo de documento del beneficiario', 'Tipo documento', 'Tipo Doc'],
        'nombre': ['Nombre completo del beneficiario', 'Nombres y apellidos', 'Nombre completo', 'Nombre', 'Nombres y apellidos del participante'],
        'primer_nombre': ['Primer Nombre del beneficiario', 'Primer nombre', 'Primer Nombre'],
        'segundo_nombre': ['Segundo Nombre del beneficiario', 'Segundo nombre', 'Segundo Nombre'],
        'primer_apellido': ['Primer apellido del beneficiario', 'Primer apellido', 'Primer Apellido'],
        'segundo_apellido': ['Segundo apellido del beneficiario', 'Segundo apellido', 'Segundo Apellido'],
        'fecha_nacimiento': ['Fecha de nacimiento del beneficiario', 'Fecha nacimiento', 'Fecha de nacimiento'],
        'edad': ['Edad del beneficiario', 'Edad', 'Edad en meses'],
        'sexo': ['Sexo del beneficiario', 'Sexo'],
        'unidad': ['Nombre de la unidad de servicio', 'Unidad de servicio', 'Unidad', 'Unidad de atención', 'UCA', 'Comunidad'],
        'docente': ['Docente', 'Agente educativo', 'Educador', 'Educadora'],
        'acudiente': ['Nombre acudiente', 'Nombre completo acudiente', 'Nombre responsable', 'Nombre completo responsable', 'Acudiente', 'Responsable'],
        'doc_acudiente': ['Número de documento del acudiente o responsable', 'Documento acudiente', 'Documento responsable', 'Cédula acudiente', 'Cedula acudiente'],
        'primer_nombre_acudiente': ['Primer nombre del acudiente o responsable', 'Primer nombre acudiente'],
        'segundo_nombre_acudiente': ['Segundo nombre del acudiente o responsable', 'Segundo nombre acudiente'],
        'primer_apellido_acudiente': ['Primer apellido del acudiente o responsable', 'Primer apellido acudiente'],
        'segundo_apellido_acudiente': ['Segundo apellido del acudiente o responsable', 'Segundo apellido acudiente'],
        'telefono': ['Teléfono del beneficiario', 'Telefono del beneficiario', 'Celular', 'Teléfono', 'Telefono', 'Contacto'],
        'direccion': ['Dirección de residencia del beneficiario', 'Direccion de residencia del beneficiario', 'Dirección', 'Direccion'],
        'cargo': ['Cargo', 'Perfil', 'Rol', 'Función', 'Funcion'],
        'tipo_beneficiario': ['Nombre Tipo de beneficiario', 'Tipo de beneficiario', 'Tipo beneficiario'],
    }
    cols = {key: buscar_columna(df, values) for key, values in aliases.items()}
    # Evitar que el nombre del acudiente se confunda con tipo/documento del acudiente.
    if cols.get('acudiente') and any(x in normalizar_texto(cols['acudiente']) for x in ['documento', 'tipo']):
        cols['acudiente'] = None
    errores: list[dict[str, Any]] = []
    for required in ['documento', 'unidad']:
        if not cols.get(required):
            errores.append({
                'tipo': 'COLUMNA_OBLIGATORIA_FALTANTE',
                'campo': required,
                'descripcion': f'No se detectó columna obligatoria: {required}.',
            })

    rows: list[dict[str, Any]] = []
    for i, row in df.iterrows():
        pn = valor_upper(row.get(cols['primer_nombre'])) if cols.get('primer_nombre') else ''
        sn = valor_upper(row.get(cols['segundo_nombre'])) if cols.get('segundo_nombre') else ''
        pa = valor_upper(row.get(cols['primer_apellido'])) if cols.get('primer_apellido') else ''
        sa = valor_upper(row.get(cols['segundo_apellido'])) if cols.get('segundo_apellido') else ''
        nombre = ' '.join([x for x in [pn, sn, pa, sa] if x]).strip()
        if not nombre and cols.get('nombre'):
            nombre = valor_upper(row.get(cols['nombre']))
        acudiente = ''
        if cols.get('acudiente'):
            acudiente = valor_upper(row.get(cols['acudiente']))
        if not acudiente:
            acudiente = ' '.join([
                valor_upper(row.get(cols.get('primer_nombre_acudiente'))) if cols.get('primer_nombre_acudiente') else '',
                valor_upper(row.get(cols.get('segundo_nombre_acudiente'))) if cols.get('segundo_nombre_acudiente') else '',
                valor_upper(row.get(cols.get('primer_apellido_acudiente'))) if cols.get('primer_apellido_acudiente') else '',
                valor_upper(row.get(cols.get('segundo_apellido_acudiente'))) if cols.get('segundo_apellido_acudiente') else '',
            ]).strip()
        fecha_nac = limpiar_valor(row.get(cols['fecha_nacimiento'])) if cols.get('fecha_nacimiento') else ''
        edad_calc = edad_meses(fecha_nac)
        edad_col = inferir_edad_meses(row.get(cols['edad'])) if cols.get('edad') else None
        rows.append({
            '_fila': int(i) + 2,
            '_hoja': limpiar_valor(row.get('_hoja_origen')),
            'documento': limpiar_valor(row.get(cols['documento'])) if cols.get('documento') else '',
            'tipo_documento': limpiar_valor(row.get(cols['tipo_documento'])) if cols.get('tipo_documento') else '',
            'nombre': nombre,
            'primer_nombre': pn,
            'segundo_nombre': sn,
            'primer_apellido': pa,
            'segundo_apellido': sa,
            'fecha_nacimiento': fecha_nac,
            'edad_meses_calculada': edad_calc,
            'edad_meses_archivo': edad_col,
            'sexo': limpiar_valor(row.get(cols['sexo'])) if cols.get('sexo') else '',
            'unidad': normalize_unidad(row.get(cols['unidad'])) if cols.get('unidad') else '',
            'docente': valor_upper(row.get(cols['docente'])) if cols.get('docente') else '',
            'acudiente': acudiente,
            'documento_acudiente': limpiar_valor(row.get(cols['doc_acudiente'])) if cols.get('doc_acudiente') else '',
            'telefono': limpiar_valor(row.get(cols['telefono'])) if cols.get('telefono') else '',
            'direccion': limpiar_valor(row.get(cols['direccion'])) if cols.get('direccion') else '',
            'cargo': valor_upper(row.get(cols['cargo'])) if cols.get('cargo') else '',
            'tipo_beneficiario': valor_upper(row.get(cols['tipo_beneficiario'])) if cols.get('tipo_beneficiario') else '',
            'rol_normalizado': clasificar_cargo(row.get(cols['cargo'])) if cols.get('cargo') else '',
        })
    return rows, errores, {k: (v or '') for k, v in cols.items()}


def hallazgo(tipo: str, row: dict[str, Any] | None, severidad: str, descripcion: str, campo: str = '', valor_actual: Any = '', valor_esperado: Any = '') -> dict[str, Any]:
    row = row or {}
    return {
        'tipo': tipo,
        'categoria': TIPOS_HALLAZGO.get(tipo, tipo),
        'severidad': severidad,
        'documento': row.get('documento') or row.get('documento_acudiente') or '',
        'nombre': row.get('nombre') or row.get('acudiente') or '',
        'unidad': row.get('unidad') or '',
        'docente': row.get('docente') or '',
        'campo': campo,
        'valor_actual': limpiar_valor(valor_actual),
        'valor_esperado': limpiar_valor(valor_esperado),
        'descripcion': descripcion,
        'fila': row.get('_fila'),
        'hoja': row.get('_hoja'),
        'datos': row,
    }


def analizar_filas(rows: list[dict[str, Any]], errores: list[dict[str, Any]], docentes_por_unidad: set[str] | None = None, docentes_sin_unidad_db: list[dict[str, Any]] | None = None, talento_contexto: bool = False) -> dict[str, Any]:
    docentes_por_unidad = docentes_por_unidad or set()
    hallazgos: list[dict[str, Any]] = []
    docs: dict[str, list[dict[str, Any]]] = {}
    unidades_presentes: set[str] = set()
    hoy = datetime.now()

    if errores:
        for err in errores:
            hallazgos.append(hallazgo('COLUMNA_OBLIGATORIA_FALTANTE', None, 'CRITICA', err.get('descripcion') or 'Columna obligatoria faltante.', err.get('campo', '')))

    for row in rows:
        doc = limpiar_valor(row.get('documento'))
        nombre = limpiar_valor(row.get('nombre'))
        unidad = normalize_unidad(row.get('unidad'))
        cargo_norm = row.get('rol_normalizado') or ''
        tipo_beneficiario_norm = normalizar_texto(row.get('tipo_beneficiario'))
        es_gestante = 'gestante' in tipo_beneficiario_norm
        es_registro_talento = bool(talento_contexto or cargo_norm)

        if unidad:
            unidades_presentes.add(unidad)

        if not doc:
            hallazgos.append(hallazgo('DOCUMENTO_VACIO', row, 'CRITICA', 'El registro no tiene documento/NUI/identificación.', 'documento'))
        else:
            docs.setdefault(doc, []).append(row)

        # Talento humano: reglas especiales para duplicados/docentes sin unidad.
        # No tratar bases de beneficiarios como talento humano solo porque el cargo esté vacío.
        if es_registro_talento:
            if cargo_norm in {'DOCENTE', 'COORDINADOR', 'PSICOSOCIAL', 'ENFERMERIA', 'NUTRICIONISTA', 'PEDAGOGIA', 'AUXILIAR_ADMINISTRATIVO'} and not unidad and cargo_norm == 'DOCENTE':
                hallazgos.append(hallazgo('DOCENTE_SIN_UNIDAD', row, 'ALTA', 'Docente/agente educativo sin unidad asignada.', 'unidad'))
            continue

        if not unidad:
            hallazgos.append(hallazgo('NINO_SIN_UNIDAD', row, 'CRITICA', 'Beneficiario sin unidad de servicio.', 'unidad'))
        if not (limpiar_valor(row.get('acudiente')) or limpiar_valor(row.get('documento_acudiente'))):
            hallazgos.append(hallazgo('NINO_SIN_ACUDIENTE', row, 'ALTA', 'Beneficiario sin acudiente/responsable registrado.', 'acudiente'))
        if not limpiar_valor(row.get('telefono')):
            hallazgos.append(hallazgo('NINO_SIN_TELEFONO', row, 'MEDIA', 'Beneficiario sin teléfono/contacto.', 'telefono'))
        if not limpiar_valor(row.get('fecha_nacimiento')):
            hallazgos.append(hallazgo('NINO_SIN_FECHA_NACIMIENTO', row, 'ALTA', 'Beneficiario sin fecha de nacimiento.', 'fecha_nacimiento'))

        edad_calc = row.get('edad_meses_calculada')
        edad_archivo = row.get('edad_meses_archivo')
        if edad_calc is not None:
            if edad_calc < 0:
                hallazgos.append(hallazgo('EDAD_INCONSISTENTE', row, 'CRITICA', 'La fecha de nacimiento produce edad negativa/futura.', 'fecha_nacimiento', row.get('fecha_nacimiento'), 'Fecha válida anterior a hoy'))
            elif edad_archivo is not None and abs(int(edad_calc) - int(edad_archivo)) > 12 and not es_gestante:
                hallazgos.append(hallazgo('EDAD_INCONSISTENTE', row, 'MEDIA', 'La edad del archivo no coincide con la fecha de nacimiento.', 'edad', f'{edad_archivo} meses', f'{edad_calc} meses'))
            if edad_calc >= 72 and not es_gestante:
                hallazgos.append(hallazgo('BENEFICIARIO_FUERA_RANGO', row, 'ALTA', 'Beneficiario no gestante con 72 meses o más. Revisar continuidad/retiro.', 'edad', f'{edad_calc} meses', '< 72 meses'))
        elif edad_archivo is not None and edad_archivo >= 72 and not es_gestante:
            hallazgos.append(hallazgo('BENEFICIARIO_FUERA_RANGO', row, 'ALTA', 'Beneficiario con edad reportada fuera del rango esperado.', 'edad', f'{edad_archivo} meses', '< 72 meses'))

    for doc, items in docs.items():
        if len(items) > 1:
            for row in items:
                tipo = 'TALENTO_DUPLICADO' if (talento_contexto or row.get('rol_normalizado')) else 'DOCUMENTO_DUPLICADO'
                hallazgos.append(hallazgo(tipo, row, 'ALTA', f'Documento duplicado: aparece {len(items)} veces.', 'documento', doc, 'Documento único'))

    # Unidades sin docente: compara unidades activas/archivo contra Talento Humano.
    for unidad in sorted(u for u in unidades_presentes if u):
        if docentes_por_unidad and unidad not in docentes_por_unidad:
            hallazgos.append(hallazgo('UNIDAD_SIN_DOCENTE', {'unidad': unidad, 'nombre': unidad}, 'ALTA', f'La unidad {unidad} no tiene docente asignado en Talento Humano.', 'docente', '', 'Docente asignado'))

    for docente in docentes_sin_unidad_db or []:
        hallazgos.append(hallazgo('DOCENTE_SIN_UNIDAD', docente, 'ALTA', 'Docente registrado en Talento Humano sin unidad asignada.', 'unidad'))

    conteos_tipo: dict[str, int] = {}
    conteos_sev: dict[str, int] = {'CRITICA': 0, 'ALTA': 0, 'MEDIA': 0, 'BAJA': 0}
    for item in hallazgos:
        conteos_tipo[item['tipo']] = conteos_tipo.get(item['tipo'], 0) + 1
        sev = item.get('severidad') or 'MEDIA'
        conteos_sev[sev] = conteos_sev.get(sev, 0) + 1

    resumen = {
        'total_registros': len(rows),
        'total_hallazgos': len(hallazgos),
        'conteos': conteos_sev,
        'conteos_tipo': conteos_tipo,
        'tipos': TIPOS_HALLAZGO,
        'fecha': datetime.now().isoformat(timespec='seconds'),
    }
    hallazgos.sort(key=lambda x: (SEVERITY_ORDER.get(x.get('severidad'), 9), x.get('tipo') or '', x.get('unidad') or '', x.get('nombre') or ''))
    return {'resumen': resumen, 'hallazgos': hallazgos}


def docentes_por_unidad_desde_db(database_path: str, fundacion_id: int = 1) -> tuple[set[str], list[dict[str, Any]]]:
    import sqlite3
    conn = sqlite3.connect(database_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    docentes_por_unidad: set[str] = set()
    docentes_sin_unidad: list[dict[str, Any]] = []

    def table_exists(t: str) -> bool:
        return cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (t,)).fetchone() is not None

    if table_exists('th_personas'):
        try:
            rows = cur.execute("""
                SELECT p.documento, p.nombre, p.rol_normalizado, a.unidad
                FROM th_personas p
                LEFT JOIN th_asignaciones a
                  ON a.persona_id = p.id
                 AND COALESCE(a.estado,'ACTIVO')='ACTIVO'
                 AND COALESCE(a.fundacion_id, 1) = ?
                WHERE UPPER(COALESCE(p.rol_normalizado,'')) IN ('DOCENTE','AGENTE EDUCATIVO')
                  AND COALESCE(p.fundacion_id, 1) = ?
            """, (fundacion_id, fundacion_id)).fetchall()
            for r in rows:
                unidad = normalize_unidad(r['unidad'])
                row = dict(r)
                row['unidad'] = unidad
                if unidad:
                    docentes_por_unidad.add(unidad)
                else:
                    docentes_sin_unidad.append(row)
        except Exception:
            pass

    if table_exists('coordinadores'):
        try:
            rows = cur.execute("""
                SELECT documento, COALESCE(nombre, nombres || ' ' || apellidos) nombre, cargo, unidad
                FROM coordinadores
                WHERE COALESCE(activo, 1) = 1
            """).fetchall()
            for r in rows:
                cargo = clasificar_cargo(r['cargo'])
                if cargo == 'DOCENTE':
                    unidad = normalize_unidad(r['unidad'])
                    row = dict(r)
                    row['unidad'] = unidad
                    if unidad:
                        docentes_por_unidad.add(unidad)
                    else:
                        docentes_sin_unidad.append(row)
        except Exception:
            pass
    conn.close()
    return docentes_por_unidad, docentes_sin_unidad


def analizar_archivo(path: str, database_path: str, fundacion_id: int = 1, tipo: str = 'auto') -> dict[str, Any]:
    df = read_tabular_file(path)
    rows, errores, columnas = normalizar_base(df)
    talento_contexto = tipo.lower() in {'talento', 'talento_humano', 'th'} or any(r.get('cargo') for r in rows)
    docentes_unidad, docentes_sin_unidad = docentes_por_unidad_desde_db(database_path, fundacion_id)
    analisis = analizar_filas(rows, errores, docentes_unidad, docentes_sin_unidad, talento_contexto=talento_contexto)
    analisis['columnas_detectadas'] = columnas
    return analisis


def analizar_base_actual(database_path: str, fundacion_id: int = 1) -> dict[str, Any]:
    import sqlite3
    conn = sqlite3.connect(database_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    def table_exists(t: str) -> bool:
        return cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (t,)).fetchone() is not None

    rows: list[dict[str, Any]] = []
    errores: list[dict[str, Any]] = []
    if table_exists('beneficiarios'):
        cols = {r['name'] for r in cur.execute("PRAGMA table_info(beneficiarios)").fetchall()}
        select_cols = ['documento', 'nui', 'nombres', 'apellidos', 'primer_nombre', 'segundo_nombre', 'primer_apellido', 'segundo_apellido', 'fecha_nacimiento', 'sexo', 'unidad', 'docente', 'nombre_acudiente', 'documento_acudiente', 'telefono', 'edad_meses', 'grupo_edad', 'tipo_beneficiario', 'estado']
        select_cols = [c for c in select_cols if c in cols]
        db_rows = cur.execute(f"SELECT {', '.join(select_cols)} FROM beneficiarios").fetchall() if select_cols else []
        for idx, r in enumerate(db_rows, start=1):
            d = dict(r)
            nombre = ' '.join([limpiar_valor(d.get(c)) for c in ['primer_nombre', 'segundo_nombre', 'primer_apellido', 'segundo_apellido'] if limpiar_valor(d.get(c))]).strip()
            if not nombre:
                nombre = ' '.join([limpiar_valor(d.get('nombres')), limpiar_valor(d.get('apellidos'))]).strip()
            edad_calc = edad_meses(d.get('fecha_nacimiento'))
            edad_archivo = inferir_edad_meses(d.get('edad_meses'))
            rows.append({
                '_fila': idx,
                'documento': limpiar_valor(d.get('documento') or d.get('nui')),
                'nombre': nombre.upper(),
                'fecha_nacimiento': limpiar_valor(d.get('fecha_nacimiento')),
                'edad_meses_calculada': edad_calc,
                'edad_meses_archivo': edad_archivo,
                'sexo': limpiar_valor(d.get('sexo')),
                'unidad': normalize_unidad(d.get('unidad')),
                'docente': valor_upper(d.get('docente')),
                'acudiente': valor_upper(d.get('nombre_acudiente')),
                'documento_acudiente': limpiar_valor(d.get('documento_acudiente')),
                'telefono': limpiar_valor(d.get('telefono')),
                'direccion': '',
                'tipo_beneficiario': limpiar_valor(d.get('tipo_beneficiario')),
                'estado': limpiar_valor(d.get('estado')),
            })
    else:
        errores.append({'tipo': 'TABLA_FALTANTE', 'descripcion': 'No existe tabla beneficiarios para análisis actual.'})

    docentes_unidad, docentes_sin_unidad = docentes_por_unidad_desde_db(database_path, fundacion_id)
    conn.close()
    return analizar_filas(rows, errores, docentes_unidad, docentes_sin_unidad)


def generar_excel(analisis_id: int, resumen: dict[str, Any], hallazgos: list[dict[str, Any]], output_folder: str) -> str:
    os.makedirs(output_folder, exist_ok=True)
    path = os.path.join(output_folder, f'CALIDAD_DATOS_{analisis_id}.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen'
    ws.append(['Indicador', 'Valor'])
    ws.append(['Total registros', resumen.get('total_registros', 0)])
    ws.append(['Total hallazgos', resumen.get('total_hallazgos', 0)])
    for sev, total in (resumen.get('conteos') or {}).items():
        ws.append([sev, total])
    ws.append([])
    ws.append(['Tipo de hallazgo', 'Cantidad'])
    for tipo, total in (resumen.get('conteos_tipo') or {}).items():
        ws.append([TIPOS_HALLAZGO.get(tipo, tipo), total])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9EAF7')

    wh = wb.create_sheet('Hallazgos')
    headers = ['Severidad', 'Tipo', 'Documento', 'Nombre', 'Unidad', 'Docente', 'Campo', 'Valor actual', 'Valor esperado', 'Descripción', 'Fila', 'Hoja']
    wh.append(headers)
    for item in hallazgos:
        wh.append([
            item.get('severidad'), TIPOS_HALLAZGO.get(item.get('tipo'), item.get('tipo')),
            item.get('documento'), item.get('nombre'), item.get('unidad'), item.get('docente'),
            item.get('campo'), item.get('valor_actual'), item.get('valor_esperado'),
            item.get('descripcion'), item.get('fila'), item.get('hoja')
        ])
    for cell in wh[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='FCE4D6')
        cell.alignment = Alignment(horizontal='center')
    for sheet in [ws, wh]:
        for col_cells in sheet.columns:
            max_len = max(len(str(c.value or '')) for c in col_cells)
            sheet.column_dimensions[col_cells[0].column_letter].width = min(max(12, max_len + 2), 60)
    wb.save(path)
    return path


def generar_pdf(analisis_id: int, resumen: dict[str, Any], hallazgos: list[dict[str, Any]], output_folder: str) -> str:
    os.makedirs(output_folder, exist_ok=True)
    path = os.path.join(output_folder, f'CALIDAD_DATOS_{analisis_id}.pdf')
    doc = SimpleDocTemplate(path, pagesize=landscape(letter), rightMargin=0.4*inch, leftMargin=0.4*inch, topMargin=0.4*inch, bottomMargin=0.4*inch)
    styles = getSampleStyleSheet()
    story = [Paragraph('Reporte de Calidad de Datos - PrimeraInfancia', styles['Title']), Spacer(1, 0.15*inch)]
    resumen_data = [
        ['Indicador', 'Valor'],
        ['Total registros', str(resumen.get('total_registros', 0))],
        ['Total hallazgos', str(resumen.get('total_hallazgos', 0))],
    ]
    for sev, total in (resumen.get('conteos') or {}).items():
        resumen_data.append([sev, str(total)])
    t = Table(resumen_data, colWidths=[3*inch, 1.5*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightblue),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.2*inch))
    story.append(Paragraph('Hallazgos principales', styles['Heading2']))
    data = [['Sev.', 'Tipo', 'Documento', 'Nombre', 'Unidad', 'Descripción']]
    for item in hallazgos[:80]:
        data.append([
            item.get('severidad', ''),
            TIPOS_HALLAZGO.get(item.get('tipo'), item.get('tipo', '')),
            item.get('documento', ''),
            (item.get('nombre') or '')[:28],
            (item.get('unidad') or '')[:20],
            (item.get('descripcion') or '')[:70],
        ])
    table = Table(data, colWidths=[0.7*inch, 1.6*inch, 1.2*inch, 1.8*inch, 1.4*inch, 4.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('FONTSIZE', (0,0), (-1,-1), 7),
    ]))
    story.append(table)
    doc.build(story)
    return path
