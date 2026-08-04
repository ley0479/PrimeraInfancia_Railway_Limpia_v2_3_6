
from __future__ import annotations

import json
import os
import re
import shutil
import sqlite3
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from services.uds_catalog import normalize_unit as catalog_normalize_unit

from openpyxl import load_workbook

from modules.print_master import aplicar_configuracion_impresion_libro, infer_print_format
from .schema import CAMPOS_CANONICOS, MOTOR_PLANTILLAS_SCHEMA_SQL


FOOD_TERMS = {
    'kilo', 'kilos', 'kg', 'libra', 'libras', 'gramo', 'gramos', 'gr', 'g',
    'racion', 'ración', 'cantidad', 'alimento', 'alimentos', 'arroz', 'aceite',
    'harina', 'huevo', 'huevos', 'verdura', 'verduras', 'bienestarina', 'bolsa',
    'bolsas', 'panela', 'frijol', 'lenteja', 'azucar', 'azúcar', 'leche'
}

PHONE_TERMS = {'telefono', 'teléfono', 'celular', 'contacto', 'movil', 'móvil', 'numero celular', 'número celular'}
DOC_TERMS = {'documento', 'identidad', 'nui', 'nuip', 'cedula', 'cédula', 'identificacion', 'identificación'}
NUMBER_TERMS = {'no', 'n', 'n°', 'nº', 'numero', 'número', 'consecutivo', 'item', 'orden'}

DATA_FIELDS = {
    'tipo_documento', 'documento_beneficiario', 'nombre_completo',
    'primer_nombre', 'segundo_nombre', 'primer_apellido', 'segundo_apellido',
    'edad_anios', 'edad_meses', 'acudiente_nombre_cedula', 'nombre_acudiente',
    'documento_acudiente', 'tipo_documento_acudiente', 'parentesco', 'telefono'
}


def now_iso() -> str:
    return datetime.now().isoformat(timespec='seconds')


def connect(database_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(database_path)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_column(conn: sqlite3.Connection, table: str, column: str, definition: str) -> None:
    try:
        cols = {r['name'] for r in conn.execute(f'PRAGMA table_info({table})').fetchall()}
        if column not in cols:
            conn.execute(f'ALTER TABLE {table} ADD COLUMN {column} {definition}')
    except Exception:
        # No interrumpir el arranque si una base antigua no tiene la tabla todavía.
        pass


def init_schema(database_path: str) -> None:
    conn = connect(database_path)
    conn.executescript(MOTOR_PLANTILLAS_SCHEMA_SQL)
    # ALPHA52: extender tablas existentes sin romper versiones previas.
    ensure_column(conn, 'mp_plantillas', 'codigo', 'TEXT')
    ensure_column(conn, 'mp_plantillas', 'fecha_vigencia', 'TEXT')
    ensure_column(conn, 'mp_plantillas', 'observaciones', 'TEXT')
    ensure_column(conn, 'mp_plantillas', 'plantilla_oficial_version_id', 'INTEGER')
    ensure_column(conn, 'mp_pruebas', 'resultado_json', 'TEXT')
    ensure_column(conn, 'plantillas_oficiales_versiones', 'fecha_vigencia_fin', 'TEXT')
    ensure_column(conn, 'plantillas_oficiales_versiones', 'estado_publicacion', 'TEXT')
    ensure_column(conn, 'plantillas_oficiales_versiones', 'hash_sha256', 'TEXT')
    ensure_column(conn, 'plantillas_oficiales_versiones', 'manual_path', 'TEXT')
    ensure_column(conn, 'plantillas_oficiales_versiones', 'reglas_json', 'TEXT')
    conn.commit()
    conn.close()


def normalizar_texto(valor: Any) -> str:
    texto = str(valor or '').strip().lower()
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.replace('º', 'o').replace('°', 'o')
    texto = re.sub(r'[^a-z0-9]+', ' ', texto)
    return ' '.join(texto.split())


def col_letter_to_index(col: str) -> int:
    value = 0
    for ch in col.upper():
        if 'A' <= ch <= 'Z':
            value = value * 26 + (ord(ch) - 64)
    return value


def index_to_col(n: int) -> str:
    result = ''
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result or 'A'


def classify_label(label: str) -> str | None:
    text = normalizar_texto(label)
    if not text:
        return None

    food = any(term in text for term in FOOD_TERMS)
    if food:
        return None

    if (text in {'no', 'n', 'n o', 'numero', 'número', 'orden'} or 'consecutivo' in text or 'no de orden' in text or 'n de orden' in text) and not any(t in text for t in DOC_TERMS):
        return 'consecutivo'

    if text in {'sexo', 'genero', 'género'} or 'sexo' in text:
        return 'sexo'
    if 'grupo etario' in text or 'rango edad' in text or 'edad grupo' in text:
        return 'grupo_etario'
    if 'unidad de servicio' in text or 'uds' in text or 'uca' in text:
        return 'unidad_servicio'
    if 'observacion' in text or 'observaciones' in text:
        return 'observaciones'
    if 'primer nombre' in text:
        return 'primer_nombre'
    if 'segundo nombre' in text:
        return 'segundo_nombre'
    if 'primer apellido' in text:
        return 'primer_apellido'
    if 'segundo apellido' in text:
        return 'segundo_apellido'
    if 'nombre completo y cedula' in text or 'nombre completo y cedula' in text or 'nombre y cedula' in text or 'quien recibe' in text:
        return 'acudiente_nombre_cedula'
    if ('acudiente' in text or 'responsable' in text) and any(t in text for t in DOC_TERMS):
        return 'documento_acudiente'
    if ('acudiente' in text or 'responsable' in text) and 'tipo' in text and 'doc' in text:
        return 'tipo_documento_acudiente'
    if ('acudiente' in text or 'responsable' in text) and ('nombre' in text or 'apellidos' in text):
        return 'nombre_acudiente'
    if any(t in text for t in PHONE_TERMS):
        if any(scope in text for scope in ['uds', 'unidad', 'servicio', 'docente', 'coordinador']):
            return None
        return 'telefono'
    if 'parentesco' in text:
        return 'parentesco'
    if ('tipo' in text and ('doc' in text or 'documento' in text)) and 'acudiente' not in text and 'responsable' not in text:
        return 'tipo_documento'
    if any(t in text for t in DOC_TERMS) and 'acudiente' not in text and 'responsable' not in text and 'tipo' not in text:
        return 'documento_beneficiario'
    if 'nombres y apellidos del participante' in text or 'nombre del participante' in text or ('nombre completo' in text and 'acudiente' not in text):
        return 'nombre_completo'
    if text in {'anos', 'anios', 'años'} or ('edad' in text and ('ano' in text or 'año' in text)):
        return 'edad_anios'
    if text in {'mes', 'meses'} or ('edad' in text and 'mes' in text):
        return 'edad_meses'
    if 'fecha de entrega' in text or ('fecha' in text and 'entrega' in text):
        return 'fecha_entrega'
    if 'lote' in text:
        return 'lote'
    if 'cantidad' in text and not food:
        return 'cantidad'
    if 'total asistencia' in text or 'total mensual' in text:
        return 'total_asistencias'
    if 'menor' in text and '6' in text:
        return 'verificacion_menor_6'
    if 'mayor' in text and '6' in text:
        return 'verificacion_mayor_6'
    if 'gestante' in text:
        return 'verificacion_gestantes'
    return None


def label_has_food_terms(label: str) -> bool:
    text = normalizar_texto(label)
    return any(term in text for term in FOOD_TERMS)


def label_has_doc_terms(label: str) -> bool:
    text = normalizar_texto(label)
    return any(term in text for term in DOC_TERMS)


def label_is_numbering(label: str) -> bool:
    text = normalizar_texto(label)
    return text in {'no', 'n', 'numero', 'número', 'orden', 'item'} or 'consecutivo' in text or 'no de orden' in text or 'n de orden' in text


def detect_template(file_path: str) -> dict:
    wb = load_workbook(file_path, data_only=False, read_only=False)
    hojas = []
    columnas_detectadas = []
    riesgos = []
    for ws in wb.worksheets:
        hoja_info = {
            'nombre': ws.title,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'columnas_detectadas': 0,
        }
        max_rows = min(ws.max_row or 1, 80)
        max_cols = min(ws.max_column or 1, 120)
        for row in range(1, max_rows + 1):
            non_empty = 0
            for col in range(1, max_cols + 1):
                val = ws.cell(row=row, column=col).value
                if val not in (None, ''):
                    non_empty += 1
                    label = str(val).strip()
                    field = classify_label(label)
                    if field:
                        item = {
                            'sheet': ws.title,
                            'row': row,
                            'col': col,
                            'col_letter': index_to_col(col),
                            'cell': f'{index_to_col(col)}{row}',
                            'label': label,
                            'normalized': normalizar_texto(label),
                            'suggested_field': field,
                            'data_start_row': row + 1,
                        }
                        columnas_detectadas.append(item)
                        hoja_info['columnas_detectadas'] += 1
                    elif label_has_food_terms(label):
                        riesgos.append({
                            'sheet': ws.title,
                            'cell': f'{index_to_col(col)}{row}',
                            'label': label,
                            'risk': 'COLUMNA_ALIMENTO',
                            'message': 'Columna de alimento/cantidad detectada. No debe recibir teléfono ni documentos.'
                        })
            # Detener temprano si la hoja es muy vacía después de muchos renglones.
            if row > 45 and non_empty == 0 and hoja_info['columnas_detectadas'] > 0:
                break
        hojas.append(hoja_info)
    return {
        'hojas': hojas,
        'columnas': columnas_detectadas,
        'riesgos': riesgos,
        'campos': CAMPOS_CANONICOS,
        'total_columnas_detectadas': len(columnas_detectadas),
    }


def validate_mapping(mapping: list[dict], strict: bool = False) -> dict:
    errores = []
    advertencias = []
    usados = {}

    if not mapping:
        errores.append({
            'code': 'MAPEO_VACIO',
            'message': 'No hay campos mapeados. Debes mapear al menos documento, nombre y/o columnas principales.'
        })

    for item in mapping or []:
        field = item.get('field') or item.get('campo')
        label = item.get('label') or ''
        row = int(item.get('row') or 0)
        col = int(item.get('col') or 0)
        data_start_row = int(item.get('data_start_row') or item.get('dataStartRow') or (row + 1))
        key = f'{item.get("sheet")}:{row}:{col}'

        if not field or field == 'ignorar':
            continue

        if field in usados:
            advertencias.append({
                'code': 'CAMPO_DUPLICADO',
                'field': field,
                'message': f'El campo {field} aparece más de una vez. Se usará el primer mapeo válido en algunas operaciones.'
            })
        usados[field] = True

        if data_start_row <= row:
            errores.append({
                'code': 'RIESGO_ENCABEZADO',
                'field': field,
                'cell': key,
                'message': 'La fila inicial de datos no puede estar sobre la fila del encabezado. Esto evita borrar encabezados oficiales.'
            })

        if field == 'telefono' and label_has_food_terms(label):
            errores.append({
                'code': 'TELEFONO_EN_ALIMENTO',
                'field': field,
                'cell': key,
                'message': 'El teléfono/celular no puede mapearse a columnas de kilos, gramos, raciones, alimentos o cantidades.'
            })

        if field == 'telefono' and any(scope in normalizar_texto(label) for scope in ['telefono uds', 'telefono unidad', 'telefono servicio', 'telefono docente', 'telefono coordinador']):
            errores.append({
                'code': 'TELEFONO_NO_BENEFICIARIO',
                'field': field,
                'cell': key,
                'message': 'Esta columna parece ser teléfono de UDS/docente/coordinador, no celular del acudiente o beneficiario.'
            })

        if field in {'documento_beneficiario', 'documento_acudiente'} and label_is_numbering(label):
            errores.append({
                'code': 'DOCUMENTO_EN_NUMERACION',
                'field': field,
                'cell': key,
                'message': 'Un documento/NUI no puede mapearse a una columna de numeración o consecutivo.'
            })

        if field == 'consecutivo' and label_has_doc_terms(label):
            errores.append({
                'code': 'CONSECUTIVO_EN_DOCUMENTO',
                'field': field,
                'cell': key,
                'message': 'El consecutivo no puede mapearse a una columna de documento/NUI.'
            })

        if field in DATA_FIELDS and label_has_food_terms(label):
            errores.append({
                'code': 'DATO_PERSONA_EN_ALIMENTO',
                'field': field,
                'cell': key,
                'message': 'Datos de personas no pueden mapearse a columnas de alimentos, kilos, gramos, raciones o cantidades.'
            })

    required_soft = ['documento_beneficiario']
    for field in required_soft:
        if field not in usados:
            advertencias.append({
                'code': 'CAMPO_RECOMENDADO_FALTANTE',
                'field': field,
                'message': f'Se recomienda mapear {field} para evitar formatos incompletos.'
            })

    if strict and errores:
        valido = False
    else:
        valido = not errores

    return {
        'valido': valido,
        'errores': errores,
        'advertencias': advertencias,
        'total_errores': len(errores),
        'total_advertencias': len(advertencias),
    }


def table_exists(conn: sqlite3.Connection, table: str) -> bool:
    row = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone()
    return row is not None


def table_columns(conn: sqlite3.Connection, table: str) -> set[str]:
    if not table_exists(conn, table):
        return set()
    return {r['name'] for r in conn.execute(f'PRAGMA table_info({table})').fetchall()}


def normalize_unit(value: Any) -> str:
    return catalog_normalize_unit(value, preserve_unknown=True)


def get_users_by_unit(database_path: str, unidad: str, limit: int = 20) -> list[dict]:
    unidad_norm = normalize_unit(unidad)
    conn = connect(database_path)
    users: list[dict] = []
    if not table_exists(conn, 'beneficiarios'):
        conn.close()
        return users
    cols = table_columns(conn, 'beneficiarios')
    select_cols = [
        'documento', 'nui', 'tipo_documento', 'nombres', 'apellidos', 'primer_nombre',
        'segundo_nombre', 'primer_apellido', 'segundo_apellido', 'fecha_nacimiento',
        'edad_meses', 'nombre_acudiente', 'documento_acudiente', 'tipo_documento_acudiente',
        'parentesco', 'telefono', 'unidad', 'estado', 'sexo', 'grupo_etario', 'observaciones'
    ]
    select_cols = [c for c in select_cols if c in cols]
    if not select_cols:
        conn.close()
        return users
    rows = conn.execute(f"SELECT {', '.join(select_cols)} FROM beneficiarios WHERE UPPER(COALESCE(unidad,'')) = ? AND UPPER(COALESCE(estado,'ACTIVO')) IN ('ACTIVO','ACTIVA','') LIMIT ?", (unidad_norm, int(limit))).fetchall()
    if not rows:
        rows = conn.execute(f"SELECT {', '.join(select_cols)} FROM beneficiarios WHERE UPPER(COALESCE(unidad,'')) LIKE ? LIMIT ?", (f'%{unidad_norm}%', int(limit))).fetchall()
    for row in rows:
        d = dict(row)
        users.append(d)
    conn.close()
    return users


def split_name(user: dict) -> dict:
    first = user.get('primer_nombre') or ''
    second = user.get('segundo_nombre') or ''
    last1 = user.get('primer_apellido') or ''
    last2 = user.get('segundo_apellido') or ''
    if not any([first, second, last1, last2]):
        parts = f"{user.get('nombres','')} {user.get('apellidos','')}".strip().split()
        first = parts[0] if len(parts) > 0 else ''
        second = parts[1] if len(parts) > 3 else ''
        if len(parts) >= 2:
            last1 = parts[-2] if len(parts) >= 3 else parts[-1]
            last2 = parts[-1] if len(parts) >= 3 else ''
    return {
        'primer_nombre': str(first or '').strip(),
        'segundo_nombre': str(second or '').strip(),
        'primer_apellido': str(last1 or '').strip(),
        'segundo_apellido': str(last2 or '').strip(),
    }


def abrev_doc(value: Any) -> str:
    t = normalizar_texto(value)
    if not t:
        return ''
    if 'registro civil' in t or t == 'rc':
        return 'RC'
    if 'tarjeta' in t or t == 'ti':
        return 'TI'
    if 'cedula de ciudadania' in t or t in {'cc', 'c c'}:
        return 'CC'
    if 'cedula de extranjeria' in t or t == 'ce':
        return 'CE'
    if 'pasaporte' in t or t in {'pa', 'p'}:
        return 'PA'
    if 'pep' in t:
        return 'PEP'
    if 'ppt' in t:
        return 'PPT'
    return str(value or '').strip().upper()


def field_value(user: dict, field: str, index: int) -> Any:
    names = split_name(user)
    edad = int(float(user.get('edad_meses') or 0))
    anios = edad // 12
    meses = edad % 12
    nombre_completo = ' '.join([names['primer_nombre'], names['segundo_nombre'], names['primer_apellido'], names['segundo_apellido']]).replace('  ', ' ').strip()
    if not nombre_completo:
        nombre_completo = f"{user.get('nombres','')} {user.get('apellidos','')}".replace('  ', ' ').strip()
    if field == 'consecutivo':
        return index + 1
    if field == 'tipo_documento':
        return abrev_doc(user.get('tipo_documento'))
    if field == 'documento_beneficiario':
        return user.get('nui') or user.get('documento') or ''
    if field == 'nombre_completo':
        return nombre_completo
    if field in names:
        return names[field]
    if field == 'edad_anios':
        return anios
    if field == 'edad_meses':
        return meses
    if field == 'nombre_acudiente':
        return user.get('nombre_acudiente') or ''
    if field == 'tipo_documento_acudiente':
        return abrev_doc(user.get('tipo_documento_acudiente'))
    if field == 'documento_acudiente':
        return user.get('documento_acudiente') or ''
    if field == 'acudiente_nombre_cedula':
        tipo = abrev_doc(user.get('tipo_documento_acudiente')) or 'CC'
        doc = user.get('documento_acudiente') or ''
        nombre = user.get('nombre_acudiente') or ''
        return f"{nombre} - {tipo} {doc}".strip(' -')
    if field == 'telefono':
        return user.get('telefono') or ''
    if field == 'parentesco':
        return user.get('parentesco') or ''
    if field == 'sexo':
        return user.get('sexo') or ''
    if field == 'grupo_etario':
        return user.get('grupo_etario') or ''
    if field == 'unidad_servicio':
        return user.get('unidad') or user.get('unidad_servicio') or ''
    if field == 'observaciones':
        return user.get('observaciones') or ''
    if field == 'total_asistencias':
        return ''
    return ''


def safe_cell(ws, row: int, col: int):
    coordinate = f'{index_to_col(col)}{row}'
    for merged in ws.merged_cells.ranges:
        if coordinate in merged:
            return ws.cell(merged.min_row, merged.min_col)
    return ws.cell(row, col)




def normalize_products(productos: list[dict] | None) -> list[dict]:
    """Normaliza catálogo RPP para guardarlo/versionarlo sin depender de columnas quemadas."""
    cleaned = []
    for idx, item in enumerate(productos or [], start=1):
        nombre = str(item.get('nombre_producto') or item.get('nombre') or '').strip()
        if not nombre:
            continue
        columna = str(item.get('columna') or item.get('col_letter') or '').strip().upper()
        col_index = int(item.get('col_index') or (col_letter_to_index(columna) if columna else 0) or 0)
        cleaned.append({
            'nombre_producto': nombre,
            'columna': columna or (index_to_col(col_index) if col_index else ''),
            'col_index': col_index,
            'unidad_medida': str(item.get('unidad_medida') or item.get('unidad') or '').strip(),
            'cantidad': str(item.get('cantidad') or '').strip(),
            'grupo_etario_aplica': str(item.get('grupo_etario_aplica') or item.get('grupo') or 'todos').strip() or 'todos',
            'orden': int(item.get('orden') or idx),
            'activo': 1 if str(item.get('activo', '1')).lower() not in {'0', 'false', 'no'} else 0,
        })
    return cleaned


def apply_products_to_workbook(wb, productos: list[dict] | None, default_sheet: str | None = None) -> dict:
    """Escribe cantidades/productos configurados solo si el catálogo indica celdas/columnas válidas.
    No inserta filas ni columnas y no toca encabezados si no se indica fila_inicio.
    """
    escritos = 0
    errores = []
    for item in normalize_products(productos):
        if not item.get('activo'):
            continue
        sheet = item.get('hoja') or default_sheet or (wb.sheetnames[0] if wb.sheetnames else '')
        if sheet not in wb.sheetnames:
            errores.append({'producto': item.get('nombre_producto'), 'error': f'Hoja {sheet} no existe'})
            continue
        col = int(item.get('col_index') or col_letter_to_index(item.get('columna') or ''))
        row = int(item.get('fila') or item.get('fila_inicio') or 0)
        # Si no hay fila explícita, solo se valida el catálogo; no se escribe para evitar dañar la plantilla.
        if not col or not row:
            continue
        try:
            safe_cell(wb[sheet], row, col).value = item.get('cantidad') or item.get('nombre_producto')
            escritos += 1
        except Exception as exc:
            errores.append({'producto': item.get('nombre_producto'), 'error': str(exc)})
    return {'productos_escritos': escritos, 'errores_productos': errores}

def apply_mapping_to_copy(database_path: str, file_path: str, mapping: list[dict], unidad: str, output_folder: str, limit: int = 20, tipo_formato: str | None = None, productos: list[dict] | None = None) -> dict:
    validation = validate_mapping(mapping, strict=True)
    if not validation['valido']:
        return {'ok': False, 'validation': validation, 'archivo': None, 'total_usuarios': 0}

    users = get_users_by_unit(database_path, unidad, limit=limit)
    wb = load_workbook(file_path)
    product_result = apply_products_to_workbook(wb, productos, wb.sheetnames[0] if wb.sheetnames else None) if productos else {'productos_escritos': 0, 'errores_productos': []}
    for item in mapping:
        field = item.get('field')
        if not field or field == 'ignorar':
            continue
        sheet = item.get('sheet') or wb.sheetnames[0]
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        col = int(item.get('col') or 0)
        header_row = int(item.get('row') or 0)
        start_row = int(item.get('data_start_row') or item.get('dataStartRow') or (header_row + 1))
        if col <= 0 or start_row <= header_row:
            continue
        # no se insertan filas ni columnas; solo se limpian las celdas mapeadas existentes.
        max_rows_to_touch = max(limit, len(users), 20)
        for i in range(max_rows_to_touch):
            row = start_row + i
            if row > ws.max_row:
                break
            cell = safe_cell(ws, row, col)
            # Evitar tocar el encabezado oficial por seguridad.
            if cell.row <= header_row:
                continue
            cell.value = field_value(users[i], field, i) if i < len(users) else ''

    out_dir = Path(output_folder) / 'motor_plantillas_pruebas'
    out_dir.mkdir(parents=True, exist_ok=True)
    base_name = Path(file_path).stem
    unidad_safe = re.sub(r'[^A-Za-z0-9_]+', '_', normalize_unit(unidad)).strip('_') or 'UNIDAD'
    out_name = f'{unidad_safe}_{base_name}_PRUEBA_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    out_path = out_dir / out_name

    tipo_impresion = infer_print_format(tipo_formato) or infer_print_format(file_path) or infer_print_format(base_name)
    if tipo_impresion:
        aplicar_configuracion_impresion_libro(wb, tipo_impresion, source_name=str(file_path))

    wb.save(out_path)
    return {
        'ok': True,
        'validation': validation,
        'archivo': str(out_path),
        'nombre_archivo': out_name,
        'total_usuarios': len(users),
        'productos': product_result,
    }
