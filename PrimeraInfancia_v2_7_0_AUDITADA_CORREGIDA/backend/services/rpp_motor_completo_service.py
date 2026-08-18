"""
ALPHA54 — utilidades de auditoría y validación del Motor RPP completo.

Este módulo no reemplaza el generador histórico. Aporta funciones pequeñas y
seguras para que la ruta real de generación pueda auditar plantilla vigente,
minuta vigente, mapeo, usuarios y contadores antes de guardar archivos vacíos.
"""
from __future__ import annotations

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def project_backend_dir() -> Path:
    return Path(__file__).resolve().parents[1]


def log_path() -> Path:
    path = project_backend_dir() / "logs" / "rpp_motor_completo.log"
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def log_rpp_event(evento: str, **detalle: Any) -> None:
    """Registra diagnóstico JSONL sin interrumpir la generación."""
    try:
        record = {"ts": now_iso(), "evento": evento, **detalle}
        with log_path().open("a", encoding="utf-8") as fh:
            fh.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")
    except Exception:
        pass


def parse_json(value: Any, default: Any = None) -> Any:
    if value in (None, ""):
        return default
    if isinstance(value, (list, dict)):
        return value
    try:
        return json.loads(str(value))
    except Exception:
        return default


def normalize_text(value: Any) -> str:
    import re
    import unicodedata
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


REQUIRED_RPP_FIELDS = {
    "consecutivo",
    "documento_beneficiario",
    "documento",
    "nombre_completo",
    "primer_nombre",
    "primer_apellido",
}


def mapping_items_from_entry(entry: dict[str, Any]) -> list[dict[str, Any]]:
    items = []
    for key in ("mapeo", "mapping"):
        if isinstance(entry.get(key), list):
            items.extend(entry.get(key) or [])
    items.extend(parse_json(entry.get("mapeo_json"), []) or [])
    return [item for item in items if isinstance(item, dict)]


def product_items_from_entry(entry: dict[str, Any]) -> list[dict[str, Any]]:
    items = []
    for key in ("productos", "products"):
        if isinstance(entry.get(key), list):
            items.extend(entry.get(key) or [])
    items.extend(parse_json(entry.get("productos_json"), []) or [])
    return [item for item in items if isinstance(item, dict)]


def summarize_mapping(entry: dict[str, Any]) -> dict[str, Any]:
    mapping = mapping_items_from_entry(entry)
    campos = set()
    secciones = set()
    obligatorios_sin_fuente = []
    for item in mapping:
        campo = item.get("campo") or item.get("field") or item.get("suggested_field")
        if campo:
            campos.add(str(campo))
        seccion = item.get("seccion") or item.get("section")
        if seccion:
            secciones.add(str(seccion))
        if item.get("obligatorio") and not (item.get("fuente") or item.get("source") or item.get("campo") or item.get("field")):
            obligatorios_sin_fuente.append(item)
    return {
        "total_mapeos": len(mapping),
        "campos": sorted(campos),
        "secciones": sorted(secciones),
        "obligatorios_sin_fuente": obligatorios_sin_fuente,
    }


def validate_rpp_context(entry: dict[str, Any], usuarios: list[dict[str, Any]], minuta: Any, categoria: str | None = None) -> tuple[bool, list[str], dict[str, Any]]:
    """Valida las precondiciones mínimas para no exportar RPP vacío."""
    errors: list[str] = []
    if not entry or not entry.get("ruta"):
        errors.append("No existe plantilla RPP vigente.")
    elif not os.path.exists(str(entry.get("ruta"))):
        errors.append(f"La ruta de plantilla RPP vigente no existe: {entry.get('ruta')}")

    mapping_summary = summarize_mapping(entry)
    # Las plantillas históricas de manifest no siempre tienen mapeo JSON. Solo se
    # exige mapeo cuando la plantilla viene del Motor de Plantillas versionado.
    if entry.get("version_id") or entry.get("plantilla_oficial_version_id"):
        if mapping_summary["total_mapeos"] <= 0:
            errors.append("La plantilla RPP vigente no tiene mapeo completo.")
        if mapping_summary["obligatorios_sin_fuente"]:
            errors.append("La plantilla RPP tiene columnas obligatorias sin fuente definida.")

    if minuta is None:
        errors.append("No existe minuta RPP vigente para el mes seleccionado.")
    if not usuarios:
        errors.append("No hay usuarios para la UDS/grupo seleccionado.")

    summary = {
        "categoria_rpp": categoria,
        "plantilla": os.path.basename(str(entry.get("ruta") or "")),
        "version_id": entry.get("version_id") or entry.get("plantilla_oficial_version_id"),
        "version": entry.get("version"),
        "usuarios": len(usuarios or []),
        "mapeo": mapping_summary,
        "productos_plantilla": len(product_items_from_entry(entry)),
    }
    return (not errors), errors, summary


def workbook_has_user_data(path: str | os.PathLike[str]) -> bool:
    """Auditoría liviana: verifica si un xlsx generado tiene valores no vacíos.

    No se usa como ruta principal de validación porque abrir de nuevo un Excel es
    costoso; queda para pruebas/diagnóstico controlado.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value not in (None, ""):
                        return True
    except Exception:
        return False
    return False
