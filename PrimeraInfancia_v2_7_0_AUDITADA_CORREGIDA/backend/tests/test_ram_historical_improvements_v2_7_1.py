from __future__ import annotations

import tempfile
from copy import copy
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = ROOT / "backend" / "seed_data" / "templates_originales" / "oficiales" / "plantilla_ram_oficial_v2_historica.xlsx"


def main() -> None:
    import sys
    sys.path.insert(0, str(ROOT / "backend"))
    from services.ram_historical_service import generate_ram_historical, sha256_file

    users = [
        {"tipo_documento":"CC","numero_documento":"1001","primer_nombre":"GESTANTE","primer_apellido":"ZETA","edad_meses":300,"tipo_beneficiario":"GESTANTE"},
        {"tipo_documento":"RC","numero_documento":"0000000000123","primer_nombre":"BEBE","primer_apellido":"ALFA","fecha_nacimiento":"2025-11-01"},
        {"tipo_documento":"TI","numero_documento":"000456","primer_nombre":"MAYOR","primer_apellido":"BETA","fecha_nacimiento":"2022-01-01"},
    ]
    metadata = {"entidad":"EAS QA","unidad":"CONONDO","mes_numero":1,"anio":2026,"nui_uds":"NUI-REAL","codigo_cuentame":"CUENTAME-REAL","agente_educativo":"DOCENTE QA","documento_agente":"000999","telefono_uds":"3000000000"}
    before = sha256_file(TEMPLATE)
    with tempfile.TemporaryDirectory(prefix="pi-ram-v2-final-") as tmp:
        output = Path(tmp) / "ram.xlsx"
        generate_ram_historical(TEMPLATE, output, users, 2026, 1, metadata=metadata, attendance_provider=lambda _: {"lunes"})
        source_wb = load_workbook(TEMPLATE, data_only=False)
        output_wb = load_workbook(output, data_only=False)
        source = source_wb["FORMATO RAM V2 HISTORICO"]
        ws = output_wb["FORMATO RAM V2 HISTORICO"]
        assert ws["A6"].value == "MES: ENERO"
        assert "DOCENTE QA" in ws["F6"].value and "000999" in ws["I6"].value
        assert "CUENTAME-REAL" in ws["F7"].value and "3000000000" in ws["T8"].value
        assert (ws["B15"].value, ws["C15"].value) == ("RC", "0000000000123")
        assert ws["C15"].data_type == "s"
        assert [ws.cell(row, 2).value for row in range(15, 18)] == ["RC", "TI", "CC"]
        assert [ws.cell(row, 1).value for row in range(15, 35)] == list(range(1, 21))
        assert ws["H14"].alignment.horizontal == "center" and ws["I14"].alignment.vertical == "center"
        assert all(ws.cell(14, col).alignment.textRotation == 90 for col in range(10, 35))
        assert ws["AI12"].alignment.textRotation == 90 and ws["AK9"].alignment.textRotation == 90
        # A = asistencia. La marca histórica H fue retirada del formato.
        marks = [ws.cell(row, col) for row in range(15, 18) for col in range(10, 35) if ws.cell(row, col).value == "A"]
        assert marks and all(copy(cell.font) == copy(ws.cell(cell.row, 10).font) for cell in marks)
        assert (ws["E36"].value, ws["E37"].value, ws["E38"].value) == (1, 1, 1)
        assert all(str(ws[c].font.color.rgb).upper().endswith("000000") for c in ("E36", "E37", "E38"))
        assert {str(x) for x in source.merged_cells.ranges} == {str(x) for x in ws.merged_cells.ranges}
        assert {k:v.width for k,v in source.column_dimensions.items()} == {k:v.width for k,v in ws.column_dimensions.items()}
        source_wb.close(); output_wb.close()
    assert sha256_file(TEMPLATE) == before
    app_source = (ROOT / "backend" / "app.py").read_text(encoding="utf-8")
    assert "if es_ram_oficial:" in app_source and "generate_ram_historical" in app_source
    print("OK: mejoras RAM V2 histórico verificadas")


if __name__ == "__main__":
    main()
