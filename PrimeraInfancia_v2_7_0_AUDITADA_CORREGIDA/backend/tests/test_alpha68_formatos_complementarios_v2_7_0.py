"""Regresión del selector Alpha68 sin importar Flask ni ejecutar DDL."""

import ast
import json
import os
import tempfile
import traceback
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from werkzeug.utils import secure_filename


APP_PATH = Path(__file__).resolve().parents[1] / "app.py"
TARGETS = {
    "_alpha68_should_generar",
    "_alpha68_valor_usuario",
    "_alpha68_usuario_nombre",
    "_alpha68_grupo_usuario",
    "_alpha68_docente_por_usuarios",
    "_alpha68_guardar_workbook_registrado",
    "_alpha68_generar_listado_usuarios",
    "_alpha68_generar_relacion_mensual",
    "_alpha68_cargar_config_distribucion",
    "_alpha68_generar_distribucion_alimentos",
    "_alpha68_generar_complementarios_formato",
}


def _load_targets():
    tree = ast.parse(APP_PATH.read_text(encoding="utf-8"), filename=str(APP_PATH))
    functions = [
        node for node in tree.body
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)) and node.name in TARGETS
    ]
    assert {node.name for node in functions} == TARGETS
    namespace = {}
    exec(compile(ast.Module(body=functions, type_ignores=[]), str(APP_PATH), "exec"), namespace)
    return namespace


def _run(selection):
    ns = _load_targets()
    calls = []
    ns.update({
        "_alpha68_log": lambda *args, **kwargs: None,
        "_alpha68_generar_listado_usuarios": lambda *args: calls.append("listado_usuarios") or "listado.xlsx",
        "_alpha68_generar_relacion_mensual": lambda *args: calls.append("relacion_mensual") or "relacion.xlsx",
        "_alpha68_generar_distribucion_alimentos": lambda *args: calls.append("distribucion_alimentos") or "distribucion.xlsx",
    })
    generated = ns["_alpha68_generar_complementarios_formato"](
        "UDS PRUEBA", [{"Documento": "1"}], {}, selection, 8, 2026
    )
    return calls, generated


def test_seleccion_vacia_genera_paquete_complementario_completo():
    calls, generated = _run(set())
    assert calls == ["listado_usuarios", "relacion_mensual", "distribucion_alimentos"]
    assert len(generated) == 3


def test_seleccion_individual_genera_solamente_lo_solicitado():
    calls, generated = _run({"relacion_mensual"})
    assert calls == ["relacion_mensual"]
    assert generated == ["relacion.xlsx"]


def test_paquete_completo_crea_y_abre_los_tres_xlsx():
    ns = _load_targets()

    def normalizar(value):
        text = unicodedata.normalize("NFKD", str(value or ""))
        return "".join(char for char in text if not unicodedata.combining(char)).lower()

    usuarios = [
        {
            "Documento": "1001",
            "TipoDocumento": "RC",
            "PrimerNombre": "ANA",
            "PrimerApellido": "PEREZ",
            "EdadMeses": 24,
            "Docente": "DOCENTE PRUEBA",
            "Estado": "ACTIVO",
        }
    ]
    with tempfile.TemporaryDirectory() as output:
        ns.update({
            "os": os,
            "json": json,
            "traceback": traceback,
            "secure_filename": secure_filename,
            "OUTPUT_FOLDER": output,
            "normalizar_texto_clave": normalizar,
            "unir_partes": lambda *parts: " ".join(str(part) for part in parts if part),
            "_project_path": lambda name: str(APP_PATH.parent) if name == "backend" else str(APP_PATH.parents[1]),
            "registrar_archivo_generado_alpha57": lambda *args, **kwargs: None,
            "_alpha68_log": lambda *args, **kwargs: None,
            "_alpha67_unidad_slug": lambda value: secure_filename(str(value)).upper(),
        })
        generated = ns["_alpha68_generar_complementarios_formato"](
            "UDS PRUEBA", usuarios, {}, set(), 8, 2026
        )
        assert len(generated) == 3
        books = [load_workbook(path, read_only=True, data_only=True) for path in generated]
        try:
            assert [book.sheetnames[0] for book in books] == [
                "Listado usuarios", "Relación mensual", "Distribución alimentos"
            ]
            assert books[0].active.max_row == 2
            assert books[1].active["B5"].value == 1
            assert books[2].active.max_row >= 6
        finally:
            for book in books:
                book.close()


if __name__ == "__main__":
    test_seleccion_vacia_genera_paquete_complementario_completo()
    test_seleccion_individual_genera_solamente_lo_solicitado()
    test_paquete_completo_crea_y_abre_los_tres_xlsx()
    print("Formatos complementarios Alpha68 v2.7.0: PASS")
