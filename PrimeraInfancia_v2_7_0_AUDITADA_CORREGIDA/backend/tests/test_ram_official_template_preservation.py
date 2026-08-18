from __future__ import annotations

import json
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
BACKEND = ROOT / "backend"
sys.path.insert(0, str(BACKEND))

from services.ram_v3_service import (  # noqa: E402
    ATTENDANCE_COLUMNS,
    DATA_ROWS,
    RAM_SHEET,
    attendance_cell_for_date,
    generate_ram_v3,
)
from services.ram_visual_qa import compare_ram_visual_integrity  # noqa: E402


TEMPLATE = BACKEND / "seed_data" / "templates_originales" / "oficiales" / "plantilla_ram_oficial_v3.xlsx"
OUTPUT = ROOT / "data" / "output" / "RAM_PRUEBA_QA_2026_08.xlsx"
TOTAL_COORDS = {"E36", "E37", "E38"}
AGE_HEADER_COORDS = {"H14", "I14"}


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def dimension_signature(ws) -> dict:
    return {
        "columns": {
            key: (dim.width, dim.hidden, dim.bestFit, dim.outlineLevel, dim.collapsed)
            for key, dim in ws.column_dimensions.items()
        },
        "rows": {
            key: (dim.height, dim.hidden, dim.outlineLevel, dim.collapsed)
            for key, dim in ws.row_dimensions.items()
        },
    }


def page_signature(ws) -> tuple:
    def header_footer(item) -> tuple:
        return tuple(
            (part.text, part.font, part.size, part.color)
            for part in (item.left, item.center, item.right)
        )

    return (
        str(ws.print_area), str(ws.print_title_rows), str(ws.print_title_cols),
        copy(ws.page_margins), copy(ws.page_setup), copy(ws.print_options),
        header_footer(ws.oddHeader), header_footer(ws.evenHeader), header_footer(ws.firstHeader),
        header_footer(ws.oddFooter), header_footer(ws.evenFooter), header_footer(ws.firstFooter),
        ws.freeze_panes, copy(ws.sheet_properties), copy(ws.sheet_format),
    )


def visual_signature(cell) -> tuple:
    return (
        copy(cell.font), copy(cell.fill), copy(cell.border), copy(cell.alignment),
        cell.number_format, copy(cell.protection), cell.quotePrefix,
    )


def font_without_color(cell) -> tuple:
    font = cell.font
    return (
        font.name, font.sz, font.b, font.i, font.vertAlign, font.underline,
        font.strike, font.outline, font.shadow, font.condense, font.extend,
        font.family, font.charset, font.scheme,
    )


def alignment_without_axes(cell) -> tuple:
    alignment = cell.alignment
    return (
        alignment.textRotation, alignment.wrapText, alignment.shrinkToFit,
        alignment.indent, alignment.relativeIndent, alignment.justifyLastLine,
        alignment.readingOrder,
    )


def run() -> dict:
    users = [
        # Entrada deliberadamente desordenada: gestante, 4 años, 18 meses,
        # 8 meses y 2 meses. El generador debe producir el orden oficial.
        {
            "tipo_documento": "CC", "numero_documento": "1000000001",
            "primer_nombre": "MARIA", "primer_apellido": "GESTANTE",
            "edad_meses": 300, "tipo_beneficiario": "MADRE GESTANTE",
            "asistencias": {"2026-08-17": "A"},
        },
        {
            "tipo_documento": "CE", "numero_documento": "0009876543210987654321",
            "primer_nombre": "SAMUEL", "primer_apellido": "LARGO",
            "fecha_nacimiento": "2022-08-01", "tipo_beneficiario": "NIÑO",
            "fecha_retiro": "2026-08-25", "codigo_retiro": "V",
            "asistencias": {"2026-08-24": "A"},
        },
        {
            "tipo_documento": "PPT", "numero_documento": "0007000000000000042",
            "primer_nombre": "EMMA", "primer_apellido": "DIECIOCHO",
            "fecha_nacimiento": "2025-02-01", "tipo_beneficiario": "NIÑA",
            "asistencias": {"2026-08-10": "A"},
        },
        {
            "tipo_documento": "TI", "numero_documento": "001234567890",
            "primer_nombre": "MATEO", "primer_apellido": "OCHO",
            "fecha_nacimiento": "2025-12-01", "tipo_beneficiario": "NIÑO",
            "asistencias": {"2026-08-06": "A"},
        },
        {
            "tipo_documento": "RC", "numero_documento": "00000001234567890123",
            "primer_nombre": "LUNA", "segundo_nombre": "SOFIA",
            "primer_apellido": "PRUEBA", "segundo_apellido": "RAM",
            "fecha_nacimiento": "2026-06-01", "tipo_beneficiario": "NIÑO",
            "asistencias": {"2026-08-03": "A", "2026-08-04": "I"},
            "novedades": {"2026-08-04": "Cita médica"},
        },
    ]
    metadata = {
        "eas_pds": "EAS QA", "nit": "900000000", "unidad": "UCA QA RAM",
        "nui_uds": "NUI-QA-001", "codigo_cuentame": "NUI-QA-001",
        "codigo_uds": "NUI-QA-001", "mes_nombre": "AGOSTO", "anio": 2026,
        "regional": "CHOCÓ", "centro_zonal": "QA", "municipio": "QUIBDÓ",
        "agente_educativo": "AGENTE EDUCATIVA QA",
        "documento_agente": "00000123456789012345",
        "telefono_uds": "3001234567",
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    result = generate_ram_v3(
        TEMPLATE, OUTPUT, users, 2026, 8, metadata=metadata,
        non_service_dates={"2026-08-05"},
    )

    source_wb = load_workbook(TEMPLATE, data_only=False, keep_links=True)
    output_wb = load_workbook(OUTPUT, data_only=False, keep_links=True)
    source = source_wb[RAM_SHEET]
    generated = output_wb[RAM_SHEET]

    qa = compare_ram_visual_integrity(TEMPLATE, OUTPUT)
    require(qa["status"] == "PASS", f"QA visual reutilizable falló: {qa['failures'][:20]}")

    require(source_wb.sheetnames == output_wb.sheetnames, "Cambió el conjunto u orden de hojas.")
    for row in DATA_ROWS:
        require(source.cell(row, 1).value == generated.cell(row, 1).value, f"Cambió la numeración/fórmula oficial A{row}.")
    require(
        {str(item) for item in source.merged_cells.ranges} == {str(item) for item in generated.merged_cells.ranges},
        "Cambió la combinación de celdas de la plantilla RAM.",
    )
    require(dimension_signature(source) == dimension_signature(generated), "Cambiaron anchos de columnas o alturas de filas.")
    require(page_signature(source) == page_signature(generated), "Cambió el área de impresión, márgenes o configuración de página.")

    style_differences: list[str] = []
    mark_coords = {
        generated.cell(row, col).coordinate
        for row in DATA_ROWS for col in ATTENDANCE_COLUMNS
        if generated.cell(row, col).value not in (None, "")
    }
    for row in source.iter_rows():
        for source_cell in row:
            coord = source_cell.coordinate
            generated_cell = generated[coord]
            if coord in TOTAL_COORDS:
                require(copy(source_cell.fill) == copy(generated_cell.fill), f"Cambió el relleno de {coord}.")
                require(copy(source_cell.border) == copy(generated_cell.border), f"Cambió el borde de {coord}.")
                require(source_cell.number_format == generated_cell.number_format, f"Cambió el formato numérico de {coord}.")
                require(font_without_color(source_cell) == font_without_color(generated_cell), f"Cambió la fuente de {coord} fuera del color.")
                require(alignment_without_axes(source_cell) == alignment_without_axes(generated_cell), f"Cambió la alineación de {coord} fuera del centrado.")
                continue
            if coord in AGE_HEADER_COORDS:
                require(copy(source_cell.fill) == copy(generated_cell.fill), f"Cambió el relleno de {coord}.")
                require(copy(source_cell.border) == copy(generated_cell.border), f"Cambió el borde de {coord}.")
                require(copy(source_cell.font) == copy(generated_cell.font), f"Cambió la fuente de {coord}.")
                require(alignment_without_axes(source_cell) == alignment_without_axes(generated_cell), f"Cambió la orientación de {coord}.")
                continue
            if coord in mark_coords:
                require(copy(source_cell.fill) == copy(generated_cell.fill), f"Cambió el relleno de marca {coord}.")
                require(copy(source_cell.border) == copy(generated_cell.border), f"Cambió el borde de marca {coord}.")
                model = generated.cell(generated_cell.row, ATTENDANCE_COLUMNS[0])
                require(copy(generated_cell.font) == copy(model.font), f"Fuente de asistencia no uniforme en {coord}.")
                require(copy(generated_cell.alignment) == copy(model.alignment), f"Alineación de asistencia no uniforme en {coord}.")
                continue
            if visual_signature(source_cell) != visual_signature(generated_cell):
                style_differences.append(coord)
    require(not style_differences, f"Diferencias visuales no autorizadas: {style_differences[:20]}")

    expected = [
        ("RC", "00000001234567890123", "LUNA", "PRUEBA", 0, 2),
        ("TI", "001234567890", "MATEO", "OCHO", 0, 8),
        ("PPT", "0007000000000000042", "EMMA", "DIECIOCHO", 1, 6),
        ("CE", "0009876543210987654321", "SAMUEL", "LARGO", 4, 0),
        ("CC", "1000000001", "MARIA", "GESTANTE", 25, 0),
    ]
    for row, (doc_type, number, first_name, surname, years, months) in zip(DATA_ROWS, expected):
        require(generated.cell(row, 2).value == doc_type, f"Tipo documental incorrecto en B{row}.")
        require(generated.cell(row, 3).value == number, f"Número documental incorrecto en C{row}.")
        require(generated.cell(row, 3).data_type == "s", f"El documento C{row} no quedó como texto.")
        require(generated.cell(row, 4).value == first_name, f"Nombre incorrecto en D{row}.")
        require(generated.cell(row, 6).value == surname, f"Apellido incorrecto en F{row}.")
        require((generated.cell(row, 8).value, generated.cell(row, 9).value) == (years, months), f"Edad incorrecta en fila {row}.")

    require(generated["A6"].value == "MES: AGOSTO", "MES no fue diligenciado desde el período.")
    require(generated["F7"].value == "Código CUENTAME UDS: NUI-QA-001", "NUI/Código Cuéntame incorrecto.")
    require(generated["F6"].value == "Nombre Agente Educativo (a): AGENTE EDUCATIVA QA", "Agente educativo incorrecto.")
    require(generated["I6"].value == "CC: 00000123456789012345", "Cédula del agente incorrecta.")
    require(generated["T8"].value == "Teléfono UDS: 3001234567", "Teléfono del agente incorrecto.")

    attendance_expectations = {
        (15, "2026-08-03"): "A", (15, "2026-08-04"): "I",
        (16, "2026-08-06"): "A", (17, "2026-08-10"): "A",
        (18, "2026-08-24"): "A", (18, "2026-08-25"): "/",
        (19, "2026-08-17"): "A",
    }
    from datetime import date
    for (row, iso), mark in attendance_expectations.items():
        _, col = attendance_cell_for_date(date.fromisoformat(iso))
        require(generated.cell(row, col).value == mark, f"Asistencia incorrecta en {generated.cell(row, col).coordinate}.")

    require(generated["AK18"].value == "V", "Causa de retiro incorrecta.")
    require((generated["E36"].value, generated["E37"].value, generated["E38"].value) == (1, 3, 1), "Totales etarios/gestantes incorrectos.")
    for coord in TOTAL_COORDS:
        cell = generated[coord]
        require(cell.alignment.horizontal == "center" and cell.alignment.vertical == "center", f"{coord} no está centrada.")
        color = cell.font.color
        require(color is not None and color.type == "rgb" and str(color.rgb).upper().endswith("000000"), f"{coord} no tiene texto negro.")

    # Los encabezados oficiales de días y sus estilos deben ser idénticos.
    for col in ATTENDANCE_COLUMNS:
        require(source.cell(14, col).value == generated.cell(14, col).value, f"Cambió el día en {source.cell(14, col).coordinate}.")
        require(visual_signature(source.cell(14, col)) == visual_signature(generated.cell(14, col)), f"Cambió el estilo del día en {source.cell(14, col).coordinate}.")
        require(generated.cell(14, col).alignment.textRotation == 90, f"El día {generated.cell(14, col).coordinate} perdió orientación vertical.")
    for coord in ("H14", "I14"):
        require(generated[coord].alignment.horizontal == "center" and generated[coord].alignment.vertical == "center", f"{coord} no quedó centrada.")
    require(generated["AI12"].alignment.textRotation == 90, "ASISTENCIAS perdió orientación vertical.")
    require(generated["AK9"].alignment.textRotation == 90, "CAUSA DE RETIRO perdió orientación vertical.")

    source_wb.close()
    output_wb.close()
    report = {
        "ok": True,
        "archivo": str(OUTPUT),
        "participantes": len(users),
        "tipos_documento": [item[0] for item in expected],
        "documentos_como_texto": True,
        "totales": {"menores_6_meses": 1, "mayores_6_meses": 3, "gestantes": 1},
        "encabezado": {"mes": "PASS", "nui_codigo_cuentame": "PASS", "agente": "PASS", "cedula": "PASS", "telefono": "PASS"},
        "orden_grupo_etario": "PASS",
        "retiro": "PASS",
        "diferencias_formato_no_autorizadas": [],
        "merges_preservados": True,
        "dimensiones_preservadas": True,
        "estilos_preservados": True,
        "impresion_preservada": True,
        "warnings": result.get("warnings") or [],
        "qa_visual": qa,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return report


if __name__ == "__main__":
    run()
