"""Persistencia y productos del Centro Inteligente de Supervisión, Auditoría y Calidad."""
from __future__ import annotations

import csv
import io
import json
import os
from modules.dbapi_compat import sqlite3
import zipfile
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from modules.seguridad.tenant_context import tenant_storage_root
from modules.motor_gestion_proyecto.services import source_key as mgp_source_key, unit_key as mgp_unit_key

from .schema import DEFAULT_CHECKLIST, SCHEMA_SQL, SCHEMA_VERSION
from .services import (
    ACTION_STATES,
    FINDING_STATES,
    PLAN_STATES,
    RISK_LEVELS,
    SUPERVISION_STATES,
    VERIFICATION_RESULTS,
    compliance,
    file_sha256,
    finding_code,
    is_overdue,
    json_dump,
    normalize,
    now_iso,
    parse_json,
    plan_code,
    safe_state,
    unit_key,
    valid_date,
)


class CentroSupervisionRepository:
    def __init__(self, database_path: str, data_dir: str, output_folder: str):
        self.database_path = database_path
        self.data_dir = Path(data_dir)
        self.output_folder = Path(output_folder)

    def connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.database_path, timeout=30)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys=ON")
        conn.execute("PRAGMA busy_timeout=30000")
        return conn

    @staticmethod
    def _row(row: sqlite3.Row | None) -> dict[str, Any] | None:
        if not row:
            return None
        data = dict(row)
        for field in ("metadata_json", "detalle_json"):
            if field in data:
                data[field[:-5] if field.endswith("_json") else field] = parse_json(data.get(field), {})
        return data

    def init_schema(self) -> None:
        now = now_iso()
        with self.connect() as conn:
            conn.executescript(SCHEMA_SQL)
            conn.execute(
                "INSERT INTO csc_schema_version(id, version, fecha_actualizacion) VALUES(1, ?, ?) "
                "ON CONFLICT(id) DO UPDATE SET version=excluded.version, fecha_actualizacion=excluded.fecha_actualizacion",
                (SCHEMA_VERSION, now),
            )
            conn.commit()

    def ensure_catalog(self, fundacion_id: int, user_id: int | None = None) -> None:
        now = now_iso()
        with self.connect() as conn:
            for codigo, componente, categoria, criterio, evidencia, riesgo, orden in DEFAULT_CHECKLIST:
                conn.execute(
                    """
                    INSERT INTO csc_checklist_catalogo
                    (fundacion_id,codigo,componente,categoria,criterio,descripcion,evidencia_sugerida,nivel_riesgo,obligatoria,activa,orden,creada_por,fecha_creacion,fecha_actualizacion)
                    VALUES(?,?,?,?,?,?,?,?,1,1,?,?,?,?)
                    ON CONFLICT(fundacion_id,codigo) DO UPDATE SET
                        componente=excluded.componente,categoria=excluded.categoria,criterio=excluded.criterio,
                        evidencia_sugerida=excluded.evidencia_sugerida,nivel_riesgo=excluded.nivel_riesgo,
                        orden=excluded.orden,activa=1,fecha_actualizacion=excluded.fecha_actualizacion
                    """,
                    (fundacion_id, codigo, componente, categoria, criterio, criterio, evidencia, riesgo, orden, user_id, now, now),
                )
            conn.commit()

    def audit(self, fundacion_id: int, user: dict[str, Any], action: str, entity: str, entity_id: int | None, detail: Any = None) -> None:
        with self.connect() as conn:
            conn.execute(
                "INSERT INTO csc_auditoria(fundacion_id,usuario_id,usuario,accion,entidad,entidad_id,detalle_json,fecha) VALUES(?,?,?,?,?,?,?,?)",
                (fundacion_id, user.get("id"), user.get("username"), action, entity, entity_id, json_dump(detail or {}), now_iso()),
            )
            conn.commit()

    def expedientes(self, fundacion_id: int) -> list[dict[str, Any]]:
        with self.connect() as conn:
            rows = conn.execute(
                "SELECT id,unidad_id,unidad_nombre,unidad_clave,codigo_unidad,contrato,vigencia,servicio_modalidad,fase_actual,coordinador_id,coordinador_nombre,porcentaje_global,semaforo "
                "FROM giu_expedientes_uca WHERE fundacion_id=? AND estado='ACTIVO' ORDER BY unidad_nombre,vigencia DESC",
                (fundacion_id,),
            ).fetchall()
        return [dict(row) for row in rows]

    def _expediente(self, fundacion_id: int, expediente_id: int | None) -> dict[str, Any] | None:
        if not expediente_id:
            return None
        with self.connect() as conn:
            row = conn.execute(
                "SELECT * FROM giu_expedientes_uca WHERE fundacion_id=? AND id=?",
                (fundacion_id, expediente_id),
            ).fetchone()
        return dict(row) if row else None

    def dashboard(self, fundacion_id: int, filters: dict[str, Any] | None = None) -> dict[str, Any]:
        filters = filters or {}
        clauses = ["fundacion_id=?", "activa=1"]
        params: list[Any] = [fundacion_id]
        if filters.get("vigencia"):
            clauses.append("vigencia=?")
            params.append(str(filters["vigencia"]))
        if filters.get("unidad"):
            clauses.append("unidad_clave=?")
            params.append(unit_key(filters["unidad"]))
        where = " AND ".join(clauses)
        with self.connect() as conn:
            supervisions = [dict(r) for r in conn.execute(f"SELECT * FROM csc_supervisiones WHERE {where} ORDER BY COALESCE(fecha_programada,fecha_creacion) DESC LIMIT 300", params).fetchall()]
            findings = [dict(r) for r in conn.execute("SELECT * FROM csc_hallazgos WHERE fundacion_id=? AND activa=1 ORDER BY fecha_deteccion DESC LIMIT 500", (fundacion_id,)).fetchall()]
            plans = [dict(r) for r in conn.execute("SELECT * FROM csc_planes_mejora WHERE fundacion_id=? AND activa=1 ORDER BY COALESCE(fecha_limite,fecha_creacion) LIMIT 500", (fundacion_id,)).fetchall()]
            actions = [dict(r) for r in conn.execute("SELECT * FROM csc_acciones_mejora WHERE fundacion_id=? ORDER BY COALESCE(fecha_limite,fecha_creacion) LIMIT 1000", (fundacion_id,)).fetchall()]
            mgp = conn.execute(
                "SELECT COUNT(*) total, SUM(CASE WHEN estado NOT IN ('APROBADA','CERRADA','CANCELADA') AND fecha_limite < ? THEN 1 ELSE 0 END) vencidas "
                "FROM mgp_tareas WHERE fundacion_id=? AND activa=1",
                (date.today().isoformat(), fundacion_id),
            ).fetchone()
        open_findings = [r for r in findings if r["estado"] not in {"CERRADO", "DESCARTADO"}]
        critical = [r for r in open_findings if r.get("nivel_riesgo") == "CRITICO"]
        overdue_actions = [r for r in actions if is_overdue(r.get("fecha_limite"), r.get("estado"))]
        state_counts = Counter(r.get("estado") for r in supervisions)
        risk_counts = Counter(r.get("nivel_riesgo") for r in open_findings)
        return {
            "resumen": {
                "supervisiones": len(supervisions),
                "supervisiones_abiertas": sum(1 for r in supervisions if r.get("estado") not in {"CERRADA", "CANCELADA"}),
                "hallazgos_abiertos": len(open_findings),
                "hallazgos_criticos": len(critical),
                "planes_abiertos": sum(1 for r in plans if r.get("estado") not in {"CERRADO", "CANCELADO"}),
                "acciones_vencidas": len(overdue_actions),
                "tareas_motor": int((mgp or {"total": 0})["total"] or 0),
                "tareas_motor_vencidas": int((mgp or {"vencidas": 0})["vencidas"] or 0),
            },
            "por_estado": dict(state_counts),
            "por_riesgo": dict(risk_counts),
            "supervisiones": supervisions,
            "hallazgos": findings,
            "planes": plans,
            "acciones_vencidas": overdue_actions,
        }

    def create_supervision(self, fundacion_id: int, payload: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        expediente = self._expediente(fundacion_id, int(payload.get("expediente_id") or 0))
        unit_name = str(payload.get("unidad_nombre") or (expediente or {}).get("unidad_nombre") or "").strip()
        if not unit_name:
            raise ValueError("Selecciona una UCA para la supervisión.")
        self.ensure_catalog(fundacion_id, user.get("id"))
        now = now_iso()
        with self.connect() as conn:
            cur = conn.execute(
                """
                INSERT INTO csc_supervisiones
                (fundacion_id,expediente_id,unidad_id,unidad_nombre,unidad_clave,contrato,vigencia,coordinador_nombre,tipo,modalidad,titulo,objetivo,alcance,fecha_programada,estado,supervisor_id,supervisor_nombre,metadata_json,creada_por,actualizada_por,fecha_creacion,fecha_actualizacion)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    fundacion_id, (expediente or {}).get("id"), (expediente or {}).get("unidad_id"), unit_name,
                    unit_key(unit_name), str(payload.get("contrato") or (expediente or {}).get("contrato") or ""),
                    str(payload.get("vigencia") or (expediente or {}).get("vigencia") or date.today().year),
                    str((expediente or {}).get("coordinador_nombre") or payload.get("coordinador_nombre") or ""),
                    safe_state(payload.get("tipo"), {"SEGUIMIENTO_INTERNO","AUDITORIA_INTERNA","PREPARACION_INTERVENTORIA","VISITA_CALIDAD","VERIFICACION_REMOTA"}, "SEGUIMIENTO_INTERNO"),
                    safe_state(payload.get("modalidad"), {"REMOTA","EN_SITIO","MIXTA"}, "REMOTA"),
                    str(payload.get("titulo") or f"Supervisión {unit_name}"), str(payload.get("objetivo") or ""), str(payload.get("alcance") or ""),
                    valid_date(payload.get("fecha_programada")), "PROGRAMADA" if payload.get("fecha_programada") else "BORRADOR",
                    payload.get("supervisor_id") or user.get("id"), str(payload.get("supervisor_nombre") or user.get("nombre_completo") or user.get("username") or ""),
                    json_dump(payload.get("metadata") or {}), user.get("id"), user.get("id"), now, now,
                ),
            )
            supervision_id = int(cur.lastrowid)
            catalog = conn.execute(
                "SELECT * FROM csc_checklist_catalogo WHERE fundacion_id=? AND activa=1 ORDER BY orden,id",
                (fundacion_id,),
            ).fetchall()
            for item in catalog:
                conn.execute(
                    """
                    INSERT INTO csc_verificaciones
                    (fundacion_id,supervision_id,criterio_id,codigo_criterio,componente,categoria,criterio,resultado,nivel_riesgo,evidencia_requerida,fecha_creacion,fecha_actualizacion)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (fundacion_id, supervision_id, item["id"], item["codigo"], item["componente"], item["categoria"], item["criterio"], "PENDIENTE", item["nivel_riesgo"], item["obligatoria"], now, now),
                )
            conn.commit()
        self._sync_motor_task(fundacion_id, "csc_supervisiones", supervision_id, unit_name, expediente, payload.get("fecha_programada"), f"Realizar {payload.get('titulo') or 'supervisión'}", user)
        self.audit(fundacion_id, user, "CREAR_SUPERVISION", "csc_supervisiones", supervision_id, {"unidad": unit_name})
        return self.supervision(fundacion_id, supervision_id) or {}

    def list_supervisions(self, fundacion_id: int, filters: dict[str, Any] | None = None) -> list[dict[str, Any]]:
        filters = filters or {}
        where = ["fundacion_id=?", "activa=1"]
        params: list[Any] = [fundacion_id]
        for field in ("estado", "vigencia", "tipo"):
            if filters.get(field):
                where.append(f"{field}=?")
                params.append(str(filters[field]))
        if filters.get("unidad"):
            where.append("unidad_clave=?")
            params.append(unit_key(filters["unidad"]))
        with self.connect() as conn:
            rows = conn.execute(f"SELECT * FROM csc_supervisiones WHERE {' AND '.join(where)} ORDER BY COALESCE(fecha_programada,fecha_creacion) DESC", params).fetchall()
        return [self._decorate_supervision(dict(row)) for row in rows]

    def supervision(self, fundacion_id: int, supervision_id: int) -> dict[str, Any] | None:
        with self.connect() as conn:
            row = conn.execute("SELECT * FROM csc_supervisiones WHERE fundacion_id=? AND id=? AND activa=1", (fundacion_id, supervision_id)).fetchone()
            if not row:
                return None
            checks = [dict(r) for r in conn.execute("SELECT * FROM csc_verificaciones WHERE fundacion_id=? AND supervision_id=? ORDER BY componente,categoria,id", (fundacion_id, supervision_id)).fetchall()]
            findings = [dict(r) for r in conn.execute("SELECT * FROM csc_hallazgos WHERE fundacion_id=? AND supervision_id=? AND activa=1 ORDER BY fecha_deteccion DESC", (fundacion_id, supervision_id)).fetchall()]
            products = [dict(r) for r in conn.execute("SELECT * FROM csc_productos WHERE fundacion_id=? AND supervision_id=? ORDER BY fecha_generacion DESC", (fundacion_id, supervision_id)).fetchall()]
        data = dict(row)
        data["verificaciones"] = checks
        data["hallazgos"] = findings
        data["productos"] = products
        data["porcentaje_cumplimiento"] = compliance(checks)
        return self._decorate_supervision(data)

    def _decorate_supervision(self, data: dict[str, Any]) -> dict[str, Any]:
        data["metadata"] = parse_json(data.get("metadata_json"), {})
        data["vencida"] = is_overdue(data.get("fecha_programada"), data.get("estado"))
        return data

    def update_supervision(self, fundacion_id: int, supervision_id: int, payload: dict[str, Any], user: dict[str, Any], allow_close: bool = False) -> dict[str, Any]:
        current = self.supervision(fundacion_id, supervision_id)
        if not current:
            raise LookupError("Supervisión no encontrada.")
        requested = safe_state(payload.get("estado"), SUPERVISION_STATES, current.get("estado") or "BORRADOR")
        if requested == "CERRADA" and not allow_close:
            raise PermissionError("El cierre requiere un rol de coordinación autorizado.")
        if requested == "CERRADA":
            pending = [r for r in current.get("verificaciones", []) if r.get("resultado") == "PENDIENTE"]
            if pending:
                raise ValueError("No se puede cerrar con criterios pendientes.")
            open_findings = [r for r in current.get("hallazgos", []) if r.get("estado") not in {"CERRADO", "DESCARTADO"}]
            if open_findings:
                raise ValueError("No se puede cerrar mientras existan hallazgos abiertos.")
        now = now_iso()
        fields = {
            "titulo": payload.get("titulo", current.get("titulo")),
            "objetivo": payload.get("objetivo", current.get("objetivo")),
            "alcance": payload.get("alcance", current.get("alcance")),
            "fecha_programada": valid_date(payload.get("fecha_programada")) if "fecha_programada" in payload else current.get("fecha_programada"),
            "estado": requested,
            "resultado_general": payload.get("resultado_general", current.get("resultado_general")),
            "supervisor_nombre": payload.get("supervisor_nombre", current.get("supervisor_nombre")),
        }
        with self.connect() as conn:
            conn.execute(
                "UPDATE csc_supervisiones SET titulo=?,objetivo=?,alcance=?,fecha_programada=?,estado=?,resultado_general=?,supervisor_nombre=?,porcentaje_cumplimiento=?,actualizada_por=?,fecha_actualizacion=?,fecha_inicio=COALESCE(fecha_inicio,CASE WHEN ?='EN_EJECUCION' THEN ? END),fecha_fin=CASE WHEN ?='CERRADA' THEN ? ELSE fecha_fin END WHERE fundacion_id=? AND id=?",
                (fields["titulo"], fields["objetivo"], fields["alcance"], fields["fecha_programada"], fields["estado"], fields["resultado_general"], fields["supervisor_nombre"], compliance(current.get("verificaciones", [])), user.get("id"), now, requested, now, requested, now, fundacion_id, supervision_id),
            )
            conn.commit()
        self.audit(fundacion_id, user, "ACTUALIZAR_SUPERVISION", "csc_supervisiones", supervision_id, {"estado": requested})
        return self.supervision(fundacion_id, supervision_id) or {}

    def update_verification(self, fundacion_id: int, verification_id: int, payload: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        with self.connect() as conn:
            row = conn.execute("SELECT * FROM csc_verificaciones WHERE fundacion_id=? AND id=?", (fundacion_id, verification_id)).fetchone()
            if not row:
                raise LookupError("Criterio de verificación no encontrado.")
            current = dict(row)
            result = safe_state(payload.get("resultado"), VERIFICATION_RESULTS, current.get("resultado") or "PENDIENTE")
            risk = safe_state(payload.get("nivel_riesgo"), RISK_LEVELS, current.get("nivel_riesgo") or "MEDIO")
            requires_finding = 1 if result == "NO_CUMPLE" else int(bool(payload.get("requiere_hallazgo", current.get("requiere_hallazgo"))))
            now = now_iso()
            conn.execute(
                "UPDATE csc_verificaciones SET resultado=?,nivel_riesgo=?,observaciones=?,requiere_hallazgo=?,evaluado_por=?,evaluado_por_nombre=?,fecha_evaluacion=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?",
                (result, risk, str(payload.get("observaciones") or ""), requires_finding, user.get("id"), user.get("nombre_completo") or user.get("username"), now, now, fundacion_id, verification_id),
            )
            supervision_id = int(current["supervision_id"])
            checks = [dict(r) for r in conn.execute("SELECT resultado FROM csc_verificaciones WHERE fundacion_id=? AND supervision_id=?", (fundacion_id, supervision_id)).fetchall()]
            conn.execute("UPDATE csc_supervisiones SET porcentaje_cumplimiento=?,actualizada_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?", (compliance(checks), user.get("id"), now, fundacion_id, supervision_id))
            conn.commit()
        self.audit(fundacion_id, user, "EVALUAR_CRITERIO", "csc_verificaciones", verification_id, {"resultado": result, "riesgo": risk})
        with self.connect() as conn:
            updated = conn.execute("SELECT * FROM csc_verificaciones WHERE fundacion_id=? AND id=?", (fundacion_id, verification_id)).fetchone()
        return dict(updated)

    def _next_sequence(self, conn: sqlite3.Connection, table: str, fundacion_id: int) -> int:
        row = conn.execute(f"SELECT COALESCE(MAX(id),0)+1 AS n FROM {table} WHERE fundacion_id=?", (fundacion_id,)).fetchone()
        return int(row["n"] or 1)

    def create_finding(self, fundacion_id: int, payload: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        supervision_id = int(payload.get("supervision_id") or 0) or None
        verification_id = int(payload.get("verificacion_id") or 0) or None
        supervision = self.supervision(fundacion_id, supervision_id) if supervision_id else None
        verification = None
        if verification_id:
            with self.connect() as conn:
                row = conn.execute("SELECT * FROM csc_verificaciones WHERE fundacion_id=? AND id=?", (fundacion_id, verification_id)).fetchone()
                verification = dict(row) if row else None
        now = now_iso()
        with self.connect() as conn:
            sequence = self._next_sequence(conn, "csc_hallazgos", fundacion_id)
            code = finding_code(supervision_id, sequence)
            cur = conn.execute(
                """
                INSERT INTO csc_hallazgos
                (fundacion_id,supervision_id,verificacion_id,expediente_id,unidad_nombre,unidad_clave,componente,categoria,codigo,titulo,descripcion,criterio_afectado,nivel_riesgo,tipo,estado,fecha_deteccion,fecha_limite,responsable_id,responsable_nombre,requiere_plan,origen_modulo,origen_id,creada_por,actualizada_por,fecha_creacion,fecha_actualizacion)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    fundacion_id, supervision_id, verification_id, (supervision or {}).get("expediente_id"),
                    str(payload.get("unidad_nombre") or (supervision or {}).get("unidad_nombre") or ""),
                    unit_key(payload.get("unidad_nombre") or (supervision or {}).get("unidad_nombre") or ""),
                    str(payload.get("componente") or (verification or {}).get("componente") or "Administrativo y de Gestión"),
                    str(payload.get("categoria") or (verification or {}).get("categoria") or "Supervisión"), code,
                    str(payload.get("titulo") or f"Hallazgo {code}"), str(payload.get("descripcion") or (verification or {}).get("observaciones") or "").strip(),
                    str(payload.get("criterio_afectado") or (verification or {}).get("criterio") or ""),
                    safe_state(payload.get("nivel_riesgo") or (verification or {}).get("nivel_riesgo"), RISK_LEVELS, "MEDIO"),
                    safe_state(payload.get("tipo"), {"NO_CONFORMIDAD","OBSERVACION","OPORTUNIDAD_MEJORA","RIESGO","INCUMPLIMIENTO"}, "NO_CONFORMIDAD"),
                    "ABIERTO", valid_date(payload.get("fecha_deteccion")) or date.today().isoformat(), valid_date(payload.get("fecha_limite")),
                    payload.get("responsable_id"), str(payload.get("responsable_nombre") or ""), 1 if payload.get("requiere_plan", True) else 0,
                    str(payload.get("origen_modulo") or "SUPERVISION"), payload.get("origen_id"), user.get("id"), user.get("id"), now, now,
                ),
            )
            finding_id = int(cur.lastrowid)
            conn.commit()
        self._sync_motor_task(fundacion_id, "csc_hallazgos", finding_id, str(payload.get("unidad_nombre") or (supervision or {}).get("unidad_nombre") or ""), supervision, payload.get("fecha_limite"), f"Atender hallazgo {code}", user, priority="CRITICA" if normalize(payload.get("nivel_riesgo") or (verification or {}).get("nivel_riesgo")) == "CRITICO" else "ALTA")
        self.audit(fundacion_id, user, "CREAR_HALLAZGO", "csc_hallazgos", finding_id, {"codigo": code})
        return self.finding(fundacion_id, finding_id) or {}

    def finding(self, fundacion_id: int, finding_id: int) -> dict[str, Any] | None:
        with self.connect() as conn:
            row = conn.execute("SELECT * FROM csc_hallazgos WHERE fundacion_id=? AND id=? AND activa=1", (fundacion_id, finding_id)).fetchone()
            if not row:
                return None
            plans = [dict(r) for r in conn.execute("SELECT * FROM csc_planes_mejora WHERE fundacion_id=? AND hallazgo_id=? AND activa=1 ORDER BY fecha_creacion", (fundacion_id, finding_id)).fetchall()]
            followups = [dict(r) for r in conn.execute("SELECT * FROM csc_seguimientos WHERE fundacion_id=? AND entidad_tipo='HALLAZGO' AND entidad_id=? ORDER BY fecha DESC", (fundacion_id, finding_id)).fetchall()]
        data = dict(row)
        data["planes"] = plans
        data["seguimientos"] = followups
        data["vencido"] = is_overdue(data.get("fecha_limite"), data.get("estado"))
        return data

    def list_findings(self, fundacion_id: int, filters: dict[str, Any] | None = None) -> list[dict[str, Any]]:
        filters = filters or {}
        where = ["fundacion_id=?", "activa=1"]
        params: list[Any] = [fundacion_id]
        for field in ("estado", "nivel_riesgo", "componente", "supervision_id"):
            if filters.get(field) not in (None, ""):
                where.append(f"{field}=?")
                params.append(filters[field])
        if filters.get("unidad"):
            where.append("unidad_clave=?")
            params.append(unit_key(filters["unidad"]))
        with self.connect() as conn:
            rows = conn.execute(f"SELECT * FROM csc_hallazgos WHERE {' AND '.join(where)} ORDER BY CASE nivel_riesgo WHEN 'CRITICO' THEN 1 WHEN 'ALTO' THEN 2 WHEN 'MEDIO' THEN 3 ELSE 4 END, fecha_deteccion DESC", params).fetchall()
        return [{**dict(r), "vencido": is_overdue(r["fecha_limite"], r["estado"])} for r in rows]

    def update_finding(self, fundacion_id: int, finding_id: int, payload: dict[str, Any], user: dict[str, Any], allow_close: bool = False) -> dict[str, Any]:
        current = self.finding(fundacion_id, finding_id)
        if not current:
            raise LookupError("Hallazgo no encontrado.")
        requested = safe_state(payload.get("estado"), FINDING_STATES, current.get("estado") or "ABIERTO")
        if requested == "CERRADO":
            if not allow_close:
                raise PermissionError("El hallazgo solo puede cerrarse mediante validación de coordinación.")
            plans = current.get("planes") or []
            if current.get("requiere_plan") and not plans:
                raise ValueError("El hallazgo requiere un plan de mejora antes del cierre.")
            if any(p.get("estado") != "CERRADO" for p in plans):
                raise ValueError("Todos los planes asociados deben estar cerrados y validados.")
            if not str(payload.get("motivo_cierre") or "").strip():
                raise ValueError("Registra el motivo y la evidencia de validación del cierre.")
        now = now_iso()
        with self.connect() as conn:
            conn.execute(
                """
                UPDATE csc_hallazgos SET estado=?,nivel_riesgo=?,fecha_limite=?,responsable_id=?,responsable_nombre=?,resolucion_propuesta=?,resolucion_validada=?,validado_por=?,fecha_validacion=?,cerrado_por=?,fecha_cierre=?,motivo_cierre=?,actualizada_por=?,fecha_actualizacion=?
                WHERE fundacion_id=? AND id=?
                """,
                (
                    requested, safe_state(payload.get("nivel_riesgo"), RISK_LEVELS, current.get("nivel_riesgo") or "MEDIO"),
                    valid_date(payload.get("fecha_limite")) if "fecha_limite" in payload else current.get("fecha_limite"),
                    payload.get("responsable_id", current.get("responsable_id")), payload.get("responsable_nombre", current.get("responsable_nombre")),
                    payload.get("resolucion_propuesta", current.get("resolucion_propuesta")), payload.get("resolucion_validada", current.get("resolucion_validada")),
                    user.get("id") if requested in {"RESUELTO_PENDIENTE_VALIDACION", "CERRADO"} else current.get("validado_por"),
                    now if requested in {"RESUELTO_PENDIENTE_VALIDACION", "CERRADO"} else current.get("fecha_validacion"),
                    user.get("id") if requested == "CERRADO" else current.get("cerrado_por"), now if requested == "CERRADO" else current.get("fecha_cierre"),
                    payload.get("motivo_cierre", current.get("motivo_cierre")), user.get("id"), now, fundacion_id, finding_id,
                ),
            )
            conn.commit()
        self._update_motor_state(fundacion_id, "csc_hallazgos", finding_id, "CERRADA" if requested in {"CERRADO", "DESCARTADO"} else "EN_PROCESO")
        self.audit(fundacion_id, user, "ACTUALIZAR_HALLAZGO", "csc_hallazgos", finding_id, {"estado": requested})
        return self.finding(fundacion_id, finding_id) or {}

    def create_plan(self, fundacion_id: int, payload: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        finding_id = int(payload.get("hallazgo_id") or 0) or None
        finding = self.finding(fundacion_id, finding_id) if finding_id else None
        now = now_iso()
        with self.connect() as conn:
            sequence = self._next_sequence(conn, "csc_planes_mejora", fundacion_id)
            code = plan_code(finding_id, sequence)
            cur = conn.execute(
                """
                INSERT INTO csc_planes_mejora
                (fundacion_id,hallazgo_id,expediente_id,unidad_nombre,unidad_clave,codigo,nombre,objetivo,alcance,responsable_id,responsable_nombre,fecha_inicio,fecha_limite,estado,progreso,indicador_resultado,meta,observaciones,creada_por,actualizada_por,fecha_creacion,fecha_actualizacion)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,'BORRADOR',0,?,?,?,?,?,?,?)
                """,
                (
                    fundacion_id, finding_id, (finding or {}).get("expediente_id"), str(payload.get("unidad_nombre") or (finding or {}).get("unidad_nombre") or ""),
                    unit_key(payload.get("unidad_nombre") or (finding or {}).get("unidad_nombre") or ""), code,
                    str(payload.get("nombre") or f"Plan de mejora {code}"), str(payload.get("objetivo") or "").strip(), str(payload.get("alcance") or ""),
                    payload.get("responsable_id"), str(payload.get("responsable_nombre") or ""), valid_date(payload.get("fecha_inicio")) or date.today().isoformat(),
                    valid_date(payload.get("fecha_limite")), str(payload.get("indicador_resultado") or ""), str(payload.get("meta") or ""), str(payload.get("observaciones") or ""),
                    user.get("id"), user.get("id"), now, now,
                ),
            )
            plan_id = int(cur.lastrowid)
            if finding_id:
                conn.execute("UPDATE csc_hallazgos SET estado='EN_PLAN',actualizada_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=? AND estado='ABIERTO'", (user.get("id"), now, fundacion_id, finding_id))
            conn.commit()
        self._sync_motor_task(fundacion_id, "csc_planes_mejora", plan_id, str(payload.get("unidad_nombre") or (finding or {}).get("unidad_nombre") or ""), finding, payload.get("fecha_limite"), f"Ejecutar plan de mejora {code}", user, priority="ALTA")
        self.audit(fundacion_id, user, "CREAR_PLAN_MEJORA", "csc_planes_mejora", plan_id, {"codigo": code, "hallazgo_id": finding_id})
        return self.plan(fundacion_id, plan_id) or {}

    def plan(self, fundacion_id: int, plan_id: int) -> dict[str, Any] | None:
        with self.connect() as conn:
            row = conn.execute("SELECT * FROM csc_planes_mejora WHERE fundacion_id=? AND id=? AND activa=1", (fundacion_id, plan_id)).fetchone()
            if not row:
                return None
            actions = [dict(r) for r in conn.execute("SELECT * FROM csc_acciones_mejora WHERE fundacion_id=? AND plan_id=? ORDER BY id", (fundacion_id, plan_id)).fetchall()]
            followups = [dict(r) for r in conn.execute("SELECT * FROM csc_seguimientos WHERE fundacion_id=? AND entidad_tipo='PLAN' AND entidad_id=? ORDER BY fecha DESC", (fundacion_id, plan_id)).fetchall()]
        data = dict(row)
        data["acciones"] = actions
        data["seguimientos"] = followups
        data["vencido"] = is_overdue(data.get("fecha_limite"), data.get("estado"))
        return data

    def list_plans(self, fundacion_id: int, filters: dict[str, Any] | None = None) -> list[dict[str, Any]]:
        filters = filters or {}
        where = ["fundacion_id=?", "activa=1"]
        params: list[Any] = [fundacion_id]
        if filters.get("estado"):
            where.append("estado=?")
            params.append(filters["estado"])
        if filters.get("unidad"):
            where.append("unidad_clave=?")
            params.append(unit_key(filters["unidad"]))
        with self.connect() as conn:
            rows = conn.execute(f"SELECT * FROM csc_planes_mejora WHERE {' AND '.join(where)} ORDER BY COALESCE(fecha_limite,fecha_creacion)", params).fetchall()
        return [{**dict(r), "vencido": is_overdue(r["fecha_limite"], r["estado"])} for r in rows]

    def update_plan(self, fundacion_id: int, plan_id: int, payload: dict[str, Any], user: dict[str, Any], allow_review: bool = False) -> dict[str, Any]:
        current = self.plan(fundacion_id, plan_id)
        if not current:
            raise LookupError("Plan no encontrado.")
        requested = safe_state(payload.get("estado"), PLAN_STATES, current.get("estado") or "BORRADOR")
        if requested in {"APROBADO", "CERRADO"} and not allow_review:
            raise PermissionError("La aprobación y el cierre requieren coordinación.")
        actions = current.get("acciones") or []
        if requested == "CERRADO":
            if not actions:
                raise ValueError("El plan debe tener al menos una acción.")
            if any(a.get("estado") != "COMPLETADA" for a in actions):
                raise ValueError("Todas las acciones deben estar completadas y validadas.")
            if any(int(a.get("evidencia_requerida") or 0) and int(a.get("evidencias_total") or 0) == 0 for a in actions):
                raise ValueError("Las acciones que requieren evidencia deben tener soportes.")
        progress = float(payload.get("progreso", current.get("progreso") or 0) or 0)
        if actions:
            progress = round(sum(float(a.get("progreso") or 0) for a in actions) / len(actions), 2)
        now = now_iso()
        with self.connect() as conn:
            conn.execute(
                """
                UPDATE csc_planes_mejora SET nombre=?,objetivo=?,alcance=?,responsable_id=?,responsable_nombre=?,fecha_inicio=?,fecha_limite=?,estado=?,progreso=?,indicador_resultado=?,meta=?,observaciones=?,revisado_por=?,fecha_revision=?,aprobado_por=?,fecha_aprobacion=?,cerrado_por=?,fecha_cierre=?,actualizada_por=?,fecha_actualizacion=?
                WHERE fundacion_id=? AND id=?
                """,
                (
                    payload.get("nombre", current.get("nombre")), payload.get("objetivo", current.get("objetivo")), payload.get("alcance", current.get("alcance")),
                    payload.get("responsable_id", current.get("responsable_id")), payload.get("responsable_nombre", current.get("responsable_nombre")),
                    valid_date(payload.get("fecha_inicio")) if "fecha_inicio" in payload else current.get("fecha_inicio"),
                    valid_date(payload.get("fecha_limite")) if "fecha_limite" in payload else current.get("fecha_limite"), requested, progress,
                    payload.get("indicador_resultado", current.get("indicador_resultado")), payload.get("meta", current.get("meta")), payload.get("observaciones", current.get("observaciones")),
                    user.get("id") if requested == "PENDIENTE_APROBACION" else current.get("revisado_por"), now if requested == "PENDIENTE_APROBACION" else current.get("fecha_revision"),
                    user.get("id") if requested == "APROBADO" else current.get("aprobado_por"), now if requested == "APROBADO" else current.get("fecha_aprobacion"),
                    user.get("id") if requested == "CERRADO" else current.get("cerrado_por"), now if requested == "CERRADO" else current.get("fecha_cierre"),
                    user.get("id"), now, fundacion_id, plan_id,
                ),
            )
            if requested == "CERRADO" and current.get("hallazgo_id"):
                conn.execute("UPDATE csc_hallazgos SET estado='RESUELTO_PENDIENTE_VALIDACION',actualizada_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=? AND estado!='CERRADO'", (user.get("id"), now, fundacion_id, current["hallazgo_id"]))
            conn.commit()
        self._update_motor_state(fundacion_id, "csc_planes_mejora", plan_id, "CERRADA" if requested == "CERRADO" else "EN_PROCESO")
        self.audit(fundacion_id, user, "ACTUALIZAR_PLAN_MEJORA", "csc_planes_mejora", plan_id, {"estado": requested})
        return self.plan(fundacion_id, plan_id) or {}

    def create_action(self, fundacion_id: int, plan_id: int, payload: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        plan = self.plan(fundacion_id, plan_id)
        if not plan:
            raise LookupError("Plan no encontrado.")
        now = now_iso()
        with self.connect() as conn:
            cur = conn.execute(
                """
                INSERT INTO csc_acciones_mejora
                (fundacion_id,plan_id,titulo,descripcion,responsable_id,responsable_nombre,fecha_inicio,fecha_limite,estado,progreso,evidencia_requerida,creada_por,actualizada_por,fecha_creacion,fecha_actualizacion)
                VALUES(?,?,?,?,?,?,?,?,'PENDIENTE',0,?,?,?,?,?)
                """,
                (fundacion_id, plan_id, str(payload.get("titulo") or "Acción de mejora"), str(payload.get("descripcion") or ""), payload.get("responsable_id"), str(payload.get("responsable_nombre") or ""), valid_date(payload.get("fecha_inicio")) or date.today().isoformat(), valid_date(payload.get("fecha_limite")), 1 if payload.get("evidencia_requerida", True) else 0, user.get("id"), user.get("id"), now, now),
            )
            action_id = int(cur.lastrowid)
            conn.commit()
        self._sync_motor_task(fundacion_id, "csc_acciones_mejora", action_id, plan.get("unidad_nombre") or "", plan, payload.get("fecha_limite"), str(payload.get("titulo") or "Acción de mejora"), user, priority="ALTA")
        self.audit(fundacion_id, user, "CREAR_ACCION_MEJORA", "csc_acciones_mejora", action_id, {"plan_id": plan_id})
        return self.action(fundacion_id, action_id) or {}

    def action(self, fundacion_id: int, action_id: int) -> dict[str, Any] | None:
        with self.connect() as conn:
            row = conn.execute("SELECT * FROM csc_acciones_mejora WHERE fundacion_id=? AND id=?", (fundacion_id, action_id)).fetchone()
        return dict(row) if row else None

    def update_action(self, fundacion_id: int, action_id: int, payload: dict[str, Any], user: dict[str, Any], allow_validate: bool = False) -> dict[str, Any]:
        current = self.action(fundacion_id, action_id)
        if not current:
            raise LookupError("Acción no encontrada.")
        requested = safe_state(payload.get("estado"), ACTION_STATES, current.get("estado") or "PENDIENTE")
        if requested == "COMPLETADA":
            if not allow_validate:
                requested = "PENDIENTE_VALIDACION"
            elif int(current.get("evidencia_requerida") or 0) and int(current.get("evidencias_total") or 0) == 0:
                raise ValueError("Carga la evidencia requerida antes de validar la acción.")
        progress = min(100.0, max(0.0, float(payload.get("progreso", 100 if requested == "COMPLETADA" else current.get("progreso") or 0))))
        now = now_iso()
        with self.connect() as conn:
            conn.execute(
                "UPDATE csc_acciones_mejora SET titulo=?,descripcion=?,responsable_id=?,responsable_nombre=?,fecha_inicio=?,fecha_limite=?,estado=?,progreso=?,resultado=?,validado_por=?,fecha_validacion=?,actualizada_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?",
                (payload.get("titulo", current.get("titulo")), payload.get("descripcion", current.get("descripcion")), payload.get("responsable_id", current.get("responsable_id")), payload.get("responsable_nombre", current.get("responsable_nombre")), valid_date(payload.get("fecha_inicio")) if "fecha_inicio" in payload else current.get("fecha_inicio"), valid_date(payload.get("fecha_limite")) if "fecha_limite" in payload else current.get("fecha_limite"), requested, progress, payload.get("resultado", current.get("resultado")), user.get("id") if requested == "COMPLETADA" else current.get("validado_por"), now if requested == "COMPLETADA" else current.get("fecha_validacion"), user.get("id"), now, fundacion_id, action_id),
            )
            actions = [dict(r) for r in conn.execute("SELECT progreso FROM csc_acciones_mejora WHERE fundacion_id=? AND plan_id=?", (fundacion_id, current["plan_id"])).fetchall()]
            plan_progress = round(sum(float(a.get("progreso") or 0) for a in actions) / len(actions), 2) if actions else 0
            conn.execute("UPDATE csc_planes_mejora SET progreso=?,estado=CASE WHEN estado='BORRADOR' THEN 'EN_EJECUCION' ELSE estado END,actualizada_por=?,fecha_actualizacion=? WHERE fundacion_id=? AND id=?", (plan_progress, user.get("id"), now, fundacion_id, current["plan_id"]))
            conn.commit()
        self._update_motor_state(fundacion_id, "csc_acciones_mejora", action_id, "CERRADA" if requested == "COMPLETADA" else "EN_PROCESO")
        self.audit(fundacion_id, user, "ACTUALIZAR_ACCION_MEJORA", "csc_acciones_mejora", action_id, {"estado": requested})
        return self.action(fundacion_id, action_id) or {}

    def add_followup(self, fundacion_id: int, entity_type: str, entity_id: int, payload: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        entity_type = safe_state(entity_type, {"HALLAZGO", "PLAN", "ACCION", "SUPERVISION"}, "HALLAZGO")
        now = now_iso()
        with self.connect() as conn:
            cur = conn.execute(
                "INSERT INTO csc_seguimientos(fundacion_id,entidad_tipo,entidad_id,tipo,fecha,descripcion,resultado,proxima_fecha,estado,usuario_id,usuario_nombre,fecha_creacion) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                (fundacion_id, entity_type, entity_id, safe_state(payload.get("tipo"), {"SEGUIMIENTO","VISITA","VERIFICACION","COMITE","LLAMADA","CIERRE"}, "SEGUIMIENTO"), valid_date(payload.get("fecha")) or date.today().isoformat(), str(payload.get("descripcion") or "").strip(), str(payload.get("resultado") or ""), valid_date(payload.get("proxima_fecha")), "REGISTRADO", user.get("id"), user.get("nombre_completo") or user.get("username"), now),
            )
            follow_id = int(cur.lastrowid)
            conn.commit()
        self.audit(fundacion_id, user, "REGISTRAR_SEGUIMIENTO", "csc_seguimientos", follow_id, {"entidad": entity_type, "entidad_id": entity_id})
        with self.connect() as conn:
            row = conn.execute("SELECT * FROM csc_seguimientos WHERE fundacion_id=? AND id=?", (fundacion_id, follow_id)).fetchone()
        return dict(row)

    def list_followups(self, fundacion_id: int, entity_type: str, entity_id: int) -> list[dict[str, Any]]:
        with self.connect() as conn:
            rows = conn.execute("SELECT * FROM csc_seguimientos WHERE fundacion_id=? AND entidad_tipo=? AND entidad_id=? ORDER BY fecha DESC,id DESC", (fundacion_id, normalize(entity_type), entity_id)).fetchall()
        return [dict(r) for r in rows]

    def add_evidence(self, fundacion_id: int, entity_type: str, entity_id: int, file, description: str, user: dict[str, Any]) -> dict[str, Any]:
        entity_type = safe_state(entity_type, {"SUPERVISION","VERIFICACION","HALLAZGO","PLAN","ACCION"}, "SUPERVISION")
        original = os.path.basename(str(file.filename or "evidencia"))
        extension = Path(original).suffix.lower()
        if extension not in {".pdf", ".doc", ".docx", ".xlsx", ".xls", ".csv", ".txt", ".png", ".jpg", ".jpeg", ".webp", ".zip"}:
            raise ValueError("Tipo de archivo no permitido.")
        root = tenant_storage_root(self.data_dir, fundacion_id) / "centro_supervision_calidad" / entity_type.lower() / str(entity_id)
        root.mkdir(parents=True, exist_ok=True)
        now = now_iso()
        safe_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{Path(original).name}"
        path = root / safe_name
        file.save(path)
        size = path.stat().st_size
        if size <= 0 or size > 50 * 1024 * 1024:
            path.unlink(missing_ok=True)
            raise ValueError("La evidencia está vacía o supera 50 MB.")
        sha = file_sha256(path)
        with self.connect() as conn:
            version_row = conn.execute("SELECT COALESCE(MAX(version),0)+1 n FROM csc_evidencias WHERE fundacion_id=? AND entidad_tipo=? AND entidad_id=? AND nombre_original=?", (fundacion_id, entity_type, entity_id, original)).fetchone()
            cur = conn.execute(
                "INSERT INTO csc_evidencias(fundacion_id,entidad_tipo,entidad_id,nombre_original,nombre_guardado,ruta_archivo,mime_type,tamano_bytes,sha256,version,descripcion,cargada_por,cargada_por_nombre,fecha_carga) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (fundacion_id, entity_type, entity_id, original, safe_name, str(path), getattr(file, "mimetype", None), size, sha, int(version_row["n"] or 1), description, user.get("id"), user.get("nombre_completo") or user.get("username"), now),
            )
            evidence_id = int(cur.lastrowid)
            if entity_type == "VERIFICACION":
                conn.execute("UPDATE csc_verificaciones SET evidencias_total=(SELECT COUNT(*) FROM csc_evidencias WHERE fundacion_id=? AND entidad_tipo='VERIFICACION' AND entidad_id=?),fecha_actualizacion=? WHERE fundacion_id=? AND id=?", (fundacion_id, entity_id, now, fundacion_id, entity_id))
            elif entity_type == "ACCION":
                conn.execute("UPDATE csc_acciones_mejora SET evidencias_total=(SELECT COUNT(*) FROM csc_evidencias WHERE fundacion_id=? AND entidad_tipo='ACCION' AND entidad_id=?),fecha_actualizacion=? WHERE fundacion_id=? AND id=?", (fundacion_id, entity_id, now, fundacion_id, entity_id))
            conn.commit()
        self.audit(fundacion_id, user, "CARGAR_EVIDENCIA", "csc_evidencias", evidence_id, {"entidad": entity_type, "entidad_id": entity_id, "sha256": sha})
        return self.evidence(fundacion_id, evidence_id) or {}

    def evidence(self, fundacion_id: int, evidence_id: int) -> dict[str, Any] | None:
        with self.connect() as conn:
            row = conn.execute("SELECT * FROM csc_evidencias WHERE fundacion_id=? AND id=?", (fundacion_id, evidence_id)).fetchone()
        return dict(row) if row else None

    def list_evidence(self, fundacion_id: int, entity_type: str, entity_id: int) -> list[dict[str, Any]]:
        with self.connect() as conn:
            rows = conn.execute("SELECT id,entidad_tipo,entidad_id,nombre_original,mime_type,tamano_bytes,sha256,version,descripcion,cargada_por_nombre,fecha_carga FROM csc_evidencias WHERE fundacion_id=? AND entidad_tipo=? AND entidad_id=? ORDER BY fecha_carga DESC", (fundacion_id, normalize(entity_type), entity_id)).fetchall()
        return [dict(r) for r in rows]

    def evidence_path(self, fundacion_id: int, evidence_id: int) -> tuple[Path, str, str | None] | None:
        row = self.evidence(fundacion_id, evidence_id)
        if not row:
            return None
        try:
            path = Path(row["ruta_archivo"]).resolve(strict=True)
            path.relative_to(tenant_storage_root(self.data_dir, fundacion_id).resolve())
        except (OSError, ValueError):
            return None
        if file_sha256(path) != row["sha256"]:
            return None
        return path, row["nombre_original"], row.get("mime_type")

    def build_products(self, fundacion_id: int, supervision_id: int, user: dict[str, Any]) -> list[dict[str, Any]]:
        supervision = self.supervision(fundacion_id, supervision_id)
        if not supervision:
            raise LookupError("Supervisión no encontrada.")
        root = tenant_storage_root(self.data_dir, fundacion_id) / "centro_supervision_calidad" / "productos" / f"supervision_{supervision_id}"
        root.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = f"supervision_{supervision_id}_{unit_key(supervision['unidad_nombre']).lower()}_{stamp}"
        xlsx_path = root / f"{base}.xlsx"
        pdf_path = root / f"{base}.pdf"
        zip_path = root / f"{base}.zip"
        self._write_xlsx(supervision, xlsx_path)
        self._write_pdf(supervision, pdf_path)
        self._write_zip(fundacion_id, supervision, zip_path, xlsx_path, pdf_path)
        products: list[dict[str, Any]] = []
        with self.connect() as conn:
            for kind, path, mime in (("MATRIZ_EXCEL", xlsx_path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"), ("INFORME_PDF", pdf_path, "application/pdf"), ("PAQUETE_ZIP", zip_path, "application/zip")):
                sha = file_sha256(path)
                cur = conn.execute(
                    "INSERT INTO csc_productos(fundacion_id,supervision_id,expediente_id,tipo_producto,nombre_archivo,ruta_archivo,mime_type,tamano_bytes,sha256,estado,generado_por,fecha_generacion,metadata_json) VALUES(?,?,?,?,?,?,?,?,?,'BORRADOR',?,?,?)",
                    (fundacion_id, supervision_id, supervision.get("expediente_id"), kind, path.name, str(path), mime, path.stat().st_size, sha, user.get("id"), now_iso(), json_dump({"unidad": supervision["unidad_nombre"], "cumplimiento": supervision.get("porcentaje_cumplimiento")})),
                )
                products.append({"id": int(cur.lastrowid), "tipo_producto": kind, "nombre_archivo": path.name, "sha256": sha, "estado": "BORRADOR"})
            conn.commit()
        self.audit(fundacion_id, user, "GENERAR_PRODUCTOS_SUPERVISION", "csc_supervisiones", supervision_id, {"productos": [p["tipo_producto"] for p in products]})
        return products

    def list_products(self, fundacion_id: int, supervision_id: int | None = None, limit: int = 300) -> list[dict[str, Any]]:
        where = ["fundacion_id=?"]
        params: list[Any] = [fundacion_id]
        if supervision_id:
            where.append("supervision_id=?")
            params.append(int(supervision_id))
        params.append(max(1, min(int(limit or 300), 1000)))
        with self.connect() as conn:
            rows = conn.execute(
                f"SELECT * FROM csc_productos WHERE {' AND '.join(where)} ORDER BY fecha_generacion DESC,id DESC LIMIT ?",
                params,
            ).fetchall()
        result: list[dict[str, Any]] = []
        for row in rows:
            item = dict(row)
            item["metadata"] = parse_json(item.get("metadata_json"), {})
            item.pop("ruta_archivo", None)
            result.append(item)
        return result

    def product_path(self, fundacion_id: int, product_id: int) -> tuple[Path, str, str | None] | None:
        with self.connect() as conn:
            row = conn.execute("SELECT * FROM csc_productos WHERE fundacion_id=? AND id=?", (fundacion_id, product_id)).fetchone()
        if not row:
            return None
        data = dict(row)
        try:
            path = Path(data["ruta_archivo"]).resolve(strict=True)
            path.relative_to(tenant_storage_root(self.data_dir, fundacion_id).resolve())
        except (OSError, ValueError):
            return None
        if file_sha256(path) != data["sha256"]:
            return None
        return path, data["nombre_archivo"], data.get("mime_type")

    def _write_xlsx(self, supervision: dict[str, Any], path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen"
        ws.append(["CENTRO DE SUPERVISIÓN, AUDITORÍA Y CALIDAD"])
        ws.append(["UCA", supervision.get("unidad_nombre")])
        ws.append(["Contrato", supervision.get("contrato")])
        ws.append(["Vigencia", supervision.get("vigencia")])
        ws.append(["Estado", supervision.get("estado")])
        ws.append(["Cumplimiento", supervision.get("porcentaje_cumplimiento")])
        ws["A1"].font = Font(bold=True, size=14)
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 65
        checks = wb.create_sheet("Lista de verificación")
        checks.append(["Código", "Componente", "Categoría", "Criterio", "Resultado", "Riesgo", "Observaciones", "Evidencias"])
        for row in supervision.get("verificaciones") or []:
            checks.append([row.get("codigo_criterio"), row.get("componente"), row.get("categoria"), row.get("criterio"), row.get("resultado"), row.get("nivel_riesgo"), row.get("observaciones"), row.get("evidencias_total")])
        findings = wb.create_sheet("Hallazgos")
        findings.append(["Código", "Título", "Componente", "Riesgo", "Estado", "Responsable", "Fecha límite", "Descripción"])
        for row in supervision.get("hallazgos") or []:
            findings.append([row.get("codigo"), row.get("titulo"), row.get("componente"), row.get("nivel_riesgo"), row.get("estado"), row.get("responsable_nombre"), row.get("fecha_limite"), row.get("descripcion")])
        for sheet in wb.worksheets:
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="0F766E")
                cell.alignment = Alignment(wrap_text=True)
            sheet.freeze_panes = "A2"
            for col in sheet.columns:
                letter = col[0].column_letter
                sheet.column_dimensions[letter].width = min(60, max(12, max(len(str(c.value or "")) for c in col[:100]) + 2))
        wb.save(path)

    def _write_pdf(self, supervision: dict[str, Any], path: Path) -> None:
        styles = getSampleStyleSheet()
        story = [Paragraph("Centro Inteligente de Supervisión, Auditoría y Calidad", styles["Title"]), Spacer(1, 0.3 * cm)]
        story.append(Paragraph(f"UCA: {supervision.get('unidad_nombre')} · Vigencia: {supervision.get('vigencia')} · Estado: {supervision.get('estado')}", styles["BodyText"]))
        story.append(Paragraph(f"Cumplimiento de la lista: {float(supervision.get('porcentaje_cumplimiento') or 0):.1f}%", styles["Heading2"]))
        data = [["Código", "Componente", "Resultado", "Riesgo"]]
        for row in supervision.get("verificaciones") or []:
            data.append([str(row.get("codigo_criterio") or ""), str(row.get("componente") or "")[:42], str(row.get("resultado") or ""), str(row.get("nivel_riesgo") or "")])
        table = Table(data, colWidths=[3 * cm, 8.5 * cm, 3 * cm, 2.5 * cm], repeatRows=1)
        table.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0F766E")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"), ("GRID", (0,0), (-1,-1), 0.25, colors.grey), ("VALIGN", (0,0), (-1,-1), "TOP"), ("FONTSIZE", (0,0), (-1,-1), 7)]))
        story += [table, Spacer(1, 0.4 * cm), Paragraph("Hallazgos", styles["Heading2"])]
        finding_data = [["Código", "Título", "Riesgo", "Estado"]]
        for row in supervision.get("hallazgos") or []:
            finding_data.append([row.get("codigo"), str(row.get("titulo") or "")[:70], row.get("nivel_riesgo"), row.get("estado")])
        ftable = Table(finding_data, colWidths=[3 * cm, 10 * cm, 2.5 * cm, 3 * cm], repeatRows=1)
        ftable.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#334155")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), 0.25, colors.grey), ("FONTSIZE", (0,0), (-1,-1), 7), ("VALIGN", (0,0), (-1,-1), "TOP")]))
        story.append(ftable)
        SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=1 * cm, leftMargin=1 * cm, topMargin=1 * cm, bottomMargin=1 * cm).build(story)

    def _write_zip(self, fundacion_id: int, supervision: dict[str, Any], path: Path, xlsx_path: Path, pdf_path: Path) -> None:
        with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.write(xlsx_path, xlsx_path.name)
            zf.write(pdf_path, pdf_path.name)
            zf.writestr("00_RESUMEN.json", json.dumps({k: supervision.get(k) for k in ("id","unidad_nombre","contrato","vigencia","estado","porcentaje_cumplimiento")}, ensure_ascii=False, indent=2, default=str))
            buffer = io.StringIO()
            writer = csv.writer(buffer)
            writer.writerow(["codigo", "componente", "categoria", "resultado", "riesgo", "observaciones"])
            for row in supervision.get("verificaciones") or []:
                writer.writerow([row.get("codigo_criterio"), row.get("componente"), row.get("categoria"), row.get("resultado"), row.get("nivel_riesgo"), row.get("observaciones")])
            zf.writestr("01_LISTA_VERIFICACION.csv", buffer.getvalue().encode("utf-8-sig"))
            manifest = io.StringIO()
            mw = csv.writer(manifest)
            mw.writerow(["entidad", "entidad_id", "archivo", "sha256", "fecha"])
            with self.connect() as conn:
                evidence = conn.execute("SELECT * FROM csc_evidencias WHERE fundacion_id=? AND ((entidad_tipo='SUPERVISION' AND entidad_id=?) OR (entidad_tipo='VERIFICACION' AND entidad_id IN (SELECT id FROM csc_verificaciones WHERE fundacion_id=? AND supervision_id=?)) OR (entidad_tipo='HALLAZGO' AND entidad_id IN (SELECT id FROM csc_hallazgos WHERE fundacion_id=? AND supervision_id=?)))", (fundacion_id, supervision["id"], fundacion_id, supervision["id"], fundacion_id, supervision["id"])).fetchall()
            root = tenant_storage_root(self.data_dir, fundacion_id).resolve()
            for row in evidence:
                data = dict(row)
                mw.writerow([data["entidad_tipo"], data["entidad_id"], data["nombre_original"], data["sha256"], data["fecha_carga"]])
                try:
                    fpath = Path(data["ruta_archivo"]).resolve(strict=True)
                    fpath.relative_to(root)
                    if file_sha256(fpath) == data["sha256"]:
                        zf.write(fpath, f"evidencias/{data['entidad_tipo'].lower()}_{data['entidad_id']}/{data['nombre_original']}")
                except (OSError, ValueError):
                    continue
            zf.writestr("02_MANIFIESTO_EVIDENCIAS.csv", manifest.getvalue().encode("utf-8-sig"))
            zf.writestr("LEEME.txt", "Paquete preparado como BORRADOR. No sustituye la validación humana ni la decisión de supervisión/interventoría.\n")

    def _sync_motor_task(self, fundacion_id: int, source_table: str, source_id: int, unit_name: str, source: dict[str, Any] | None, due_date: Any, title: str, user: dict[str, Any], priority: str = "MEDIA") -> None:
        now = now_iso()
        key = mgp_source_key(source_table, source_id)
        expediente_id = (source or {}).get("expediente_id") or (source or {}).get("id") if source_table == "csc_supervisiones" else (source or {}).get("expediente_id")
        with self.connect() as conn:
            try:
                legacy_key = f"{source_table}:{source_id}"
                if legacy_key != key:
                    conn.execute("DELETE FROM mgp_tareas WHERE fundacion_id=? AND fuente_tabla=? AND fuente_clave=?", (fundacion_id, source_table, legacy_key))
                conn.execute(
                    """
                    INSERT INTO mgp_tareas
                    (fundacion_id,expediente_id,unidad_id,unidad_nombre,unidad_clave,fuente_modulo,fuente_tabla,fuente_id,fuente_clave,tipo_tarea,componente,titulo,descripcion,fecha_inicio,fecha_limite,estado,prioridad,puntaje_prioridad,responsable_id,responsable_nombre,requiere_evidencia,metadata_json,activa,creada_por,actualizada_por,fecha_creacion,fecha_actualizacion)
                    VALUES(?,?,?,?,?,'CENTRO_SUPERVISION',?,?,?,?, 'Administrativo y de Gestión',?,?,?,?,'PENDIENTE',?,?,?, ?,1,?,1,?,?,?,?)
                    ON CONFLICT(fundacion_id,fuente_tabla,fuente_clave) DO UPDATE SET titulo=excluded.titulo,fecha_limite=excluded.fecha_limite,prioridad=excluded.prioridad,actualizada_por=excluded.actualizada_por,fecha_actualizacion=excluded.fecha_actualizacion
                    """,
                    (fundacion_id, expediente_id, (source or {}).get("unidad_id"), unit_name, mgp_unit_key(unit_name), source_table, source_id, key, "SUPERVISION" if source_table == "csc_supervisiones" else "PLAN_MEJORA", title, "Tarea generada por el Centro de Supervisión. Requiere revisión humana.", date.today().isoformat(), valid_date(due_date), priority, 75 if priority == "CRITICA" else 45 if priority == "ALTA" else 20, user.get("id"), user.get("nombre_completo") or user.get("username"), json_dump({"source": source_table, "source_id": source_id}), user.get("id"), user.get("id"), now, now),
                )
                conn.commit()
            except sqlite3.OperationalError:
                conn.rollback()

    def _update_motor_state(self, fundacion_id: int, source_table: str, source_id: int, state_value: str) -> None:
        with self.connect() as conn:
            try:
                conn.execute("UPDATE mgp_tareas SET estado=?,fecha_finalizacion=CASE WHEN ?='CERRADA' THEN ? ELSE fecha_finalizacion END,fecha_actualizacion=? WHERE fundacion_id=? AND fuente_tabla=? AND fuente_id=?", (state_value, state_value, date.today().isoformat(), now_iso(), fundacion_id, source_table, source_id))
                conn.commit()
            except sqlite3.OperationalError:
                conn.rollback()
