from __future__ import annotations

import json
import os
import re
from modules.dbapi_compat import sqlite3
from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .schema import RG_SCHEMA_SQL

MESES_ES = {
    1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
    5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
    9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE',
}


def now_iso() -> str:
    return datetime.now().isoformat(timespec='seconds')


def normalizar_texto(valor: Any) -> str:
    import unicodedata
    texto = str(valor or '').strip().lower()
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r'[^a-z0-9]+', ' ', texto)
    return ' '.join(texto.split())


def safe_filename(valor: Any) -> str:
    texto = str(valor or '').strip()
    texto = re.sub(r'[^A-Za-z0-9_\-.]+', '_', texto)
    texto = re.sub(r'_+', '_', texto).strip('_')
    return texto or 'reporte'


def periodo_key(anio: int, mes: int) -> str:
    return f'{int(anio)}-{int(mes):02d}'


class ReportesGerencialesService:
    """Generador de informes gerenciales con redacción ejecutiva, tablas y gráficos.

    El módulo consulta información existente de forma tolerante. Si una tabla no existe
    o aún no tiene datos, genera secciones con alertas y recomendaciones en lugar de fallar.
    """

    def __init__(self, database_path: str, output_folder: str):
        self.database_path = database_path
        self._output_folder = output_folder

    @property
    def output_folder(self) -> Path:
        path = Path(os.fspath(self._output_folder))
        path.mkdir(parents=True, exist_ok=True)
        return path

    @property
    def reportes_folder(self) -> Path:
        path = self.output_folder / 'reportes_gerenciales'
        path.mkdir(parents=True, exist_ok=True)
        return path

    def connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.database_path)
        conn.row_factory = sqlite3.Row
        return conn

    def init_schema(self) -> None:
        conn = self.connect()
        conn.executescript(RG_SCHEMA_SQL)
        conn.commit()
        conn.close()

    def table_exists(self, cur: sqlite3.Cursor, table: str) -> bool:
        row = cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone()
        return bool(row)

    def table_columns(self, cur: sqlite3.Cursor, table: str) -> set[str]:
        if not self.table_exists(cur, table):
            return set()
        return {row['name'] for row in cur.execute(f"PRAGMA table_info({table})").fetchall()}

    def fetch_all(self, table: str, sql: str, params: tuple = ()) -> list[dict[str, Any]]:
        conn = self.connect()
        cur = conn.cursor()
        if not self.table_exists(cur, table):
            conn.close()
            return []
        try:
            rows = [dict(r) for r in cur.execute(sql, params).fetchall()]
        except Exception:
            rows = []
        conn.close()
        return rows

    def get_fundacion_nombre(self, fundacion_id: int | None) -> str:
        rows = self.fetch_all('fundaciones', 'SELECT nombre FROM fundaciones WHERE id=?', (fundacion_id or 1,))
        return (rows[0].get('nombre') if rows else 'Fundación Principal') or 'Fundación Principal'

    def _where_fundacion(self, cur: sqlite3.Cursor, table: str, fundacion_id: int | None) -> tuple[str, list[Any]]:
        cols = self.table_columns(cur, table)
        if fundacion_id and 'fundacion_id' in cols:
            return 'COALESCE(fundacion_id, ?) = ?', [fundacion_id, fundacion_id]
        return '1=1', []

    def get_beneficiarios(self, fundacion_id: int | None = None) -> list[dict[str, Any]]:
        conn = self.connect()
        cur = conn.cursor()
        if not self.table_exists(cur, 'master_ninos'):
            conn.close()
            return []
        where, params = self._where_fundacion(cur, 'master_ninos', fundacion_id)
        rows = [dict(r) for r in cur.execute(
            f'''SELECT *, unidad_servicio AS unidad, documento AS nui
                FROM master_ninos WHERE activo=1 AND {where}''', tuple(params)
        ).fetchall()]
        conn.close()
        return rows

    def get_talento(self, fundacion_id: int | None = None) -> list[dict[str, Any]]:
        conn = self.connect()
        cur = conn.cursor()
        if not self.table_exists(cur, 'master_talento_humano'):
            conn.close()
            return []
        where, params = self._where_fundacion(cur, 'master_talento_humano', fundacion_id)
        rows = [dict(r) for r in cur.execute(
            f'''SELECT *, nombre_completo AS nombre, unidad_servicio AS unidad
                FROM master_talento_humano WHERE activo=1 AND {where}
                ORDER BY unidad_servicio, cargo, nombre_completo''', tuple(params)
        ).fetchall()]
        conn.close()
        return rows

    def _active_beneficiarios(self, beneficiarios: list[dict[str, Any]]) -> list[dict[str, Any]]:
        activos = []
        for b in beneficiarios:
            estado = normalizar_texto(b.get('estado'))
            if not estado or estado in {'activo', 'activa'}:
                activos.append(b)
        return activos

    def _edad_meses(self, b: dict[str, Any]) -> int:
        try:
            return int(float(b.get('edad_meses') or 0))
        except Exception:
            return 0

    def _grupo_edad(self, b: dict[str, Any]) -> str:
        tipo = normalizar_texto(b.get('tipo_beneficiario'))
        edad = self._edad_meses(b)
        if 'gestante' in tipo:
            return 'Gestantes'
        if edad < 6:
            return 'Menores de 6 meses'
        if 6 <= edad <= 11:
            return '6 a 11 meses'
        if 12 <= edad <= 35:
            return '1 a 2 años 11 meses'
        if 36 <= edad <= 71:
            return '3 a 5 años 11 meses'
        return '5 años o más'

    def _estado_entregable_normalizado(self, estado: Any, fecha_limite: Any = None) -> str:
        estado_norm = normalizar_texto(estado).upper().replace(' ', '_')
        if estado_norm in {'APROBADO', 'CUMPLIDO', 'CARGADO', 'VALIDADO'}:
            return 'CUMPLIDO'
        if estado_norm in {'VENCIDO'}:
            return 'VENCIDO'
        if fecha_limite:
            try:
                fecha = datetime.fromisoformat(str(fecha_limite)[:10]).date()
                if fecha < date.today() and estado_norm not in {'APROBADO', 'CUMPLIDO'}:
                    return 'VENCIDO'
            except Exception:
                pass
        if estado_norm in {'EN_PROCESO', 'EN_REVISION', 'CARGADO'}:
            return 'EN PROCESO'
        return 'PENDIENTE'

    def recopilar_datos(self, mes: int, anio: int, fundacion_id: int | None = None) -> dict[str, Any]:
        mes = max(1, min(12, int(mes)))
        anio = int(anio)
        periodo = periodo_key(anio, mes)
        fundacion_nombre = self.get_fundacion_nombre(fundacion_id)
        beneficiarios = self.get_beneficiarios(fundacion_id)
        activos = self._active_beneficiarios(beneficiarios)
        talento = self.get_talento(fundacion_id)

        unidades: dict[str, int] = {}
        grupos: dict[str, int] = {}
        sexo: dict[str, int] = {}
        unidades_incompletas: list[dict[str, Any]] = []
        retiros_edad: list[dict[str, Any]] = []
        for b in activos:
            unidad = (b.get('unidad') or 'SIN UNIDAD').strip() or 'SIN UNIDAD'
            unidades[unidad] = unidades.get(unidad, 0) + 1
            grupo = self._grupo_edad(b)
            grupos[grupo] = grupos.get(grupo, 0) + 1
            s = (b.get('sexo') or 'Sin dato').strip() or 'Sin dato'
            sexo[s] = sexo.get(s, 0) + 1
            edad = self._edad_meses(b)
            if edad >= 60 and 'gestante' not in normalizar_texto(b.get('tipo_beneficiario')):
                retiros_edad.append({
                    'documento': b.get('documento') or b.get('nui'),
                    'nombre': ' '.join(str(x or '').strip() for x in [b.get('nombres') or b.get('nombre'), b.get('apellidos')] if str(x or '').strip()),
                    'unidad': unidad,
                    'edad_meses': edad,
                })
        for unidad, total in unidades.items():
            if 0 < total < 20:
                unidades_incompletas.append({'unidad': unidad, 'total': total, 'faltan': 20 - total})

        roles_talento: dict[str, int] = {}
        for t in talento:
            cargo = (t.get('cargo') or t.get('tipo_equipo') or 'Sin cargo').strip() or 'Sin cargo'
            key = cargo.upper()
            roles_talento[key] = roles_talento.get(key, 0) + 1

        entregables = self._recopilar_entregables(periodo, fundacion_id)
        salud = self._recopilar_salud(fundacion_id)
        cruce = self._recopilar_cruce(fundacion_id)
        auditoria = self._recopilar_auditoria(fundacion_id)

        indicadores = {
            'Beneficiarios activos': len(activos),
            'Total beneficiarios históricos': len(beneficiarios),
            'Unidades con participantes': len(unidades),
            'Unidades con cobertura incompleta': len(unidades_incompletas),
            'Talento humano registrado': len(talento),
            'Coordinadores': sum(v for k, v in roles_talento.items() if 'COORD' in normalizar_texto(k).upper()),
            'Docentes / agentes educativos': sum(v for k, v in roles_talento.items() if any(x in normalizar_texto(k) for x in ['docente', 'agente'])),
            'Entregables del periodo': entregables['total'],
            'Entregables pendientes': entregables['pendientes'],
            'Entregables vencidos': entregables['vencidos'],
            'Documentos por revisar': entregables['por_revisar'],
            'Casos nutricionales críticos': salud['criticos'],
            'Valoraciones nutricionales': salud['valorados'],
            'Alertas abiertas': len(salud['alertas']) + len(entregables['alertas']),
            'Novedades nuevos': cruce.get('nuevos', 0),
            'Novedades retirados': cruce.get('retirados', 0),
            'Novedades cambios': cruce.get('cambios', 0),
        }

        hallazgos = self._generar_hallazgos(indicadores, unidades_incompletas, salud, entregables, cruce)
        alertas = self._generar_alertas(unidades_incompletas, salud, entregables, retiros_edad)
        recomendaciones = self._generar_recomendaciones(indicadores, hallazgos, alertas)
        pendientes = entregables['pendientes_detalle'] + salud['pendientes_detalle']
        responsables = self._resumen_responsables(talento, entregables['items'])
        resumen = self._generar_resumen_ejecutivo(fundacion_nombre, mes, anio, indicadores, hallazgos, alertas)
        conclusion = self._generar_conclusion(fundacion_nombre, mes, anio, indicadores, hallazgos)

        return {
            'fundacion_id': fundacion_id or 1,
            'fundacion_nombre': fundacion_nombre,
            'periodo': periodo,
            'mes': mes,
            'anio': anio,
            'mes_nombre': MESES_ES[mes],
            'indicadores': indicadores,
            'distribucion_unidades': unidades,
            'distribucion_grupos': grupos,
            'distribucion_sexo': sexo,
            'distribucion_talento': roles_talento,
            'unidades_incompletas': unidades_incompletas,
            'retiros_edad': retiros_edad,
            'salud': salud,
            'entregables': entregables,
            'cruce': cruce,
            'auditoria': auditoria,
            'hallazgos': hallazgos,
            'alertas': alertas,
            'recomendaciones': recomendaciones,
            'pendientes': pendientes,
            'responsables': responsables,
            'resumen_ejecutivo': resumen,
            'conclusion': conclusion,
        }

    def _recopilar_entregables(self, periodo: str, fundacion_id: int | None) -> dict[str, Any]:
        conn = self.connect()
        cur = conn.cursor()
        items: list[dict[str, Any]] = []
        por_revisar = 0
        if self.table_exists(cur, 'gp_entregables'):
            cols = self.table_columns(cur, 'gp_entregables')
            where = '1=1'
            params: list[Any] = []
            if 'periodo' in cols:
                where += ' AND (periodo=? OR periodo IS NULL OR periodo="")'
                params.append(periodo)
            if fundacion_id and 'fundacion_id' in cols:
                where += ' AND COALESCE(fundacion_id, ?) = ?'
                params.extend([fundacion_id, fundacion_id])
            try:
                items = [dict(r) for r in cur.execute(f'SELECT * FROM gp_entregables WHERE {where} ORDER BY fecha_limite, estado', tuple(params)).fetchall()]
            except Exception:
                items = []
        if self.table_exists(cur, 'gp_documentos'):
            cols = self.table_columns(cur, 'gp_documentos')
            where = "UPPER(COALESCE(estado,'')) IN ('CARGADO','EN_REVISION','EN REVISIÓN','PENDIENTE')"
            params = []
            if fundacion_id and 'fundacion_id' in cols:
                where += ' AND COALESCE(fundacion_id, ?) = ?'
                params.extend([fundacion_id, fundacion_id])
            try:
                por_revisar = cur.execute(f'SELECT COUNT(*) AS c FROM gp_documentos WHERE {where}', tuple(params)).fetchone()['c']
            except Exception:
                por_revisar = 0
        conn.close()

        total = len(items)
        pendientes = 0
        vencidos = 0
        cumplidos = 0
        pendientes_detalle: list[dict[str, Any]] = []
        alertas: list[dict[str, Any]] = []
        for item in items:
            estado = self._estado_entregable_normalizado(item.get('estado'), item.get('fecha_limite'))
            if estado == 'CUMPLIDO':
                cumplidos += 1
            elif estado == 'VENCIDO':
                vencidos += 1
                alertas.append({'nivel': 'ROJO', 'tipo': 'Entregable vencido', 'mensaje': item.get('titulo') or item.get('tipo') or 'Entregable vencido', 'responsable': item.get('responsable') or '', 'unidad': item.get('unidad') or ''})
                pendientes_detalle.append({'tipo': 'Entregable vencido', 'titulo': item.get('titulo') or item.get('tipo'), 'responsable': item.get('responsable'), 'unidad': item.get('unidad'), 'fecha_limite': item.get('fecha_limite'), 'estado': item.get('estado')})
            else:
                pendientes += 1
                pendientes_detalle.append({'tipo': 'Entregable pendiente', 'titulo': item.get('titulo') or item.get('tipo'), 'responsable': item.get('responsable'), 'unidad': item.get('unidad'), 'fecha_limite': item.get('fecha_limite'), 'estado': item.get('estado')})
        return {
            'items': items,
            'total': total,
            'pendientes': pendientes,
            'vencidos': vencidos,
            'cumplidos': cumplidos,
            'por_revisar': int(por_revisar or 0),
            'pendientes_detalle': pendientes_detalle[:500],
            'alertas': alertas[:200],
        }

    def _recopilar_salud(self, fundacion_id: int | None) -> dict[str, Any]:
        conn = self.connect()
        cur = conn.cursor()
        valoraciones: list[dict[str, Any]] = []
        alertas: list[dict[str, Any]] = []
        if self.table_exists(cur, 'master_salud_nutricion'):
            cols = self.table_columns(cur, 'master_salud_nutricion')
            where = '1=1'
            params: list[Any] = []
            if fundacion_id and 'fundacion_id' in cols:
                where += ' AND COALESCE(fundacion_id, ?) = ?'
                params.extend([fundacion_id, fundacion_id])
            try:
                valoraciones = [dict(r) for r in cur.execute(
                    f'''SELECT *, diagnostico_nutricional AS diagnostico_global,
                               fecha_toma AS fecha_valoracion
                        FROM master_salud_nutricion
                        WHERE activo=1 AND {where} ORDER BY fecha_toma DESC LIMIT 5000''', tuple(params)
                ).fetchall()]
            except Exception:
                valoraciones = []
        if self.table_exists(cur, 'sn_alertas'):
            cols = self.table_columns(cur, 'sn_alertas')
            where = 'COALESCE(atendida,0)=0'
            params = []
            if fundacion_id and 'fundacion_id' in cols:
                where += ' AND COALESCE(fundacion_id, ?) = ?'
                params.extend([fundacion_id, fundacion_id])
            try:
                alertas = [dict(r) for r in cur.execute(f'SELECT * FROM sn_alertas WHERE {where} ORDER BY fecha_alerta DESC LIMIT 500', tuple(params)).fetchall()]
            except Exception:
                alertas = []
        conn.close()

        diagnosticos: dict[str, int] = {}
        criticos = 0
        pendientes = 0
        pendientes_detalle: list[dict[str, Any]] = []
        for v in valoraciones:
            diag = (v.get('diagnostico_global') or v.get('diagnostico') or v.get('estado_nutricional') or 'Sin diagnóstico').strip() or 'Sin diagnóstico'
            diagnosticos[diag] = diagnosticos.get(diag, 0) + 1
            nivel = normalizar_texto(v.get('nivel_alerta') or diag)
            if any(x in nivel for x in ['critico', 'severa', 'desnutricion', 'rojo']):
                criticos += 1
            if any(not v.get(c) for c in ['peso_kg', 'talla_cm', 'fecha_valoracion']):
                pendientes += 1
                pendientes_detalle.append({'tipo': 'Nutrición pendiente', 'titulo': v.get('nombre_completo') or v.get('nombre'), 'responsable': v.get('docente'), 'unidad': v.get('unidad'), 'fecha_limite': v.get('proximo_control'), 'estado': 'PENDIENTE'})
        return {
            'valoraciones': valoraciones,
            'valorados': len(valoraciones),
            'criticos': criticos,
            'pendientes': pendientes,
            'diagnosticos': diagnosticos,
            'alertas': [{'nivel': a.get('nivel'), 'tipo': a.get('tipo'), 'mensaje': a.get('mensaje'), 'unidad': a.get('unidad'), 'responsable': ''} for a in alertas],
            'pendientes_detalle': pendientes_detalle[:300],
        }

    def _recopilar_cruce(self, fundacion_id: int | None) -> dict[str, Any]:
        conn = self.connect()
        cur = conn.cursor()
        result = {'nuevos': 0, 'retirados': 0, 'reemplazados': 0, 'trasladados': 0, 'cambios': 0, 'ultimo': None}
        if self.table_exists(cur, 'cb_cruces'):
            cols = self.table_columns(cur, 'cb_cruces')
            where = '1=1'
            params: list[Any] = []
            if fundacion_id and 'fundacion_id' in cols:
                where += ' AND COALESCE(fundacion_id, ?) = ?'
                params.extend([fundacion_id, fundacion_id])
            try:
                row = cur.execute(f'SELECT * FROM cb_cruces WHERE {where} ORDER BY id DESC LIMIT 1', tuple(params)).fetchone()
                if row:
                    d = dict(row)
                    result['ultimo'] = d
                    for key in ['nuevos', 'retirados', 'reemplazados', 'trasladados', 'cambios']:
                        for col in [f'total_{key}', key]:
                            if col in d:
                                result[key] = int(d.get(col) or 0)
                                break
            except Exception:
                pass
        conn.close()
        return result

    def _recopilar_auditoria(self, fundacion_id: int | None) -> dict[str, Any]:
        conn = self.connect()
        cur = conn.cursor()
        total = 0
        acciones: dict[str, int] = {}
        for table in ['auditoria', 'auditoria_seguridad', 'pm_auditoria', 'backups_auditoria']:
            if not self.table_exists(cur, table):
                continue
            try:
                count = cur.execute(f'SELECT COUNT(*) AS c FROM {table}').fetchone()['c']
                total += int(count or 0)
            except Exception:
                continue
        conn.close()
        return {'total_eventos': total, 'acciones': acciones}

    def _generar_hallazgos(self, indicadores: dict[str, Any], unidades_incompletas: list[dict[str, Any]], salud: dict[str, Any], entregables: dict[str, Any], cruce: dict[str, Any]) -> list[dict[str, str]]:
        hallazgos: list[dict[str, str]] = []
        if unidades_incompletas:
            hallazgos.append({'nivel': 'ALTO', 'titulo': 'Cobertura incompleta en unidades', 'detalle': f'{len(unidades_incompletas)} unidad(es) tienen menos de 20 participantes activos.'})
        if salud.get('criticos'):
            hallazgos.append({'nivel': 'CRITICO', 'titulo': 'Casos nutricionales críticos', 'detalle': f"Se identificaron {salud.get('criticos')} caso(s) críticos o de riesgo nutricional."})
        if entregables.get('vencidos'):
            hallazgos.append({'nivel': 'CRITICO', 'titulo': 'Entregables vencidos', 'detalle': f"Hay {entregables.get('vencidos')} entregable(s) vencido(s) en el periodo."})
        if entregables.get('pendientes'):
            hallazgos.append({'nivel': 'MEDIO', 'titulo': 'Entregables pendientes', 'detalle': f"Hay {entregables.get('pendientes')} entregable(s) pendiente(s) de gestión o evidencia."})
        if cruce.get('nuevos') or cruce.get('retirados') or cruce.get('cambios'):
            hallazgos.append({'nivel': 'MEDIO', 'titulo': 'Novedades de base Cuéntame', 'detalle': f"Último cruce: {cruce.get('nuevos')} nuevos, {cruce.get('retirados')} retirados y {cruce.get('cambios')} cambios."})
        if not hallazgos:
            hallazgos.append({'nivel': 'BAJO', 'titulo': 'Operación sin hallazgos críticos registrados', 'detalle': 'No se identificaron hallazgos críticos con la información disponible.'})
        return hallazgos

    def _generar_alertas(self, unidades_incompletas: list[dict[str, Any]], salud: dict[str, Any], entregables: dict[str, Any], retiros_edad: list[dict[str, Any]]) -> list[dict[str, Any]]:
        alertas: list[dict[str, Any]] = []
        for u in unidades_incompletas[:30]:
            alertas.append({'nivel': 'AMARILLO', 'tipo': 'Cobertura', 'mensaje': f"{u['unidad']} tiene {u['total']}/20 participantes. Faltan {u['faltan']}.", 'responsable': 'Coordinador de unidad'})
        alertas.extend(salud.get('alertas', [])[:50])
        alertas.extend(entregables.get('alertas', [])[:50])
        for r in retiros_edad[:30]:
            alertas.append({'nivel': 'AMARILLO', 'tipo': 'Edad de seguimiento', 'mensaje': f"{r.get('nombre')} tiene {r.get('edad_meses')} meses.", 'responsable': r.get('unidad')})
        return alertas

    def _generar_recomendaciones(self, indicadores: dict[str, Any], hallazgos: list[dict[str, Any]], alertas: list[dict[str, Any]]) -> list[dict[str, str]]:
        recomendaciones = []
        niveles = {h.get('nivel') for h in hallazgos}
        if 'CRITICO' in niveles:
            recomendaciones.append({'prioridad': 'Alta', 'recomendacion': 'Realizar comité operativo de seguimiento dentro de las siguientes 48 horas para cerrar alertas críticas.'})
        if indicadores.get('Unidades con cobertura incompleta', 0):
            recomendaciones.append({'prioridad': 'Alta', 'recomendacion': 'Validar cupos, reemplazos e ingresos en unidades con cobertura incompleta antes del cierre mensual.'})
        if indicadores.get('Casos nutricionales críticos', 0):
            recomendaciones.append({'prioridad': 'Alta', 'recomendacion': 'Priorizar seguimiento nutricional, remisión y evidencia de atención para casos críticos.'})
        if indicadores.get('Entregables vencidos', 0):
            recomendaciones.append({'prioridad': 'Media', 'recomendacion': 'Solicitar plan de choque por coordinador para normalizar entregables vencidos.'})
        recomendaciones.append({'prioridad': 'Media', 'recomendacion': 'Conservar el paquete mensual completo como soporte de auditoría y revisión gerencial.'})
        return recomendaciones

    def _resumen_responsables(self, talento: list[dict[str, Any]], entregables: list[dict[str, Any]]) -> list[dict[str, Any]]:
        resumen: dict[str, dict[str, Any]] = {}
        for t in talento:
            nombre = t.get('coordinador') or (t.get('nombre') if 'coord' in normalizar_texto(t.get('cargo')) else '') or 'Sin coordinador'
            if not nombre:
                nombre = 'Sin coordinador'
            resumen.setdefault(nombre, {'responsable': nombre, 'unidades': set(), 'talento': 0, 'pendientes': 0, 'vencidos': 0})
            if t.get('unidad'):
                resumen[nombre]['unidades'].add(t.get('unidad'))
            resumen[nombre]['talento'] += 1
        for e in entregables:
            responsable = e.get('responsable') or 'Sin responsable'
            resumen.setdefault(responsable, {'responsable': responsable, 'unidades': set(), 'talento': 0, 'pendientes': 0, 'vencidos': 0})
            if e.get('unidad'):
                resumen[responsable]['unidades'].add(e.get('unidad'))
            estado = self._estado_entregable_normalizado(e.get('estado'), e.get('fecha_limite'))
            if estado == 'VENCIDO':
                resumen[responsable]['vencidos'] += 1
            elif estado != 'CUMPLIDO':
                resumen[responsable]['pendientes'] += 1
        rows = []
        for v in resumen.values():
            rows.append({
                'responsable': v['responsable'],
                'unidades': ', '.join(sorted(v['unidades'])) if v['unidades'] else '',
                'talento': v['talento'],
                'pendientes': v['pendientes'],
                'vencidos': v['vencidos'],
            })
        rows.sort(key=lambda x: (-int(x['vencidos']), -int(x['pendientes']), x['responsable']))
        return rows

    def _generar_resumen_ejecutivo(self, fundacion: str, mes: int, anio: int, indicadores: dict[str, Any], hallazgos: list[dict[str, Any]], alertas: list[dict[str, Any]]) -> str:
        return (
            f"Durante {MESES_ES[mes]} de {anio}, la operación de {fundacion} registra "
            f"{indicadores.get('Beneficiarios activos', 0)} beneficiarios activos distribuidos en "
            f"{indicadores.get('Unidades con participantes', 0)} unidad(es). Se identifican "
            f"{indicadores.get('Entregables pendientes', 0)} entregable(s) pendiente(s), "
            f"{indicadores.get('Entregables vencidos', 0)} vencido(s), "
            f"{indicadores.get('Casos nutricionales críticos', 0)} caso(s) nutricionales críticos y "
            f"{len(alertas)} alerta(s) operativas. El presente informe consolida indicadores, hallazgos, "
            f"pendientes, responsables y recomendaciones para la toma de decisiones gerenciales."
        )

    def _generar_conclusion(self, fundacion: str, mes: int, anio: int, indicadores: dict[str, Any], hallazgos: list[dict[str, Any]]) -> str:
        criticos = sum(1 for h in hallazgos if h.get('nivel') == 'CRITICO')
        if criticos:
            return f"La operación de {fundacion} requiere seguimiento gerencial prioritario durante {MESES_ES[mes]} de {anio}, especialmente en entregables vencidos, cobertura y alertas nutricionales."
        if indicadores.get('Entregables pendientes', 0) or indicadores.get('Unidades con cobertura incompleta', 0):
            return f"La operación presenta pendientes controlables. Se recomienda cerrar novedades antes del corte mensual para fortalecer el cumplimiento ICBF."
        return f"La operación no presenta alertas críticas registradas con la información disponible. Se recomienda mantener el seguimiento preventivo y conservar soportes mensuales."

    def _list_to_rows(self, data: list[dict[str, Any]], keys: list[str]) -> list[list[Any]]:
        return [[item.get(k, '') for k in keys] for item in data]

    def generar_excel(self, data: dict[str, Any], path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Resumen Ejecutivo'
        title_fill = PatternFill('solid', fgColor='1F2937')
        header_fill = PatternFill('solid', fgColor='D9EAF7')
        soft_fill = PatternFill('solid', fgColor='EEF6FF')
        thin = Side(style='thin', color='CBD5E1')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_title(sheet, title, subtitle=''):
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
            c = sheet.cell(1, 1)
            c.value = title
            c.fill = title_fill
            c.font = Font(bold=True, color='FFFFFF', size=14)
            c.alignment = Alignment(horizontal='center')
            if subtitle:
                sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
                sheet.cell(2, 1).value = subtitle
                sheet.cell(2, 1).alignment = Alignment(horizontal='center')

        def write_table(sheet, start_row, headers, rows):
            for col, h in enumerate(headers, 1):
                cell = sheet.cell(start_row, col, h)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            for r, row in enumerate(rows, start_row + 1):
                for c, value in enumerate(row, 1):
                    cell = sheet.cell(r, c, value)
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    cell.border = border
            for col in range(1, max(2, len(headers)) + 1):
                sheet.column_dimensions[get_column_letter(col)].width = min(36, max(14, len(str(headers[col - 1])) + 4 if col <= len(headers) else 16))

        style_title(ws, f"Reporte gerencial mensual - {data['mes_nombre']} {data['anio']}", data['fundacion_nombre'])
        ws['A4'] = 'Resumen ejecutivo'
        ws['A4'].font = Font(bold=True, size=12)
        ws.merge_cells('A5:H8')
        ws['A5'] = data['resumen_ejecutivo']
        ws['A5'].alignment = Alignment(wrap_text=True, vertical='top')
        ws['A5'].fill = soft_fill
        row = 10
        indicadores_rows = [[k, v] for k, v in data['indicadores'].items()]
        write_table(ws, row, ['Indicador', 'Valor'], indicadores_rows)

        ws2 = wb.create_sheet('Indicadores y Gráficos')
        style_title(ws2, 'Indicadores y gráficos', data['periodo'])
        write_table(ws2, 4, ['Indicador', 'Valor'], indicadores_rows)
        if indicadores_rows:
            chart = BarChart()
            chart.title = 'Indicadores principales'
            chart.y_axis.title = 'Valor'
            chart.x_axis.title = 'Indicador'
            data_ref = Reference(ws2, min_col=2, min_row=4, max_row=4 + len(indicadores_rows))
            cats_ref = Reference(ws2, min_col=1, min_row=5, max_row=4 + len(indicadores_rows))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.height = 8
            chart.width = 18
            ws2.add_chart(chart, 'D4')

        # Distribuciones.
        ws3 = wb.create_sheet('Distribuciones')
        style_title(ws3, 'Distribuciones operativas', data['periodo'])
        current = 4
        for title, dist in [
            ('Distribución por grupo de edad', data['distribucion_grupos']),
            ('Distribución por sexo', data['distribucion_sexo']),
            ('Diagnósticos nutricionales', data['salud'].get('diagnosticos', {})),
            ('Talento humano por cargo', data['distribucion_talento']),
        ]:
            ws3.cell(current, 1, title).font = Font(bold=True, size=12)
            current += 1
            rows = [[k, v] for k, v in (dist or {}).items()]
            write_table(ws3, current, ['Categoría', 'Total'], rows)
            if rows:
                pie = PieChart()
                pie.title = title
                data_ref = Reference(ws3, min_col=2, min_row=current, max_row=current + len(rows))
                cats_ref = Reference(ws3, min_col=1, min_row=current + 1, max_row=current + len(rows))
                pie.add_data(data_ref, titles_from_data=True)
                pie.set_categories(cats_ref)
                pie.height = 7
                pie.width = 9
                ws3.add_chart(pie, f'D{current}')
            current += max(6, len(rows) + 4)

        sheet_specs = [
            ('Hallazgos', ['Nivel', 'Título', 'Detalle'], [[h.get('nivel'), h.get('titulo'), h.get('detalle')] for h in data['hallazgos']]),
            ('Alertas', ['Nivel', 'Tipo', 'Mensaje', 'Responsable', 'Unidad'], [[a.get('nivel'), a.get('tipo'), a.get('mensaje'), a.get('responsable'), a.get('unidad')] for a in data['alertas']]),
            ('Recomendaciones', ['Prioridad', 'Recomendación'], [[r.get('prioridad'), r.get('recomendacion')] for r in data['recomendaciones']]),
            ('Pendientes', ['Tipo', 'Título', 'Responsable', 'Unidad', 'Fecha límite', 'Estado'], [[p.get('tipo'), p.get('titulo'), p.get('responsable'), p.get('unidad'), p.get('fecha_limite'), p.get('estado')] for p in data['pendientes']]),
            ('Responsables', ['Responsable', 'Unidades', 'Talento', 'Pendientes', 'Vencidos'], [[r.get('responsable'), r.get('unidades'), r.get('talento'), r.get('pendientes'), r.get('vencidos')] for r in data['responsables']]),
            ('Unidades', ['Unidad', 'Total participantes'], [[k, v] for k, v in sorted(data['distribucion_unidades'].items())]),
        ]
        for sheet_name, headers, rows in sheet_specs:
            sh = wb.create_sheet(sheet_name[:31])
            style_title(sh, sheet_name, data['periodo'])
            write_table(sh, 4, headers, rows)

        ws_end = wb.create_sheet('Conclusión')
        style_title(ws_end, 'Conclusión y cierre gerencial', data['periodo'])
        ws_end.merge_cells('A4:H8')
        ws_end['A4'] = data['conclusion']
        ws_end['A4'].alignment = Alignment(wrap_text=True, vertical='top')
        ws_end['A4'].fill = soft_fill
        for sheet in wb.worksheets:
            sheet.freeze_panes = 'A4'
        wb.save(path)

    def _pdf_table(self, headers: list[str], rows: list[list[Any]], col_widths: list[int] | None = None):
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle
        safe_rows = [[str(x or '')[:110] for x in row] for row in rows]
        data = [headers] + (safe_rows or [['Sin registros'] + [''] * (len(headers) - 1)])
        table = Table(data, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E2E8F0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#111827')),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
        ]))
        return table

    def _bar_chart_pdf(self, title: str, data: dict[str, int]):
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.graphics.shapes import Drawing, String
        drawing = Drawing(420, 190)
        drawing.add(String(10, 170, title, fontSize=10))
        items = list(data.items())[:8]
        if not items:
            drawing.add(String(10, 90, 'Sin datos disponibles', fontSize=9))
            return drawing
        chart = VerticalBarChart()
        chart.x = 35
        chart.y = 35
        chart.height = 110
        chart.width = 350
        chart.data = [[int(v or 0) for _, v in items]]
        chart.categoryAxis.categoryNames = [str(k)[:12] for k, _ in items]
        chart.valueAxis.valueMin = 0
        chart.valueAxis.valueMax = max(1, max(int(v or 0) for _, v in items)) + 1
        chart.valueAxis.valueStep = max(1, int(chart.valueAxis.valueMax / 5) or 1)
        chart.categoryAxis.labels.angle = 35
        chart.categoryAxis.labels.fontSize = 6
        drawing.add(chart)
        return drawing

    def generar_pdf(self, data: dict[str, Any], path: Path) -> None:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer

        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(str(path), pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
        story = []
        story.append(Paragraph(f"Reporte gerencial mensual - {data['mes_nombre']} {data['anio']}", styles['Title']))
        story.append(Paragraph(data['fundacion_nombre'], styles['Heading2']))
        story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 12))

        for title, content in [
            ('Resumen ejecutivo', data['resumen_ejecutivo']),
            ('Conclusión', data['conclusion']),
        ]:
            story.append(Paragraph(title, styles['Heading2']))
            story.append(Paragraph(str(content), styles['BodyText']))
            story.append(Spacer(1, 10))

        story.append(Paragraph('Indicadores', styles['Heading2']))
        story.append(self._pdf_table(['Indicador', 'Valor'], [[k, v] for k, v in data['indicadores'].items()], [300, 120]))
        story.append(Spacer(1, 12))
        story.append(self._bar_chart_pdf('Distribución por grupos de edad', data.get('distribucion_grupos', {})))
        story.append(Spacer(1, 12))
        story.append(self._bar_chart_pdf('Diagnósticos nutricionales', data.get('salud', {}).get('diagnosticos', {})))
        story.append(PageBreak())

        story.append(Paragraph('Hallazgos', styles['Heading2']))
        story.append(self._pdf_table(['Nivel', 'Título', 'Detalle'], [[h.get('nivel'), h.get('titulo'), h.get('detalle')] for h in data['hallazgos']], [70, 160, 300]))
        story.append(Spacer(1, 12))

        story.append(Paragraph('Alertas', styles['Heading2']))
        story.append(self._pdf_table(['Nivel', 'Tipo', 'Mensaje', 'Responsable'], [[a.get('nivel'), a.get('tipo'), a.get('mensaje'), a.get('responsable')] for a in data['alertas'][:80]], [60, 90, 250, 130]))
        story.append(Spacer(1, 12))

        story.append(Paragraph('Recomendaciones', styles['Heading2']))
        story.append(self._pdf_table(['Prioridad', 'Recomendación'], [[r.get('prioridad'), r.get('recomendacion')] for r in data['recomendaciones']], [80, 450]))
        story.append(Spacer(1, 12))

        story.append(Paragraph('Pendientes', styles['Heading2']))
        story.append(self._pdf_table(['Tipo', 'Título', 'Responsable', 'Unidad', 'Fecha', 'Estado'], [[p.get('tipo'), p.get('titulo'), p.get('responsable'), p.get('unidad'), p.get('fecha_limite'), p.get('estado')] for p in data['pendientes'][:90]], [80, 130, 100, 90, 60, 60]))
        story.append(PageBreak())

        story.append(Paragraph('Responsables', styles['Heading2']))
        story.append(self._pdf_table(['Responsable', 'Unidades', 'Talento', 'Pendientes', 'Vencidos'], [[r.get('responsable'), r.get('unidades'), r.get('talento'), r.get('pendientes'), r.get('vencidos')] for r in data['responsables'][:80]], [120, 190, 60, 70, 70]))
        story.append(Spacer(1, 12))

        story.append(Paragraph('Tablas de soporte', styles['Heading2']))
        unidades_rows = [[k, v] for k, v in sorted(data['distribucion_unidades'].items())]
        story.append(self._pdf_table(['Unidad', 'Participantes'], unidades_rows[:80], [300, 100]))
        doc.build(story)

    def registrar_reporte(self, data: dict[str, Any], user: dict[str, Any], ruta_excel: Path, ruta_pdf: Path) -> int:
        conn = self.connect()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO rg_reportes
            (fundacion_id, usuario_id, periodo, mes, anio, tipo, titulo, estado,
             resumen_ejecutivo, indicadores_json, hallazgos_json, alertas_json,
             recomendaciones_json, pendientes_json, responsables_json, conclusion,
             ruta_pdf, ruta_excel, nombre_pdf, nombre_excel, total_indicadores,
             total_hallazgos, total_alertas, total_pendientes, fecha_generacion)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'GENERADO', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data.get('fundacion_id'), user.get('id'), data.get('periodo'), data.get('mes'), data.get('anio'),
            'MENSUAL', f"Reporte gerencial mensual - {data.get('mes_nombre')} {data.get('anio')}",
            data.get('resumen_ejecutivo'),
            json.dumps(data.get('indicadores'), ensure_ascii=False),
            json.dumps(data.get('hallazgos'), ensure_ascii=False),
            json.dumps(data.get('alertas'), ensure_ascii=False),
            json.dumps(data.get('recomendaciones'), ensure_ascii=False),
            json.dumps(data.get('pendientes'), ensure_ascii=False),
            json.dumps(data.get('responsables'), ensure_ascii=False),
            data.get('conclusion'), str(ruta_pdf), str(ruta_excel), ruta_pdf.name, ruta_excel.name,
            len(data.get('indicadores') or {}), len(data.get('hallazgos') or []),
            len(data.get('alertas') or []), len(data.get('pendientes') or []), now_iso()
        ))
        reporte_id = int(cur.lastrowid)
        cur.execute("""
            INSERT INTO rg_auditoria (reporte_id, fundacion_id, usuario_id, accion, detalle, datos_json, fecha)
            VALUES (?, ?, ?, 'GENERAR_REPORTE', ?, ?, ?)
        """, (reporte_id, data.get('fundacion_id'), user.get('id'), f"Reporte gerencial {data.get('periodo')} generado.", json.dumps({'excel': ruta_excel.name, 'pdf': ruta_pdf.name}, ensure_ascii=False), now_iso()))
        conn.commit()
        conn.close()
        return reporte_id

    def generar_reporte_ejecutivo(self, mes: int, anio: int, user: dict[str, Any] | None = None, output_dir: str | Path | None = None, registrar: bool = True) -> dict[str, Any]:
        # Cuando se invoca desde Paquete Mensual ya existe una transacción abierta.
        # En ese caso evitamos DDL para no bloquear SQLite; el reporte se genera sin registrar rg_reportes.
        if registrar:
            self.init_schema()
        user = user or {}
        fundacion_id = int(user.get('fundacion_id') or 1)
        mes = max(1, min(12, int(mes)))
        anio = int(anio)
        data = self.recopilar_datos(mes, anio, fundacion_id)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        out_dir = Path(output_dir) if output_dir else self.reportes_folder / data['periodo']
        out_dir.mkdir(parents=True, exist_ok=True)
        base = f"REPORTE_GERENCIAL_EJECUTIVO_{data['periodo']}_{ts}"
        ruta_excel = out_dir / f'{base}.xlsx'
        ruta_pdf = out_dir / f'{base}.pdf'
        self.generar_excel(data, ruta_excel)
        self.generar_pdf(data, ruta_pdf)
        reporte_id = None
        if registrar:
            reporte_id = self.registrar_reporte(data, user, ruta_excel, ruta_pdf)
        return {
            'id': reporte_id,
            'periodo': data['periodo'],
            'excel': str(ruta_excel),
            'pdf': str(ruta_pdf),
            'nombre_excel': ruta_excel.name,
            'nombre_pdf': ruta_pdf.name,
            'data': data,
        }

    def dashboard(self, fundacion_id: int | None = None) -> dict[str, Any]:
        hoy = datetime.now()
        data = self.recopilar_datos(hoy.month, hoy.year, fundacion_id or 1)
        reportes = self.historial(fundacion_id=fundacion_id, limit=10)
        return {
            'periodo': data['periodo'],
            'indicadores': data['indicadores'],
            'hallazgos': data['hallazgos'][:5],
            'alertas': data['alertas'][:8],
            'pendientes': data['pendientes'][:8],
            'recomendaciones': data['recomendaciones'][:5],
            'historial': reportes,
        }

    def historial(self, fundacion_id: int | None = None, limit: int = 50) -> list[dict[str, Any]]:
        self.init_schema()
        conn = self.connect()
        cur = conn.cursor()
        where = '1=1'
        params: list[Any] = []
        if fundacion_id:
            where += ' AND COALESCE(fundacion_id, ?) = ?'
            params.extend([fundacion_id, fundacion_id])
        rows = [dict(r) for r in cur.execute(f"SELECT * FROM rg_reportes WHERE {where} ORDER BY id DESC LIMIT ?", (*params, int(limit))).fetchall()]
        conn.close()
        return rows

    def obtener_reporte(self, reporte_id: int) -> dict[str, Any] | None:
        self.init_schema()
        conn = self.connect()
        row = conn.execute('SELECT * FROM rg_reportes WHERE id=?', (reporte_id,)).fetchone()
        conn.close()
        return dict(row) if row else None
