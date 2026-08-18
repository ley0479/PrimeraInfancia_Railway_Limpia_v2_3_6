"""Persistencia e integración del Motor Inteligente de Gestión del Proyecto."""
from __future__ import annotations

import csv
import io
import json
import os
from modules.dbapi_compat import sqlite3
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .schema import DEFAULT_RULES, SCHEMA_SQL, SCHEMA_VERSION
from .services import (
    COMPLETED_STATES,
    calculate_priority,
    compliance_percentage,
    file_sha256,
    is_complete,
    json_dump,
    normalize_text,
    now_iso,
    parse_json,
    safe_period,
    source_key,
    state,
    unit_key,
    valid_date,
)


class MotorGestionRepository:
    """Orquesta referencias a fuentes existentes sin reemplazarlas."""

    def __init__(self, database_path: str, data_dir: str, output_folder: str):
        self.database_path = str(database_path)
        self.data_dir = Path(data_dir).resolve()
        self.output_folder = Path(output_folder).resolve()
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.output_folder.mkdir(parents=True, exist_ok=True)

    def connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.database_path, timeout=30)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys=ON")
        conn.execute("PRAGMA busy_timeout=30000")
        return conn

    @staticmethod
    def _table_exists(conn: sqlite3.Connection, table: str) -> bool:
        return bool(conn.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone())

    @staticmethod
    def _columns(conn: sqlite3.Connection, table: str) -> set[str]:
        if not MotorGestionRepository._table_exists(conn, table):
            return set()
        return {str(row[1]) for row in conn.execute(f'PRAGMA table_info("{table}")').fetchall()}

    def init_schema(self) -> None:
        conn = self.connect()
        try:
            conn.executescript(SCHEMA_SQL)
            now = now_iso()
            for foundation_id in self._known_foundations(conn):
                self._seed_rules(conn, foundation_id, now)
            conn.execute(
                "INSERT INTO mgp_schema_version(id,version,fecha_actualizacion) VALUES(1,?,?) "
                "ON CONFLICT(id) DO UPDATE SET version=excluded.version,fecha_actualizacion=excluded.fecha_actualizacion",
                (SCHEMA_VERSION, now),
            )
            conn.commit()
        finally:
            conn.close()

    def _known_foundations(self, conn: sqlite3.Connection) -> list[int]:
        if self._table_exists(conn, "fundaciones"):
            rows = conn.execute("SELECT id FROM fundaciones WHERE COALESCE(estado,'ACTIVA')<>'ELIMINADA'").fetchall()
            return [int(row[0]) for row in rows] or [1]
        return [1]

    def _seed_rules(self, conn: sqlite3.Connection, fundacion_id: int, now: str | None = None) -> None:
        timestamp = now or now_iso()
        for item in DEFAULT_RULES:
            conn.execute(
                """
                INSERT INTO mgp_reglas
                (fundacion_id,codigo,nombre,descripcion,tipo_regla,condicion_json,accion_json,prioridad_base,activa,orden,fecha_creacion,fecha_actualizacion)
                VALUES(?,?,?,?,?,?,?,?,1,?,?,?)
                ON CONFLICT(fundacion_id,codigo) DO UPDATE SET
                    nombre=excluded.nombre,descripcion=excluded.descripcion,tipo_regla=excluded.tipo_regla,
                    condicion_json=excluded.condicion_json,accion_json=excluded.accion_json,
                    prioridad_base=excluded.prioridad_base,orden=excluded.orden,fecha_actualizacion=excluded.fecha_actualizacion
                """,
                (
                    fundacion_id, item["codigo"], item["nombre"], item["descripcion"], item["tipo_regla"],
                    json_dump(item.get("condicion") or {}), json_dump(item.get("accion") or {}),
                    int(item.get("prioridad_base") or 0), int(item.get("orden") or 100), timestamp, timestamp,
                ),
            )

    def audit(self, fundacion_id: int, user: dict[str, Any], action: str, entity: str, entity_id: int | None, detail: dict[str, Any] | None = None) -> None:
        with self.connect() as conn:
            conn.execute(
                "INSERT INTO mgp_auditoria(fundacion_id,usuario_id,usuario,accion,entidad,entidad_id,detalle_json,fecha) VALUES(?,?,?,?,?,?,?,?)",
                (
                    fundacion_id, user.get("id"), user.get("username") or user.get("email") or "sistema",
                    action, entity, entity_id, json_dump(detail or {}), now_iso(),
                ),
            )

    @staticmethod
    def _field(row: dict[str, Any], *names: str, default: Any = None) -> Any:
        for name in names:
            if name in row and row.get(name) not in (None, ""):
                return row.get(name)
        return default

    def _foundation_rows(self, conn: sqlite3.Connection, table: str, fundacion_id: int) -> list[dict[str, Any]]:
        if not self._table_exists(conn, table):
            return []
        cols = self._columns(conn, table)
        where = " WHERE fundacion_id=?" if "fundacion_id" in cols else ""
        params: tuple[Any, ...] = (fundacion_id,) if where else ()
        try:
            rows = conn.execute(f'SELECT * FROM "{table}"{where}', params).fetchall()
        except sqlite3.DatabaseError:
            return []
        return [dict(row) for row in rows]

    def _expedient_map(self, conn: sqlite3.Connection, fundacion_id: int) -> tuple[dict[str, dict[str, Any]], dict[int, dict[str, Any]]]:
        if not self._table_exists(conn, "giu_expedientes_uca"):
            return {}, {}
        rows = conn.execute(
            "SELECT * FROM giu_expedientes_uca WHERE fundacion_id=? AND COALESCE(estado,'ACTIVO')<>'ELIMINADO'",
            (fundacion_id,),
        ).fetchall()
        by_unit: dict[str, dict[str, Any]] = {}
        by_id: dict[int, dict[str, Any]] = {}
        for row in rows:
            data = dict(row)
            key = unit_key(data.get("unidad_nombre") or data.get("codigo_unidad"))
            if key:
                by_unit[key] = data
            if data.get("codigo_unidad"):
                by_unit[unit_key(data["codigo_unidad"])] = data
            by_id[int(data["id"])] = data
        return by_unit, by_id

    def _route_source(self, conn: sqlite3.Connection, fundacion_id: int, exp_by_id: dict[int, dict[str, Any]]) -> Iterable[dict[str, Any]]:
        if not self._table_exists(conn, "giu_ruta_instancias"):
            return []
        catalog: dict[int, dict[str, Any]] = {}
        if self._table_exists(conn, "giu_ruta_catalogo"):
            catalog = {int(row["id"]): dict(row) for row in conn.execute("SELECT * FROM giu_ruta_catalogo").fetchall()}
        evidence_counts: dict[int, int] = defaultdict(int)
        if self._table_exists(conn, "giu_ruta_evidencias"):
            for row in conn.execute("SELECT instancia_id,COUNT(*) total FROM giu_ruta_evidencias WHERE fundacion_id=? AND COALESCE(activo,1)=1 GROUP BY instancia_id", (fundacion_id,)).fetchall():
                evidence_counts[int(row[0])] = int(row[1])
        rows = conn.execute("SELECT * FROM giu_ruta_instancias WHERE fundacion_id=?", (fundacion_id,)).fetchall()
        result = []
        for raw in rows:
            row = dict(raw)
            exp = exp_by_id.get(int(row.get("expediente_id") or 0)) or {}
            cat = catalog.get(int(row.get("catalogo_id") or 0)) or {}
            result.append({
                "source_module": "GESTION_INTEGRAL_UCA",
                "source_table": "giu_ruta_instancias",
                "source_id": row.get("id"),
                "source_extra": row.get("actividad_codigo"),
                "expediente_id": row.get("expediente_id"),
                "unidad_id": exp.get("unidad_id"),
                "unidad_nombre": exp.get("unidad_nombre"),
                "componente": cat.get("componente"),
                "tipo_tarea": "RUTA_OPERATIVA",
                "titulo": cat.get("titulo") or row.get("actividad_codigo") or "Actividad de Ruta Operativa",
                "descripcion": cat.get("descripcion"),
                "fecha_inicio": row.get("fecha_inicio"),
                "fecha_limite": row.get("fecha_limite"),
                "fecha_finalizacion": row.get("fecha_finalizacion"),
                "estado": row.get("estado"),
                "prioridad": "ALTA" if int(cat.get("obligatoria") or 0) else "MEDIA",
                "responsable_id": row.get("responsable_id"),
                "responsable_nombre": row.get("responsable_nombre"),
                "requiere_evidencia": int(cat.get("requiere_evidencia") or 0),
                "evidencias_total": evidence_counts.get(int(row.get("id") or 0), 0),
                "metadata": {"actividad_codigo": row.get("actividad_codigo"), "fase": cat.get("fase"), "catalogo_id": row.get("catalogo_id")},
            })
        return result

    def _plans_source(self, conn: sqlite3.Connection, fundacion_id: int, exp_by_id: dict[int, dict[str, Any]]) -> Iterable[dict[str, Any]]:
        if not self._table_exists(conn, "giu_planes_uca"):
            return []
        rows = conn.execute("SELECT * FROM giu_planes_uca WHERE fundacion_id=?", (fundacion_id,)).fetchall()
        result = []
        for raw in rows:
            row = dict(raw)
            exp = exp_by_id.get(int(row.get("expediente_id") or 0)) or {}
            result.append({
                "source_module": "GESTION_INTEGRAL_UCA", "source_table": "giu_planes_uca", "source_id": row.get("id"),
                "source_extra": row.get("codigo"), "expediente_id": row.get("expediente_id"), "unidad_id": exp.get("unidad_id"),
                "unidad_nombre": exp.get("unidad_nombre"), "componente": "ADMINISTRATIVO_GESTION", "tipo_tarea": "PLAN_OPERATIVO",
                "titulo": row.get("nombre") or row.get("codigo") or "Plan operativo", "descripcion": row.get("observaciones"),
                "fecha_inicio": row.get("fecha_inicio"), "fecha_limite": row.get("fecha_fin"), "estado": row.get("estado"),
                "prioridad": "MEDIA", "responsable_id": row.get("responsable_id"), "responsable_nombre": row.get("responsable_nombre"),
                "requiere_evidencia": 1, "evidencias_total": 0,
                "metadata": {"codigo_plan": row.get("codigo"), "progreso": row.get("progreso")},
            })
        return result

    def _generic_source(
        self,
        conn: sqlite3.Connection,
        fundacion_id: int,
        table: str,
        *,
        source_module: str,
        task_type: str,
        component: str,
        exp_by_unit: dict[str, dict[str, Any]],
        title_fields: tuple[str, ...],
        unit_fields: tuple[str, ...],
        due_fields: tuple[str, ...],
        start_fields: tuple[str, ...] = (),
        state_fields: tuple[str, ...] = ("estado",),
        responsible_fields: tuple[str, ...] = ("responsable_nombre", "responsable", "coordinador"),
        description_fields: tuple[str, ...] = ("descripcion", "observaciones"),
        priority_fields: tuple[str, ...] = ("prioridad",),
        evidence_fields: tuple[str, ...] = ("archivo_evidencia", "ruta_archivo", "archivo"),
    ) -> list[dict[str, Any]]:
        result = []
        for row in self._foundation_rows(conn, table, fundacion_id):
            unit_name = self._field(row, *unit_fields, default="")
            exp = exp_by_unit.get(unit_key(unit_name)) or {}
            title = str(self._field(row, *title_fields, default=f"Registro {table} #{row.get('id') or ''}") or "").strip()
            due = self._field(row, *due_fields)
            if table == "sn_entregables_mes" and not due:
                year = int(row.get("anio") or date.today().year)
                month = max(1, min(12, int(row.get("mes") or date.today().month)))
                due = f"{year:04d}-{month:02d}-28"
            evidence_value = self._field(row, *evidence_fields)
            result.append({
                "source_module": source_module, "source_table": table, "source_id": row.get("id"),
                "source_extra": None if table in {"csc_supervisiones", "csc_hallazgos", "csc_planes_mejora", "fcr_actividades", "fcr_compromisos", "fcr_alertas"} else self._field(row, "clave_unica", "codigo", "periodo"),
                "expediente_id": exp.get("id"), "unidad_id": exp.get("unidad_id"), "unidad_nombre": unit_name or exp.get("unidad_nombre"),
                "componente": component, "tipo_tarea": task_type, "titulo": title,
                "descripcion": self._field(row, *description_fields), "fecha_inicio": self._field(row, *start_fields),
                "fecha_limite": due, "fecha_finalizacion": self._field(row, "fecha_entrega", "fecha_finalizacion"),
                "estado": self._field(row, *state_fields, default="PENDIENTE"), "prioridad": self._field(row, *priority_fields, default="MEDIA"),
                "responsable_id": self._field(row, "responsable_id", "usuario_creador_id"),
                "responsable_nombre": self._field(row, *responsible_fields), "requiere_evidencia": int(bool(row.get("requiere_evidencia"))),
                "evidencias_total": 1 if evidence_value else 0, "metadata": {"source_row": row},
            })
        return result

    def collect_sources(self, conn: sqlite3.Connection, fundacion_id: int) -> list[dict[str, Any]]:
        exp_by_unit, exp_by_id = self._expedient_map(conn, fundacion_id)
        collected: list[dict[str, Any]] = []
        collected.extend(self._route_source(conn, fundacion_id, exp_by_id))
        collected.extend(self._plans_source(conn, fundacion_id, exp_by_id))
        collected.extend(self._generic_source(
            conn, fundacion_id, "calendario_entregables", source_module="CALENDARIO", task_type="ACTIVIDAD_CALENDARIO",
            component="TRANSVERSAL", exp_by_unit=exp_by_unit, title_fields=("titulo",), unit_fields=("unidad",),
            due_fields=("fecha_limite",), start_fields=("fecha_inicio",),
        ))
        collected.extend(self._generic_source(
            conn, fundacion_id, "gp_entregables", source_module="GESTION_PEDAGOGICA", task_type="ENTREGABLE_PEDAGOGICO",
            component="PROCESO_PEDAGOGICO", exp_by_unit=exp_by_unit, title_fields=("titulo", "tipo"), unit_fields=("unidad",),
            due_fields=("fecha_limite",),
        ))
        collected.extend(self._generic_source(
            conn, fundacion_id, "sn_entregables_mes", source_module="SALUD_NUTRICION", task_type="ENTREGABLE_SALUD",
            component="SALUD_NUTRICION", exp_by_unit=exp_by_unit, title_fields=("codigo",), unit_fields=("uds", "unidad"),
            due_fields=("fecha_limite",), description_fields=("observaciones",), priority_fields=("prioridad",),
        ))
        collected.extend(self._generic_source(
            conn, fundacion_id, "sn_actividades_integrales", source_module="SALUD_NUTRICION", task_type="ACTIVIDAD_SALUD_NUTRICION",
            component="SALUD_NUTRICION", exp_by_unit=exp_by_unit, title_fields=("titulo", "tipo_actividad"), unit_fields=("unidad_nombre",),
            due_fields=("fecha_programada",), start_fields=("fecha_programada",), description_fields=("objetivo", "metodologia"),
            responsible_fields=("responsable_nombre",), priority_fields=(), evidence_fields=(),
        ))
        collected.extend(self._generic_source(
            conn, fundacion_id, "sn_canalizaciones", source_module="SALUD_NUTRICION", task_type="CANALIZACION_SALUD",
            component="SALUD_NUTRICION", exp_by_unit=exp_by_unit, title_fields=("tipo_ruta",), unit_fields=("unidad_nombre",),
            due_fields=("fecha_limite",), start_fields=("fecha_activacion",), description_fields=("motivo", "resultado_cierre"),
            responsible_fields=("responsable_nombre",), priority_fields=("prioridad",), evidence_fields=("evidencia_cierre",),
        ))
        collected.extend(self._generic_source(
            conn, fundacion_id, "fcr_actividades", source_module="FAMILIAS_REDES", task_type="ACTIVIDAD_FAMILIAR_COMUNITARIA",
            component="FAMILIA_COMUNIDAD_REDES", exp_by_unit=exp_by_unit, title_fields=("titulo", "tipo"), unit_fields=("unidad_nombre",),
            due_fields=("fecha_limite_cierre", "fecha_programada"), start_fields=("fecha_programada",),
            description_fields=("objetivo", "metodologia", "observaciones"), responsible_fields=("profesional_nombre",),
            priority_fields=(),
        ))
        collected.extend(self._generic_source(
            conn, fundacion_id, "fcr_compromisos", source_module="FAMILIAS_REDES", task_type="COMPROMISO_FAMILIAR",
            component="FAMILIA_COMUNIDAD_REDES", exp_by_unit=exp_by_unit, title_fields=("titulo",), unit_fields=("unidad_nombre",),
            due_fields=("fecha_limite",), start_fields=("fecha_compromiso",), responsible_fields=("responsable_nombre",),
        ))
        collected.extend(self._generic_source(
            conn, fundacion_id, "fcr_alertas", source_module="FAMILIAS_REDES", task_type="SEGUIMIENTO_ALERTA_FAMILIAR",
            component="FAMILIA_COMUNIDAD_REDES", exp_by_unit=exp_by_unit, title_fields=("tipo",), unit_fields=("unidad_nombre",),
            due_fields=("fecha_proximo_seguimiento",), start_fields=("fecha_identificacion",), description_fields=("descripcion",),
            responsible_fields=("responsable_nombre",), priority_fields=("nivel",),
        ))
        collected.extend(self._generic_source(
            conn, fundacion_id, "csc_supervisiones", source_module="SUPERVISION_CALIDAD", task_type="SUPERVISION",
            component="ADMINISTRATIVO_GESTION", exp_by_unit=exp_by_unit, title_fields=("titulo", "tipo"), unit_fields=("unidad_nombre",),
            due_fields=("fecha_programada",), description_fields=("objetivo", "alcance"), responsible_fields=("supervisor_nombre",),
            priority_fields=(),
        ))
        collected.extend(self._generic_source(
            conn, fundacion_id, "csc_hallazgos", source_module="SUPERVISION_CALIDAD", task_type="HALLAZGO",
            component="CALIDAD", exp_by_unit=exp_by_unit, title_fields=("titulo", "codigo"), unit_fields=("unidad_nombre",),
            due_fields=("fecha_limite",), start_fields=("fecha_deteccion",), description_fields=("descripcion", "criterio_afectado"),
            responsible_fields=("responsable_nombre",), priority_fields=("nivel_riesgo",),
        ))
        collected.extend(self._generic_source(
            conn, fundacion_id, "csc_planes_mejora", source_module="SUPERVISION_CALIDAD", task_type="PLAN_MEJORA",
            component="CALIDAD", exp_by_unit=exp_by_unit, title_fields=("nombre", "codigo"), unit_fields=("unidad_nombre",),
            due_fields=("fecha_limite",), start_fields=("fecha_inicio",), description_fields=("objetivo", "alcance"),
            responsible_fields=("responsable_nombre",), priority_fields=(),
        ))
        return collected

    def _upsert_task(self, conn: sqlite3.Connection, fundacion_id: int, item: dict[str, Any], user: dict[str, Any]) -> tuple[int, bool]:
        key = source_key(str(item.get("source_table")), item.get("source_id"), item.get("source_extra"))
        existing = conn.execute(
            "SELECT * FROM mgp_tareas WHERE fundacion_id=? AND fuente_tabla=? AND fuente_clave=?",
            (fundacion_id, str(item.get("source_table")), key),
        ).fetchone()
        current = dict(existing) if existing else {}
        incoming_state = state(item.get("estado"))
        if current and state(current.get("estado")) in COMPLETED_STATES and incoming_state not in COMPLETED_STATES:
            incoming_state = state(current.get("estado"))
        unit_name = str(item.get("unidad_nombre") or "").strip() or None
        base_task = {
            "estado": incoming_state,
            "prioridad": normalize_text(item.get("prioridad")) or "MEDIA",
            "fecha_limite": valid_date(item.get("fecha_limite")),
            "requiere_evidencia": int(bool(item.get("requiere_evidencia"))),
            "evidencias_total": int(item.get("evidencias_total") or 0),
        }
        priority = calculate_priority(base_task)
        now = now_iso()
        values = (
            fundacion_id, item.get("expediente_id"), item.get("unidad_id"), unit_name, unit_key(unit_name),
            str(item.get("source_module") or "DESCONOCIDO")[:80], str(item.get("source_table") or "desconocido")[:80],
            int(item.get("source_id") or 0) or None, key, str(item.get("tipo_tarea") or "ENTREGABLE")[:80],
            str(item.get("componente") or "TRANSVERSAL")[:120], str(item.get("titulo") or "Tarea operativa")[:500],
            str(item.get("descripcion") or "")[:4000] or None, valid_date(item.get("fecha_inicio")), valid_date(item.get("fecha_limite")),
            valid_date(item.get("fecha_finalizacion")), incoming_state, priority["prioridad"], priority["puntaje"],
            item.get("responsable_id"), str(item.get("responsable_nombre") or "")[:250] or None,
            base_task["requiere_evidencia"], base_task["evidencias_total"], int(priority["bloqueada"]),
            json_dump(item.get("metadata") or {}), user.get("id"), user.get("id"), now, now,
        )
        conn.execute(
            """
            INSERT INTO mgp_tareas
            (fundacion_id,expediente_id,unidad_id,unidad_nombre,unidad_clave,fuente_modulo,fuente_tabla,fuente_id,fuente_clave,
             tipo_tarea,componente,titulo,descripcion,fecha_inicio,fecha_limite,fecha_finalizacion,estado,prioridad,puntaje_prioridad,
             responsable_id,responsable_nombre,requiere_evidencia,evidencias_total,bloqueada,metadata_json,creada_por,actualizada_por,
             fecha_creacion,fecha_actualizacion)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(fundacion_id,fuente_tabla,fuente_clave) DO UPDATE SET
              expediente_id=excluded.expediente_id,unidad_id=excluded.unidad_id,unidad_nombre=excluded.unidad_nombre,
              unidad_clave=excluded.unidad_clave,fuente_modulo=excluded.fuente_modulo,fuente_id=excluded.fuente_id,
              tipo_tarea=excluded.tipo_tarea,componente=excluded.componente,titulo=excluded.titulo,descripcion=excluded.descripcion,
              fecha_inicio=excluded.fecha_inicio,fecha_limite=excluded.fecha_limite,fecha_finalizacion=excluded.fecha_finalizacion,
              estado=excluded.estado,prioridad=excluded.prioridad,puntaje_prioridad=excluded.puntaje_prioridad,
              responsable_id=COALESCE(excluded.responsable_id,mgp_tareas.responsable_id),
              responsable_nombre=COALESCE(excluded.responsable_nombre,mgp_tareas.responsable_nombre),
              requiere_evidencia=excluded.requiere_evidencia,evidencias_total=excluded.evidencias_total,
              metadata_json=excluded.metadata_json,activa=1,actualizada_por=excluded.actualizada_por,fecha_actualizacion=excluded.fecha_actualizacion
            """,
            values,
        )
        row = conn.execute("SELECT id FROM mgp_tareas WHERE fundacion_id=? AND fuente_tabla=? AND fuente_clave=?", (fundacion_id, item["source_table"], key)).fetchone()
        return int(row[0]), not bool(existing)

    def synchronize(self, fundacion_id: int, user: dict[str, Any]) -> dict[str, Any]:
        conn = self.connect()
        created = 0
        updated = 0
        source_counts: Counter[str] = Counter()
        try:
            self._seed_rules(conn, fundacion_id)
            items = self.collect_sources(conn, fundacion_id)
            for item in items:
                _, was_created = self._upsert_task(conn, fundacion_id, item, user)
                source_counts[str(item.get("source_module") or "OTRO")] += 1
                created += int(was_created)
                updated += int(not was_created)
            conn.commit()
        finally:
            conn.close()
        priority = self.recalculate_priorities(fundacion_id)
        reminders = self.generate_reminders(fundacion_id)
        result = {
            "fuentes_leidas": sum(source_counts.values()), "creadas": created, "actualizadas": updated,
            "por_fuente": dict(source_counts), "prioridades_recalculadas": priority, "recordatorios": reminders,
        }
        self.audit(fundacion_id, user, "SINCRONIZAR_MOTOR_PROYECTO", "mgp_tareas", None, result)
        return result

    def _dependency_open(self, conn: sqlite3.Connection, fundacion_id: int, task_id: int) -> bool:
        rows = conn.execute(
            """
            SELECT t.estado,d.obligatoria FROM mgp_dependencias d
            JOIN mgp_tareas t ON t.id=d.depende_de_tarea_id AND t.fundacion_id=d.fundacion_id
            WHERE d.fundacion_id=? AND d.tarea_id=?
            """,
            (fundacion_id, task_id),
        ).fetchall()
        return any(int(row["obligatoria"] or 0) and not is_complete(row["estado"]) for row in rows)

    def recalculate_priorities(self, fundacion_id: int) -> int:
        conn = self.connect()
        count = 0
        try:
            rows = conn.execute("SELECT * FROM mgp_tareas WHERE fundacion_id=? AND activa=1", (fundacion_id,)).fetchall()
            for raw in rows:
                task = dict(raw)
                dependency_open = self._dependency_open(conn, fundacion_id, int(task["id"]))
                result = calculate_priority(task, dependency_open=dependency_open)
                reason = "Dependencia obligatoria pendiente" if dependency_open else None
                conn.execute(
                    "UPDATE mgp_tareas SET prioridad=?,puntaje_prioridad=?,bloqueada=?,motivo_bloqueo=?,fecha_actualizacion=? WHERE id=? AND fundacion_id=?",
                    (result["prioridad"], result["puntaje"], int(result["bloqueada"]), reason, now_iso(), task["id"], fundacion_id),
                )
                count += 1
            conn.commit()
        finally:
            conn.close()
        return count

    def generate_reminders(self, fundacion_id: int) -> int:
        conn = self.connect()
        generated = 0
        try:
            rows = conn.execute("SELECT * FROM mgp_tareas WHERE fundacion_id=? AND activa=1", (fundacion_id,)).fetchall()
            now = now_iso()
            today = date.today().isoformat()
            for raw in rows:
                task = dict(raw)
                if is_complete(task.get("estado")):
                    continue
                priority = calculate_priority(task, dependency_open=bool(task.get("bloqueada")))
                if priority["puntaje"] < 25:
                    continue
                kind = "VENCIDA" if priority["vencida"] else ("BLOQUEADA" if priority["bloqueada"] else "PROXIMO_VENCIMIENTO")
                title = f"{priority['prioridad']}: {task['titulo']}"
                due = task.get("fecha_limite") or "sin fecha límite"
                message = f"La tarea requiere atención. Estado {task.get('estado')}; fecha límite {due}."
                cursor = conn.execute(
                    """
                    INSERT OR IGNORE INTO mgp_recordatorios
                    (fundacion_id,tarea_id,tipo,nivel,titulo,mensaje,fecha_programada,estado,destinatario_id,destinatario_nombre,leido,creado_en,actualizado_en)
                    VALUES(?,?,?,?,?,?,?,'PENDIENTE',?,?,0,?,?)
                    """,
                    (fundacion_id, task["id"], kind, priority["prioridad"], title[:500], message[:2000], today, task.get("responsable_id"), task.get("responsable_nombre"), now, now),
                )
                generated += int(cursor.rowcount or 0)
            conn.commit()
        finally:
            conn.close()
        return generated

    def list_tasks(self, fundacion_id: int, filters: dict[str, Any] | None = None, limit: int = 1000) -> list[dict[str, Any]]:
        filters = filters or {}
        where = ["fundacion_id=?", "activa=1"]
        params: list[Any] = [fundacion_id]
        if filters.get("expediente_id"):
            where.append("expediente_id=?")
            params.append(int(filters["expediente_id"]))
        if filters.get("periodo"):
            where.append("substr(COALESCE(fecha_limite,fecha_inicio,fecha_creacion),1,7)=?")
            params.append(safe_period(filters["periodo"]))
        for field in ("estado", "prioridad", "fuente_modulo", "componente", "tipo_tarea"):
            if filters.get(field):
                where.append(f"{field}=?")
                params.append(normalize_text(filters[field]))
        if filters.get("unidad"):
            where.append("unidad_clave=?")
            params.append(unit_key(filters["unidad"]))
        if filters.get("responsable_id"):
            where.append("responsable_id=?")
            params.append(int(filters["responsable_id"]))
        if filters.get("solo_abiertas"):
            placeholders = ",".join("?" for _ in COMPLETED_STATES)
            where.append(f"estado NOT IN ({placeholders})")
            params.extend(sorted(COMPLETED_STATES))
        if filters.get("buscar"):
            where.append("(LOWER(titulo) LIKE LOWER(?) OR LOWER(COALESCE(descripcion,'')) LIKE LOWER(?))")
            term = f"%{filters['buscar']}%"
            params.extend([term, term])
        params.append(max(1, min(5000, int(limit))))
        with self.connect() as conn:
            rows = conn.execute(
                f"SELECT * FROM mgp_tareas WHERE {' AND '.join(where)} ORDER BY puntaje_prioridad DESC,COALESCE(fecha_limite,'9999-12-31'),id LIMIT ?",
                params,
            ).fetchall()
            result = []
            for row in rows:
                data = dict(row)
                data["metadata"] = parse_json(data.pop("metadata_json", None), {})
                data.update({k: v for k, v in calculate_priority(data, dependency_open=bool(data.get("bloqueada"))).items() if k in {"vencida", "dias_restantes"}})
                result.append(data)
            return result

    def task(self, fundacion_id: int, task_id: int) -> dict[str, Any] | None:
        rows = self.list_tasks(fundacion_id, {"buscar": None}, limit=5000)
        return next((row for row in rows if int(row["id"]) == int(task_id)), None)

    def create_manual_task(self, fundacion_id: int, payload: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
        title = str(payload.get("titulo") or "").strip()
        if not title:
            raise ValueError("El título es obligatorio.")
        manual_key = str(payload.get("clave") or f"MANUAL-{datetime.now().strftime('%Y%m%d%H%M%S%f')}")
        item = {
            "source_module": "MOTOR_GESTION", "source_table": "mgp_manual", "source_id": None, "source_extra": manual_key,
            "expediente_id": int(payload.get("expediente_id") or 0) or None, "unidad_id": int(payload.get("unidad_id") or 0) or None,
            "unidad_nombre": payload.get("unidad_nombre"), "componente": payload.get("componente") or "TRANSVERSAL",
            "tipo_tarea": payload.get("tipo_tarea") or "TAREA_MANUAL", "titulo": title, "descripcion": payload.get("descripcion"),
            "fecha_inicio": payload.get("fecha_inicio"), "fecha_limite": payload.get("fecha_limite"), "estado": "PENDIENTE",
            "prioridad": payload.get("prioridad") or "MEDIA", "responsable_id": payload.get("responsable_id"),
            "responsable_nombre": payload.get("responsable_nombre"), "requiere_evidencia": int(bool(payload.get("requiere_evidencia"))),
            "evidencias_total": 0, "metadata": {"creacion": "MANUAL", "plantilla_documento_id": payload.get("plantilla_documento_id")},
        }
        with self.connect() as conn:
            task_id, _ = self._upsert_task(conn, fundacion_id, item, user)
            conn.commit()
        self.audit(fundacion_id, user, "CREAR_TAREA_MANUAL", "mgp_tareas", task_id, {"titulo": title})
        return self.task(fundacion_id, task_id) or {}

    def update_task(self, fundacion_id: int, task_id: int, payload: dict[str, Any], user: dict[str, Any], allow_review: bool = False) -> dict[str, Any]:
        allowed_states = {"PENDIENTE", "EN_PROCESO", "PENDIENTE_EVIDENCIA", "PENDIENTE_REVISION", "DEVUELTA", "APROBADA", "CERRADA", "NO_APLICA", "CANCELADA"}
        review_states = {"DEVUELTA", "APROBADA", "CERRADA", "NO_APLICA"}
        conn = self.connect()
        try:
            raw = conn.execute("SELECT * FROM mgp_tareas WHERE id=? AND fundacion_id=? AND activa=1", (task_id, fundacion_id)).fetchone()
            if not raw:
                raise LookupError("Tarea no encontrada.")
            current = dict(raw)
            new_state = state(payload.get("estado") or current.get("estado"))
            if new_state not in allowed_states:
                raise ValueError("Estado de tarea no permitido.")
            if new_state in review_states and not allow_review:
                raise PermissionError("Solo coordinación puede revisar, aprobar o cerrar tareas.")
            if new_state in {"APROBADA", "CERRADA"} and int(current.get("requiere_evidencia") or 0) and int(current.get("evidencias_total") or 0) <= 0:
                raise ValueError("La tarea requiere evidencia antes de aprobarse o cerrarse.")
            completed_date = date.today().isoformat() if new_state in COMPLETED_STATES else current.get("fecha_finalizacion")
            conn.execute(
                """
                UPDATE mgp_tareas SET estado=?,fecha_inicio=?,fecha_limite=?,fecha_finalizacion=?,responsable_id=?,responsable_nombre=?,
                    revisor_id=?,revisor_nombre=?,aprobador_id=?,aprobador_nombre=?,descripcion=?,actualizada_por=?,fecha_actualizacion=?
                WHERE id=? AND fundacion_id=?
                """,
                (
                    new_state, valid_date(payload.get("fecha_inicio")) or current.get("fecha_inicio"),
                    valid_date(payload.get("fecha_limite")) or current.get("fecha_limite"), completed_date,
                    payload.get("responsable_id") if payload.get("responsable_id") is not None else current.get("responsable_id"),
                    payload.get("responsable_nombre") if payload.get("responsable_nombre") is not None else current.get("responsable_nombre"),
                    user.get("id") if new_state in review_states else current.get("revisor_id"),
                    user.get("username") if new_state in review_states else current.get("revisor_nombre"),
                    user.get("id") if new_state in {"APROBADA", "CERRADA"} else current.get("aprobador_id"),
                    user.get("username") if new_state in {"APROBADA", "CERRADA"} else current.get("aprobador_nombre"),
                    str(payload.get("descripcion") if payload.get("descripcion") is not None else current.get("descripcion") or "")[:4000] or None,
                    user.get("id"), now_iso(), task_id, fundacion_id,
                ),
            )
            conn.commit()
        finally:
            conn.close()
        self.recalculate_priorities(fundacion_id)
        self.audit(fundacion_id, user, "ACTUALIZAR_TAREA", "mgp_tareas", task_id, {"estado": new_state})
        return self.task(fundacion_id, task_id) or {}

    def add_dependency(self, fundacion_id: int, task_id: int, depends_on: int, user: dict[str, Any], required: bool = True) -> list[dict[str, Any]]:
        if task_id == depends_on:
            raise ValueError("Una tarea no puede depender de sí misma.")
        with self.connect() as conn:
            valid = conn.execute("SELECT COUNT(*) FROM mgp_tareas WHERE fundacion_id=? AND id IN (?,?) AND activa=1", (fundacion_id, task_id, depends_on)).fetchone()[0]
            if int(valid) != 2:
                raise LookupError("Una de las tareas no existe.")
            conn.execute(
                "INSERT OR IGNORE INTO mgp_dependencias(fundacion_id,tarea_id,depende_de_tarea_id,tipo,obligatoria,creada_por,fecha_creacion) VALUES(?,?,?,'FIN_A_INICIO',?,?,?)",
                (fundacion_id, task_id, depends_on, int(required), user.get("id"), now_iso()),
            )
        self.recalculate_priorities(fundacion_id)
        self.audit(fundacion_id, user, "AGREGAR_DEPENDENCIA", "mgp_dependencias", task_id, {"depende_de": depends_on})
        return self.list_dependencies(fundacion_id, task_id)

    def list_dependencies(self, fundacion_id: int, task_id: int) -> list[dict[str, Any]]:
        with self.connect() as conn:
            rows = conn.execute(
                """
                SELECT d.*,t.titulo,t.estado,t.fecha_limite FROM mgp_dependencias d
                JOIN mgp_tareas t ON t.id=d.depende_de_tarea_id AND t.fundacion_id=d.fundacion_id
                WHERE d.fundacion_id=? AND d.tarea_id=? ORDER BY d.id
                """,
                (fundacion_id, task_id),
            ).fetchall()
            return [dict(row) for row in rows]

    def reminders(self, fundacion_id: int, user: dict[str, Any] | None = None, unread_only: bool = False) -> list[dict[str, Any]]:
        where = ["fundacion_id=?"]
        params: list[Any] = [fundacion_id]
        if unread_only:
            where.append("leido=0")
        if user and normalize_text(user.get("rol")) not in {"SUPERADMIN", "GERENTE", "COORDINADOR", "AUXILIAR_ADMINISTRATIVO"}:
            where.append("(destinatario_id IS NULL OR destinatario_id=?)")
            params.append(int(user.get("id") or 0))
        with self.connect() as conn:
            rows = conn.execute(f"SELECT * FROM mgp_recordatorios WHERE {' AND '.join(where)} ORDER BY leido ASC,fecha_programada,nivel DESC,id DESC LIMIT 1000", params).fetchall()
            return [dict(row) for row in rows]

    def mark_reminder_read(self, fundacion_id: int, reminder_id: int, user: dict[str, Any]) -> bool:
        with self.connect() as conn:
            cur = conn.execute("UPDATE mgp_recordatorios SET leido=1,estado='LEIDO',fecha_lectura=?,actualizado_en=? WHERE id=? AND fundacion_id=?", (now_iso(), now_iso(), reminder_id, fundacion_id))
        self.audit(fundacion_id, user, "LEER_RECORDATORIO", "mgp_recordatorios", reminder_id)
        return bool(cur.rowcount)

    def dashboard(self, fundacion_id: int, period: str | None = None, expediente_id: int | None = None, user: dict[str, Any] | None = None) -> dict[str, Any]:
        period_value = safe_period(period)
        filters: dict[str, Any] = {"periodo": period_value}
        if expediente_id:
            filters["expediente_id"] = expediente_id
        tasks = self.list_tasks(fundacion_id, filters, limit=5000)
        if user and normalize_text(user.get("rol")) not in {"SUPERADMIN", "GERENTE", "COORDINADOR", "AUXILIAR_ADMINISTRATIVO"}:
            user_id = int(user.get("id") or 0)
            assigned = [task for task in tasks if not task.get("responsable_id") or int(task.get("responsable_id") or 0) == user_id]
            tasks = assigned
        priorities = Counter(task.get("prioridad") or "MEDIA" for task in tasks)
        statuses = Counter(task.get("estado") or "PENDIENTE" for task in tasks)
        sources = Counter(task.get("fuente_modulo") or "OTRO" for task in tasks)
        overdue = sum(1 for task in tasks if task.get("vencida") and not is_complete(task.get("estado")))
        blocked = sum(1 for task in tasks if task.get("bloqueada"))
        evidence_missing = sum(1 for task in tasks if int(task.get("requiere_evidencia") or 0) and int(task.get("evidencias_total") or 0) <= 0 and not is_complete(task.get("estado")))
        completion = compliance_percentage(tasks)
        return {
            "periodo": period_value,
            "resumen": {
                "total": len(tasks), "completas": sum(1 for task in tasks if is_complete(task.get("estado"))),
                "pendientes": sum(1 for task in tasks if not is_complete(task.get("estado"))), "vencidas": overdue,
                "bloqueadas": blocked, "sin_evidencia": evidence_missing, "cumplimiento": completion,
                "semaforo": "VERDE" if completion >= 85 and not overdue else ("AMARILLO" if completion >= 60 and overdue <= 2 else "ROJO"),
            },
            "prioridades": dict(priorities), "estados": dict(statuses), "fuentes": dict(sources),
            "criticas": [task for task in tasks if task.get("prioridad") == "CRITICA"][:20],
            "proximas": sorted([task for task in tasks if task.get("dias_restantes") is not None and int(task["dias_restantes"]) >= 0 and not is_complete(task.get("estado"))], key=lambda item: item["dias_restantes"])[:20],
            "recordatorios_no_leidos": len(self.reminders(fundacion_id, user=user, unread_only=True)),
        }

    def _product_dir(self, fundacion_id: int, period: str) -> Path:
        path = self.data_dir / "tenants" / str(fundacion_id) / "archivos_actualizados" / "motor_gestion" / period
        path.mkdir(parents=True, exist_ok=True)
        return path

    @staticmethod
    def _rows_for_export(tasks: list[dict[str, Any]]) -> list[list[Any]]:
        rows = [["ID", "UCA", "Componente", "Título", "Fuente", "Estado", "Prioridad", "Fecha límite", "Responsable", "Evidencias", "Bloqueada"]]
        for item in tasks:
            rows.append([
                item.get("id"), item.get("unidad_nombre") or "", item.get("componente") or "", item.get("titulo") or "",
                item.get("fuente_modulo") or "", item.get("estado") or "", item.get("prioridad") or "", item.get("fecha_limite") or "",
                item.get("responsable_nombre") or "", item.get("evidencias_total") or 0, "SI" if item.get("bloqueada") else "NO",
            ])
        return rows

    def _create_excel(self, path: Path, tasks: list[dict[str, Any]], summary: dict[str, Any], library_docs: list[dict[str, Any]]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Seguimiento"
        rows = self._rows_for_export(tasks)
        for row in rows:
            ws.append(row)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        widths = [8, 24, 25, 45, 22, 20, 14, 14, 24, 12, 12]
        for index, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + index)].width = width
        ws.freeze_panes = "A2"
        rs = wb.create_sheet("Resumen")
        for key, value in summary.items():
            rs.append([key, value])
        rs.column_dimensions["A"].width = 30
        rs.column_dimensions["B"].width = 22
        bs = wb.create_sheet("Biblioteca aplicable")
        bs.append(["Código", "Documento", "Componente", "Versión vigente", "Estado"])
        for doc in library_docs:
            version = doc.get("version_vigente") or {}
            bs.append([doc.get("codigo"), doc.get("nombre"), doc.get("componente"), version.get("version"), version.get("estado")])
        for cell in bs[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
        wb.save(path)

    def _create_pdf(self, path: Path, period: str, tasks: list[dict[str, Any]], summary: dict[str, Any], library_docs: list[dict[str, Any]]) -> None:
        styles = getSampleStyleSheet()
        story = [Paragraph("Borrador de informe mensual — Motor de Gestión", styles["Title"]), Spacer(1, 0.3 * cm)]
        story.append(Paragraph(f"Periodo: {period}. Documento generado automáticamente con información registrada; requiere revisión y aprobación humana.", styles["BodyText"]))
        story.append(Spacer(1, 0.25 * cm))
        summary_rows = [["Indicador", "Valor"]] + [[str(key), str(value)] for key, value in summary.items()]
        table = Table(summary_rows, colWidths=[8 * cm, 5 * cm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"), ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.extend([table, Spacer(1, 0.4 * cm), Paragraph("Tareas críticas y pendientes", styles["Heading2"])])
        selected = [task for task in tasks if not is_complete(task.get("estado"))][:60]
        task_rows = [["UCA", "Tarea", "Estado", "Prioridad", "Fecha"]]
        for item in selected:
            task_rows.append([str(item.get("unidad_nombre") or "")[:22], str(item.get("titulo") or "")[:58], item.get("estado"), item.get("prioridad"), item.get("fecha_limite") or ""])
        task_table = Table(task_rows, repeatRows=1, colWidths=[3.5 * cm, 11.5 * cm, 3.2 * cm, 2.5 * cm, 2.5 * cm])
        task_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5), ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(task_table)
        story.extend([PageBreak(), Paragraph("Documentos oficiales relacionados", styles["Heading2"])])
        if library_docs:
            for doc in library_docs[:80]:
                version = doc.get("version_vigente") or {}
                story.append(Paragraph(f"• {doc.get('codigo')} — {doc.get('nombre')} (versión {version.get('version') or 'sin versión vigente'})", styles["BodyText"]))
        else:
            story.append(Paragraph("No se identificaron documentos vigentes relacionados.", styles["BodyText"]))
        document = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=1 * cm, leftMargin=1 * cm, topMargin=1 * cm, bottomMargin=1 * cm)
        document.build(story)

    def _library_documents(self, conn: sqlite3.Connection, fundacion_id: int) -> list[dict[str, Any]]:
        if not self._table_exists(conn, "biblioteca_icbf_documentos"):
            return []
        docs = conn.execute("SELECT * FROM biblioteca_icbf_documentos WHERE fundacion_id=? AND COALESCE(activo,1)=1 ORDER BY componente,nombre", (fundacion_id,)).fetchall()
        result = []
        for raw in docs:
            doc = dict(raw)
            version = None
            if self._table_exists(conn, "biblioteca_icbf_versiones"):
                row = conn.execute("SELECT * FROM biblioteca_icbf_versiones WHERE fundacion_id=? AND documento_id=? AND estado='VIGENTE' ORDER BY id DESC LIMIT 1", (fundacion_id, doc["id"])).fetchone()
                version = dict(row) if row else None
            doc["version_vigente"] = version
            result.append(doc)
        return result

    def prepare_monthly_products(self, fundacion_id: int, period: str | None, expediente_id: int | None, user: dict[str, Any]) -> dict[str, Any]:
        period_value = safe_period(period)
        filters: dict[str, Any] = {"periodo": period_value}
        if expediente_id:
            filters["expediente_id"] = expediente_id
        tasks = self.list_tasks(fundacion_id, filters, limit=5000)
        dashboard = self.dashboard(fundacion_id, period_value, expediente_id, user=None)
        summary = dashboard["resumen"]
        with self.connect() as conn:
            library_docs = self._library_documents(conn, fundacion_id)
        folder = self._product_dir(fundacion_id, period_value)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        scope = f"UCA_{expediente_id}" if expediente_id else "FUNDACION"
        base = f"BORRADOR_GESTION_{scope}_{period_value}_{stamp}"
        xlsx = folder / f"{base}.xlsx"
        pdf = folder / f"{base}.pdf"
        package = folder / f"{base}.zip"
        self._create_excel(xlsx, tasks, summary, library_docs)
        self._create_pdf(pdf, period_value, tasks, summary, library_docs)

        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer)
        for row in self._rows_for_export(tasks):
            writer.writerow(row)
        with zipfile.ZipFile(package, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.write(xlsx, xlsx.name)
            archive.write(pdf, pdf.name)
            archive.writestr("01_TAREAS.csv", csv_buffer.getvalue())
            archive.writestr("02_RESUMEN.json", json.dumps(dashboard, ensure_ascii=False, indent=2, default=str))
            archive.writestr("03_BIBLIOTECA_APLICABLE.json", json.dumps(library_docs, ensure_ascii=False, indent=2, default=str))
            archive.writestr("LEEME.txt", "Producto generado automáticamente con datos registrados y validados. Es un BORRADOR y requiere revisión y aprobación humana.\n")

        records = []
        now = now_iso()
        with self.connect() as conn:
            for product_type, path, mime in (
                ("MATRIZ_EXCEL", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ("INFORME_PDF", pdf, "application/pdf"),
                ("PAQUETE_MENSUAL", package, "application/zip"),
            ):
                digest = file_sha256(path)
                cur = conn.execute(
                    """
                    INSERT INTO mgp_productos
                    (fundacion_id,expediente_id,periodo,tipo_producto,nombre_archivo,ruta_archivo,mime_type,tamano_bytes,sha256,estado,
                     requiere_revision,resumen_json,generado_por,fecha_generacion)
                    VALUES(?,?,?,?,?,?,?,?,?,'BORRADOR',1,?,?,?)
                    """,
                    (fundacion_id, expediente_id, period_value, product_type, path.name, str(path), mime, path.stat().st_size, digest, json_dump(summary), user.get("id"), now),
                )
                records.append({"id": cur.lastrowid, "tipo_producto": product_type, "nombre_archivo": path.name, "tamano_bytes": path.stat().st_size, "sha256": digest, "estado": "BORRADOR"})
            conn.commit()
        closure = self.prepare_closure(fundacion_id, period_value, expediente_id, user)
        self.audit(fundacion_id, user, "GENERAR_PRODUCTOS_MENSUALES", "mgp_productos", None, {"periodo": period_value, "productos": len(records), "expediente_id": expediente_id})
        return {"periodo": period_value, "productos": records, "cierre": closure, "advertencia": "Productos en estado BORRADOR; requieren revisión y aprobación humana."}

    def list_products(self, fundacion_id: int, period: str | None = None, expediente_id: int | None = None) -> list[dict[str, Any]]:
        where = ["fundacion_id=?"]
        params: list[Any] = [fundacion_id]
        if period:
            where.append("periodo=?")
            params.append(safe_period(period))
        if expediente_id:
            where.append("expediente_id=?")
            params.append(expediente_id)
        with self.connect() as conn:
            rows = conn.execute(f"SELECT * FROM mgp_productos WHERE {' AND '.join(where)} ORDER BY fecha_generacion DESC,id DESC", params).fetchall()
            return [dict(row) for row in rows]

    def product_path(self, fundacion_id: int, product_id: int) -> tuple[Path, str, str] | None:
        with self.connect() as conn:
            row = conn.execute("SELECT * FROM mgp_productos WHERE id=? AND fundacion_id=?", (product_id, fundacion_id)).fetchone()
        if not row:
            return None
        path = Path(row["ruta_archivo"]).resolve()
        root = (self.data_dir / "tenants" / str(fundacion_id)).resolve()
        try:
            path.relative_to(root)
        except ValueError:
            return None
        if not path.is_file() or file_sha256(path) != row["sha256"]:
            return None
        return path, str(row["nombre_archivo"]), str(row["mime_type"] or "application/octet-stream")

    def review_product(self, fundacion_id: int, product_id: int, action: str, user: dict[str, Any]) -> dict[str, Any]:
        normalized = normalize_text(action)
        if normalized not in {"REVISAR", "APROBAR", "DEVOLVER"}:
            raise ValueError("Acción no permitida.")
        now = now_iso()
        with self.connect() as conn:
            row = conn.execute("SELECT * FROM mgp_productos WHERE id=? AND fundacion_id=?", (product_id, fundacion_id)).fetchone()
            if not row:
                raise LookupError("Producto no encontrado.")
            if normalized == "REVISAR":
                conn.execute("UPDATE mgp_productos SET estado='PENDIENTE_APROBACION',revisado_por=?,fecha_revision=? WHERE id=? AND fundacion_id=?", (user.get("id"), now, product_id, fundacion_id))
            elif normalized == "APROBAR":
                conn.execute("UPDATE mgp_productos SET estado='APROBADO',revisado_por=COALESCE(revisado_por,?),aprobado_por=?,fecha_revision=COALESCE(fecha_revision,?),fecha_aprobacion=? WHERE id=? AND fundacion_id=?", (user.get("id"), user.get("id"), now, now, product_id, fundacion_id))
            else:
                conn.execute("UPDATE mgp_productos SET estado='DEVUELTO',revisado_por=?,fecha_revision=? WHERE id=? AND fundacion_id=?", (user.get("id"), now, product_id, fundacion_id))
            conn.commit()
            updated = conn.execute("SELECT * FROM mgp_productos WHERE id=?", (product_id,)).fetchone()
        self.audit(fundacion_id, user, f"{normalized}_PRODUCTO", "mgp_productos", product_id)
        return dict(updated)

    def prepare_closure(self, fundacion_id: int, period: str, expediente_id: int | None, user: dict[str, Any]) -> dict[str, Any]:
        tasks = self.list_tasks(fundacion_id, {"periodo": period, **({"expediente_id": expediente_id} if expediente_id else {})}, limit=5000)
        total = len(tasks)
        complete = sum(1 for task in tasks if is_complete(task.get("estado")))
        overdue = sum(1 for task in tasks if task.get("vencida") and not is_complete(task.get("estado")))
        pending = total - complete
        completion = compliance_percentage(tasks)
        product_total = len(self.list_products(fundacion_id, period, expediente_id))
        now = now_iso()
        with self.connect() as conn:
            conn.execute(
                """
                INSERT INTO mgp_cierres_mensuales
                (fundacion_id,expediente_id,periodo,estado,tareas_total,tareas_completas,tareas_pendientes,tareas_vencidas,alertas_total,productos_total,
                 porcentaje_cumplimiento,preparado_por,fecha_preparacion)
                VALUES(?,?,?,'BORRADOR',?,?,?,?,?,?,?, ?,?)
                ON CONFLICT(fundacion_id,expediente_id,periodo) DO UPDATE SET
                  tareas_total=excluded.tareas_total,tareas_completas=excluded.tareas_completas,tareas_pendientes=excluded.tareas_pendientes,
                  tareas_vencidas=excluded.tareas_vencidas,alertas_total=excluded.alertas_total,productos_total=excluded.productos_total,
                  porcentaje_cumplimiento=excluded.porcentaje_cumplimiento,preparado_por=excluded.preparado_por,fecha_preparacion=excluded.fecha_preparacion
                """,
                (fundacion_id, expediente_id, period, total, complete, pending, overdue, overdue, product_total, completion, user.get("id"), now),
            )
            conn.commit()
            row = conn.execute("SELECT * FROM mgp_cierres_mensuales WHERE fundacion_id=? AND expediente_id IS ? AND periodo=?", (fundacion_id, expediente_id, period)).fetchone()
        return dict(row)

    def list_closures(self, fundacion_id: int, period: str | None = None) -> list[dict[str, Any]]:
        where = ["fundacion_id=?"]
        params: list[Any] = [fundacion_id]
        if period:
            where.append("periodo=?")
            params.append(safe_period(period))
        with self.connect() as conn:
            rows = conn.execute(f"SELECT * FROM mgp_cierres_mensuales WHERE {' AND '.join(where)} ORDER BY periodo DESC,id DESC", params).fetchall()
            return [dict(row) for row in rows]

    def review_closure(self, fundacion_id: int, closure_id: int, action: str, user: dict[str, Any], observations: str | None = None) -> dict[str, Any]:
        normalized = normalize_text(action)
        if normalized not in {"REVISAR", "APROBAR", "CERRAR", "DEVOLVER"}:
            raise ValueError("Acción de cierre no permitida.")
        now = now_iso()
        with self.connect() as conn:
            row = conn.execute("SELECT * FROM mgp_cierres_mensuales WHERE id=? AND fundacion_id=?", (closure_id, fundacion_id)).fetchone()
            if not row:
                raise LookupError("Cierre mensual no encontrado.")
            if normalized in {"APROBAR", "CERRAR"} and int(row["tareas_vencidas"] or 0) > 0:
                raise ValueError("No se puede aprobar o cerrar mientras existan tareas vencidas.")
            status = {"REVISAR": "PENDIENTE_APROBACION", "APROBAR": "APROBADO", "CERRAR": "CERRADO", "DEVOLVER": "DEVUELTO"}[normalized]
            conn.execute(
                """
                UPDATE mgp_cierres_mensuales SET estado=?,observaciones=?,revisado_por=?,aprobado_por=?,fecha_revision=?,fecha_aprobacion=?,fecha_cierre=?
                WHERE id=? AND fundacion_id=?
                """,
                (
                    status, str(observations or row["observaciones"] or "")[:4000] or None,
                    user.get("id") if normalized in {"REVISAR", "DEVOLVER", "APROBAR", "CERRAR"} else row["revisado_por"],
                    user.get("id") if normalized in {"APROBAR", "CERRAR"} else row["aprobado_por"],
                    now if normalized in {"REVISAR", "DEVOLVER", "APROBAR", "CERRAR"} else row["fecha_revision"],
                    now if normalized in {"APROBAR", "CERRAR"} else row["fecha_aprobacion"],
                    now if normalized == "CERRAR" else row["fecha_cierre"], closure_id, fundacion_id,
                ),
            )
            conn.commit()
            updated = conn.execute("SELECT * FROM mgp_cierres_mensuales WHERE id=?", (closure_id,)).fetchone()
        self.audit(fundacion_id, user, f"{normalized}_CIERRE_MENSUAL", "mgp_cierres_mensuales", closure_id)
        return dict(updated)

    def rules(self, fundacion_id: int) -> list[dict[str, Any]]:
        with self.connect() as conn:
            self._seed_rules(conn, fundacion_id)
            conn.commit()
            rows = conn.execute("SELECT * FROM mgp_reglas WHERE fundacion_id=? ORDER BY orden,codigo", (fundacion_id,)).fetchall()
            result = []
            for row in rows:
                item = dict(row)
                item["condicion"] = parse_json(item.pop("condicion_json", None), {})
                item["accion"] = parse_json(item.pop("accion_json", None), {})
                result.append(item)
            return result
