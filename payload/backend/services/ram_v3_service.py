"""Generador seguro para el formato oficial RAM Versión 3.

La plantilla fuente nunca se modifica. El servicio trabaja sobre una copia,
conserva la hoja de instrucciones y escribe únicamente valores en celdas de
datos del formato oficial F27.MT1.PP Versión 3.
"""
from __future__ import annotations

import calendar
import hashlib
import json
import re
import unicodedata
from copy import copy, deepcopy
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

RAM_SHEET = "FORMATO RAM"
INSTRUCTIONS_SHEET = "INSTRUCCIONES DILIGENCIAMIENTO"
RAM_CODE = "F27.MT1.PP"
RAM_VERSION = "3"
DATA_ROWS = tuple(range(15, 35))
ATTENDANCE_COLUMNS = tuple(range(10, 35))  # J:AH, cinco semanas x lunes-viernes.
ALLOWED_DOCUMENT_TYPES = {"RC", "TI", "CC", "CE", "PA", "SD"}
ALLOWED_WITHDRAWAL_CODES = {"D", "V", "T", "S", "I", "M", "O"}
WEEKDAY_NAMES = {0: "lunes", 1: "martes", 2: "miercoles", 3: "jueves", 4: "viernes"}
MONTHS_ES = {1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"}
NON_SERVICE_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")


class RamV3Error(ValueError):
    """Error verificable de la plantilla o de los datos RAM V3."""


def _norm(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(re.sub(r"[^a-z0-9]+", " ", text).split())


def _value(user: dict[str, Any], *keys: str, default: Any = "") -> Any:
    for key in keys:
        value = user.get(key)
        if value not in (None, ""):
            return value
    return default


def _text(user: dict[str, Any], *keys: str) -> str:
    return str(_value(user, *keys, default="") or "").strip()


def sha256_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    raw = raw.split("T", 1)[0].strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def age_at_period_start(user: dict[str, Any], year: int, month: int) -> tuple[int | str, int | str, str | None]:
    """Edad cumplida al primer día del mes, según la instrucción oficial."""
    dob = parse_date(_value(user, "FechaNacimiento", "fecha_nacimiento", "fechaNacimiento"))
    if dob:
        cutoff = date(year, month, 1)
        if dob > cutoff:
            return "", "", "La fecha de nacimiento es posterior al mes reportado."
        months = (cutoff.year - dob.year) * 12 + cutoff.month - dob.month
        if cutoff.day < dob.day:
            months -= 1
        months = max(0, months)
        return months // 12, months % 12, None
    stored = _value(user, "EdadMeses", "edad_meses", "edadMeses", default=None)
    if stored not in (None, ""):
        try:
            total = max(0, int(float(stored)))
            return total // 12, total % 12, "Se usó la edad almacenada porque no existe fecha de nacimiento verificable."
        except (TypeError, ValueError):
            pass
    return "", "", "Falta fecha de nacimiento y edad almacenada."


def normalize_document_type(value: Any) -> str:
    raw = _norm(value)
    aliases = {
        "rc": "RC", "registro civil": "RC",
        "ti": "TI", "tarjeta de identidad": "TI",
        "cc": "CC", "c c": "CC", "cedula de ciudadania": "CC",
        "ce": "CE", "cedula de extranjeria": "CE",
        "pa": "PA", "pasaporte": "PA",
        "sd": "SD", "sin documento": "SD",
    }
    return aliases.get(raw, str(value or "").strip().upper())


def participant_document_number(user: dict[str, Any]) -> tuple[str, str | None]:
    """Obtiene el número sin confundirlo con el tipo documental.

    Algunas bases históricas de la plataforma contienen textos como
    ``REGISTRO CIVIL`` en la columna ``documento``. Esos valores no se escriben
    en el RAM; se busca primero un campo específico de número y, si no existe,
    se deja vacío con una advertencia verificable.
    """
    candidates = (
        "NumeroDocumento", "NúmeroDocumento", "numero_documento",
        "DocumentoBeneficiario", "documento_beneficiario",
        "NUI", "nui", "NUIP", "nuip", "Documento", "documento",
    )
    rejected: list[str] = []
    for key in candidates:
        raw = _text(user, key)
        if not raw:
            continue
        normalized = _norm(raw)
        as_type = normalize_document_type(raw)
        if as_type in ALLOWED_DOCUMENT_TYPES and normalized in {
            "rc", "registro civil", "ti", "tarjeta de identidad",
            "cc", "c c", "cedula de ciudadania", "ce",
            "cedula de extranjeria", "pa", "pasaporte", "sd",
            "sin documento",
        }:
            rejected.append(raw)
            continue
        return raw, None
    if rejected:
        return "", "El valor disponible corresponde al tipo de documento, no al número."
    return "", "Falta el número de documento del participante."


def normalize_withdrawal_code(user: dict[str, Any]) -> str:
    raw = _text(user, "CodigoRetiro", "codigo_retiro", "CausaRetiro", "causa_retiro", "MotivoRetiro", "motivo_retiro").upper()
    if raw in ALLOWED_WITHDRAWAL_CODES:
        return raw
    normalized = _norm(raw)
    aliases = {
        "cambio de domicilio": "D",
        "retiro voluntario": "V",
        "transicion a educacion formal": "T",
        "cambio de servicio": "S",
        "inasistencia reiterada": "I",
        "muerte": "M",
        "otro": "O",
    }
    return aliases.get(normalized, "")


def participant_names(user: dict[str, Any]) -> tuple[str, str, str, str]:
    first = _text(user, "PrimerNombre", "primer_nombre")
    second = _text(user, "SegundoNombre", "segundo_nombre")
    surname1 = _text(user, "PrimerApellido", "primer_apellido")
    surname2 = _text(user, "SegundoApellido", "segundo_apellido")

    if not first:
        names = _text(user, "Nombres", "nombres")
        if names:
            parts = names.split()
            first = parts[0]
            second = " ".join(parts[1:])
    if not surname1:
        surnames = _text(user, "Apellidos", "apellidos")
        if surnames:
            parts = surnames.split()
            surname1 = parts[0]
            surname2 = " ".join(parts[1:])
    if not any((first, second, surname1, surname2)):
        full = _text(user, "Nombre", "nombre")
        if full:
            parts = full.split()
            first = parts[0]
            if len(parts) == 2:
                surname1 = parts[1]
            elif len(parts) == 3:
                second, surname1 = parts[1], parts[2]
            elif len(parts) >= 4:
                second = " ".join(parts[1:-2])
                surname1, surname2 = parts[-2], parts[-1]
    return first, second, surname1, surname2


def participant_key(user: dict[str, Any]) -> str:
    internal = _text(user, "Id", "id", "BeneficiarioId", "beneficiario_id", "UsuarioId", "usuario_id")
    if internal:
        return f"id:{internal}"
    document, _ = participant_document_number(user)
    if document:
        return f"doc:{_norm(document)}"
    return "name:" + _norm(" ".join(participant_names(user)))


def deduplicate_users(users: Iterable[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    result: list[dict[str, Any]] = []
    warnings: list[str] = []
    seen: set[str] = set()
    for user in users or []:
        key = participant_key(user)
        if key in seen:
            warnings.append(f"Participante repetido omitido: {key}")
            continue
        seen.add(key)
        result.append(dict(user))
    return result, warnings


def attendance_cell_for_date(day: date) -> tuple[int, int] | None:
    """Devuelve (semana 1..5, columna J..AH) para lunes-viernes."""
    if day.weekday() > 4:
        return None
    occurrence = ((day.day - 1) // 7) + 1
    column = 10 + (occurrence - 1) * 5 + day.weekday()
    if column not in ATTENDANCE_COLUMNS:
        return None
    return occurrence, column


def _copy_sheet_visuals(source, target) -> None:
    """Completa elementos que openpyxl.copy_worksheet no duplica."""
    target.sheet_format = copy(source.sheet_format)
    target.sheet_properties = copy(source.sheet_properties)
    target.page_margins = copy(source.page_margins)
    target.page_setup = copy(source.page_setup)
    target.print_options = copy(source.print_options)
    target.freeze_panes = source.freeze_panes
    try:
        target.print_area = source.print_area
    except Exception:
        pass
    try:
        target.print_title_rows = source.print_title_rows
        target.print_title_cols = source.print_title_cols
    except Exception:
        pass
    for image in getattr(source, "_images", []):
        try:
            target.add_image(deepcopy(image), deepcopy(image.anchor))
        except Exception:
            try:
                target.add_image(copy(image), copy(image.anchor))
            except Exception:
                pass


def _validate_template(wb) -> None:
    missing = [name for name in (RAM_SHEET, INSTRUCTIONS_SHEET) if name not in wb.sheetnames]
    if missing:
        raise RamV3Error("Faltan hojas oficiales: " + ", ".join(missing))
    ws = wb[RAM_SHEET]
    if str(ws["H1"].value or "").strip() != RAM_CODE:
        raise RamV3Error(f"Código RAM inesperado en H1: {ws['H1'].value!r}")
    if "3" not in str(ws["H2"].value or ""):
        raise RamV3Error(f"La plantilla no identifica la Versión 3 en H2: {ws['H2'].value!r}")
    if ws.max_row < 41 or ws.max_column < 37:
        raise RamV3Error("La estructura RAM V3 está incompleta; se esperaba al menos A1:AK41.")


def _clear_data_rows(ws, page_number: int = 1) -> None:
    for row in DATA_ROWS:
        for col in range(1, 38):
            ws.cell(row=row, column=col).value = ""
        ws.cell(row=row, column=1).value = (page_number - 1) * 20 + row - DATA_ROWS[0] + 1
    for coord in ("AI35", "AJ35", "E36", "E37", "E38"):
        ws[coord].value = ""


def _set_page_header(ws, page: int, total_pages: int) -> None:
    ws["AG4"] = f"Hoja {page} de {total_pages}"
    # X2 pertenece al control documental del formato (página 1 de 2) y no se cambia.


def _fill_header(ws, metadata: dict[str, Any]) -> None:
    def label(prefix: str, value: Any) -> str:
        return f"{prefix}: {str(value or '').strip()}" if str(value or "").strip() else f"{prefix}:"

    values = {
        "A4": label("EAS o PDS", metadata.get("eas_pds") or metadata.get("eas")),
        "A5": label("NIT", re.sub(r"[.\-\s]", "", str(metadata.get("nit") or ""))),
        "D5": label("NÚMERO DE CONTRATO", metadata.get("contrato")),
        "F5": label("Regional", metadata.get("regional")),
        "H5": label("Centro Zonal", metadata.get("centro_zonal")),
        "U5": label("Municipio", metadata.get("municipio")),
        "A6": label("MES", metadata.get("mes_nombre")),
        "D6": label("AÑO", metadata.get("anio")),
        "F6": label("Nombre Agente Educativo (a)", metadata.get("agente_educativo")),
        "I6": label("CC", metadata.get("documento_agente")),
        "A7": label("MODALIDAD DE ATENCIÓN", metadata.get("modalidad")),
        "F7": label("Código CUENTAME UDS", metadata.get("codigo_uds")),
        "K7": label("Nombre Unidad de Servicio/ Unidad de Atención", metadata.get("unidad")),
        "A8": label("SERVICIO DE ATENCIÓN", metadata.get("servicio_atencion")),
        "F8": label("Dirección UDS", metadata.get("direccion_uds")),
        "T8": label("Teléfono UDS", metadata.get("telefono_uds")),
    }
    for coord, value in values.items():
        ws[coord].value = value


def _date_in_month(value: Any, year: int, month: int) -> date | None:
    parsed = parse_date(value)
    if parsed and parsed.year == year and parsed.month == month:
        return parsed
    return parsed


def _mapping_from_user(user: dict[str, Any], *keys: str) -> dict[str, Any]:
    raw = _value(user, *keys, default={})
    if isinstance(raw, dict):
        return {str(k).strip(): v for k, v in raw.items()}
    if isinstance(raw, str) and raw.strip():
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                return {str(k).strip(): v for k, v in parsed.items()}
        except Exception:
            return {}
    return {}


def _attendance_records(user: dict[str, Any]) -> dict[str, str]:
    raw = _mapping_from_user(user, "Asistencias", "asistencias", "ControlAsistencia", "control_asistencia")
    result: dict[str, str] = {}
    for key, value in raw.items():
        parsed = parse_date(key)
        if not parsed:
            continue
        mark = str(value or "").strip().upper()
        result[parsed.isoformat()] = mark
    return result


def _novelty_records(user: dict[str, Any]) -> dict[str, Any]:
    raw = _mapping_from_user(user, "Novedades", "novedades", "RegistroNovedades", "registro_novedades")
    result: dict[str, Any] = {}
    for key, value in raw.items():
        parsed = parse_date(key)
        if parsed:
            result[parsed.isoformat()] = value
    return result


def _fill_user_row(
    ws,
    row: int,
    user: dict[str, Any],
    year: int,
    month: int,
    attendance_weekdays: set[str],
    non_service_dates: set[str],
    warnings: list[str],
) -> tuple[int, int]:
    first, second, surname1, surname2 = participant_names(user)
    doc_type = normalize_document_type(_value(user, "TipoDocumento", "tipo_documento"))
    document, document_warning = participant_document_number(user)
    if doc_type not in ALLOWED_DOCUMENT_TYPES:
        warnings.append(f"Fila {row}: tipo de documento no reconocido ({doc_type or 'vacío'}).")
    if document_warning:
        warnings.append(f"Fila {row}: {document_warning}")
    if doc_type == "SD" and not document:
        warnings.append(f"Fila {row}: participante SD sin número CUÉNTAME.")

    years, months, age_warning = age_at_period_start(user, year, month)
    if age_warning:
        warnings.append(f"Fila {row}: {age_warning}")

    ws.cell(row=row, column=2).value = doc_type
    ws.cell(row=row, column=3).value = document
    ws.cell(row=row, column=4).value = first
    ws.cell(row=row, column=5).value = second
    ws.cell(row=row, column=6).value = surname1
    ws.cell(row=row, column=7).value = surname2
    ws.cell(row=row, column=8).value = years
    ws.cell(row=row, column=9).value = months

    entry = parse_date(_value(user, "FechaIngreso", "fecha_ingreso", "fechaIngreso"))
    withdrawal = parse_date(_value(user, "FechaRetiro", "fecha_retiro", "fechaRetiro"))
    total_a = 0
    total_i = 0
    attendance_records = _attendance_records(user)
    novelty_records = _novelty_records(user)

    for day_number in range(1, calendar.monthrange(year, month)[1] + 1):
        current = date(year, month, day_number)
        position = attendance_cell_for_date(current)
        if not position:
            continue
        _, col = position
        iso = current.isoformat()
        cell = ws.cell(row=row, column=col)
        if iso in non_service_dates:
            cell.value = ""
            continue
        if entry and current < entry:
            cell.value = ""
            continue
        if withdrawal and current >= withdrawal:
            cell.value = "/"
            continue
        explicit_mark = attendance_records.get(iso, "")
        if explicit_mark:
            if explicit_mark not in {"A", "I", "/"}:
                warnings.append(f"Fila {row}, {iso}: marca de asistencia no permitida ({explicit_mark}).")
                cell.value = ""
            else:
                cell.value = explicit_mark
                if explicit_mark == "A":
                    total_a += 1
                elif explicit_mark == "I":
                    total_i += 1
                    if not str(novelty_records.get(iso) or "").strip():
                        warnings.append(f"Fila {row}, {iso}: inasistencia sin novedad asociada.")
            continue
        weekday = WEEKDAY_NAMES[current.weekday()]
        if weekday in attendance_weekdays:
            # Compatibilidad con el flujo estable: cuando no existe registro diario,
            # se proyectan únicamente los días programados como asistencia. Nunca se
            # inventa una inasistencia.
            cell.value = "A"
            total_a += 1
        else:
            cell.value = ""

    ws.cell(row=row, column=35).value = total_a
    ws.cell(row=row, column=36).value = total_i
    withdrawal_code = normalize_withdrawal_code(user)
    ws.cell(row=row, column=37).value = withdrawal_code
    return total_a, total_i


def _fill_non_service_columns(ws, year: int, month: int, dates: set[str]) -> None:
    for iso in dates:
        parsed = parse_date(iso)
        if not parsed or parsed.year != year or parsed.month != month:
            continue
        position = attendance_cell_for_date(parsed)
        if not position:
            continue
        _, col = position
        for row in range(9, 36):
            ws.cell(row=row, column=col).fill = copy(NON_SERVICE_FILL)


def _is_gestant(user: dict[str, Any]) -> bool:
    text = _norm(_value(user, "TipoBeneficiario", "tipo_beneficiario", "GrupoEdad", "grupo_edad"))
    return "gestante" in text


def _fill_page_totals(ws, page_users: list[dict[str, Any]], attendance_counts: dict[str, int], year: int, month: int) -> None:
    total_a = sum(int(ws.cell(row=row, column=35).value or 0) for row in DATA_ROWS)
    total_i = sum(int(ws.cell(row=row, column=36).value or 0) for row in DATA_ROWS)
    ws["AI35"] = total_a
    ws["AJ35"] = total_i

    under_six = 0
    over_six = 0
    gestants = 0
    for user in page_users:
        key = participant_key(user)
        if attendance_counts.get(key, 0) <= 0:
            continue
        if _is_gestant(user):
            gestants += 1
            continue
        years, months, _ = age_at_period_start(user, year, month)
        if isinstance(years, int) and isinstance(months, int):
            total_months = years * 12 + months
            if total_months < 6:
                under_six += 1
            else:
                over_six += 1
    ws["E36"] = under_six
    ws["E37"] = over_six
    ws["E38"] = gestants


def generate_ram_v3(
    template_path: str | Path,
    output_path: str | Path,
    users: Iterable[dict[str, Any]],
    year: int,
    month: int,
    metadata: dict[str, Any] | None = None,
    attendance_provider: Callable[[dict[str, Any]], set[str]] | None = None,
    non_service_dates: Iterable[str] | None = None,
    expected_sha256: str | None = None,
) -> dict[str, Any]:
    """Genera RAM V3 y retorna diagnóstico de la operación."""
    source = Path(template_path)
    destination = Path(output_path)
    if not source.exists():
        raise FileNotFoundError(f"No existe la plantilla RAM V3: {source}")
    source_hash_before = sha256_file(source)
    if expected_sha256 and source_hash_before.lower() != expected_sha256.lower():
        raise RamV3Error("La plantilla RAM V3 no coincide con el hash oficial registrado.")

    workbook = load_workbook(source, data_only=False, keep_links=True)
    _validate_template(workbook)
    participants, warnings = deduplicate_users(users)
    metadata = dict(metadata or {})
    required_header_fields = {
        "eas_pds": "EAS o PDS",
        "nit": "NIT o consecutivo de Operación Directa",
        "unidad": "Nombre de la UDS",
        "mes_nombre": "Mes",
        "anio": "Año",
    }
    for key, label in required_header_fields.items():
        if not str(metadata.get(key) or metadata.get("eas" if key == "eas_pds" else key) or "").strip():
            warnings.append(f"Encabezado pendiente: {label} no está disponible en la fuente de datos.")
    total_pages = max(1, (len(participants) + 19) // 20)
    base_sheet = workbook[RAM_SHEET]
    instructions = workbook[INSTRUCTIONS_SHEET]

    # Eliminar únicamente páginas RAM adicionales de una generación previa; la fuente
    # oficial no debe traerlas, pero esto mantiene idempotencia si se prueba una copia.
    for sheet in list(workbook.worksheets):
        if sheet is not base_sheet and sheet is not instructions and sheet.title.startswith("FORMATO RAM "):
            workbook.remove(sheet)

    page_sheets = [base_sheet]
    for page_number in range(2, total_pages + 1):
        cloned = workbook.copy_worksheet(base_sheet)
        cloned.title = f"FORMATO RAM {page_number}"
        _copy_sheet_visuals(base_sheet, cloned)
        workbook._sheets.remove(cloned)
        instruction_index = workbook._sheets.index(instructions)
        workbook._sheets.insert(instruction_index, cloned)
        page_sheets.append(cloned)

    metadata.setdefault("anio", year)
    metadata.setdefault("mes_nombre", MONTHS_ES[month])
    non_service = {str(v).strip() for v in (non_service_dates or []) if str(v).strip()}
    attendance_counts: dict[str, int] = {}

    for page_index, ws in enumerate(page_sheets, start=1):
        _clear_data_rows(ws, page_index)
        _fill_header(ws, metadata)
        _set_page_header(ws, page_index, total_pages)
        _fill_non_service_columns(ws, year, month, non_service)
        page_users = participants[(page_index - 1) * 20: page_index * 20]
        for offset, user in enumerate(page_users):
            row = DATA_ROWS[0] + offset
            allowed = set()
            if attendance_provider:
                try:
                    allowed = {_norm(v) for v in (attendance_provider(user) or set())}
                except Exception as exc:
                    warnings.append(f"Fila {row}: no se pudo calcular la programación de asistencia: {exc}")
            total_a, _ = _fill_user_row(ws, row, user, year, month, allowed, non_service, warnings)
            attendance_counts[participant_key(user)] = total_a
        _fill_page_totals(ws, page_users, attendance_counts, year, month)

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    source_hash_after = sha256_file(source)
    if source_hash_after != source_hash_before:
        raise RamV3Error("La plantilla maestra fue modificada durante la generación.")

    # Validación de apertura y estructura del archivo generado.
    check = load_workbook(destination, data_only=False, read_only=False, keep_links=True)
    expected_sheets = [RAM_SHEET] + [f"FORMATO RAM {n}" for n in range(2, total_pages + 1)] + [INSTRUCTIONS_SHEET]
    if check.sheetnames != expected_sheets:
        raise RamV3Error(f"Orden de hojas inesperado: {check.sheetnames}")
    if str(check[RAM_SHEET]["H1"].value or "").strip() != RAM_CODE:
        raise RamV3Error("El archivo generado perdió el código oficial RAM.")
    check.close()

    return {
        "ok": True,
        "archivo": str(destination),
        "total_participantes": len(participants),
        "paginas_ram": total_pages,
        "hojas": expected_sheets,
        "hash_plantilla": source_hash_before,
        "warnings": warnings,
        "non_service_dates": sorted(non_service),
    }


def load_instruction_catalog(path: str | Path) -> dict[str, Any]:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def find_instruction(catalog: dict[str, Any], field: str) -> dict[str, Any] | None:
    target = _norm(field)
    for item in catalog.get("campos", []):
        haystack = {_norm(item.get("id")), _norm(item.get("titulo"))}
        haystack.update(_norm(alias) for alias in item.get("aliases", []))
        if target in haystack or any(target and target in value for value in haystack):
            return item
    return None
